# OpenMASS
# An open source python software for mass photometry analysis including standard landing assays and dynamic tracking.
# Copyright (C) 2024-2025 Maximilian F. K. Wills
# Code is cursed due to rapid prototyping and iteration under time pressure. read at your own risk to sanity.


import _tkinter
import tkinter as tk
from tkinter import ttk
from PIL import Image, ImageTk
from idlelib.tooltip import Hovertip
from copy import deepcopy

loading_win = tk.Tk()
loading_win.iconbitmap("icons/luxmp_logo.ico")
loading_win.attributes('-topmost', True)
loading_win.attributes('-alpha', 0.9)
style = ttk.Style()
style.tk.call('source', 'azure.tcl')
style.tk.call('set_theme', 'dark')
loading_win.geometry('640x300+640+320')
loading_win['bg'] = '#222222'
loading_win.overrideredirect(True)

TITLE = 'OpenMASS'
SUBTITLE = 'Open Mass Analysis Software Suite'
VERSION = 'v2025r'

logo = ImageTk.PhotoImage(Image.open('icons/luxmp_logo.png'))
tk.Label(master=loading_win, image=logo, bg='#222222').place(x=2, y=2)
tk.Label(master=loading_win, text=TITLE + ' ' + VERSION, font='calibri 37 bold', fg='#007fff', bg='#222222').place(x=220, y=30)
tk.Label(master=loading_win, text=SUBTITLE, font='calibri 20', fg='#ED7D31', bg='#222222').place(x=168, y=126)
loading_label = tk.Label(master=loading_win, text='Loading components...', font='calibri 12', fg='#aaaaaa', bg='#222222')
loading_label.place(x=240, y=200)
loading_progress = ttk.Progressbar(master=loading_win, length=600, maximum=100, mode='determinate', orient='horizontal')
loading_progress.place(x=20, y=240)
loading_win.update()

loading_progress.step(6)
loading_progress.update()
import tifffile
loading_progress.step(10)
loading_progress.update()
from py_modules import skew_GMM as sgmm, motion_correction as motion, foci_detection, simple_step_detection as steps, ISCAMS_lib as iscam, openmass_third_party_licences as licences
import easygui
loading_progress.step(5)
loading_progress.update()
import numpy as np
loading_progress.step(12)
loading_progress.update()
import matplotlib as mplib
mplib.use("TkAgg")
import matplotlib.style as mplstyle
from matplotlib.ticker import MultipleLocator
loading_progress.step(8)
loading_progress.update()
import matplotlib.pyplot as plt
loading_progress.step(12)
loading_progress.update()
import cv2
loading_progress.step(5)
loading_progress.update()
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
loading_progress.step(2)
loading_progress.update()
import traceback
loading_progress.step(5)
loading_progress.update()
import sys
import os
import io
from copy import deepcopy
import h5py
loading_progress.step(5)
loading_progress.update()
from sklearn.metrics import r2_score
loading_progress.step(5)
loading_progress.update()
import pickle
import openpyxl
loading_progress.step(4)
loading_progress.update()
loading_progress.step(10)
loading_progress.update()
loading_progress.step(10)
loading_progress.update()

class Root:
    def __init__(self):
        self.window = tk.Toplevel()
        self.window.title(f'OpenMASS Mass Photometry Analysis {VERSION}')
        self.window.protocol("WM_DELETE_WINDOW", self.handle_close)
        self.cwd = os.getcwd()

        # self.window.tk.call("source", "azure.tcl")
        # self.window.tk.call("set_theme", "dark")
        # self.style = ttk.Style()
        # # self.style.tk.call('source', 'azure.tcl')
        # self.style.tk.call('set_theme', 'dark')

        self.window.geometry('1920x1036')
        self.window.state('zoomed')
        # self.window.resizable(False, False)
        self.window['bg'] = '#333333'
        self.window.attributes('-topmost', True)
        self.window.update_idletasks()
        self.window.attributes('-topmost', False)
        self.window.iconbitmap("icons/luxmp_logo.ico")

        self.progress_win = None
        self.drift_win = None
        self.mass_calib_win = None
        self.profile_win = None
        self.preferences_win = None
        self.gauss_opt_win = None
        self.gauss_preview_win = None
        self.ratio_popout_win = None
        self.about_win = None

        self.is_popped_out = False

        # self.style.configure('TButton', background='#333333', foreground='#ff8800', borderwidth=2, focuscolor='none')
        # self.style.configure('TCheckbutton', background='#444444', foreground='#ff8800', borderwidth=1, focuscolor='none')
        # self.style.configure('TMenubutton', background='#444444', foreground='#ff8800', borderwidth=1, focuscolor='none')
        # self.style.map('TButton', background=[('pressed', '#ff6600'), ('active', '#ff8800')], foreground=[('pressed', 'white'), ('active', '#cccccc')])
        # self.style.map('TMenubutton', background=[('pressed', '#ff6600'), ('active', '#ff8800')], foreground=[('pressed', 'white'), ('active', '#cccccc')])
        # self.style.configure('TScale', background='#404040')
        # self.style.configure('TSpinbox', foreground='#cccccc', borderwidth=2, focuscolor='none')

        tk.Frame(master=self.window, bg='#777777').place(x=50, y=60, width=1860, height=1)
        tk.Label(master=self.window, bg='#333333', fg='#cccccc', text='Native').place(x=5, y=48, width=50)
        tk.Label(master=self.window, bg='#333333', fg='#cccccc', text='Ratiometric').place(x=638, y=48, width=72)
        tk.Label(master=self.window, bg='#333333', fg='#cccccc', text='Analysis Z-Projection').place(x=1278, y=48, width=128)
        self.canvas_native = tk.Canvas(master=self.window, width=512, height=136, bg='#222222', highlightthickness=1, highlightbackground='#111111')
        self.canvas_native.place(x=10, y=70)
        hover(self.canvas_native, "Native images acquired by your mass photometer.", hover_delay=1000)
        self.canvas_ratiometric = tk.Canvas(master=self.window, width=512, height=136, bg='#222222', highlightthickness=1, highlightbackground='#111111')
        self.canvas_ratiometric.place(x=640, y=70)
        hover(self.canvas_ratiometric, "Ratiometric stack view. Click on events or tracks to find them in the 'Ratiometric Traces' or 'Tracking' tabs respectively.", hover_delay=1000)
        self.canvas_projection = tk.Canvas(master=self.window, width=512, height=136, bg='#222222', highlightthickness=1, highlightbackground='#111111')
        self.canvas_projection.place(x=1280, y=70)
        hover(self.canvas_projection, "Enhanced Z-projection image used to find and fit events. Detected events will be outlined.", hover_delay=1000)

        self.load_button = ttk.Button(master=self.window, text='Load Data', command=self.load_data)
        self.load_button.place(x=160, y=10, width=160)
        hover(self.load_button, "Load data.\nData must be in uncompressed Refeyn .mp, .h5 (h5py), or .tiff format.", hover_delay=1000)
        self.file_path = None

        self.native_frame_var = tk.IntVar(master=self.window)
        self.native_frame_var.set(0)
        self.slider_native = ttk.Scale(master=self.window, length=512, from_=0, to=0, variable=self.native_frame_var, command=self.display_frame_native)
        self.slider_native.place(x=10, y=216, width=512)
        self.native_frame_label = tk.Label(master=self.window, bg='#333333', fg='#cccccc', text='', anchor=tk.NW)
        self.native_frame_label.place(x=528, y=214, width=50)
        hover(self.slider_native, "Scroll frames. left click in the tray to play frame by frame. right click to jump to frame.", hover_delay=1000)

        self.ratio_frame_var = tk.IntVar(master=self.window)
        self.ratio_frame_var.set(0)
        self.slider_ratio = ttk.Scale(master=self.window, length=512, from_=0, to=0, variable=self.ratio_frame_var, command=self.display_frame_ratio)
        self.slider_ratio.place(x=640, y=216, width=512)
        self.ratio_frame_label = tk.Label(master=self.window, bg='#333333', fg='#cccccc', text='', anchor=tk.NW)
        self.ratio_frame_label.place(x=1160, y=214, width=50)
        hover(self.slider_ratio, "Scroll frames. left click in the tray to play frame by frame. right click to jump to frame.", hover_delay=1000)

        self.ratio_popout_btn = ttk.Button(master=self.window, text='^', command=self.ratio_popout)
        self.ratio_popout_btn.place(x=600, y=210, width=35, height=28)
        hover(self.ratio_popout_btn, "Pop out ratiometric image canvas into a floating window. The window will always stay on top.", hover_delay=1000)

        self.proj_frame_var = tk.IntVar(master=self.window)
        self.proj_frame_var.set(0)
        self.slider_proj = ttk.Scale(master=self.window, length=512, from_=0, to=0, variable=self.proj_frame_var, command=self.display_frame_proj)
        self.slider_proj.place(x=1280, y=216, width=512)
        self.proj_frame_label = tk.Label(master=self.window, bg='#333333', fg='#cccccc', text='', anchor=tk.NW)
        self.proj_frame_label.place(x=1808, y=214, width=100)
        hover(self.slider_proj, "Scroll Z-slices. left click in the tray to play frame by frame. right click to jump to frame.", hover_delay=1000)

        self.process_button = ttk.Button(master=self.window, text='Compute Ratiometric', command=self.process_new)
        self.process_button.place(x=640, y=10, width=160)
        hover(self.process_button, "Process native stack to compute the ratiometric stack. For landing assays uncheck 'Median'; for tracking experiments, "
                                   "check 'Median' and set window size to at least 50", hover_delay=1000)
        self.events_button = ttk.Button(master=self.window, text='Detect Contrast Events', command=self.get_contrasts_and_masses)
        self.events_button.place(x=810, y=10, width=160)
        hover(self.events_button, "Detect events in images for landing assays and tracking.\nIf tracking, Z-stack should be set to 1 and trace offset 0.", hover_delay=1000)
        self.event_label = tk.Label(master=self.window, bg='#333333', fg='#cccccc', text='', anchor=tk.NW)
        self.event_label.place(x=980, y=15, width=280)

        tk.Label(master=self.window, bg='#333333', fg='#cccccc', text='Window length').place(x=340, y=15)
        self.pre_avg = ttk.Spinbox(master=self.window, from_=1, to=500, increment=1)
        self.pre_avg.set(5)
        self.pre_avg.place(x=444, y=10, width=80)
        hover(self.pre_avg, "Half window size for calculating the ratiometric stack. For landing assays the default is 5, which means 2 windows of "
                            "5 frames are independently averaged. The means are divided by each other (second / first in frame order) and the windows move right "
                            "by 1 frame and the process repeats for the entire movie. In median mode, the current frame is divided by the median of all frames "
                            "± the half window size.", hover_delay=1000)

        self.use_median = tk.IntVar(master=self.window)
        self.use_median.set(0)
        self.use_median_cb = ttk.Checkbutton(master=self.window, text='Median', offvalue=0, onvalue=1, variable=self.use_median)
        self.use_median_cb.place(x=540, y=13)
        hover(self.use_median_cb, "Determines the use of median mode. If enabled, ratiometric images are computed using a median window. This should be "
                                  "enabled for tracking experiments with mobile PSFs only.", hover_delay=1000)

        self.show_boxes = tk.IntVar(master=self.window)
        self.show_boxes.set(1)
        self.show_boxes_cb = ttk.Checkbutton(master=self.window, text='Show boxes', offvalue=0, onvalue=1, variable=self.show_boxes,
                                             command=lambda: self.display_frame_ratio(int(float(self.ratio_frame_var.get()))))
        self.show_boxes_cb.place(x=530, y=70)
        hover(self.show_boxes_cb, "Display box overlays around detected events.", hover_delay=1000)

        self.correct_motion_var = tk.IntVar(master=self.window)
        self.correct_motion_var.set(0)
        self.correct_motion_cb = ttk.Checkbutton(master=self.window, text='\nMotion\nCorrection', offvalue=0, onvalue=1, variable=self.correct_motion_var, command=self.motion_enabled)
        self.correct_motion_cb.place(x=530, y=100)
        hover(self.correct_motion_cb, "Motion correction tries to suppress artifacts from vibrational motion by notch filtering the dominant resonance in fourier space. "
                                  "This may cause ringing for large mass objects. Ringing suppression using event kernel apodisation can be enabled in preferences but can "
                                  "cause stretching of event traces as if a larger window size was used. Motion correction will only take effect the next time the ratiometric "
                                  "stack is computed.", hover_delay=1000)

        self.correct_plots_var = tk.IntVar(master=self.window)
        self.correct_plots_var.set(0)
        self.correct_plots_cb = ttk.Checkbutton(master=self.window, text='Plots', offvalue=0, onvalue=1, variable=self.correct_plots_var)
        # self.correct_plots_cb.place(x=530, y=160)

        self.invert_ratiometric = tk.IntVar(master=self.window)
        self.invert_ratiometric.set(0)
        self.invert_ratiometric_cb = ttk.Checkbutton(master=self.window, text='Invert', offvalue=0, onvalue=1, variable=self.invert_ratiometric)
        self.invert_ratiometric_cb.place(x=530, y=160)
        hover(self.invert_ratiometric_cb, "Invert the ratiometric stack if focus was set to incorrect maximum and events have inverted contrast. The "
                                      "inversion only takes effect after the ratiometric stack is computed again.", hover_delay=1000)

        # tk.Label(master=self.window, bg='#333333', fg='#cccccc', text='Post Window').place(x=494, y=15)
        # self.post_avg = ttk.Spinbox(master=self.window, from_=1, to=100, increment=1)
        # self.post_avg.set(8)
        # self.post_avg.place(x=580, y=10, width=50)

        tk.Label(master=self.window, bg='#333333', fg='#cccccc', text='Z-Stack Size').place(x=1280, y=15)
        self.window_size = ttk.Spinbox(master=self.window, from_=1, to=400, increment=1)
        self.window_size.set(50)
        self.window_size.place(x=1360, y=10, width=80)
        hover(self.window_size, "The Z-stack size determines the window size of the Z-max projection when finding events in a landing assay. "
                                      "This value should be set to 1 when tracking to ensure events are detected individually in all frames.", hover_delay=1000)

        self.dynamic_updates = tk.IntVar(master=self.window)
        self.dynamic_updates.set(1)
        self.dynamic_updates_cb = ttk.Checkbutton(master=self.window, variable=self.dynamic_updates, text='Dynamically update graph', onvalue=1, offvalue=0)
        self.dynamic_updates_cb.place(x=1660, y=14)

        tk.Label(master=self.window, bg='#333333', fg='#cccccc', text='Trace Offset').place(x=1460, y=15)
        self.offset = ttk.Spinbox(master=self.window, from_=0, to=20, increment=1)
        self.offset.set(5)
        self.offset.place(x=1542, y=10, width=80)
        hover(self.offset, "Trace offset determines the extra padding around each Z-stack in frames to prevent event duplication of events which overlap "
                           "Z-stack edges.", hover_delay=1000)

        self.internal_window = 40
        self.internal_frame_average = 5

        tk.Label(master=self.window, bg='#333333', fg='#cccccc', text='Contrast Range').place(x=1164, y=65)
        self.ratio_norm = ttk.Spinbox(master=self.window, from_=0.001, to=10, increment=0.001, command=self.update_ratio_contrast)
        self.ratio_norm.set(0.01)
        self.ratio_norm.bind("<Return>", lambda a: self.display_frame_ratio(self.ratio_frame_var.get()))
        self.ratio_norm.place(x=1160, y=90, width=110)
        hover(self.ratio_norm, "Controls image clipping for display. Images will be clipped between this value and its negative counterpart.", hover_delay=1000)

        tk.Label(master=self.window, bg='#333333', fg='#cccccc', text='Label Persistence').place(x=1160, y=140)
        self.persistence = ttk.Spinbox(master=self.window, from_=1, to=50, increment=1)
        self.persistence.set(int(float(self.pre_avg.get())))
        self.persistence.place(x=1160, y=165, width=110)
        hover(self.persistence, "Controls the persistence in frames of rendered event boxes around their minimum contrast points. This value should "
                                "be the same as the ratiometric 'Window length' unless median mode is used in which case it should be 1 to prevent ghosting.", hover_delay=1000)

        self.frame = tk.Frame(master=self.window)
        self.frame.place(x=1152, y=368)
        self.figure = plt.Figure(figsize=(7.6, 6), dpi=100)
        self.figure.set_facecolor("#333333")
        self.figure.subplots_adjust(top=0.97, bottom=0.06, left=0.1, right=0.98)
        self.plotter = None

        self.canvas = FigureCanvasTkAgg(self.figure, master=self.frame)
        self.canvas.draw()
        self.canvas.get_tk_widget().pack(side="top", fill="both", expand=1)
        self.toolbar = NavigationToolbar2Tk(self.canvas, self.frame)
        self.toolbar.update()
        self.canvas.get_tk_widget().pack()
        hover(self.canvas.get_tk_widget(), "Contrast / Mass histogram. Click and drag over peaks with the mouse to fit them.", hover_delay=1000)
        tk.Label(master=self.window, bg='#333333', fg='#cccccc', text='Minimum').place(x=1457, y=978)
        self.min_entry = ttk.Entry(master=self.window)
        self.min_entry.place(x=1522, y=974, width=80, height=30)
        self.min_entry.insert(tk.END, '-0.1')
        tk.Label(master=self.window, bg='#333333', fg='#cccccc', text='Maximum').place(x=1625, y=978)
        self.max_entry = ttk.Entry(master=self.window)
        self.max_entry.place(x=1692, y=974, width=80, height=30)
        self.max_entry.insert(tk.END, '0.1')
        self.min_entry.bind("<Return>", lambda a: self.plot_histogram(None))
        self.max_entry.bind("<Return>", lambda a: self.plot_histogram(None))
        hover(self.min_entry, "Minimum histogram range. This value is automatically converted to mass if mass mode is enabled.", hover_delay=1000)
        hover(self.max_entry, "Maximum histogram range. This value is automatically converted to mass if mass mode is enabled.", hover_delay=1000)

        tk.Frame(master=self.window, bg='#777777').place(x=1152, y=247, width=758, height=1)
        self.use_mass = tk.IntVar(master=self.window)
        self.use_mass.set(0)
        self.mass_checkbox = ttk.Checkbutton(master=self.window, variable=self.use_mass, text='Mass', onvalue=1, offvalue=0, command=self.convert)
        self.mass_checkbox.place(x=1310, y=263)
        hover(self.mass_checkbox, "Toggle mass mode. If mass mode is enabled event or track mass distribution are converted to kDa and bin width / plot ranges "
                              "will be recalculated to masses from ratiometric contrast and the current mas calibration.", hover_delay=1000)

        self.enable_adsorption = tk.IntVar(master=self.window)
        self.enable_adsorption.set(1)
        self.enable_adsorption_cb = ttk.Checkbutton(master=self.window, variable=self.enable_adsorption, text='Adsorption', onvalue=1, offvalue=0,
                                                 command=lambda: self.display_frame_proj(int(float(self.slider_proj.get()))))
        self.enable_adsorption_cb.place(x=1800, y=70)
        hover(self.enable_adsorption_cb, "If checked, will show binding (adsorption) Z-stacks.", hover_delay=1000)

        self.enable_desorption = tk.IntVar(master=self.window)
        self.enable_desorption.set(0)
        self.enable_desorption_cb = ttk.Checkbutton(master=self.window, variable=self.enable_desorption, text='Desorption', onvalue=1, offvalue=0,
                                                    command=lambda: self.display_frame_proj(int(float(self.slider_proj.get()))))
        self.enable_desorption_cb.place(x=1800, y=110)
        hover(self.enable_desorption_cb, "If checked, will show unbinding (desorption) Z-stacks.", hover_delay=1000)

        self.matplotlib_button = ttk.Button(master=self.window, text='MatPlotLib', command=self.matplotlib_display)
        self.matplotlib_button.place(x=1800, y=150, width=110)
        hover(self.matplotlib_button, "Plot Z-stack in external matplotlib canvas.", hover_delay=1000)

        self.calibration = [1, 0]
        self.mass_entry = ttk.Entry(master=self.window)
        self.mass_entry.place(x=1160, y=260, width=140, height=33)
        self.mass_entry.insert(0, 'No Mass Calibration')
        self.mass_entry['state'] = tk.DISABLED
        hover(self.mass_entry, "Gradient and intercept of current mass calibration. Calibrations persist until a new one is created or loaded or when a workspace is loaded. "
                               "Calibrations are saved with the workspace when saving .lmp files (File --> Save Experiment)", hover_delay=1000)

        self.main_bins_mode = tk.IntVar(master=self.window)
        self.main_bins_mode.set(2)
        self.main_bin_count_cb = ttk.Radiobutton(master=self.window, text=' Bin Count', variable=self.main_bins_mode, value=1, command=self.bins_mode)
        self.main_bin_count_cb.place(x=1160, y=318)
        self.main_bin_width_cb = ttk.Radiobutton(master=self.window, text=' Bin Width', variable=self.main_bins_mode, value=2, command=self.bins_mode)
        self.main_bin_width_cb.place(x=1260, y=318)
        hover(self.main_bin_width_cb, "Convert histogram bin calculations to use bin width. Numeric quantity will update with defaults from preferences when "
                                      "switching between contrast and mass mode.", hover_delay=1000)
        hover(self.main_bin_count_cb, "Convert histogram bin calculations to use a fixed bin count. The default can be edited in preferences.", hover_delay=1000)

        self.fit_mode = tk.IntVar(master=self.window)
        self.fit_mode.set(preferences['hist']['fit mode'])
        self.fit_mode1_cb = ttk.Radiobutton(master=self.window, text='Fit multiple peaks', variable=self.fit_mode, value=1)
        self.fit_mode1_cb.place(x=1634, y=318)
        self.fit_mode2_cb = ttk.Radiobutton(master=self.window, text='Fit single peaks', variable=self.fit_mode, value=2)
        self.fit_mode2_cb.place(x=1780, y=318)
        self.fit_mode3_cb = ttk.Radiobutton(master=self.window, text='Advanced preview', variable=self.fit_mode, value=3)
        self.fit_mode3_cb.place(x=1480, y=318)
        hover(self.fit_mode1_cb, "Dragging the mouse over peaks in the histogram will attempt to fit a mixture to multiple peaks.", hover_delay=1000)
        hover(self.fit_mode2_cb, "Dragging the mouse over peaks in the histogram will  fit an individual peak.", hover_delay=1000)
        hover(self.fit_mode3_cb, "Dragging the mouse over peaks in the histogram will open the selected range in the advanced preview window allowing full "
                                 "user control of fitting parameters, Gaussian symmetry, or fitting an exact number of components.", hover_delay=1000)

        # tk.Label(master=self.window, bg='#333333', fg='#cccccc', text='Bins', anchor=tk.NW).place(x=1408, y=320, width=50)
        self.bins = ttk.Spinbox(master=self.window, from_=10, to=5000, increment=10, command=lambda: self.plot_histogram(None))
        self.bins.set(10)
        self.bins.bind("<Return>", lambda a: self.plot_histogram(None))
        self.bins.place(x=1360, y=317, width=90, height=30)
        hover(self.bins, "Number of bins or bin width in contrast or mass (kDa) units depending on histogram mode and bin calculation type.", hover_delay=1000)

        if preferences['hist']['bin type'] == "count":
            self.main_bins_mode.set(1)
            self.bins.set(preferences['hist']['default count'])
        elif preferences['hist']['bin type'] == "width":
            self.main_bins_mode.set(2)
            self.bins.set(preferences['hist']['default contrast'])

        self.replot_button = ttk.Button(master=self.window, text='Replot', command=lambda: self.plot_histogram(None))
        self.replot_button.place(x=1380, y=973, width=62, height=32)
        hover(self.replot_button, "Refresh the histogram plot. This is not normally necessary unless you have updated track mass distribution filters or refitted plateaus which doesn't "
                                  "automatically update the main histogram with the newly filtered contrast or mass distribution.", hover_delay=1000)

        tk.Frame(master=self.window, bg='#777777').place(x=1380, y=256, width=1, height=40)

        self.fit_manually = tk.IntVar(master=self.window)
        self.fit_manually.set(0)
        self.fit_manually_cb = ttk.Checkbutton(master=self.window, text='Manual Gaussian fit', variable=self.fit_manually, onvalue=1, offvalue=0, command=self.update_fit_mode)
        self.fit_manually_cb.place(x=1390, y=263)
        hover(self.fit_manually_cb, "Draw Gaussians manually. Not recommended except qualitatively when peak fitting fails. When enabled, select a component from the dropdown "
                                    "menu and then click and drag over the peak from -2 std. devs to +2 std. devs at the height of the mode of the peak.", hover_delay=1000)

        self.manual_component = tk.StringVar(master=self.window)
        self.field_list = ['Component 1,', 'Component 2,', 'Component 3,', 'Component 4,', 'Component 5,', 'Component 6,', 'Component 7,', 'Component 8,']
        self.manual_fit_selection = ttk.OptionMenu(self.window, self.manual_component, 'select component', *self.field_list, command=self.bind_mouse)
        self.manual_fit_selection['menu'].configure(bg='#555555', relief=tk.RIDGE, fg='white')
        self.manual_fit_selection.place(x=1550, y=260, width=142, height=33)
        hover(self.manual_fit_selection, "When drawing manual Gaussians, select a component first.", hover_delay=1000)

        self.auto_fits = []

        self.number_of_fits = tk.StringVar(master=self.window)
        numbers = ['1', '2', '3', '4', '5', '6', '7', '8']
        self.number_selection = ttk.OptionMenu(self.window, self.number_of_fits, '1', *numbers, command=self.plot_histogram)
        self.number_selection['menu'].configure(bg='#555555', relief=tk.RIDGE, fg='white')
        self.number_selection.place(x=1702, y=260, width=60, height=33)
        hover(self.number_selection, "Maximum number of manual components to display (in order they exist in component drop-down).", hover_delay=1000)

        self.manual_fit_selection['state'] = tk.DISABLED
        self.number_selection['state'] = tk.DISABLED

        self.delete_fits_button = ttk.Button(master=self.window, text='Delete All Fits', command=self.delete_fits)
        self.delete_fits_button.place(x=1772, y=260, width=138)
        hover(self.delete_fits_button, "Delete all fits. GMM Gaussians and manually drawn Gaussians are deleted separately.", hover_delay=1000)

        self.parameter_frame = ttk.LabelFrame(master=self.window, text=" Event Detection Parameters ", padding=(6, 5))
        self.parameter_frame.place(x=640, y=238, width=502, height=310)

        # now set as contrast threshold
        tk.Label(master=self.parameter_frame, bg='#333333', fg='#cccccc', text='Contrast threshold', anchor=tk.NW).place(x=0, y=4, width=190)
        self.detection_threshold = ttk.Spinbox(master=self.parameter_frame, from_=0.001, to=0.010, increment=0.0001)
        self.detection_threshold.set(0.0025)
        self.detection_threshold.place(x=150, y=0, width=75)
        hover(self.detection_threshold, "Threshold for the maximum filter. A candidate event must have at least one pixel where the absolute magnitude is greater or"
                                        "equal to this threshold to be considered.", hover_delay=1000)

        # parameter now represents minimum distance between points but has not been renamed to avoid breaking other code. should be renamed in future
        tk.Label(master=self.parameter_frame, bg='#333333', fg='#cccccc', text='Nearest neighbours', anchor=tk.NW).place(x=0, y=54, width=190)
        self.nearest_neighbours = ttk.Spinbox(master=self.parameter_frame, from_=3, to=10, increment=0.5)
        self.nearest_neighbours.set(4)
        self.nearest_neighbours.place(x=150, y=50, width=75)
        hover(self.nearest_neighbours, "Nearest neighbour threshold. Any two PSFs with centroids within this distance (pixels) from each other will be rejected.", hover_delay=1000)

        tk.Label(master=self.parameter_frame, bg='#333333', fg='#cccccc', text='Minimum PSF width (σ)', anchor=tk.NW).place(x=0, y=104, width=190)
        self.min_sigma = ttk.Spinbox(master=self.parameter_frame, from_=0.5, to=3, increment=0.02)
        self.min_sigma.set(0.5)
        self.min_sigma.place(x=150, y=100, width=75)
        hover(self.min_sigma, "Minimum PSF width threshold. PSFs with narrower standard deviation (pixels) will be rejected.", hover_delay=1000)

        tk.Label(master=self.parameter_frame, bg='#333333', fg='#cccccc', text='Maximum PSF width (σ)', anchor=tk.NW).place(x=0, y=154, width=190)
        self.max_sigma = ttk.Spinbox(master=self.parameter_frame, from_=1, to=5, increment=0.02)
        self.max_sigma.set(1.7)
        self.max_sigma.place(x=150, y=150, width=75)
        hover(self.max_sigma, "Maximum PSF width threshold. PSFs with wider standard deviation (pixels) will be rejected.", hover_delay=1000)

        # parameter is no longer used and widget is hidden but not removed to avoid breaking code
        # tk.Label(master=self.parameter_frame, bg='#333333', fg='#cccccc', text='Minimum ABS intensity', anchor=tk.NW).place(x=0, y=164, width=190)
        self.min_intensity = ttk.Spinbox(master=self.parameter_frame, from_=1, to=255, increment=1)
        self.min_intensity.set(8)
        # self.min_intensity.place(x=170, y=160, width=60)

        tk.Label(master=self.parameter_frame, bg='#333333', fg='#cccccc', text='Minimum eccentricity', anchor=tk.NW).place(x=0, y=204, width=190)
        self.eccentricity = ttk.Spinbox(master=self.parameter_frame, from_=0.1, to=1, increment=0.01)
        self.eccentricity.set(0.7)
        self.eccentricity.place(x=150, y=200, width=75)
        hover(self.eccentricity, "Eccentricity threshold. PSFs with min. / maj. width projected in x and y lower than this threshold will be rejected.", hover_delay=1000)

        # parameter is no longer used and widget is hidden but not removed to avoid breaking code
        # tk.Label(master=self.parameter_frame, bg='#333333', fg='#cccccc', text='Minimum GaussFit residual', anchor=tk.NW).place(x=0, y=244, width=190)
        self.min_gauss = ttk.Spinbox(master=self.parameter_frame, from_=1, to=20, increment=0.25)
        self.min_gauss.set(4)
        # self.min_gauss.place(x=170, y=240, width=60)

        # parameter is no longer used and widget is hidden but not removed to avoid breaking code
        self.use_global_norm = tk.IntVar(master=self.window)
        self.use_global_norm.set(0)
        self.use_global_norm_cb = ttk.Checkbutton(master=self.parameter_frame, variable=self.use_global_norm, text='  Global normalization', onvalue=1, offvalue=0, style='Switch.TCheckbutton')
        # self.use_global_norm_cb.place(x=260, y=0)

        self.use_low_pass = tk.IntVar(master=self.window)
        self.use_low_pass.set(1)
        self.use_low_pass_cb = ttk.Checkbutton(master=self.parameter_frame, variable=self.use_low_pass, text=' Low-pass subtraction', onvalue=1, offvalue=0, style='Switch.TCheckbutton')
        self.use_low_pass_cb.place(x=260, y=0)
        hover(self.use_low_pass_cb, "Subtract a low pass filtered image from the original enhanced Z-stack to clean up background (Effectively a High pass filter).", hover_delay=1000)

        # parameter is no longer used and widget is hidden but not removed to avoid breaking code
        self.average_mode = tk.IntVar(master=self.window)
        self.average_mode.set(0)
        self.average_mode_cb = ttk.Checkbutton(master=self.parameter_frame, variable=self.average_mode, text='  Average Mode', onvalue=1, offvalue=0, style='Switch.TCheckbutton')
        # self.average_mode_cb.place(x=260, y=30)

        self.load_calibration_button = ttk.Button(master=self.parameter_frame, text='Load Calibration', command=self.load_calibration)
        self.load_calibration_button.place(x=256, y=200, width=122)

        self.save_calibration_button = ttk.Button(master=self.parameter_frame, text='Save Current', command=self.save_calibration)
        self.save_calibration_button.place(x=384, y=200, width=98)

        self.mass_calibration_button = ttk.Button(master=self.parameter_frame, text='Create Mass Calibration', command=self.create_calibration)
        self.mass_calibration_button.place(x=256, y=240, width=226)

        hover(self.load_calibration_button, "Load a mass calibration (.mc) file.", hover_delay=1000)
        hover(self.save_calibration_button, "Save the currently applied mass calibration as a .mc file.", hover_delay=1000)
        hover(self.mass_calibration_button, "Create a mass calibration from current fitted peaks of a calibrant movie.", hover_delay=1000)

        # parameter is no longer used and widget is hidden but not removed to avoid breaking code
        self.extended_trace = tk.IntVar(master=self.window)
        self.extended_trace.set(0)
        self.extended_trace_cb = ttk.Checkbutton(master=self.parameter_frame, variable=self.extended_trace, text='Extend traces', onvalue=1, offvalue=0)
        # self.extended_trace_cb.place(x=260, y=100)

        self.binary_mask = tk.IntVar(master=self.window)
        self.binary_mask.set(0)
        self.binary_mask_cb = ttk.Checkbutton(master=self.parameter_frame, variable=self.binary_mask, text=' Subpixel mask convolution', onvalue=1, offvalue=0, style='Switch.TCheckbutton')
        self.binary_mask_cb.place(x=260, y=80)
        hover(self.binary_mask_cb, "Corrects contrast measurements by shifting and weighting the 3×3 PSF mask to match the PSF’s sub-pixel centroid, ensuring consistent "
                                   "values even when the PSF lies between pixels.", hover_delay=1000)

        # parameter is no longer used and widget is hidden but not removed to avoid breaking code
        self.extension_amount = ttk.Spinbox(master=self.parameter_frame, from_=1, to=50, increment=1)
        self.extension_amount.set(5)
        # self.extension_amount.place(x=380, y=98, width=50)
        # ttk.Label(master=self.parameter_frame, text='frames').place(x=435, y=134)

        self.mask_frame = ttk.LabelFrame(master=self.window, text=" Signal Masking ", padding=(6, 5))
        self.mask_frame.place(x=10, y=238, width=620, height=310)
        self.mask_notebook = ttk.Notebook(master=self.mask_frame)
        self.mask_notebook.pack(fill='both', expand=True)
        self.mask_tab1 = ttk.Frame(master=self.mask_notebook)
        self.mask_notebook.add(self.mask_tab1, text='Temporal Mask')
        self.mask_tab2 = ttk.Frame(master=self.mask_notebook)
        self.mask_notebook.add(self.mask_tab2, text='Spatial Mask')
        self.mask_tab3 = ttk.Frame(master=self.mask_notebook)
        self.mask_notebook.add(self.mask_tab3, text='Ratiometric Traces')
        self.mask_tab4 = ttk.Frame(master=self.mask_notebook)
        self.mask_notebook.add(self.mask_tab4, text='Event Refinement')
        self.mask_tab5 = ttk.Frame(master=self.mask_notebook)
        self.mask_notebook.add(self.mask_tab5, text='LapGauss Filter')

        self.analysis_frame = ttk.LabelFrame(master=self.window, text=" Analysis ", padding=(6, 5))
        self.analysis_frame.place(x=10, y=549, width=1132, height=455)
        self.analysis_notebook = ttk.Notebook(master=self.analysis_frame)
        self.analysis_notebook.pack(fill='both', expand=True)
        self.analysis_tab1 = ttk.Frame(master=self.analysis_notebook)
        self.analysis_notebook.add(self.analysis_tab1, text='Tracking')
        self.analysis_tab2 = ttk.Frame(master=self.analysis_notebook)
        self.analysis_notebook.add(self.analysis_tab2, text='Datasets')
        self.analysis_tab3 = ttk.Frame(master=self.analysis_notebook)
        self.analysis_notebook.add(self.analysis_tab3, text='Figures')
        self.analysis_tab4 = ttk.Frame(master=self.analysis_notebook)
        self.analysis_notebook.add(self.analysis_tab4, text='Export and Save')

        self.temp_frame = tk.Frame(master=self.mask_tab1)
        self.temp_frame.place(x=5, y=5)
        self.temp_figure = plt.Figure(figsize=(5.95, 2), dpi=100)
        self.temp_figure.set_facecolor("#333333")
        self.temp_figure.subplots_adjust(top=0.99, bottom=0.22, left=0.12, right=0.96)
        self.temp_plotter = None

        self.temp_canvas = FigureCanvasTkAgg(self.temp_figure, master=self.temp_frame)
        self.temp_canvas.draw()
        self.temp_canvas.get_tk_widget().pack(side="top", fill="both", expand=1)
        hover(self.temp_canvas.get_tk_widget(), "Click and drag over the temporal range to mask out signal. Multiple selections are possible", hover_delay=1000)
        self.temp_bind_mouse = None
        self.temp_bind_motion = None
        self.temp_rect = None
        self.temporal_mask = []
        self.temporal_mask_timeline = []

        self.temp_invert = tk.IntVar(master=self.window)
        self.temp_invert.set(0)
        self.temp_invert_cb = ttk.Checkbutton(master=self.mask_tab1, text='Invert', variable=self.temp_invert, offvalue=0, onvalue=1, command=self.invert_temp_command)
        self.temp_invert_cb.place(x=10, y=212)
        hover(self.temp_invert_cb, "Invert the selected mask, keeping only highlighted regions.", hover_delay=1000)

        self.delete_temp_button = ttk.Button(master=self.mask_tab1, text='Reset', command=self.delete_temp)
        self.delete_temp_button.place(x=100, y=210, width=100)
        hover(self.delete_temp_button, "Delete all mask selections.", hover_delay=1000)

        self.disp_sharpness = tk.IntVar(master=self.window)
        self.disp_sharpness.set(0)
        self.disp_sharpness_cb = ttk.Checkbutton(master=self.mask_tab1, text='Plot sharpness', variable=self.disp_sharpness, offvalue=0, onvalue=1, command=self.plot_temporal)
        self.disp_sharpness_cb.place(x=320, y=212)
        hover(self.disp_sharpness_cb, "Display the sharpness of the native movie (normalised std. dev. per frame).", hover_delay=1000)

        self.sharpness_mc = tk.IntVar(master=self.window)
        self.sharpness_mc.set(0)
        self.sharpness_mc_cb = ttk.Checkbutton(master=self.mask_tab1, text='Mass Correction', variable=self.sharpness_mc, offvalue=0, onvalue=1, command=lambda: self.plot_histogram(None))
        self.sharpness_mc_cb.place(x=455, y=212)
        hover(self.sharpness_mc_cb, "Enable mass correction to try and compensate for drifting defocus. This function assumes the maximum "
                                    "normalised sharpness value is in focus and "
                                    "corrects the contrast of events detected by the proportion of decreased sharpness relative to the maximum. "
                                    "Do not use mass correction if there are large step-like jumps in the sharpness from heavy particles landing "
                                    "as this will skew results of mass correction.", hover_delay=1000)

        self.spatial_frame = tk.Frame(master=self.mask_tab2)
        self.spatial_frame.place(x=5, y=5)
        self.spatial_figure = plt.Figure(figsize=(5.95, 2), dpi=100)
        self.spatial_figure.set_facecolor("#333333")
        self.spatial_figure.subplots_adjust(top=0.99, bottom=0.22, left=0.11, right=0.96)
        self.spatial_plotter = None

        self.spatial_canvas = FigureCanvasTkAgg(self.spatial_figure, master=self.spatial_frame)
        self.spatial_canvas.draw()
        self.spatial_canvas.get_tk_widget().pack(side="top", fill="both", expand=1)
        hover(self.spatial_canvas.get_tk_widget(), "Click and drag over the Z-projection of the Movie to mask out regions in space. Multiple selections are possible.", hover_delay=1000)
        self.spatial_bind_mouse = None
        self.spatial_bind_motion = None
        self.spatial_rect = None
        self.spatial_mask = []
        self.spatial_mask_image = None
        self.spatial_mask_binary = None

        self.spatial_invert = tk.IntVar(master=self.window)
        self.spatial_invert.set(0)
        self.spatial_invert_cb = ttk.Checkbutton(master=self.mask_tab2, text='Invert', variable=self.spatial_invert, offvalue=0, onvalue=1, command=self.invert_spatial_command)
        self.spatial_invert_cb.place(x=10, y=212)
        hover(self.spatial_invert_cb, "Invert the selected mask, keeping only highlighted regions.", hover_delay=1000)

        self.delete_spatial_button = ttk.Button(master=self.mask_tab2, text='Reset', command=self.delete_spatial)
        self.delete_spatial_button.place(x=100, y=210, width=100)
        hover(self.delete_spatial_button, "Delete all mask selections.", hover_delay=1000)

        self.trace_frame = tk.Frame(master=self.mask_tab3)
        self.trace_frame.place(x=5, y=5)
        self.trace_figure = plt.Figure(figsize=(5.95, 2), dpi=100)
        self.trace_figure.set_facecolor("#333333")
        self.trace_figure.subplots_adjust(top=0.96, bottom=0.22, left=0.13, right=0.96)
        self.trace_plotter = None

        self.trace_canvas = FigureCanvasTkAgg(self.trace_figure, master=self.trace_frame)
        self.trace_canvas.draw()
        self.trace_canvas.get_tk_widget().pack(side="top", fill="both", expand=1)
        hover(self.trace_canvas.get_tk_widget(), "Ratiometric trace view for events.", hover_delay=1000)
        self.trace_bind_mouse = None
        self.current_trace_event = None
        self.canvas_ratiometric.bind("<Button-1>", self.event_clicked)

        self.include = tk.IntVar(master=self.window)
        self.include.set(1)
        self.include_cb = ttk.Checkbutton(master=self.mask_tab3, text='Include', onvalue=1, offvalue=0, variable=self.include, command=self.update_include)
        self.include_cb.place(x=10, y=212)
        hover(self.include_cb, "Manually toggle inclusion or exclusion of the currently selected event.", hover_delay=1000)


        self.trace_info_entry = ttk.Entry(master=self.mask_tab3)
        self.trace_info_entry['state'] = tk.DISABLED
        self.trace_info_entry.place(x=100, y=210, width=490)
        hover(self.trace_info_entry, "Metrics of the current event trace.", hover_delay=1000)

        self.trace_label1 = tk.Label(master=self.mask_tab3, bg='#333333', fg='white', anchor=tk.NW, text=f'Event: []', font='calibri 9')
        self.trace_label2 = tk.Label(master=self.mask_tab3, bg='#333333', fg='white', anchor=tk.NW, text=f'Cont.: []', font='calibri 9')
        self.trace_label3 = tk.Label(master=self.mask_tab3, bg='#333333', fg='white', anchor=tk.NW, text=f'Mass: []', font='calibri 9')
        self.trace_label4 = tk.Label(master=self.mask_tab3, bg='#333333', fg='white', anchor=tk.NW, text=f'StD: []', font='calibri 9')
        self.trace_label5 = tk.Label(master=self.mask_tab3, bg='#333333', fg='white', anchor=tk.NW, text=f'SNR: []', font='calibri 9')
        self.trace_label1.place(x=110, y=215, width=90)
        self.trace_label2.place(x=190, y=215, width=90)
        self.trace_label3.place(x=295, y=215, width=100)
        self.trace_label4.place(x=410, y=215, width=90)
        self.trace_label5.place(x=510, y=215, width=75)

        self.filter_frame = tk.Frame(master=self.mask_tab4)
        self.filter_frame.place(x=2, y=1)
        self.filter_figure = plt.Figure(figsize=(4.7, 2.1), dpi=100)
        self.filter_figure.set_facecolor("#333333")
        self.filter_figure.subplots_adjust(top=0.98, bottom=0.2, left=0.15, right=0.96)
        self.filter_plotter = None

        self.filter_canvas = FigureCanvasTkAgg(self.filter_figure, master=self.filter_frame)
        self.filter_canvas.draw()
        self.filter_canvas.get_tk_widget().pack(side="top", fill="both", expand=1)
        self.filter_toolbar = NavigationToolbar2Tk(self.filter_canvas, self.filter_frame)
        self.filter_toolbar.update()
        self.filter_canvas.get_tk_widget().pack(side='top', fill='x', expand=1)
        hover(self.filter_canvas.get_tk_widget(), "Threshold histogram. Click to place threshold.", hover_delay=1000)
        self.filter_toolbar.winfo_children()[-1].pack_forget()
        self.filter_toolbar.winfo_children()[-2].pack_forget()

        self.filter_bind_mouse = None

        self.filter_plot_mode = tk.IntVar(master=self.window)
        self.filter_plot_mode.set(1)
        self.filter_r2_plot_rb = ttk.Radiobutton(master=self.mask_tab4, text='Gradient r²', variable=self.filter_plot_mode, value=1, command=self.filter_plot_decide)
        self.filter_r2_plot_rb.place(x=470, y=10)
        self.filter_gradient_plot_rb = ttk.Radiobutton(master=self.mask_tab4, text='Gradient diff', variable=self.filter_plot_mode, value=2, command=self.filter_plot_decide)
        self.filter_gradient_plot_rb.place(x=470, y=40)
        self.filter_snr_plot_rb = ttk.Radiobutton(master=self.mask_tab4, text='Signal / Noise', variable=self.filter_plot_mode, value=3, command=self.filter_plot_decide)
        self.filter_snr_plot_rb.place(x=470, y=70)
        hover(self.filter_r2_plot_rb, "Switch to gradient fit r² threshold.", hover_delay=1000)
        hover(self.filter_gradient_plot_rb, "Switch to gradient difference threshold.", hover_delay=1000)
        hover(self.filter_snr_plot_rb, "Switch to contrast signal / noise ratio threshold.", hover_delay=1000)

        self.filter_apply_button = ttk.Button(master=self.mask_tab4, text='Apply', command=lambda: self.get_filter_cursor(None))
        self.filter_apply_button.place(x=310, y=213, width=60, height=32)
        self.filter_reset_button = ttk.Button(master=self.mask_tab4, text='Reset', command=self.filter_reset)
        self.filter_reset_button.place(x=380, y=213, width=70, height=32)

        hover(self.filter_apply_button, "Apply current thresholds to refine events.", hover_delay=1000)
        hover(self.filter_reset_button, "Remove all thresholds and re-include all events.", hover_delay=1000)

        self.filter_r2_threshold = None
        self.filter_grad_threshold = None
        self.filter_snr_threshold = None

        # self.use_contrast_gradients = tk.IntVar(master=self.window)
        # self.use_contrast_gradients.set(0)
        # self.use_contrast_gradients_cb = ttk.Checkbutton(master=self.mask_tab4, text='Mass gradients', onvalue=1, offvalue=0, variable=self.use_contrast_gradients, command=self.update_contrast_gradient)
        # self.use_contrast_gradients_cb.place(x=310, y=212)

        self.laplacian_canvas_adsorption = tk.Canvas(master=self.mask_tab5, width=130, height=130, bg='#222222', highlightthickness=1, highlightbackground='#111111')
        self.laplacian_canvas_adsorption.place(x=2, y=2)
        self.laplacian_canvas_desorption = tk.Canvas(master=self.mask_tab5, width=130, height=130, bg='#222222', highlightthickness=1, highlightbackground='#111111')
        self.laplacian_canvas_desorption.place(x=140, y=2)
        hover(self.laplacian_canvas_adsorption, "DoG filter (adsorption PSF).", hover_delay=1000)
        hover(self.laplacian_canvas_desorption, "DoG filter (desorption PSF).", hover_delay=1000)

        self.lap_graph_frame_ad = tk.Frame(master=self.mask_tab5)
        self.lap_graph_frame_ad.place(x=2, y=138)
        self.lgf_ad_figure = plt.Figure(figsize=(1.32, 1.06), dpi=100)
        self.lgf_ad_figure.set_facecolor("#222222")
        self.lgf_ad_figure.subplots_adjust(top=1, bottom=0, left=0, right=1)
        self.lgf_ad_canvas = FigureCanvasTkAgg(self.lgf_ad_figure, master=self.lap_graph_frame_ad)
        self.lgf_ad_canvas.draw()
        self.lgf_ad_canvas.get_tk_widget().pack(side="top", fill="both", expand=1)
        hover(self.lgf_ad_canvas.get_tk_widget(), "Adsorption PSF cross-sectional profile.", hover_delay=1000)

        self.lap_graph_frame_des = tk.Frame(master=self.mask_tab5)
        self.lap_graph_frame_des.place(x=140, y=138)
        self.lgf_des_figure = plt.Figure(figsize=(1.32, 1.06), dpi=100)
        self.lgf_des_figure.set_facecolor("#222222")
        self.lgf_des_figure.subplots_adjust(top=1, bottom=0, left=0, right=1)
        self.lgf_des_canvas = FigureCanvasTkAgg(self.lgf_des_figure, master=self.lap_graph_frame_des)
        self.lgf_des_canvas.draw()
        self.lgf_des_canvas.get_tk_widget().pack(side="top", fill="both", expand=1)
        hover(self.lgf_des_canvas.get_tk_widget(), "Desorption PSF cross-sectional profile.", hover_delay=1000)

        tk.Label(master=self.mask_tab5, bg='#333333', fg='#cccccc', text='Centre width (σ)', anchor=tk.NW).place(x=360, y=15, width=180)
        self.lap_centre_sigma = ttk.Spinbox(master=self.mask_tab5, from_=0.5, to=3, increment=0.01, command=self.calculate_laplacian_kernel)
        self.lap_centre_sigma.set(1.65)
        self.lap_centre_sigma.bind("<Return>", lambda a:self.calculate_laplacian_kernel())
        self.lap_centre_sigma.place(x=480, y=12, width=100)
        hover(self.lap_centre_sigma, "Centre width (std. dev. in pixels) of Difference of Gaussians filter.", hover_delay=1000)

        tk.Label(master=self.mask_tab5, bg='#333333', fg='#cccccc', text='Outer width (σ)', anchor=tk.NW).place(x=360, y=65, width=180)
        self.lap_outer_sigma = ttk.Spinbox(master=self.mask_tab5, from_=1, to=4, increment=0.01, command=self.calculate_laplacian_kernel)
        self.lap_outer_sigma.set(2.2)
        self.lap_outer_sigma.bind("<Return>", lambda a:self.calculate_laplacian_kernel())
        self.lap_outer_sigma.place(x=480, y=62, width=100)
        hover(self.lap_outer_sigma, "Outer width (std. dev. in pixels) of Difference of Gaussians filter.", hover_delay=1000)

        tk.Label(master=self.mask_tab5, bg='#333333', fg='#cccccc', text='Rel. Magnitude', anchor=tk.NW).place(x=360, y=115, width=180)
        self.lap_magnitude = ttk.Spinbox(master=self.mask_tab5, from_=1.01, to=4, increment=0.01, command=self.calculate_laplacian_kernel)
        self.lap_magnitude.set(1.5)
        self.lap_magnitude.bind("<Return>", lambda a:self.calculate_laplacian_kernel())
        self.lap_magnitude.place(x=480, y=112, width=100)
        hover(self.lap_magnitude, "Relative weighting between central and outer Gaussian.", hover_delay=1000)

        tk.Label(master=self.mask_tab5, bg='#333333', fg='#cccccc', text='Contrast lim.', anchor=tk.NW).place(x=360, y=165, width=180)
        self.lap_contrast = ttk.Spinbox(master=self.mask_tab5, from_=0.05, to=3, increment=0.05, command=self.display_laplacian)
        self.lap_contrast.set(0.5)
        self.lap_contrast.bind("<Return>", lambda a: self.display_laplacian())
        self.lap_contrast.place(x=480, y=162, width=100)
        hover(self.lap_contrast, "Clipping amount for display (does not affect behaviour of kernel).", hover_delay=1000)

        self.use_laplacian = tk.IntVar(master=self.window)
        self.use_laplacian.set(0)
        self.use_laplacian_cb = ttk.Checkbutton(master=self.parameter_frame, variable=self.use_laplacian, text=' Laplacian of Gaussian Filter', onvalue=1, offvalue=0, style='Switch.TCheckbutton')
        self.use_laplacian_cb.place(x=260, y=30)
        hover(self.use_laplacian_cb, "Enable use of Difference of Gaussians approximation to Laplacian of Gaussian kernel for event feature enhancement. "
                                     "If this setting is disabled, no LoG filtering is used.", hover_delay=1000)

        self.laplacian_filter = None
        self.calculate_laplacian_kernel()

        self.tracking_settings_frame = ttk.LabelFrame(master=self.analysis_tab1, text=" Tracking Parameters ", padding=(6, 5))
        self.tracking_settings_frame.place(x=4, y=4, width=310, height=240)

        tk.Label(master=self.tracking_settings_frame, bg='#333333', fg='#cccccc', text='Max. Spatial Displacement', anchor=tk.NW).place(x=2, y=3, width=170)
        self.max_displacement = ttk.Entry(master=self.tracking_settings_frame)
        self.max_displacement.insert(tk.END, 5)
        self.max_displacement.place(x=190, y=0, width=50)
        tk.Label(master=self.tracking_settings_frame, bg='#333333', fg='#cccccc', text='pixels', anchor=tk.NW).place(x=248, y=3, width=48)
        hover(self.max_displacement, "Trajectory linking parameter. Defines the maximum jump distance in pixels a PSF can move from frame to frame to be linked to the same track "
                                     "Higher values are needed for faster particles but can increase the likelihood of track mixing in dense environments.", hover_delay=1000)

        tk.Label(master=self.tracking_settings_frame, bg='#333333', fg='#cccccc', text='Minimum Path Duration', anchor=tk.NW).place(x=2, y=43, width=170)
        self.min_path = ttk.Entry(master=self.tracking_settings_frame)
        self.min_path.insert(tk.END, 10)
        self.min_path.place(x=190, y=40, width=50)
        tk.Label(master=self.tracking_settings_frame, bg='#333333', fg='#cccccc', text='frames', anchor=tk.NW).place(x=248, y=43, width=48)
        hover(self.min_path, "Trajectory linking parameter. Trajectories must have at least this many time points or be discarded.", hover_delay=1000)

        tk.Label(master=self.tracking_settings_frame, bg='#333333', fg='#cccccc', text='Max. Temporal Displacement', anchor=tk.NW).place(x=2, y=83, width=180)
        self.max_dark = ttk.Entry(master=self.tracking_settings_frame)
        self.max_dark.insert(tk.END, 2)
        self.max_dark.place(x=190, y=80, width=50)
        tk.Label(master=self.tracking_settings_frame, bg='#333333', fg='#cccccc', text='frames', anchor=tk.NW).place(x=248, y=83, width=48)
        hover(self.max_dark, "Trajectory linking parameter. Maximum allowed time delta between events linked to the same trajectory.", hover_delay=1000)

        tk.Label(master=self.tracking_settings_frame, bg='#333333', fg='#cccccc', text='Pixel Size Calibration', anchor=tk.NW).place(x=2, y=138, width=180)
        self.pixel_size = ttk.Entry(master=self.tracking_settings_frame)
        self.pixel_size.insert(tk.END, 84.4)
        self.pixel_size.place(x=190, y=135, width=50)
        tk.Label(master=self.tracking_settings_frame, bg='#333333', fg='#cccccc', text='nm', anchor=tk.NW).place(x=248, y=138, width=48)
        hover(self.pixel_size, "Set the correct pixel size for your movie. This includes pixel binning when saved from the instrument. For example if the native pixel "
                                    "size of your mass photometer CMOS camera is 20 nm, but pixel binning is set to 4, then the pixel size entered here would be 80. This "
                                    "parameter is required for accurate calculation of diffusivity.", hover_delay=1000)

        tk.Label(master=self.tracking_settings_frame, bg='#333333', fg='#cccccc', text='Frame Interval', anchor=tk.NW).place(x=2, y=178, width=180)
        self.frame_interval = ttk.Entry(master=self.tracking_settings_frame)
        self.frame_interval.insert(tk.END, 10)
        self.frame_interval.place(x=190, y=175, width=50)
        tk.Label(master=self.tracking_settings_frame, bg='#333333', fg='#cccccc', text='ms', anchor=tk.NW).place(x=248, y=178, width=48)
        hover(self.frame_interval, "Time between frames. Calculate from recording time and number of frames acquired. The default recording time is usually 60 s.", hover_delay=1000)

        self.tracking_mask_frame = ttk.LabelFrame(master=self.analysis_tab1, text=" Signal Masking Behaviour ", padding=(6, 5))
        self.tracking_mask_frame.place(x=4, y=250, width=310, height=160)

        self.track_use_temporal = tk.IntVar(master=self.window)
        self.track_use_temporal.set(1)
        self.track_use_temporal_cb = ttk.Checkbutton(master=self.tracking_mask_frame, text='Ignore events based on temporal mask', variable=self.track_use_temporal, offvalue=0, onvalue=1)
        self.track_use_temporal_cb.place(x=5, y=0)
        hover(self.track_use_temporal_cb, "Enable if you wish events that are linked into trajectories to use the temporal masking rules. Disable to ignore temporal mask.", hover_delay=1000)

        self.track_use_spatial = tk.IntVar(master=self.window)
        self.track_use_spatial.set(1)
        self.track_use_spatial_cb = ttk.Checkbutton(master=self.tracking_mask_frame, text='Ignore events based on spatial mask', variable=self.track_use_spatial, offvalue=0, onvalue=1)
        self.track_use_spatial_cb.place(x=5, y=41)
        hover(self.track_use_spatial_cb, "Enable if you wish events that are linked into trajectories to use the spatial masking rules. Disable to ignore spatial mask.", hover_delay=1000)

        self.track_use_custom = tk.IntVar(master=self.window)
        self.track_use_custom.set(1)
        self.track_use_custom_cb = ttk.Checkbutton(master=self.tracking_mask_frame, text='Include manual', variable=self.track_use_custom, offvalue=0, onvalue=1)
        self.track_use_custom_cb.place(x=5, y=82)
        hover(self.track_use_custom_cb, "Enable if you wish events that are linked into trajectories to be affected by manual selection or deselection. Disable to ignore user defined "
                                 "selection of events.", hover_delay=1000)

        self.start_tracking_button = ttk.Button(master=self.tracking_mask_frame, text='Track Events', command=self.track_events)
        self.start_tracking_button.place(x=195, y=80, width=100)
        hover(self.start_tracking_button, "Begin linking events into trajectories.", hover_delay=1000)

        self.tracking_results_frame = ttk.LabelFrame(master=self.analysis_tab1, text=" Results ", padding=(6, 0))
        self.tracking_results_frame.place(x=322, y=4, width=790, height=420)

        self.tracking_results_notebook = ttk.Notebook(master=self.tracking_results_frame)
        self.tracking_results_notebook.place(x=2, y=11, width=775, height=358)
        self.results_tab1 = ttk.Frame(master=self.tracking_results_notebook)
        self.tracking_results_notebook.add(self.results_tab1, text='Tracks & Diffusion')
        self.results_tab2 = ttk.Frame(master=self.tracking_results_notebook)
        self.tracking_results_notebook.add(self.results_tab2, text='Mass Distributions')
        self.results_tab3 = ttk.Frame(master=self.tracking_results_notebook)
        self.tracking_results_notebook.add(self.results_tab3, text='Time Series')

        tk.Label(master=self.tracking_results_frame, bg='#333333', fg='#cccccc', text='Track number', anchor=tk.NW).place(x=394, y=4, width=90)
        self.current_track = ttk.Spinbox(master=self.tracking_results_frame, from_=1, to=1, increment=1, command=self.update_tracks)
        self.current_track.set('N/A')
        self.current_track['state'] = tk.DISABLED
        self.current_track.bind("<Return>", lambda a:self.update_tracks())
        self.current_track.place(x=490, y=0, width=80, height=32)
        hover(self.current_track, "Current selected track. You can cycle through tracks or enter a value here. Clicking a track in the ratiometric "
                                  "view will select the respective track here too.", hover_delay=1000)

        self.overlay_tracks = tk.IntVar(master=self.window)
        self.overlay_tracks.set(1)
        self.overlay_tracks_cb = ttk.Checkbutton(master=self.tracking_results_frame, text='Overlay tracks on movie', variable=self.overlay_tracks,
                                                 offvalue=0, onvalue=1, command=lambda: self.display_frame_ratio(int(float(self.ratio_frame_var.get()))))
        self.overlay_tracks_cb.place(x=590, y=2)
        hover(self.overlay_tracks_cb, "Overlay trajectory renders on the ratiometric and native canvases.", hover_delay=1000)

        self.track_frame = tk.Frame(master=self.results_tab1)
        self.track_frame.place(x=2, y=2)
        self.track_figure = plt.Figure(figsize=(3.2, 2.9), dpi=100)
        self.track_figure.set_facecolor("#2f2f2f")
        self.track_figure.subplots_adjust(top=0.98, bottom=0.15, left=0.15, right=0.98)
        self.track_plotter = None

        self.track_canvas = FigureCanvasTkAgg(self.track_figure, master=self.track_frame)
        self.track_canvas.draw()
        self.track_canvas.get_tk_widget().pack(side="top", fill="both", expand=1)
        hover(self.track_canvas.get_tk_widget(), "Trajectory display.", hover_delay=1000)
        self.track_toolbar = NavigationToolbar2Tk(self.track_canvas, self.track_frame)
        self.track_toolbar.update()
        self.track_canvas.get_tk_widget().pack(side='top', fill='x', expand=1)
        self.track_toolbar.winfo_children()[-1].pack_forget()
        self.track_toolbar.winfo_children()[-2].pack_forget()

        self.track_type_label = ttk.Label(master=self.results_tab1, text='')
        self.track_type_label.place(x=232, y=298, width=100)

        ttk.Entry(master=self.results_tab1, state=tk.DISABLED).place(x=330, y=2, width=434, height=40)
        # ttk.Entry(master=self.results_tab1, state=tk.DISABLED).place(x=330, y=2, width=150, height=40)
        # ttk.Label(master=self.results_tab1).place(x=470, y=3, width=20, height=38)
        # tk.Frame(master=self.results_tab1, bg='#777777').place(x=470, y=2, width=20, height=1)
        # tk.Frame(master=self.results_tab1, bg='#777777').place(x=470, y=41, width=10, height=1)
        self.track_plot_mode = tk.IntVar(master=self.window)
        self.track_plot_mode.set(1)
        self.track_plot_selected_cb = ttk.Radiobutton(master=self.results_tab1, text='Plot selected track', variable=self.track_plot_mode, value=1, command=self.plot_track)
        self.track_plot_selected_cb.place(x=335, y=8)
        hover(self.track_plot_selected_cb, "Plot only the trajectory of the currently selected track.", hover_delay=1000)

        self.track_plot_all_cb = ttk.Radiobutton(master=self.results_tab1, text='Plot all tracks', variable=self.track_plot_mode, value=2, command=self.plot_track)
        self.track_plot_all_cb.place(x=500, y=8)
        hover(self.track_plot_all_cb, "Plot all tracks in the same canvas (it is recommended to turn off annotations for this).", hover_delay=1000)

        self.track_plot_diff_cb = ttk.Radiobutton(master=self.results_tab1, text='Plot diffusion', variable=self.track_plot_mode, value=3, command=self.plot_track)
        self.track_plot_diff_cb.place(x=640, y=8)
        hover(self.track_plot_diff_cb, "Calculate and plot the diffusion coefficient distribution of all tracks. Compute takes time. Do not leave in this mode "
                                           "as it will cause lag when selecting tracks.", hover_delay=1000)

        self.track_show_binding = tk.IntVar(master=self.window)
        self.track_show_binding.set(1)
        self.track_show_binding_cb = ttk.Checkbutton(master=self.results_tab1, text='Adsorption', variable=self.track_show_binding, offvalue=0, onvalue=1, command=self.plot_track)
        # self.track_show_binding_cb.place(x=500, y=43)

        self.track_show_unbinding = tk.IntVar(master=self.window)
        self.track_show_unbinding.set(0)
        self.track_show_unbinding_cb = ttk.Checkbutton(master=self.results_tab1, text='Desorption', variable=self.track_show_unbinding, offvalue=0, onvalue=1, command=self.plot_track)
        # self.track_show_unbinding_cb.place(x=640, y=43)
        self.track_cols = ['blue', 'orange', 'green', 'brown', 'pink', 'yellow', 'purple', 'red']

        self.annotate_tracks = tk.IntVar(master=self.window)
        self.annotate_tracks.set(1)
        self.annotate_tracks_cb = ttk.Checkbutton(master=self.results_tab1, text='Annotate Tracks', variable=self.annotate_tracks, offvalue=0, onvalue=1, command=self.plot_track)
        self.annotate_tracks_cb.place(x=335, y=46)
        hover(self.annotate_tracks_cb, "Annotate trajectory with its start and end frame.", hover_delay=1000)

        self.mass_frame = tk.Frame(master=self.results_tab2)
        self.mass_frame.place(x=2, y=2)
        self.mass_figure = plt.Figure(figsize=(4.2, 2.84), dpi=100)
        self.mass_figure.set_facecolor("#2f2f2f")
        self.mass_figure.subplots_adjust(top=0.98, bottom=0.15, left=0.15, right=0.98)
        self.mass_plotter = None

        self.mass_canvas = FigureCanvasTkAgg(self.mass_figure, master=self.mass_frame)
        self.mass_canvas.draw()
        self.mass_canvas.get_tk_widget().pack(side="top", fill="both", expand=1)
        hover(self.mass_canvas.get_tk_widget(), "Plot of track-contrast or track-mass distribution.", hover_delay=1000)
        self.mass_toolbar = NavigationToolbar2Tk(self.mass_canvas, self.mass_frame)
        self.mass_toolbar.update()
        self.mass_canvas.get_tk_widget().pack(side='top', fill='x', expand=1)

        tk.Label(master=self.results_tab2, bg='#333333', fg='#cccccc', text='Bins', anchor=tk.NW).place(x=430, y=7, width=50)
        self.mass_bins = ttk.Spinbox(master=self.results_tab2, from_=10, to=400, increment=5, command=self.plot_track_masses)
        self.mass_bins.set(60)
        self.mass_bins.bind("<Return>", lambda a:self.plot_track_masses())
        self.mass_bins.place(x=470, y=2, width=80)
        hover(self.mass_bins, "Number of histogram bins.", hover_delay=1000)

        self.track_use_mass = tk.IntVar(master=self.window)
        self.track_use_mass.set(0)
        self.track_mass_cb = ttk.Checkbutton(master=self.results_tab2, variable=self.track_use_mass, text='Mass', onvalue=1, offvalue=0, command=self.plot_track_masses)
        self.track_mass_cb.place(x=570, y=5)
        hover(self.track_mass_cb, "Display track-masses (kDa) instead of contrasts if a mass calibration is loaded.", hover_delay=1000)

        self.track_override = tk.IntVar(master=self.window)
        self.track_override.set(0)
        self.track_override_cb = ttk.Checkbutton(master=self.results_tab2, variable=self.track_override, text='Plot on main', onvalue=1, offvalue=0, command=lambda: self.plot_histogram(None))
        self.track_override_cb.place(x=650, y=5)
        hover(self.track_override_cb, "Plot track-contrasts or track-masses on main histogram instead of events. They can then be fitted.", hover_delay=1000)

        self.mass_filtering_frame = ttk.LabelFrame(master=self.results_tab2, text='  Mass series filtering  ')
        self.mass_filtering_frame.place(x=430, y=40, width=330, height=150)
        hover(self.mass_filtering_frame, "Mass series filtering rejects tracks with values outside the thresholds. This includes the track display area when "
                                         "set to all tracks, and diffusion coefficients. It also excludes those tracks from being added to datasets for example for D vs Mass "
                                         "so you can use these thresholds for filtering tracks used to make figures.", hover_delay=200)

        self.mass_replot_btn = ttk.Button(master=self.results_tab2, text='Refresh Histogram', command=self.plot_track_masses)
        self.mass_replot_btn.place(x=430, y=220, width=330)
        hover(self.mass_replot_btn, "Refresh histogram with changes such as if time-series fits were recalculated and are currently being plotted.", hover_delay=1000)

        tk.Label(master=self.mass_filtering_frame, bg='#333333', fg='#cccccc', text='Minimum displacement threshold', anchor=tk.NW).place(x=10, y=10, width=220)
        self.mass_min_disp = ttk.Spinbox(master=self.mass_filtering_frame, from_=0.0, to=10, increment=0.1, command=self.plot_track_masses)
        self.mass_min_disp.set(0.3)
        self.mass_min_disp.bind("<Return>", lambda a:self.plot_track_masses())
        self.mass_min_disp.place(x=230, y=5, width=80)
        hover(self.mass_min_disp, "Exclude tracks with lower mean displacement distances (pixels).", hover_delay=1000)

        tk.Label(master=self.mass_filtering_frame, bg='#333333', fg='#cccccc', text='Standard deviation limit', anchor=tk.NW).place(x=10, y=50, width=200)
        self.mass_std_lim = ttk.Spinbox(master=self.mass_filtering_frame, from_=0.5, to=4, increment=0.0175, command=self.plot_track_masses)
        self.mass_std_lim.set(1.1675)
        self.mass_std_lim.bind("<Return>", lambda a:self.plot_track_masses())
        self.mass_std_lim.place(x=230, y=45, width=80)
        hover(self.mass_std_lim, "Calculate track mean contrasts from data points inside this range, rejecting outliers to improve precision.", hover_delay=1000)

        tk.Label(master=self.mass_filtering_frame, bg='#333333', fg='#cccccc', text='Minimum track length', anchor=tk.NW).place(x=10, y=90, width=200)
        self.mass_min_track = ttk.Spinbox(master=self.mass_filtering_frame, from_=5, to=10000, increment=1, command=self.plot_track_masses)
        self.mass_min_track.set(10)
        self.mass_min_track.bind("<Return>", lambda a:self.plot_track_masses())
        self.mass_min_track.place(x=230, y=85, width=80)
        hover(self.mass_min_track, "Exclude tracks with fewer data points.", hover_delay=1000)

        self.time_frame = tk.Frame(master=self.results_tab3)
        self.time_frame.place(x=2, y=2)
        self.time_figure = plt.Figure(figsize=(5, 2.84), dpi=100)
        self.time_figure.set_facecolor("#2f2f2f")
        self.time_figure.subplots_adjust(top=0.97, bottom=0.16, left=0.14, right=0.98, hspace=0.2)
        self.time_plotter1, self.time_plotter2 = None, None

        self.time_canvas = FigureCanvasTkAgg(self.time_figure, master=self.time_frame)
        self.time_canvas.draw()
        self.time_canvas.get_tk_widget().pack(side="top", fill="both", expand=1)
        hover(self.time_canvas.get_tk_widget(), "Trajectory Time series plots.", hover_delay=1000)
        self.time_toolbar = NavigationToolbar2Tk(self.time_canvas, self.time_frame)
        self.time_toolbar.update()
        self.time_canvas.get_tk_widget().pack(side='top', fill='x', expand=1)

        self.time_filtering_frame = ttk.LabelFrame(master=self.results_tab3, text=' Time series dynamics ')
        self.time_filtering_frame.place(x=510, y=0, width=250, height=320)

        self.use_plateau_contrasts = tk.IntVar(master=self.window)
        self.use_plateau_contrasts.set(0)
        self.use_plateau_contrasts_cb = ttk.Checkbutton(master=self.time_filtering_frame, variable=self.use_plateau_contrasts, text=' Display plateau contrasts',
                                                        onvalue=1, offvalue=0, style='Switch.TCheckbutton', command=lambda: self.plot_histogram(None))
        self.use_plateau_contrasts_cb.place(x=14, y=10)
        hover(self.use_plateau_contrasts_cb, "Display fitted plateau contrast or mass distribution instead of mean track-contrasts or masses.", hover_delay=1000)

        self.use_chung = tk.IntVar(master=self.window)
        self.use_chung.set(1)
        self.use_chung_cb = ttk.Checkbutton(master=self.time_filtering_frame, variable=self.use_chung, text=' Use Chung-Kennedy filtering', onvalue=1, offvalue=0)
        self.use_chung_cb.place(x=24, y=55)
        hover(self.use_chung_cb, "Prefilter time series data with a Chung-Kennedy filter.", hover_delay=1000)

        tk.Label(master=self.time_filtering_frame, bg='#333333', fg='#cccccc', text='Rolling avg. win.', anchor=tk.NW).place(x=14, y=95, width=130)
        self.time_rolling_win_size = ttk.Spinbox(master=self.time_filtering_frame, from_=1, to=20, increment=1)
        self.time_rolling_win_size.set(4)
        self.time_rolling_win_size.place(x=150, y=90, width=80)
        hover(self.time_rolling_win_size, "Rolling mean window size.", hover_delay=1000)

        tk.Label(master=self.time_filtering_frame, bg='#333333', fg='#cccccc', text='Δ Plateau Penalty', anchor=tk.NW).place(x=14, y=135, width=130)
        self.time_contrast_threshold = ttk.Spinbox(master=self.time_filtering_frame, from_=0.0001, to=0.1, increment=0.0001)
        self.time_contrast_threshold.set(0.001)
        self.time_contrast_threshold.place(x=150, y=130, width=80)
        hover(self.time_contrast_threshold, "Penalty for ruptures PELT model changepoint detector. Higher penalty means more robust to noise but may miss "
                                            "real plateau chnages and vice versa.", hover_delay=1000)

        tk.Label(master=self.time_filtering_frame, bg='#333333', fg='#cccccc', text='Min. plateau len.', anchor=tk.NW).place(x=14, y=175, width=130)
        self.time_min_plateau = ttk.Spinbox(master=self.time_filtering_frame, from_=3, to=200, increment=1)
        self.time_min_plateau.set(8)
        self.time_min_plateau.place(x=150, y=170, width=80)
        hover(self.time_min_plateau, "Fitted plateaus cannot be shorter than this value.", hover_delay=1000)

        tk.Label(master=self.time_filtering_frame, bg='#333333', fg='#cccccc', text='Plateau deviation cor.', anchor=tk.NW).place(x=14, y=215, width=130)
        self.time_plateau_std_filter = ttk.Spinbox(master=self.time_filtering_frame, from_=0.001, to=1, increment=0.001)
        self.time_plateau_std_filter.set(0.01)
        self.time_plateau_std_filter.place(x=150, y=210, width=80)
        hover(self.time_plateau_std_filter, "Decides whether dynamics are present and should be fitted vs a static contrast vs time to which a mean will "
                                  "be fitted based on the standard deviation of the signal as a proportion of its magnitude.", hover_delay=1000)

        self.time_recalc_button = ttk.Button(master=self.time_filtering_frame, text='Recalculate Plateaus', command=self.recalculate_plateaus)
        self.time_recalc_button.place(x=10, y=260, width=150, height=34)
        hover(self.time_recalc_button, "Recalculate time series plateau fits.", hover_delay=1000)

        self.time_plot_filtered_trace_btn = ttk.Button(master=self.time_filtering_frame, text='Filtered', command=self.time_plot_filtered_trace)
        self.time_plot_filtered_trace_btn.place(x=166, y=260, width=70, height=34)
        hover(self.time_plot_filtered_trace_btn, "Plot filtered time series.", hover_delay=1000)

        self.time_plot_multimer_btn = ttk.Button(master=self.results_tab3, text='Match Multimers', command=self.match_multimers)
        self.time_plot_multimer_btn.place(x=232, y=289, width=126, height=34)
        hover(self.time_plot_multimer_btn, "Enter a monomer mass and guesses for the number of multimers using current peak fits to see if peaks are multiples of "
                                           "the monomer mass.", hover_delay=1000)

        self.hist_image = tk.PhotoImage(file='icons/lux_hist.png')
        self.time_plot_plats_btn = ttk.Button(master=self.results_tab3, image=self.hist_image, command=self.plot_time_hist)
        self.time_plot_plats_btn.place(x=10, y=245, width=44, height=35)
        hover(self.time_plot_plats_btn, "Show distribution of plateaus for current time series as a histogram.", hover_delay=1000)

        self.tree_pane = ttk.PanedWindow(master=self.analysis_tab2)
        self.tree_pane.place(x=0, y=0, width=740, height=390)
        self.pane_1 = ttk.Frame(self.tree_pane, padding=6)
        self.tree_pane.add(self.pane_1, weight=1)

        self.scrollbar = ttk.Scrollbar(master=self.pane_1)
        self.scrollbar.pack(side="right", fill="y")

        self.treeview = ttk.Treeview(
            master=self.pane_1,
            selectmode="browse",
            yscrollcommand=self.scrollbar.set,
            columns=(1, 2),
            height=10,
        )
        self.treeview.pack(expand=True, fill="both")
        self.scrollbar.config(command=self.treeview.yview)

        self.treeview.column('#0', anchor="w", width=160)
        self.treeview.column(1, anchor='w', width=220)
        self.treeview.column(2, anchor='w', width=160)
        self.treeview.heading('#0', text="Dataset", anchor="center")
        self.treeview.heading(1, text='Filename / Data type', anchor='nw')
        self.treeview.heading(2, text='Information', anchor='nw')

        self.add_dataset_button = ttk.Button(master=self.analysis_tab2, text='Add Dataset', command=self.add_dataset)
        self.add_dataset_button.place(x=760, y=5, width=160)
        self.remove_dataset_button = ttk.Button(master=self.analysis_tab2, text='Remove Dataset', command=self.delete_dataset)
        self.remove_dataset_button.place(x=930, y=5, width=160)
        self.select_data_button = ttk.Button(master=self.analysis_tab2, text='Select Data', command=self.treeview_select)
        self.select_data_button.place(x=760, y=45, width=160)

        self.cancel_fig_select_data_btn = ttk.Button(master=self.analysis_tab2, text='Cancel Data Selection', command=self.cancel_fig_select_data)

        self.clear_datasets_btn = ttk.Button(master=self.analysis_tab2, text='Clear all data', command=self.clear_all_datasets)
        self.clear_datasets_btn.place(x=760, y=350, width=160)

        self.dataset_index = 0
        self.all_datasets = []
        self.add_data_win = None

        self.bind_mouse_tree = self.treeview.bind('<<TreeviewSelect>>', self.treeview_clicked)
        self.bind_double_tree = self.treeview.bind('<Double-1>', self.treeview_double_clicked)


        self.list_pane = ttk.PanedWindow(master=self.analysis_tab3)
        self.list_pane.place(x=0, y=0, width=310, height=300)
        self.pane_2 = ttk.Frame(self.list_pane, padding=6)
        self.list_pane.add(self.pane_2, weight=1)

        self.list_scrollbar = ttk.Scrollbar(master=self.pane_2)
        self.list_scrollbar.pack(side="right", fill="y")

        self.figure_listbox = ttk.Treeview(
            master=self.pane_2,
            selectmode="browse",
            yscrollcommand=self.scrollbar.set,
            columns=(1),
            height=4,
        )

        self.figure_listbox.pack(expand=True, fill="both")
        self.list_scrollbar.config(command=self.figure_listbox.yview)

        self.figure_listbox.column('#0', anchor='w', width=120)
        self.figure_listbox.column(1, anchor='w', width=160)
        self.figure_listbox.heading('#0', text='Figure / Source', anchor='nw')
        self.figure_listbox.heading(1, text='Title / Data', anchor='nw')

        self.all_figures = []
        self.bind_mouse_fig = self.figure_listbox.bind('<<TreeviewSelect>>', self.fig_list_clicked)
        self.bind_double_fig = self.figure_listbox.bind('<Double-1>', self.fig_list_double_clicked)
        self.bind_right_fig = self.figure_listbox.bind('<Button-3>', self.fig_list_right_clicked)

        self.create_figure_button = ttk.Button(master=self.analysis_tab3, text='New Figure', command=self.create_figure)
        self.create_figure_button.place(x=5, y=353, width=100)
        self.create_figure_button = ttk.Button(master=self.analysis_tab3, text='Delete Figure', command=self.delete_figure)
        self.create_figure_button.place(x=111, y=353, width=104)
        self.clear_figures_btn = ttk.Button(master=self.analysis_tab3, text='Clear all', command=self.clear_all_figures)
        self.clear_figures_btn.place(x=222, y=353, width=72)

        self.figure_select_data_btn = ttk.Button(master=self.analysis_tab3, text='Select Data', command=lambda: self.fig_list_right_clicked(None))
        self.figure_select_data_btn.place(x=111, y=310, width=183)

        self.figure_export_hist_btn = ttk.Button(master=self.analysis_tab3, text='Export Data', command=self.export_hist_data)
        self.figure_export_hist_btn.place(x=5, y=310, width=100)

        self.fig_frame = tk.Frame(master=self.analysis_tab3)
        self.fig_frame.place(x=682, y=2)
        self.fig_figure = plt.Figure(figsize=(8.4, 7.06), dpi=50)
        self.fig_figure.set_facecolor("#2f2f2f")
        self.fig_figure.subplots_adjust(top=0.92, bottom=0.12, left=0.15, right=0.96)
        self.fig_plotter = None

        self.fig_canvas = FigureCanvasTkAgg(self.fig_figure, master=self.fig_frame)
        self.fig_canvas.draw()
        self.fig_canvas.get_tk_widget().pack(side="top", fill="both", expand=1)
        self.fig_toolbar = NavigationToolbar2Tk(self.fig_canvas, self.fig_frame)
        self.fig_toolbar.update()
        self.fig_canvas.get_tk_widget().pack(side='top', fill='x', expand=1)

        ttk.Label(master=self.analysis_tab3, text='Title', anchor=tk.NW).place(x=320, y=12, width=50)
        self.figure_title = ttk.Entry(master=self.analysis_tab3)
        self.figure_title.bind("<Return>", lambda a:self.update_figure())
        self.figure_title.place(x=360, y=5, width=310, height=34)

        ttk.Label(master=self.analysis_tab3, text='x-axis', anchor=tk.NW).place(x=320, y=52, width=50)
        self.xaxis_title = ttk.Entry(master=self.analysis_tab3)
        self.xaxis_title.bind("<Return>", lambda a:self.update_figure())
        self.xaxis_title.place(x=360, y=45, width=120, height=34)

        ttk.Label(master=self.analysis_tab3, text='y-axis', anchor=tk.NW).place(x=510, y=52, width=50)
        self.yaxis_title = ttk.Entry(master=self.analysis_tab3)
        self.yaxis_title.bind("<Return>", lambda a:self.update_figure())
        self.yaxis_title.place(x=550, y=45, width=120, height=34)

        ttk.Label(master=self.analysis_tab3, text='Title font size', anchor=tk.NW).place(x=320, y=92, width=80)
        self.title_size = ttk.Spinbox(master=self.analysis_tab3, from_=5, to=40, increment=1, command=self.update_figure)
        self.title_size.set(28)
        self.title_size.bind("<Return>", lambda a:self.update_figure())
        self.title_size.place(x=410, y=85, width=70, height=34)

        ttk.Label(master=self.analysis_tab3, text='Axis font size', anchor=tk.NW).place(x=510, y=92, width=80)
        self.axis_size = ttk.Spinbox(master=self.analysis_tab3, from_=5, to=40, increment=1, command=self.update_figure)
        self.axis_size.set(22)
        self.axis_size.bind("<Return>", lambda a:self.update_figure())
        self.axis_size.place(x=600, y=85, width=70, height=34)

        tk.Frame(master=self.analysis_tab3, bg='#777777').place(x=310, y=130, height=1, width=360)
        tk.Frame(master=self.analysis_tab3, bg='#777777').place(x=430, y=136, height=180, width=1)

        ttk.Label(master=self.analysis_tab3, text='Alignment', anchor=tk.NW).place(x=360, y=136, width=69)
        ttk.Label(master=self.analysis_tab3, text='Left', anchor=tk.NW).place(x=330, y=166, width=50)
        self.fig_left_align = ttk.Spinbox(master=self.analysis_tab3, from_=0.0, to=0.8, increment=0.01, command=self.update_figure)
        self.fig_left_align.set(0.150)
        self.fig_left_align.bind("<Return>", lambda a:self.update_figure())
        self.fig_left_align.place(x=360, y=160, width=64, height=34)

        ttk.Label(master=self.analysis_tab3, text='Right', anchor=tk.NW).place(x=322, y=206, width=50)
        self.fig_right_align = ttk.Spinbox(master=self.analysis_tab3, from_=0.2, to=1.0, increment=0.01, command=self.update_figure)
        self.fig_right_align.set(0.960)
        self.fig_right_align.bind("<Return>", lambda a:self.update_figure())
        self.fig_right_align.place(x=360, y=200, width=64, height=34)

        ttk.Label(master=self.analysis_tab3, text='Top', anchor=tk.NW).place(x=330, y=246, width=50)
        self.fig_top_align = ttk.Spinbox(master=self.analysis_tab3, from_=0.2, to=1.0, increment=0.01, command=self.update_figure)
        self.fig_top_align.set(0.920)
        self.fig_top_align.bind("<Return>", lambda a:self.update_figure())
        self.fig_top_align.place(x=360, y=240, width=64, height=34)

        ttk.Label(master=self.analysis_tab3, text='Bottom', anchor=tk.NW).place(x=310, y=286, width=50)
        self.fig_bottom_align = ttk.Spinbox(master=self.analysis_tab3, from_=0.0, to=0.80, increment=0.01, command=self.update_figure)
        self.fig_bottom_align.set(0.120)
        self.fig_bottom_align.bind("<Return>", lambda a:self.update_figure())
        self.fig_bottom_align.place(x=360, y=280, width=64, height=34)

        ttk.Label(master=self.analysis_tab3, text='Min', anchor=tk.NW).place(x=440, y=142, width=50)
        self.fig_xmin = ttk.Spinbox(master=self.analysis_tab3, from_=-1.0, to=1.0, increment=0.0001, command=self.update_figure)
        self.fig_xmin.set(0.0)
        self.fig_xmin.bind("<Return>", lambda a: self.update_figure())
        self.fig_xmin.place(x=470, y=136, width=76, height=34)

        ttk.Label(master=self.analysis_tab3, text='Max', anchor=tk.NW).place(x=560, y=142, width=50)
        self.fig_xmax = ttk.Spinbox(master=self.analysis_tab3, from_=-1.0, to=1.0, increment=0.0001, command=self.update_figure)
        self.fig_xmax.set(1.0)
        self.fig_xmax.bind("<Return>", lambda a:self.update_figure())
        self.fig_xmax.place(x=594, y=136, width=76, height=34)

        ttk.Label(master=self.analysis_tab3, text='Bin width', anchor=tk.NW).place(x=490, y=182, width=60)
        self.fig_bin_width = ttk.Spinbox(master=self.analysis_tab3, from_=0.0001, to=1.0, increment=0.0001, command=self.update_figure)
        self.fig_bin_width.bind("<Return>", lambda a:self.update_figure())
        self.fig_bin_width.place(x=550, y=176, width=120, height=34)

        self.fig_use_density = tk.IntVar(master=self.analysis_tab3)
        self.fig_use_density.set(1)
        self.fig_use_density_cb = ttk.Checkbutton(master=self.analysis_tab3, text='Density > Counts', variable=self.fig_use_density, onvalue=1, offvalue=0, style='Switch.TCheckbutton', command=self.update_figure)
        self.fig_use_density_cb.place(x=440, y=220)

        self.fig_annotations = tk.IntVar(master=self.analysis_tab3)
        self.fig_annotations.set(1)
        self.fig_annotations_cb = ttk.Checkbutton(master=self.analysis_tab3, text='Annotations', variable=self.fig_annotations, onvalue=1, offvalue=0, style='Switch.TCheckbutton', command=self.update_figure)
        self.fig_annotations_cb.place(x=440, y=250)

        ttk.Label(master=self.analysis_tab3, text='size', anchor=tk.NW).place(x=580, y=253, width=50)
        self.fig_an_size = ttk.Spinbox(master=self.analysis_tab3, from_=1, to=40, increment=1, command=self.update_figure)
        self.fig_an_size.set(20)
        self.fig_an_size.bind("<Return>", lambda a:self.update_figure())
        self.fig_an_size.place(x=610, y=248, width=60, height=34)

        self.fig_grid = tk.IntVar(master=self.analysis_tab3)
        self.fig_grid.set(1)
        self.fig_grid_cb = ttk.Checkbutton(master=self.analysis_tab3, text='Grid lines', variable=self.fig_grid, onvalue=1, offvalue=0, style='Switch.TCheckbutton',
                                           command=self.update_figure)
        self.fig_grid_cb.place(x=440, y=280)

        self.fig_legend = tk.IntVar(master=self.analysis_tab3)
        self.fig_legend.set(0)
        self.fig_legend_cb = ttk.Checkbutton(master=self.analysis_tab3, text='Figure legend', variable=self.fig_legend, onvalue=1, offvalue=0, style='Switch.TCheckbutton',
                                             command=self.update_figure)
        self.fig_legend_cb.place(x=360, y=320)

        self.disp_large_fig = ttk.Button(master=self.analysis_tab3, text='Display Fullsize Figure', command=self.plot_fullsize_fig)
        self.disp_large_fig.place(x=520, y=353, width=150)

        self.disp_3d_fig = ttk.Button(master=self.analysis_tab3, text='Open 3D Histogram', command=self.plot_3D_hist)
        # self.disp_large_fig.place(x=520, y=312, width=150)

        self.apply_fig_annotations_button = ttk.Button(master=self.analysis_tab3, text='Apply', command=self.update_figure)
        self.apply_fig_annotations_button.place(x=320, y=353, width=190)

        self.mov_export_frame = ttk.LabelFrame(master=self.analysis_tab4, text=" Export Movies ", padding=(6, 0))
        self.mov_export_frame.place(x=4, y=0, width=400, height=388)

        self.export_ratiometric_mov_btn = ttk.Button(master=self.mov_export_frame, text='Export Ratiometric', command=self.export_ratiometric_mov)
        self.export_ratiometric_mov_btn.place(x=4, y=10)

        self.export_ratiometric_track_btn = ttk.Button(master=self.mov_export_frame, text='Export Ratiometric with current track', command=self.export_ratiometric_movtrack)
        self.export_ratiometric_track_btn.place(x=144, y=10, width=240)

        tk.Frame(master=self.mov_export_frame, bg='#777777').place(x=4, y=51, width=380, height=1)

        self.export_frames = tk.IntVar(master=self.window)
        self.export_frames.set(1)
        self.export_frames_cb = ttk.Checkbutton(master=self.mov_export_frame, text='Frame label', variable=self.export_frames, onvalue=1, offvalue=0)
        self.export_frames_cb.place(x=4, y=58)

        self.export_masses = tk.IntVar(master=self.window)
        self.export_masses.set(1)
        self.export_masses_cb = ttk.Checkbutton(master=self.mov_export_frame, text='Mass labels', variable=self.export_masses, onvalue=1, offvalue=0)
        self.export_masses_cb.place(x=4, y=98)

        ttk.Label(master=self.mov_export_frame, text='Font size').place(x=180, y=101)
        self.export_mass_size = ttk.Spinbox(master=self.mov_export_frame, from_=1, to=50, increment=1)
        self.export_mass_size.set(14)
        self.export_mass_size.place(x=250, y=96, width=134)

        ttk.Label(master=self.mov_export_frame, text='Start frame').place(x=4, y=141)
        self.export_start_frame = ttk.Spinbox(master=self.mov_export_frame, from_=0, to=0, increment=1)
        self.export_start_frame.set(0)
        self.export_start_frame.place(x=80, y=136, width=100)

        ttk.Label(master=self.mov_export_frame, text='End frame').place(x=210, y=141)
        self.export_end_frame = ttk.Spinbox(master=self.mov_export_frame, from_=0, to=0, increment=1)
        self.export_end_frame.set(0)
        self.export_end_frame.place(x=284, y=136, width=100)

        ttk.Label(master=self.mov_export_frame, text='Frame rate').place(x=206, y=181)
        self.export_fps = ttk.Spinbox(master=self.mov_export_frame, from_=10, to=120, increment=5)
        self.export_fps.set(30)
        self.export_fps.place(x=284, y=176, width=100)

        ttk.Label(master=self.mov_export_frame, text='Track Opacity').place(x=4, y=181)
        self.export_opacity = ttk.Spinbox(master=self.mov_export_frame, from_=0, to=1, increment=0.05)
        self.export_opacity.set(0.6)
        self.export_opacity.place(x=100, y=176, width=80)

        self.export_apply_filter = tk.IntVar(master=self.window)
        self.export_apply_filter.set(1)
        self.export_apply_filter_cb = ttk.Checkbutton(master=self.mov_export_frame, text='Apply filters', variable=self.export_apply_filter, onvalue=1, offvalue=0)
        self.export_apply_filter_cb.place(x=4, y=220)

        self.export_histogram_vid = tk.IntVar(master=self.window)
        self.export_histogram_vid.set(0)
        self.export_histogram_vid_cb = ttk.Checkbutton(master=self.mov_export_frame, text='Add histogram', variable=self.export_histogram_vid, onvalue=1, offvalue=0)
        self.export_histogram_vid_cb.place(x=4, y=250)

        ttk.Label(master=self.mov_export_frame, text='Bin width', anchor=tk.NW).place(x=4, y=285, width=60)
        self.mov_bin_width = ttk.Spinbox(master=self.mov_export_frame, from_=1, to=20, increment=0.1)
        self.mov_bin_width.set(8.0)
        self.mov_bin_width.place(x=68, y=280, width=60, height=34)

        tk.Frame(master=self.mov_export_frame, bg='#777777').place(x=140, y=220, width=1, height=200)
        ttk.Label(master=self.mov_export_frame, text='Histogram alignment', anchor=tk.NW).place(x=200, y=220, width=150)

        ttk.Label(master=self.mov_export_frame, text='x off.', anchor=tk.NW).place(x=150, y=255, width=50)
        self.export_left_align = ttk.Spinbox(master=self.mov_export_frame, from_=0.0, to=0.8, increment=0.01)
        self.export_left_align.set(0.06)
        self.export_left_align.place(x=195, y=250, width=64, height=34)

        ttk.Label(master=self.mov_export_frame, text='y off.', anchor=tk.NW).place(x=276, y=255, width=50)
        self.export_bottom_align = ttk.Spinbox(master=self.mov_export_frame, from_=0.0, to=0.8, increment=0.01)
        self.export_bottom_align.set(0.17)
        self.export_bottom_align.place(x=320, y=250, width=64, height=34)

        ttk.Label(master=self.mov_export_frame, text='x scale', anchor=tk.NW).place(x=150, y=295, width=50)
        self.export_xsize = ttk.Spinbox(master=self.mov_export_frame, from_=0.2, to=0.6, increment=0.01)
        self.export_xsize.set(0.4)
        self.export_xsize.place(x=195, y=290, width=64, height=34)

        ttk.Label(master=self.mov_export_frame, text='y scale', anchor=tk.NW).place(x=276, y=295, width=50)
        self.export_ysize = ttk.Spinbox(master=self.mov_export_frame, from_=0.2, to=0.6, increment=0.01)
        self.export_ysize.set(0.4)
        self.export_ysize.place(x=320, y=290, width=64, height=34)

        self.export_ratiometric_tracks_btn = ttk.Button(master=self.mov_export_frame, text='Export Ratiometric with all tracks', command=self.export_ratiometric_movtracks)
        self.export_ratiometric_tracks_btn.place(x=144, y=58, width=240)

        self.mov_preview_btn = ttk.Button(master=self.mov_export_frame, text="Preview frame", command=self.movie_preview)
        self.mov_preview_btn.place(x=4, y=320, width=126)

        self.datafig_export_frame = ttk.LabelFrame(master=self.analysis_tab4, text=" Save Figure Data ", padding=(6, 0))
        self.datafig_export_frame.place(x=420, y=0, width=400, height=70)

        self.save_datafigs_btn = ttk.Button(master=self.datafig_export_frame, text='Save datasets and figures', command=self.save_datafigs)
        self.save_datafigs_btn.place(x=4, y=10)

        self.load_datafigs_btn = ttk.Button(master=self.datafig_export_frame, text='Load datasets and figures', command=self.load_datafigs)
        self.load_datafigs_btn.place(x=190, y=10, width=190)

        self.data_export_frame = ttk.LabelFrame(master=self.analysis_tab4, text=" Export Analysed Data ", padding=(6, 0))
        self.data_export_frame.place(x=420, y=85, width=685, height=303)

        self.export_results_notebook = ttk.Notebook(master=self.data_export_frame)
        self.export_results_notebook.place(x=2, y=3, width=670, height=270)
        self.export_tab1 = ttk.Frame(master=self.export_results_notebook)
        self.export_results_notebook.add(self.export_tab1, text='Events and Tracks')
        self.export_tab2 = ttk.Frame(master=self.export_results_notebook)
        self.export_results_notebook.add(self.export_tab2, text='Processed Datasets')

        self.events_export_frame = ttk.LabelFrame(master=self.export_tab1, text=" Export Events ", padding=(6, 0))
        self.events_export_frame.place(x=5, y=3, width=650, height=110)

        self.tracks_export_frame = ttk.LabelFrame(master=self.export_tab1, text=" Export Tracks ", padding=(6, 0))
        self.tracks_export_frame.place(x=5, y=120, width=650, height=110)

        ttk.Label(master=self.events_export_frame, text='Select data').place(x=5, y=10)
        self.export_events_selection = tk.StringVar(master=self.window)
        field_list = ['Events', 'Contrasts', 'Masses']
        self.export_events_selection_menu = ttk.OptionMenu(self.events_export_frame, self.export_events_selection, 'Events', *field_list)
        self.export_events_selection_menu['menu'].configure(bg='#555555', relief=tk.RIDGE, fg='white')
        self.export_events_selection_menu.place(x=85, y=5, width=130)

        ttk.Label(master=self.events_export_frame, text='Select export format').place(x=255, y=10)
        self.export_events_format = tk.StringVar(master=self.window)
        field_list = ['Excel (.xlsx)', 'Pickled python object (.pydat)']
        self.export_events_format_menu = ttk.OptionMenu(self.events_export_frame, self.export_events_format, 'Excel (.xlsx)', *field_list)
        self.export_events_format_menu['menu'].configure(bg='#555555', relief=tk.RIDGE, fg='white')
        self.export_events_format_menu.place(x=390, y=5, width=240)

        self.export_events_button = ttk.Button(master=self.events_export_frame, text='Export', command=self.export_events_binding)
        self.export_events_button.place(x=390, y=50, width=240)

        ttk.Label(master=self.tracks_export_frame, text='Select data').place(x=5, y=10)
        self.export_tracks_selection = tk.StringVar(master=self.window)
        field_list = ['Tracks', 'Track contrasts', 'Track masses']
        self.export_tracks_selection_menu = ttk.OptionMenu(self.tracks_export_frame, self.export_tracks_selection, 'Tracks', *field_list)
        self.export_tracks_selection_menu['menu'].configure(bg='#555555', relief=tk.RIDGE, fg='white')
        self.export_tracks_selection_menu.place(x=85, y=5, width=130)

        ttk.Label(master=self.tracks_export_frame, text='Select export format').place(x=255, y=10)
        self.export_tracks_format = tk.StringVar(master=self.window)
        field_list = ['Excel (.xlsx)', 'Pickled python object (.pydat)']
        self.export_tracks_format_menu = ttk.OptionMenu(self.tracks_export_frame, self.export_tracks_format, 'Excel (.xlsx)', *field_list)
        self.export_tracks_format_menu['menu'].configure(bg='#555555', relief=tk.RIDGE, fg='white')
        self.export_tracks_format_menu.place(x=390, y=5, width=240)

        self.export_tracks_button = ttk.Button(master=self.tracks_export_frame, text='Export', command=self.export_tracks_binding)
        self.export_tracks_button.place(x=390, y=50, width=240)

        ttk.Label(master=self.export_tab2, text='Functionality not available in this version.').place(x=40, y=30)

        self.capture_motion_temporal = self.temp_canvas.mpl_connect('motion_notify_event', self.temporal_vline)
        self.temporal_vline_widget = None
        self.temporal_vline_text = None
        self.temporal_vline_text2 = None
        self.max_contrast = None

        self.capture_mouse = None
        self.capture_motion = self.figure.canvas.mpl_connect('motion_notify_event', self.figure_vline)
        self.figure_vline_widget = None
        self.drag_rect = None
        self.x_axis = np.linspace(0, 1, 1000)

        self.resolution = tuple()
        self.scale_factor = 4
        self.canvas_offset = 1

        self.native_stack = None
        self.ratiometric_stack = None
        self.zstack_adsorption = None
        self.zstack_desorption = None

        self.sharpness = []

        self.events = []
        self.contrasts = []
        self.masses = []
        self.tracks = []
        self.track_contrasts = []
        self.track_masses = []

        self.fits = []
        self.temp_xmin, self.temp_xmax = 0, 1

        self.restore_minimized_task_button = tk.Button(master=self.analysis_frame, bg='#007fff', fg='#cccccc', text='Restore progress bar', padx=10, pady=0, relief='flat',
                                                       command=self.restore_task_win)
        self.cancel_flag = False
        self.imported_data = False
        self.is_busy = False

        try:
            self.window.update()
            loading_win.withdraw()
        except Exception:
            'Failed'

        self.bins_mode(plot_type='initialisation')

        self.menu_state = tk.IntVar(master=self.window)
        self.menu_state.set(0)
        self.menu_image = tk.PhotoImage(file='icons/menu.png')
        self.menu_button = ttk.Checkbutton(master=self.window, image=self.menu_image, style='Toggle.TButton', offvalue=0, onvalue=1, command=self.set_menu_state, variable=self.menu_state)
        self.menu_button.place(x=10, y=10, width=120, height=33)
        hover(self.menu_button, "Toggle the menu ribbon.", hover_delay=1000)

        self.menu_tile = tk.Frame(master=self.window, bg='#282828')
        self.file_menu_button = ttk.Menubutton(master=self.menu_tile, text='File')
        self.file_menu = tk.Menu(master=self.window)
        self.file_menu.add_command(label='New Experiment', command=self.new_exp)
        self.file_menu.add_command(label='Open Experiment', command=self.open_experiment)
        self.file_menu.add_command(label='Save Experiment', command=self.save_experiment)
        self.file_menu.add_separator()
        self.file_menu.add_command(label='Open .dsf Workspace', command=self.load_datafigs)
        self.file_menu.add_command(label='Save .dsf Workspace', command=self.save_datafigs)
        self.file_menu.add_separator()
        self.file_menu.add_command(label='Analysis Profiles', command=self.analysis_profiles)
        self.file_menu.add_command(label='Reboot OpenMASS', command=reset)
        self.file_menu_button['menu'] = self.file_menu
        self.file_menu_button.place(x=10, y=5)

        self.edit_menu_button = ttk.Menubutton(master=self.menu_tile, text='Edit')
        self.edit_menu = tk.Menu(master=self.window)
        self.edit_menu.add_command(label='Preferences', command=self.open_preferences)
        self.edit_menu_button['menu'] = self.edit_menu
        self.edit_menu_button.place(x=69, y=5)

        self.tools_menu_button = ttk.Menubutton(master=self.menu_tile, text='Tools')
        self.tools_menu = tk.Menu(master=self.window)
        self.tools_menu.add_command(label='Drift Correction', command=self.correct_drift)
        self.tools_menu.add_command(label='Optimise Window Size', command=None)
        self.tools_menu.entryconfig("Optimise Window Size", state=tk.DISABLED)
        self.tools_menu_button['menu'] = self.tools_menu
        self.tools_menu_button.place(x=131, y=5)

        self.data_menu_button = ttk.Menubutton(master=self.menu_tile, text='Data')
        self.data_menu = tk.Menu(master=self.window)
        self.data_menu.add_command(label='Show Gaussian Mixture Model Options', command=self.gaussfit_options)
        self.data_menu.add_separator()
        self.data_menu.add_command(label='Import Ratiometric Stack', command=self.import_ratiometric)
        self.data_menu.add_separator()
        self.data_menu.add_command(label='Import Contrasts (.pydat)', command=self.import_contrast_data)
        self.data_menu.add_separator()
        self.data_menu.add_command(label='Export Ratiometric (.tiff)', command=self.export_ratiometric_tiff)
        self.data_menu.add_command(label='Export Z-Stacks', command=None)
        self.data_menu.entryconfig("Export Z-Stacks", state=tk.DISABLED)
        self.data_menu_button['menu'] = self.data_menu
        self.data_menu_button.place(x=203, y=5)

        self.about_menu_button = ttk.Menubutton(master=self.menu_tile, text='About')
        self.about_menu = tk.Menu(master=self.window)
        self.about_menu.add_command(label='Versions / Licence', command=self.about)
        self.about_menu.add_command(label='Third Party Licences', command=self.about_third)
        self.about_menu_button['menu'] = self.about_menu
        self.about_menu_button.place(x=270, y=5)

        self.plot_histogram('initialisation')
        if preferences['start'] == 'experiment':
            self.select_experiment_type()
        elif preferences['start'] == 'profile':
            self.analysis_profiles()
        elif preferences['start'] == 'landing':
            try:
                self.load_analysis_profile(self.cwd + r"\Profiles\landing_assay_default.dprf")
            except Exception:
                print(traceback.format_exc())
                easygui.msgbox(title='Error!', msg='Could not find profile for landing assay preset.')
        elif preferences['start'] == 'tracking':
            try:
                self.load_analysis_profile(self.cwd + r"\Profiles\dynamic_tracking_default.dprf")
            except Exception:
                print(traceback.format_exc())
                easygui.msgbox(title='Error!', msg='Could not find profile for dynamic tracking preset.')

    def select_experiment_type(self):
        choice = easygui.indexbox(title='Experiment Type', msg='Please select the kind of experiment you wish to analyse. This preset will optimise settings for the selected experiment.',
                                  choices=['Landing Assay', 'Dynamic Tracking', 'Custom', 'Keep previous settings'], default_choice=0, cancel_choice=3)
        if choice == 0:
            try:
                self.load_analysis_profile(self.cwd + r"\Profiles\landing_assay_default.dprf")
            except Exception:
                print(traceback.format_exc())
                self.detection_threshold.set(1.1)
                self.nearest_neighbours.set(8)
                self.eccentricity.set(0.65)
                self.use_median.set(0)
                self.pre_avg.set(5)
                self.window_size.set(50)
                self.offset.set(5)
                self.track_override.set(0)
                self.persistence.set(5)
                self.show_boxes.set(1)
                self.use_laplacian.set(0)
                self.use_global_norm.set(0)
                self.average_mode.set(0)
                self.use_low_pass.set(1)

        elif choice == 1:
            try:
                self.load_analysis_profile(self.cwd + r"\Profiles\dynamic_tracking_default.dprf")
            except Exception:
                print(traceback.format_exc())
                self.detection_threshold.set(4)
                self.nearest_neighbours.set(8)
                self.eccentricity.set(0.6)
                self.use_median.set(1)
                self.pre_avg.set(100)
                self.window_size.set(1)
                self.offset.set(0)
                self.persistence.set(1)
                self.use_laplacian.set(0)
                self.use_global_norm.set(0)
                self.average_mode.set(0)
                self.use_low_pass.set(1)
        elif choice == 2:
            self.analysis_profiles()

    def save_analysis_profile(self, path):
        print(path)
        try:
            with open(path, 'wb') as file:
                pickle.dump(int(self.pre_avg.get()), file)
                pickle.dump(int(self.use_median.get()), file)
                pickle.dump(int(self.window_size.get()), file)
                pickle.dump(int(self.offset.get()), file)
                pickle.dump(float(self.ratio_norm.get()), file)
                pickle.dump(int(self.persistence.get()), file)
                pickle.dump(self.enable_adsorption.get(), file)
                pickle.dump(self.enable_desorption.get(), file)
                pickle.dump(self.show_boxes.get(), file)

                pickle.dump(self.lap_centre_sigma.get(), file)
                pickle.dump(self.lap_outer_sigma.get(), file)
                pickle.dump(self.lap_magnitude.get(), file)
                pickle.dump(self.lap_contrast.get(), file)

                pickle.dump(float(self.detection_threshold.get()), file)
                pickle.dump(float(self.nearest_neighbours.get()), file)
                pickle.dump(float(self.min_sigma.get()), file)
                pickle.dump(float(self.max_sigma.get()), file)
                pickle.dump(float(self.min_intensity.get()), file)
                pickle.dump(float(self.eccentricity.get()), file)
                pickle.dump(float(self.min_gauss.get()), file)

                pickle.dump(self.use_global_norm.get(), file)
                pickle.dump(self.use_low_pass.get(), file)
                pickle.dump(self.average_mode.get(), file)
                pickle.dump(self.use_laplacian.get(), file)
                pickle.dump(self.extended_trace.get(), file)
                pickle.dump(int(self.extension_amount.get()), file)

                pickle.dump(self.max_displacement.get(), file)
                pickle.dump(self.min_path.get(), file)
                pickle.dump(self.max_dark.get(), file)
                pickle.dump(self.pixel_size.get(), file)
                pickle.dump(self.frame_interval.get(), file)

                pickle.dump(self.track_use_temporal.get(), file)
                pickle.dump(self.track_use_spatial.get(), file)
                pickle.dump(self.track_use_custom.get(), file)

                pickle.dump(self.mass_min_disp.get(), file)
                pickle.dump(self.mass_std_lim.get(), file)
                pickle.dump(self.mass_min_track.get(), file)

                pickle.dump(self.invert_ratiometric.get(), file)
                pickle.dump(self.binary_mask.get(), file)
                pickle.dump(self.correct_motion_var.get(), file)

        except Exception:
            easygui.msgbox(title='Error', msg=f"Error saving analysis profile:\n{traceback.format_exc()}")

    def load_analysis_profile(self, path):
        print(path)
        try:
            with open(path, 'rb') as file:
                self.pre_avg.set(pickle.load(file))
                self.use_median.set(pickle.load(file))
                self.window_size.set(pickle.load(file))
                self.offset.set(pickle.load(file))
                self.ratio_norm.set(pickle.load(file))
                self.persistence.set(pickle.load(file))
                self.enable_adsorption.set(pickle.load(file))
                self.enable_desorption.set(pickle.load(file))
                self.show_boxes.set(pickle.load(file))

                self.lap_centre_sigma.set(pickle.load(file))
                self.lap_outer_sigma.set(pickle.load(file))
                self.lap_magnitude.set(pickle.load(file))
                self.lap_contrast.set(pickle.load(file))

                self.detection_threshold.set(pickle.load(file))
                self.nearest_neighbours.set(pickle.load(file))
                self.min_sigma.set(pickle.load(file))
                self.max_sigma.set(pickle.load(file))
                self.min_intensity.set(pickle.load(file))
                self.eccentricity.set(pickle.load(file))
                self.min_gauss.set(pickle.load(file))

                self.use_global_norm.set(pickle.load(file))
                self.use_low_pass.set(pickle.load(file))
                self.average_mode.set(pickle.load(file))
                self.use_laplacian.set(pickle.load(file))
                self.extended_trace.set(pickle.load(file))
                self.extension_amount.set(pickle.load(file))

                self.max_displacement.delete(0, tk.END)
                self.max_displacement.insert(0, pickle.load(file))
                self.min_path.delete(0, tk.END)
                self.min_path.insert(0, pickle.load(file))
                self.max_dark.delete(0, tk.END)
                self.max_dark.insert(0, pickle.load(file))
                self.pixel_size.delete(0, tk.END)
                self.pixel_size.insert(0, pickle.load(file))
                self.frame_interval.delete(0, tk.END)
                self.frame_interval.insert(0, pickle.load(file))

                self.track_use_temporal.set(pickle.load(file))
                self.track_use_spatial.set(pickle.load(file))
                self.track_use_custom.set(pickle.load(file))

                self.mass_min_disp.set(pickle.load(file))
                self.mass_std_lim.set(pickle.load(file))
                self.mass_min_track.set(pickle.load(file))

                self.invert_ratiometric.set(pickle.load(file))
                self.binary_mask.set(pickle.load(file))
                self.correct_motion_var.set(pickle.load(file))
            self.calculate_laplacian_kernel()
            self.display_laplacian()

        except Exception:
            easygui.msgbox(title='Error', msg=f"Error loading analysis profile:\n{traceback.format_exc()}")

    def analysis_profiles(self):
        try:
            self.profile_win.window.destroy()
        except Exception:
            pass  # window doesn't exist
        self.profile_win = ProfileWin(cwd=self.cwd)

    def new_exp(self, load=False):
        self.window.title(f'OpenMASS Mass Photometry Analysis {VERSION}')
        self.imported_data = False
        self.track_override.set(0)
        self.overlay_tracks.set(0)
        self.resolution = tuple()
        self.scale_factor = 4
        self.canvas_offset = 1

        self.native_stack = None
        self.ratiometric_stack = None
        self.zstack_adsorption = None
        self.zstack_desorption = None

        self.events = []
        self.contrasts = []
        self.masses = []
        self.tracks = []
        self.track_contrasts = []
        self.track_masses = []

        self.fits = []
        self.auto_fits = []

        self.field_list = ['Component 1,', 'Component 2,', 'Component 3,', 'Component 4,', 'Component 5,', 'Component 6,', 'Component 7,', 'Component 8,']
        self.manual_fit_selection.set_menu('select component', *self.field_list)
        self.plot_histogram('initialisation')

        self.native_frame_var.set(0)
        self.ratio_frame_var.set(0)
        self.proj_frame_var.set(0)
        self.slider_native['to'] = 0
        self.slider_ratio['to'] = 0
        self.slider_proj['to'] = 0
        self.canvas_native.delete('all')
        self.canvas_ratiometric.delete('all')
        self.canvas_projection.delete('all')

        self.temporal_mask_timeline = []
        self.sharpness = []
        self.temporal_mask = []
        self.spatial_mask_binary = None
        self.spatial_mask = []
        try:
            self.plot_temporal()
        except Exception:
            ''' Plot has been reset '''
        try:
            self.spatial_figure.clf()
            self.spatial_canvas.draw()
        except Exception:
            ''' Plot has been reset '''
        try:
            self.trace_figure.clf()
            self.trace_canvas.draw()
        except Exception:
            print(traceback.format_exc())

        self.trace_label1['text'] = f'Event: []'
        self.trace_label2['text'] = f'Cont.: []'
        self.trace_label3['text'] = f'Mass: []'
        self.trace_label4['text'] = f'StD: []'
        self.trace_label5['text'] = f'SNR: []'

        self.temp_invert.set(0)
        self.spatial_invert.set(0)
        self.filter_figure.clf()
        self.filter_canvas.draw()
        self.track_figure.clf()
        self.track_canvas.draw()
        self.mass_figure.clf()
        self.mass_canvas.draw()
        self.time_figure.clf()
        self.time_canvas.draw()
        if not load:
            if preferences['start'] == 'experiment':
                self.select_experiment_type()
            elif preferences['start'] == 'profile':
                self.analysis_profiles()
            elif preferences['start'] == 'landing':
                try:
                    self.load_analysis_profile(self.cwd + r"\Profiles\landing_assay_default.dprf")
                except Exception:
                    print(traceback.format_exc())
                    easygui.msgbox(title='Error!', msg='Could not find profile for landing assay preset.')
            elif preferences['start'] == 'tracking':
                try:
                    self.load_analysis_profile(self.cwd + r"\Profiles\dynamic_tracking_default.dprf")
                except Exception:
                    print(traceback.format_exc())
                    easygui.msgbox(title='Error!', msg='Could not find profile for dynamic tracking preset.')

    def save_experiment(self):
        path = easygui.filesavebox(title='Save Experiment', filetypes=['*.lmp'], default='N://*.lmp')
        if not path:
            # path = 'masses.msf'
            # with open(path, 'wb') as file:
            #     pickle.dump(self.masses, file)
            #     pickle.dump(self.track_masses, file)
            return
        if path[-4:] != '.lmp':
            path = path + '.lmp'
        try:
            self.progress_win.window.destroy()
        except Exception:
            pass
        self.progress_win = ProgressWin(title='Saving Experiment', msg='Dumping native stack...')
        self.progress_win.progress['maximum'] = 100
        self.progress_win.window.update()
        with open(path, 'wb') as file:
            pickle.dump(self.native_stack, file)
            self.progress_win.progress.step(30)
            self.progress_win.label['text'] = 'Dumping ratiometric stack...'
            self.progress_win.window.update()
            pickle.dump(self.ratiometric_stack, file)
            self.progress_win.progress.step(30)
            self.progress_win.label['text'] = 'Dumping z-stacks...'
            self.progress_win.window.update()
            pickle.dump(self.zstack_adsorption, file)
            self.progress_win.progress.step(15)
            self.progress_win.window.update()
            pickle.dump(self.zstack_desorption, file)
            self.progress_win.progress.step(15)
            self.progress_win.label['text'] = 'Dumping events and parameters...'
            self.progress_win.window.update()
            pickle.dump(self.events, file)
            pickle.dump(self.resolution, file)
            pickle.dump(self.canvas_offset, file)
            pickle.dump(self.scale_factor, file)
            pickle.dump(self.slider_native['to'], file)
            pickle.dump(self.slider_ratio['to'], file)
            pickle.dump(self.slider_proj['to'], file)
            pickle.dump(self.native_frame_var.get(), file)
            pickle.dump(self.ratio_frame_var.get(), file)
            pickle.dump(self.proj_frame_var.get(), file)
            pickle.dump(self.internal_window, file)
            pickle.dump(self.internal_frame_average, file)

            pickle.dump(self.spatial_mask_image, file)
            pickle.dump(self.spatial_mask_binary, file)
            pickle.dump(self.temporal_mask_timeline, file)
            pickle.dump(self.spatial_mask, file)
            pickle.dump(self.temporal_mask, file)

            pickle.dump(self.filter_r2_threshold, file)
            pickle.dump(self.filter_grad_threshold, file)
            pickle.dump(self.filter_snr_threshold, file)
            pickle.dump(self.current_trace_event, file)

            pickle.dump(int(self.pre_avg.get()), file)
            pickle.dump(int(self.use_median.get()), file)
            pickle.dump(int(self.window_size.get()), file)
            pickle.dump(int(self.offset.get()), file)
            pickle.dump(float(self.ratio_norm.get()), file)
            pickle.dump(int(self.persistence.get()), file)
            pickle.dump(self.enable_adsorption.get(), file)
            pickle.dump(self.enable_desorption.get(), file)
            pickle.dump(self.show_boxes.get(), file)

            pickle.dump(float(self.min_entry.get()), file)
            pickle.dump(float(self.max_entry.get()), file)
            pickle.dump(float(self.bins.get()), file)
            pickle.dump(self.use_mass.get(), file)
            pickle.dump(self.field_list, file)
            pickle.dump(self.number_of_fits.get(), file)

            pickle.dump(self.calibration, file)
            pickle.dump(self.lap_centre_sigma.get(), file)
            pickle.dump(self.lap_outer_sigma.get(), file)
            pickle.dump(self.lap_magnitude.get(), file)
            pickle.dump(self.lap_contrast.get(), file)

            pickle.dump(float(self.detection_threshold.get()), file)
            pickle.dump(float(self.nearest_neighbours.get()), file)
            pickle.dump(float(self.min_sigma.get()), file)
            pickle.dump(float(self.max_sigma.get()), file)
            pickle.dump(float(self.min_intensity.get()), file)
            pickle.dump(float(self.eccentricity.get()), file)
            pickle.dump(float(self.min_gauss.get()), file)

            pickle.dump(self.use_global_norm.get(), file)
            pickle.dump(self.use_low_pass.get(), file)
            pickle.dump(self.average_mode.get(), file)
            pickle.dump(self.use_laplacian.get(), file)
            pickle.dump(self.extended_trace.get(), file)
            pickle.dump(int(self.extension_amount.get()), file)

            # Tracking data and UI params
            pickle.dump(self.tracks, file)
            pickle.dump(self.track_contrasts, file)
            pickle.dump(self.track_masses, file)
            if len(self.tracks) > 0:
                pickle.dump(int(float(self.current_track.get())), file)
            else:
                pickle.dump(self.current_track.get(), file)

            pickle.dump(self.max_displacement.get(), file)
            pickle.dump(self.min_path.get(), file)
            pickle.dump(self.max_dark.get(), file)
            pickle.dump(self.pixel_size.get(), file)
            pickle.dump(self.frame_interval.get(), file)

            pickle.dump(self.track_use_temporal.get(), file)
            pickle.dump(self.track_use_spatial.get(), file)
            pickle.dump(self.track_use_custom.get(), file)

            pickle.dump(self.overlay_tracks.get(), file)
            pickle.dump(self.mass_bins.get(), file)
            pickle.dump(self.track_use_mass.get(), file)
            pickle.dump(self.track_override.get(), file)

            pickle.dump(self.mass_min_disp.get(), file)
            pickle.dump(self.mass_std_lim.get(), file)
            pickle.dump(self.mass_min_track.get(), file)

            pickle.dump(self.file_path, file)

            pickle.dump(self.invert_ratiometric.get(), file)
            pickle.dump(self.binary_mask.get(), file)

            pickle.dump(self.sharpness_mc.get(), file)

            pickle.dump(self.auto_fits, file)
            pickle.dump(self.main_bins_mode.get(), file)
            pickle.dump(self.fit_manually.get(), file)

            pickle.dump(self.correct_motion_var.get(), file)

            pickle.dump(self.use_plateau_contrasts.get(), file)
            pickle.dump(self.use_chung.get(), file)
            pickle.dump(self.time_rolling_win_size.get(), file)
            pickle.dump(self.time_contrast_threshold.get(), file)
            pickle.dump(self.time_min_plateau.get(), file)
            pickle.dump(self.time_plateau_std_filter.get(), file)

            self.progress_win.progress.step(9)
            self.progress_win.window.update()
        self.progress_win.handle_close()


    def open_experiment(self):
        path = easygui.fileopenbox(title='Open Experiment', filetypes=['*.lmp'], default='N://*.lmp')
        if not path:
            return
        self.new_exp(load=True)
        try:
            self.progress_win.window.destroy()
        except Exception:
            pass
        self.progress_win = ProgressWin(title='Loading Experiment', msg='Retrieving native stack...')
        self.progress_win.progress['maximum'] = 100
        self.progress_win.window.update()
        with open(path, 'rb') as file:
            self.native_stack = pickle.load(file)
            self.progress_win.progress.step(30)
            self.progress_win.label['text'] = 'Retrieving ratiometric stack...'
            self.progress_win.window.update()
            self.ratiometric_stack = pickle.load(file)
            self.progress_win.progress.step(30)
            self.progress_win.label['text'] = 'Retrieving z-stacks...'
            self.progress_win.window.update()
            self.zstack_adsorption = pickle.load(file)
            self.progress_win.progress.step(15)
            self.progress_win.window.update()
            self.zstack_desorption = pickle.load(file)
            self.progress_win.progress.step(15)
            self.progress_win.label['text'] = 'Retrieving events and parameters...'
            self.progress_win.window.update()
            try:
                sharp = []
                for idx in range(np.shape(self.ratiometric_stack)[2] - 1):
                    sharp.append(np.std(self.native_stack[:, :, idx]))
                sharp = np.convolve(sharp, [1, 1, 1], mode="valid")
                sharp = sharp / np.max(sharp)
                sharp = sharp ** 2
                sharp = list(sharp) + [1, 1, 1]
                self.sharpness = sharp
            except Exception:
                print(traceback.format_exc())
            self.max_contrast = np.max(np.abs(self.ratiometric_stack))
            self.events = pickle.load(file)
            self.resolution = pickle.load(file)
            self.canvas_offset = pickle.load(file)
            self.scale_factor = pickle.load(file)
            self.slider_native['to'] = pickle.load(file)
            self.slider_ratio['to'] = pickle.load(file)
            self.slider_proj['to'] = pickle.load(file)
            nfv = pickle.load(file)
            rfv = pickle.load(file)
            pfv = pickle.load(file)
            self.internal_window = pickle.load(file)
            self.internal_frame_average = pickle.load(file)

            self.spatial_mask_image = pickle.load(file)
            self.spatial_mask_binary = pickle.load(file)
            self.temporal_mask_timeline = pickle.load(file)
            self.spatial_mask = pickle.load(file)
            self.temporal_mask = pickle.load(file)

            self.filter_r2_threshold = pickle.load(file)
            self.filter_grad_threshold = pickle.load(file)
            self.filter_snr_threshold = pickle.load(file)
            self.current_trace_event = pickle.load(file)

            self.pre_avg.set(pickle.load(file))
            self.use_median.set(pickle.load(file))
            self.window_size.set(pickle.load(file))
            self.offset.set(pickle.load(file))
            self.ratio_norm.set(pickle.load(file))
            self.persistence.set(pickle.load(file))
            self.enable_adsorption.set(pickle.load(file))
            self.enable_desorption.set(pickle.load(file))
            self.show_boxes.set(pickle.load(file))

            self.min_entry.delete(0, tk.END)
            self.min_entry.insert(0, pickle.load(file))
            self.max_entry.delete(0, tk.END)
            self.max_entry.insert(0, pickle.load(file))
            _bins_ = pickle.load(file)
            self.bins.set(_bins_)
            self.use_mass.set(pickle.load(file))
            self.field_list = pickle.load(file)
            self.manual_fit_selection.set_menu(self.field_list[0], *self.field_list)
            self.number_of_fits.set(str(pickle.load(file)))

            self.calibration = pickle.load(file)
            root.mass_entry['state'] = tk.NORMAL
            root.mass_entry.delete(0, tk.END)
            root.mass_entry.insert(0, f'G: {round(self.calibration[0], 1)}, I: {round(self.calibration[1], 1)}')
            root.mass_entry['state'] = tk.DISABLED

            self.lap_centre_sigma.set(pickle.load(file))
            self.lap_outer_sigma.set(pickle.load(file))
            self.lap_magnitude.set(pickle.load(file))
            self.lap_contrast.set(pickle.load(file))

            self.detection_threshold.set(pickle.load(file))
            self.nearest_neighbours.set(pickle.load(file))
            self.min_sigma.set(pickle.load(file))
            self.max_sigma.set(pickle.load(file))
            self.min_intensity.set(pickle.load(file))
            self.eccentricity.set(pickle.load(file))
            self.min_gauss.set(pickle.load(file))

            self.use_global_norm.set(pickle.load(file))
            self.use_low_pass.set(pickle.load(file))
            self.average_mode.set(pickle.load(file))
            self.use_laplacian.set(pickle.load(file))
            self.extended_trace.set(pickle.load(file))
            self.extension_amount.set(pickle.load(file))

            # Tracks and UI parameters
            self.tracks = pickle.load(file)
            self.track_contrasts = pickle.load(file)
            self.track_masses = pickle.load(file)
            if len(self.tracks) > 0:
                self.current_track['state'] = tk.NORMAL
                self.current_track.set(pickle.load(file))
                self.current_track['to'] = len(self.tracks)
                self.update_tracks()
                self.plot_track_masses()
            else:
                null = pickle.load(file)

            self.max_displacement.delete(0, tk.END)
            self.max_displacement.insert(0, pickle.load(file))
            self.min_path.delete(0, tk.END)
            self.min_path.insert(0, pickle.load(file))
            self.max_dark.delete(0, tk.END)
            self.max_dark.insert(0, pickle.load(file))
            self.pixel_size.delete(0, tk.END)
            self.pixel_size.insert(0, pickle.load(file))
            self.frame_interval.delete(0, tk.END)
            self.frame_interval.insert(0, pickle.load(file))

            self.track_use_temporal.set(pickle.load(file))
            self.track_use_spatial.set(pickle.load(file))
            self.track_use_custom.set(pickle.load(file))

            self.overlay_tracks.set(pickle.load(file))
            self.mass_bins.set(pickle.load(file))
            self.track_use_mass.set(pickle.load(file))
            self.track_override.set(pickle.load(file))

            self.mass_min_disp.set(pickle.load(file))
            self.mass_std_lim.set(pickle.load(file))
            self.mass_min_track.set(pickle.load(file))
            self.file_path = pickle.load(file)
            try:
                self.invert_ratiometric.set(pickle.load(file))
            except Exception:
                print('Old file type - state of ratiometric stack inversion not loaded. Defaulting to disabled.')
                self.invert_ratiometric.set(0)
            try:
                self.binary_mask.set(pickle.load(file))
            except Exception:
                print('Old file type - state of subpixel convolution setting not loaded. Defaulting to disabled.')
                self.binary_mask.set(0)


            try:
                self.sharpness_mc.set(pickle.load(file))
                self.auto_fits = pickle.load(file)
                self.main_bins_mode.set(pickle.load(file))
                self.fit_manually.set(pickle.load(file))
                self.correct_motion_var.set(pickle.load(file))
                self.update_fit_mode(load=True)
            except Exception:
                print('Old file type detected. Defaulting to old bin and fitting settings')
                self.sharpness_mc.set(0)
                self.auto_fits = []
                self.main_bins_mode.set(1)
                self.fit_manually.set(1)
                self.correct_motion_var.set(0)
                self.update_fit_mode(load=True)

            try:
                self.use_plateau_contrasts.set(pickle.load(file))
                self.use_chung.set(pickle.load(file))
                self.time_rolling_win_size.set(pickle.load(file))
                self.time_contrast_threshold.set(pickle.load(file))
                self.time_min_plateau.set(pickle.load(file))
                self.time_plateau_std_filter.set(pickle.load(file))
            except Exception:
                self.use_plateau_contrasts.set(0)
                self.use_chung.set(1)
                self.time_rolling_win_size.set(4)
                self.time_contrast_threshold.set(0.001)
                self.time_min_plateau.set(8)
                self.time_plateau_std_filter.set(0.01)
                print('Old file type - Cannot load time series filtering params.')

            self.bins_mode(plot_type='initialisation')
            self.bins.set(_bins_)

            self.window.title(f"OpenMASS Mass Photometry Analysis {VERSION} - '{self.file_path}'")

            if len(self.tracks) > 0:
                self.update_tracks()
                self.plot_track_masses()

            self.progress_win.progress.step(9)
            self.progress_win.window.update()
        try:
            print("ratiometric size for export:", np.shape(self.ratiometric_stack))
            self.export_start_frame['to'] = np.shape(self.ratiometric_stack)[2] - 2
            self.export_end_frame['from_'] = 1
            self.export_end_frame['to'] = np.shape(self.ratiometric_stack)[2]
            self.export_end_frame.set(np.shape(self.ratiometric_stack)[2])
        except Exception:
            print(traceback.format_exc())

        self.native_frame_var.set(nfv)
        self.ratio_frame_var.set(rfv)
        self.proj_frame_var.set(pfv)
        self.display_frame_native(nfv)
        self.display_frame_ratio(rfv)
        self.display_frame_proj(pfv)

        sharp = []
        for idx in range(np.shape(self.ratiometric_stack)[2] - 1):
            sharp.append(np.std(self.native_stack[:, :, idx]))
        sharp = np.convolve(sharp, [1, 1, 1], mode='valid')
        sharp = sharp / np.max(sharp)
        sharp = sharp ** 2
        sharp = list(sharp) + [1, 1, 1]
        self.sharpness = sharp

        self.plot_spatial()
        self.plot_temporal()
        try:
            self.filter_plot_mode.set(1)
            self.filter_plot_decide()
        except Exception:
            self.filter_figure.clf()
            self.filter_canvas.draw()

        for idx in range(len(self.spatial_mask)):
            x1, y1 = self.spatial_mask[idx][1][0] - self.spatial_mask[idx][0][0], self.spatial_mask[idx][1][1] - self.spatial_mask[idx][0][1]
            spatial_rect = mplib.patches.Rectangle(tuple(self.spatial_mask[idx][0]), x1, y1, linewidth=1, edgecolor=(0.35, 0.08, 1, 0.7), facecolor=(0.45, 0, 1, 0.5))
            self.spatial_plotter.add_patch(spatial_rect)
        self.spatial_canvas.draw()
        self.plot_trace()

        self.convert_masses()
        self.plot_histogram(None)
        self.progress_win.handle_close()
        self.calculate_laplacian_kernel()
        self.display_laplacian()

    def set_menu_state(self):
        if self.menu_state.get() == 0:
            self.menu_tile.place_forget()
        else:
            self.menu_tile.place(x=140, y=5, width=1770, height=44)

    def correct_drift(self):
        if preferences['warn']['drift']:
            choice = easygui.indexbox(title="Warning!", msg="Drift correction causes systematic mass shift. Only analyse drift corrected movies using a drift corrected mass calibration.",
                                      choices=['Proceed to dirft correction', 'Cancel'], default_choice=1, cancel_choice=1)
        else:
            choice = 0
        if choice == 0 and self.native_stack is not None:
            self.drift_win = DriftCorrectionWin()

    def create_calibration(self):
        try:
            self.mass_calib_win.handle_close()
        except Exception:
            'window not open'
        means = []
        if self.fit_manually.get() == 1:
            number = int(float(self.number_of_fits.get()))
            for idx in range(number):
                try:
                    if float(self.field_list[idx].split(",")[1][1:]) > 0:
                        means.append(float(self.field_list[idx].split(",")[1][1:]))
                except Exception:
                    'Failed'
        else:
            for idx in range(len(self.auto_fits)):
                if self.auto_fits[idx][1][0] > 0:
                    means.append(self.auto_fits[idx][1][0])
            means = sorted(means)
        self.mass_calib_win = MassCalibWin(means, len(means))

    def update_calibration_dsf(self):
        try:
            root.mass_entry['state'] = tk.NORMAL
            root.mass_entry.delete(0, tk.END)
            root.mass_entry.insert(0, f'G: {round(self.calibration[0], 1)}, I: {round(self.calibration[1], 1)}')
            root.mass_entry['state'] = tk.DISABLED
        except Exception:
            traceback.format_exc()
            easygui.msgbox(title='Error!', msg='Unable to load mass calibration from dsf file.')
            return
        try:
            self.convert_masses()
            if self.use_mass.get() == 1:
                self.min_entry.delete(0, tk.END)
                self.max_entry.delete(0, tk.END)
                self.min_entry.insert(0, round(np.min(self.masses) - 50, 2))
                self.max_entry.insert(0, round(np.max(self.masses) + 50, 2))
            self.update_trace_labels()
            self.display_frame_ratio(index=int(float(self.ratio_frame_var.get())))
        except Exception:
            traceback.format_exc()
        try:
            self.plot_track_masses()
        except Exception:
            ''' Failed '''

    def load_calibration(self):
        path = easygui.fileopenbox(title='Load mass calibration', filetypes=['*.mc'], default='N://*.mc')
        if path:
            try:
                with open(path, 'rb') as file:
                    calib = pickle.load(file)
                    try:
                        calib_type = pickle.load(file)
                    except Exception:
                        ''' No binary mask  / subpixel convolution state '''
                        calib_type = 0
                if len(self.events) > 0 and calib_type != self.binary_mask.get():
                    if calib_type == 1:
                        choice = easygui.indexbox(title='Warning!', msg="This mass calibration was created with subpixel convolution masking enabled. The current data has been analysed "
                                                                        "using the standard mode. Using this mass calibration may indroduce error if the data is not reanalysed with subpixel "
                                                                        "convolution masking enabled.", choices=['Apply and Reanalyse', 'Cancel'], cancel_choice=1)
                    elif calib_type == 0:
                        choice = easygui.indexbox(title='Warning!', msg="This mass calibration was created without subpixel convolution masking enabled. The current data has been analysed "
                                                                        "with subpixel convolution masking enabled. Using this mass calibration may indroduce error if the data is not reanalysed "
                                                                        "with subpixel convolution masking disabled.", choices=['Apply and Reanalyse', 'Cancel'], cancel_choice=1)
                    if choice == 1:
                        return
                    else:
                        self.delete_fits()
                        self.binary_mask.set(calib_type)
                        self.get_contrasts_and_masses()
                self.calibration = calib
                root.mass_entry['state'] = tk.NORMAL
                root.mass_entry.delete(0, tk.END)
                root.mass_entry.insert(0, f'G: {round(self.calibration[0], 1)}, I: {round(self.calibration[1], 1)}')
                root.mass_entry['state'] = tk.DISABLED
            except Exception:
                print(traceback.format_exc())
                easygui.msgbox(title='Error!', msg='Unable to load mass calibration file. Only mass calibrations created by this program can be loaded. Mass '
                                                   'calibrations created by Refeyn DiscoverMP cannot be loaded in OpenMASS.')
                return
            try:
                self.convert_masses()
                if self.use_mass.get() == 1:
                    self.min_entry.delete(0, tk.END)
                    self.max_entry.delete(0, tk.END)
                    self.min_entry.insert(0, round(np.min(self.masses)-50, 2))
                    self.max_entry.insert(0, round(np.max(self.masses)+50, 2))
                self.plot_histogram(None)
                self.update_trace_labels()
                self.display_frame_ratio(index=int(float(self.ratio_frame_var.get())))
            except Exception:
                traceback.format_exc()
            try:
                self.plot_track_masses()
            except Exception:
                ''' Failed '''
            easygui.msgbox(title='Success!', msg='Mass calibration applied!')

    def save_calibration(self):
        if self.calibration == [1, 0]:
            easygui.msgbox(title='Unable to save!', msg='No mass calibration has been created this session. Please create one to proceed.')
            return
        path = easygui.filesavebox(title='Save mass calibration', filetypes=['*.mc'], default='N://*.mc')
        if not path[-3:] == '.mc':
            path = path + '.mc'
        if path:
            with open(path, 'wb') as file:
                pickle.dump(self.calibration, file)
                pickle.dump(self.binary_mask.get(), file)

    def convert(self):
        if self.use_mass.get() == 1:
            cmin, cmax = float(self.min_entry.get()), float(self.max_entry.get())
            cmin, cmax = round(cmin*self.calibration[0]-self.calibration[1], 1), round(cmax*self.calibration[0] + self.calibration[1], 1)
            self.min_entry.delete(0, tk.END)
            self.max_entry.delete(0, tk.END)
            self.min_entry.insert(tk.END, cmin)
            self.max_entry.insert(tk.END, cmax)
        else:
            cmin, cmax = float(self.min_entry.get()), float(self.max_entry.get())
            cmin, cmax = round((cmin+self.calibration[1]) / self.calibration[0], 6), round((cmax-self.calibration[1]) / self.calibration[0], 6)
            self.min_entry.delete(0, tk.END)
            self.max_entry.delete(0, tk.END)
            self.min_entry.insert(tk.END, cmin)
            self.max_entry.insert(tk.END, cmax)
        self.bins_mode()

    def bind_mouse(self, field):
        if len(self.events) > 0:
            self.capture_mouse = self.figure.canvas.mpl_connect('button_press_event', self.get_init_fit_point)

    def figure_vline(self, event):
        try:
            self.figure_vline_widget.remove()
        except Exception:
            """ Can't """
        if len(self.events) > 0 or self.imported_data:
            try:
                self.figure_vline_widget = self.plotter.axvline(event.xdata, linestyle='--', linewidth=1, color='orange')
            except Exception:
                """ Failed to plot """
        self.canvas.draw()

    def draw_rectangle(self, event):
        x = event.xdata
        try:
            self.drag_rect.remove()
        except Exception:
            """ failed """
        name = self.manual_component.get()
        index = self.field_list.index(name)
        try:
            x1, y1 = float(self.field_list[index].split(",")[1][1:]), float(self.field_list[index].split(",")[2][1:])
            if self.use_mass.get() == 1:
                if x1 > 0:
                    x1 = x1 * self.calibration[0] + self.calibration[1]
                else:
                    x1 = x1 * self.calibration[0] - self.calibration[1]
            self.drag_rect = self.plotter.axvspan(x1, x, facecolor='#44aaff', alpha=0.5)
            self.canvas.draw()
        except Exception:
            """ failed """

    def get_init_fit_point(self, event):
        self.figure.canvas.mpl_disconnect(self.capture_mouse)
        self.figure.canvas.mpl_disconnect(self.capture_motion)
        self.capture_mouse = self.figure.canvas.mpl_connect('button_release_event', self.get_final_fit_point)
        self.capture_motion = self.figure.canvas.mpl_connect('motion_notify_event', self.draw_rectangle)
        mouse_xpos, mouse_ypos = event.xdata, event.ydata
        if self.use_mass.get() == 1:
            if mouse_xpos > 0:
                mouse_xpos = (mouse_xpos-self.calibration[1]) / self.calibration[0]
                mouse_ypos = mouse_ypos * self.calibration[0]
            else:
                mouse_xpos = (mouse_xpos + self.calibration[1]) / self.calibration[0]
                mouse_ypos = mouse_ypos * self.calibration[0]
        mouse_xpos, mouse_ypos = round(mouse_xpos, 6), round(mouse_ypos, 6)

        name = self.manual_component.get()
        index = self.field_list.index(name)
        self.field_list[index] = f"Component {index + 1}, {mouse_xpos}, {mouse_ypos}"
        self.manual_fit_selection.set_menu(self.field_list[index], *self.field_list)

    def get_final_fit_point(self, event):
        self.figure.canvas.mpl_disconnect(self.capture_mouse)
        self.figure.canvas.mpl_disconnect(self.capture_motion)
        self.capture_motion = self.figure.canvas.mpl_connect('motion_notify_event', self.figure_vline)
        mouse_xpos, mouse_ypos = event.xdata, event.ydata
        if self.use_mass.get() == 1:
            if mouse_xpos > 0:
                mouse_xpos = (mouse_xpos-self.calibration[1]) / self.calibration[0]
                mouse_ypos = mouse_ypos * self.calibration[0]
            else:
                mouse_xpos = (mouse_xpos + self.calibration[1]) / self.calibration[0]
                mouse_ypos = mouse_ypos * self.calibration[0]
        mouse_xpos, mouse_ypos = round(mouse_xpos, 6), round(mouse_ypos, 6)

        name = self.manual_component.get()
        index = self.field_list.index(name)

        x1, y1 = float(self.field_list[index].split(",")[1][1:]), float(self.field_list[index].split(",")[2][1:])
        mean = round((x1 + mouse_xpos)/2, 6)
        amp = round((y1 + mouse_ypos)/2, 6)
        std = round(abs(mean - x1)/2, 6)

        self.field_list[index] = f"Component {index + 1}, {mean}, {std}, {amp}"
        self.manual_fit_selection.set_menu(self.field_list[index], *self.field_list)
        self.plot_histogram(None)

    def get_skew_init_point(self, event):
        if self.plotter.get_navigate_mode() is not None:
            return
        self.figure.canvas.mpl_disconnect(self.capture_motion)
        self.figure.canvas.mpl_disconnect(self.capture_mouse)
        self.capture_mouse = self.figure.canvas.mpl_connect('button_release_event', self.get_skew_final_point)
        self.capture_motion = self.figure.canvas.mpl_connect('motion_notify_event', self.draw_skew_rectangle)
        self.temp_xmin = event.xdata

    def get_skew_final_point(self, event):
        self.figure.canvas.mpl_disconnect(self.capture_mouse)
        self.figure.canvas.mpl_disconnect(self.capture_motion)
        self.capture_motion = self.figure.canvas.mpl_connect('motion_notify_event', self.figure_vline)
        self.capture_mouse = self.figure.canvas.mpl_connect('button_press_event', self.get_skew_init_point)
        self.temp_xmax = event.xdata
        if self.temp_xmax < self.temp_xmin:
            self.temp_xmin, self.temp_xmax = self.temp_xmax, self.temp_xmin
        try:
            self.drag_rect.remove()
        except Exception:
            """ failed """

        if self.use_mass.get() == 1:
            if self.temp_xmin > 0:
                self.temp_xmin = (self.temp_xmin - self.calibration[1]) / self.calibration[0]
            else:
                self.temp_xmin = (self.temp_xmin + self.calibration[1]) / self.calibration[0]
            if self.temp_xmax > 0:
                self.temp_xmax = (self.temp_xmax - self.calibration[1]) / self.calibration[0]
            else:
                self.temp_xmax = (self.temp_xmax + self.calibration[1]) / self.calibration[0]
        data_to_fit = []
        if self.track_override.get() == 0:
            for idx, datum in enumerate(self.contrasts):
                if self.temp_xmin <= datum <= self.temp_xmax:
                    data_to_fit.append(datum)
        else:
            for idx, datum in enumerate(self.track_contrasts):
                if self.temp_xmin <= datum <= self.temp_xmax:
                    data_to_fit.append(datum)
        if self.main_bins_mode.get() == 2:
            bin_width = float(self.bins.get())
            if self.use_mass.get() == 1:
                bin_width = bin_width / self.calibration[0]
        else:
            bin_width = 0.0002
        bins = int((self.temp_xmax - self.temp_xmin) / bin_width)
        if self.track_override.get() == 0:
            amp_factor = len(data_to_fit) / len(self.contrasts)
        else:
            amp_factor = len(data_to_fit) / len(self.track_contrasts)

        if self.fit_mode.get() == 1:
            self.fit_multiple(data_to_fit, bins, amp_factor)
        elif self.fit_mode.get() == 2:
            self.fit_single(data_to_fit, bins, amp_factor)
        elif self.fit_mode.get() == 3:
            self.advanced_skew_fit(data_to_fit, bins, amp_factor)

    def gaussfit_options(self):
        self.gauss_opt_win = GaussOptWin()

    def advanced_skew_fit(self, data, bins, amp_factor):
        try:
            self.gauss_preview_win.handle_close(auto=True)
        except Exception:
            print(traceback.format_exc())
        try:
            params = self.gauss_preview_win.UI_params
        except Exception:
            print(traceback.format_exc())
            params = ('auto', gauss_settings['max fits'], gauss_settings['maxiter'], gauss_settings['penalty'], gauss_settings['optimizer'], gauss_settings['n_init'])
        self.gauss_preview_win = GaussPreviewWin(data, bins, amp_factor, params)

    def fit_multiple(self, data, bins, amp_factor):
        self.check_fit_replace()
        error = False
        err_trace = ""
        fits = []

        try:
            params = self.fit_hist_data(data, comps='auto', bins=bins, penalty=gauss_settings['penalty'], callback=None,
                                        max_iter=gauss_settings['maxiter'],
                                        n_init=gauss_settings['n_init'],
                                        max_comp=gauss_settings['max fits'],
                                        optimizer=gauss_settings['optimizer'],
                                        )
            for idx in range(len(params)):
                fits.append(params[idx] + [[1, 1, 1, 1]])
                fits[-1][0][2] = fits[-1][0][2] * amp_factor
        except Exception:
            err_trace = traceback.format_exc()
            error = True

        for fit in fits:
            self.auto_fits.append(fit)
        self.plot_histogram(None)

        if error:
            print(err_trace)
            if preferences['warn']['fiterr']:
                easygui.msgbox(title='Error', msg=f'Error occurred while trying to autofit histogram:\n{str(err_trace)}')

        for idx in range(len(self.auto_fits)):
            print(self.auto_fits[idx])

    def fit_single(self, data, bins, amp_factor):
        self.check_fit_replace()
        error = False
        err_trace = ""
        fits = []

        try:
            params = self.fit_hist_data(data, comps=1, bins=bins, penalty=gauss_settings['penalty'], callback=None,
                                        max_iter=gauss_settings['maxiter'],
                                        n_init=gauss_settings['n_init'],
                                        max_comp=gauss_settings['max fits'],
                                        optimizer=gauss_settings['optimizer'],
                                        )
            for idx in range(len(params)):
                fits.append(params[idx])
                fits[-1][0][2] = fits[-1][0][2] * amp_factor
        except Exception:
            err_trace = traceback.format_exc()
            error = True

        for fit in fits:
            self.auto_fits.append(fit)
        self.plot_histogram(None)

        if error:
            print(err_trace)
            if preferences['warn']['fiterr']:
                easygui.msgbox(title='Error', msg=f'Error occurred while trying to fit histogram:\n{str(err_trace)}')

        for idx in range(len(self.auto_fits)):
            print(self.auto_fits[idx])

    def check_fit_replace(self):
        new_list = []
        for idx in range(len(self.auto_fits)):
            if -3 <= self.auto_fits[idx][0][3] <= 3:
                fit_range = [self.auto_fits[idx][1][0] - 2*self.auto_fits[idx][1][1], self.auto_fits[idx][1][0] + 2*self.auto_fits[idx][1][1]]
            elif self.auto_fits[idx][0][3] < -3:
                fit_range = [self.auto_fits[idx][1][0] - 3 * self.auto_fits[idx][1][1], self.auto_fits[idx][1][0] + 1 * self.auto_fits[idx][1][1]]
            elif self.auto_fits[idx][0][3] > 3:
                fit_range = [self.auto_fits[idx][1][0] - 1 * self.auto_fits[idx][1][1], self.auto_fits[idx][1][0] + 3 * self.auto_fits[idx][1][1]]
            if not self.temp_xmin < self.auto_fits[idx][1][0] < self.temp_xmax and not self.temp_xmin < self.auto_fits[idx][0][0] < self.temp_xmax\
                    and not ((fit_range[0] < self.temp_xmin < fit_range[0]) or (fit_range[0] < self.temp_xmax < fit_range[0])):
                new_list.append(self.auto_fits[idx])
        self.auto_fits = []
        self.auto_fits = new_list

    def draw_skew_rectangle(self, event):
        x = event.xdata
        try:
            self.drag_rect.remove()
        except Exception:
            """ failed """
        self.drag_rect = self.plotter.axvspan(self.temp_xmin, x, facecolor='#44aaff', alpha=0.5)
        self.canvas.draw()

    def delete_fits(self, auto=False):
        if self.fit_manually.get() == 1:
            self.field_list = ['Component 1,', 'Component 2,', 'Component 3,', 'Component 4,', 'Component 5,', 'Component 6,', 'Component 7,', 'Component 8,']
            self.manual_fit_selection.set_menu('select component', *self.field_list)
        else:
            self.auto_fits = []
        if not auto:
            self.plot_histogram(None)

    def update_fit_mode(self, load=False):
        self.figure.canvas.mpl_disconnect(self.capture_mouse)
        self.figure.canvas.mpl_disconnect(self.capture_motion)
        if self.fit_manually.get() == 0:
            if len(self.events) > 0 or self.imported_data:
                self.capture_mouse = self.figure.canvas.mpl_connect('button_press_event', self.get_skew_init_point)
                self.manual_fit_selection['state'] = tk.DISABLED
                self.number_selection['state'] = tk.DISABLED
        else:
            if len(self.events) > 0 or self.imported_data:
                self.manual_fit_selection['state'] = tk.NORMAL
                self.number_selection['state'] = tk.NORMAL
            if not load:
                if preferences['warn']['gauss']:
                    easygui.msgbox(title="Warning!", msg="Caution: Manually fitting Gaussians isn't true fitting. It draws a Gaussian fit from -2 to 2 standard deviations according to the region selected and position of the mouse and does not actually fit the data. This method is not objective and we do not recommend using it unless the skewed GMM fit fails completely, and even then only illustratively.")
        if not load:
            self.plot_histogram(None)

    def return_moments(self):
        if self.fit_manually.get() == 1:
            components = int(float(self.number_of_fits.get()))
            moments = []
            for idx in range(components):
                try:
                    mean, std, amp = float(self.field_list[idx].split(",")[1][1:]), float(self.field_list[idx].split(",")[2][1:]), float(self.field_list[idx].split(",")[3][1:])
                    if mean > 0:
                        moments.append([(amp) / self.calibration[0], mean * self.calibration[0] + self.calibration[1], std * self.calibration[0]])
                    else:
                        moments.append([(amp) / self.calibration[0], mean * self.calibration[0] - self.calibration[1], std * self.calibration[0]])
                except Exception:
                    'fit not available'
                    print(traceback.format_exc())
        else:
            moments = []
            for idx in range(len(self.auto_fits)):
                mean, std, amp, skew = self.auto_fits[idx][0]
                if mean > 0:
                    moments.append([(amp) / self.calibration[0], mean * self.calibration[0] + self.calibration[1], std * self.calibration[0], skew])
                else:
                    moments.append([(amp) / self.calibration[0], mean * self.calibration[0] - self.calibration[1], std * self.calibration[0], skew])
        return moments

    def return_fits(self, track=True):
        if self.fit_manually.get() == 1:
            components = int(float(self.number_of_fits.get()))
            moments = []
            for idx in range(components):
                try:
                    mean, std, amp = float(self.field_list[idx].split(",")[1][1:]), float(self.field_list[idx].split(",")[2][1:]), float(self.field_list[idx].split(",")[3][1:])
                    if mean > 0:
                        moments.append([mean*self.calibration[0]+self.calibration[1], std*self.calibration[0], (amp)/self.calibration[0]])
                    else:
                        moments.append([mean * self.calibration[0] - self.calibration[1], std * self.calibration[0], (amp) / self.calibration[0]])
                except Exception:
                    'fit not available'
                    print(traceback.format_exc())
        else:
            moments = []
            for idx in range(len(self.auto_fits)):
                mean, std, amp, skew = self.auto_fits[idx][0]
                if mean > 0:
                    moments.append([mean * self.calibration[0] + self.calibration[1], std * self.calibration[0], (amp) / self.calibration[0], skew])
                else:
                    moments.append([mean * self.calibration[0] - self.calibration[1], std * self.calibration[0], (amp) / self.calibration[0], skew])
        try:
            if self.use_mass.get():
                start = float(self.min_entry.get())
                end = float(self.max_entry.get())
            else:
                start = float(self.min_entry.get()) * self.calibration[0]
                end = float(self.max_entry.get()) * self.calibration[0]
            x_axis = np.linspace(start, end, 2000)
        except ValueError:
            print("No histogram data, cannot get min / max in 'get_fits()'")
        fits = []
        fit_func = None
        for fit_index in range(len(moments)):
            if self.fit_manually.get() == 1:
                fit_func = self.gauss
                fit = self.gauss(x_axis, moments[fit_index][2], moments[fit_index][0], moments[fit_index][1])
            else:
                fit_func = sgmm.skew_gaussian
                fit = sgmm.skew_gaussian(x_axis, moments[fit_index][2], moments[fit_index][0], moments[fit_index][1], moments[fit_index][3])
            fits.append(fit)
        if len(fits) >= 1:
            total_fit = np.copy(fits[0])
            for idx in range(1, len(fits)):
                total_fit += fits[idx]
            fits.append(total_fit)
        return x_axis, fits, fit_func

    def get_fits(self):
        if self.fit_manually.get() == 1:
            components = int(float(self.number_of_fits.get()))
            moments = []
            for idx in range(components):
                try:
                    mean, std, amp = float(self.field_list[idx].split(",")[1][1:]), float(self.field_list[idx].split(",")[2][1:]), float(self.field_list[idx].split(",")[3][1:])
                    if self.use_mass.get() == 1:
                        if mean > 0:
                            moments.append([mean*self.calibration[0]+self.calibration[1], std*self.calibration[0], (amp)/self.calibration[0]])
                        else:
                            moments.append([mean * self.calibration[0] - self.calibration[1], std * self.calibration[0], (amp) / self.calibration[0]])
                    else:
                        moments.append([mean, std, amp])
                except Exception:
                    'fit not available'
                    print(traceback.format_exc())
        else:
            moments = []
            for idx in range(len(self.auto_fits)):
                if self.use_mass.get() == 0:
                    moments.append(self.auto_fits[idx][0])
                else:
                    mean, std, amp, skew = self.auto_fits[idx][0]
                    if mean > 0:
                        moments.append([mean * self.calibration[0] + self.calibration[1], std * self.calibration[0], (amp) / self.calibration[0], skew])
                    else:
                        moments.append([mean * self.calibration[0] - self.calibration[1], std * self.calibration[0], (amp) / self.calibration[0], skew])
        try:
            if self.use_mass.get() == 1:
                start = np.min(self.masses)
                end = np.max(self.masses)
            else:
                start = np.min(self.contrasts)
                end = np.max(self.contrasts)
            self.x_axis = np.linspace(start, end, 10000)
        except ValueError:
            print("No histogram data, cannot get min / max in 'get_fits()'")
        self.fits = []
        for fit_index in range(len(moments)):
            if self.fit_manually.get() == 1:
                fit = self.gauss(self.x_axis, moments[fit_index][2], moments[fit_index][0], moments[fit_index][1])
            else:
                fit = sgmm.skew_gaussian(self.x_axis, moments[fit_index][2], moments[fit_index][0], moments[fit_index][1], moments[fit_index][3])
            self.fits.append(fit)
        if len(self.fits) >= 1:
            total_fit = np.copy(self.fits[0])
            for idx in range(1, len(self.fits)):
                total_fit += self.fits[idx]
            self.fits.append(total_fit)

    def gauss(self, x, a, x0, sigma):
        return a * np.exp(-(x - x0) ** 2 / (2 * sigma ** 2))

    def import_contrast_data(self):
        paths = easygui.fileopenbox(msg="Open contrast data files (.pydat).", default="NN:/*.pydat", multiple=True)
        if not paths:
            return
        self.new_exp(load=True)
        self.imported_data = True
        for path in paths:
            with open(path, 'rb') as file:
                data = pickle.load(file)
                assert isinstance(data, list), TypeError("Data must be a list of contrast values.")
                self.contrasts += data
        for c in self.contrasts:
            self.masses.append(c * self.calibration[0] + self.calibration[1])
        self.update_fit_mode(load=True)
        self.analysis_notebook.select(self.analysis_tab2)
        self.plot_histogram(None)

    def import_ratiometric(self):
        filetypes = [
            ["*.tiff;*.tif", "Supported Files (*.tiff *.tif)"]
            ]
        path = easygui.fileopenbox(msg="Open Mass Photometry file.", filetypes=filetypes, default="NN:/*.tiff;*.tif")
        if path:
            try:
                self.new_exp(load=True)
            except Exception:
                print('An error occured while clearing data and interface:')
                print(traceback.format_exc())
            if path[-5:] == '.tiff' or path[-4:] == '.tif':
                self.file_path = path
                self.window.title(f"OpenMASS Mass Photometry Analysis {VERSION} - '{path}'")
                try:
                    self.progress_win.window.destroy()
                except Exception:
                    pass
                self.progress_win = ProgressWin(title='Loading data', msg='Opening IFDs...')
                self.event_label['text'] = 'Opening IFDs...'
                self.event_label.update()
                raw_tif = Image.open(path)
                h, w = np.shape(raw_tif)
                self.resolution = (h, w)
                self.scale_factor = 136 / h
                self.canvas_offset = int(256 - w * self.scale_factor / 2)
                tif_array = np.zeros((h, w, raw_tif.n_frames), dtype='float16')
                self.progress_win.label['text'] = 'Reading frames...'
                self.progress_win.progress['maximum'] = raw_tif.n_frames + 1
                for index in range(raw_tif.n_frames):
                    raw_tif.seek(index)
                    tif_array[:, :, index] = np.array(raw_tif)
                    if index % 10 == 0:
                        self.progress_win.progress.step(10)
                        self.event_label['text'] = f'Reading... ({index}/{raw_tif.n_frames})'
                        self.event_label.update()
                self.progress_win.handle_close()

                self.ratiometric_stack = tif_array.astype(np.float32)
                start, end = 0, -1
                for idx in range(1, raw_tif.n_frames):
                    if np.sum(self.ratiometric_stack[:, :, idx - 1]) == 0 and np.sum(self.ratiometric_stack[:, :, idx]) != 0:
                        start = idx
                    if np.sum(self.ratiometric_stack[:, :, idx]) == 0 and np.sum(self.ratiometric_stack[:, :, idx - 1]) != 0:
                        end = idx
                self.ratiometric_stack = self.ratiometric_stack[:, :, start:end]
                self.event_label['text'] = ''
                self.event_label.update()
                self.slider_ratio['to'] = np.shape(self.ratiometric_stack)[2] - 1
                self.display_frame_ratio(0)
                self.ratio_frame_var.set(0)

                shape = np.shape(self.ratiometric_stack)[:2]
                h, w = shape
                projection = np.zeros(shape=shape)
                tif_limit = np.abs(np.clip(self.ratiometric_stack, -100, 0))
                max_index_matrix = np.argmax(tif_limit, axis=2)
                for x in range(w):
                    for y in range(h):
                        projection[y, x] = tif_limit[y, x, max_index_matrix[y, x]]
                tif_limit = np.abs(np.clip(self.ratiometric_stack, 0, 100))
                max_index_matrix = np.argmax(tif_limit, axis=2)
                for x in range(w):
                    for y in range(h):
                        projection[y, x] += tif_limit[y, x, max_index_matrix[y, x]]
                self.spatial_mask_image = projection

                self.export_start_frame['to'] = np.shape(self.ratiometric_stack)[2] - 2
                self.export_end_frame['from_'] = 1
                self.export_end_frame['to'] = np.shape(self.ratiometric_stack)[2] - 1
                self.temporal_mask_timeline = np.zeros(shape=np.shape(self.ratiometric_stack)[2])
                self.temporal_mask = []
                self.spatial_mask_binary = np.zeros(shape=shape)
                self.spatial_mask = []
                self.plot_temporal()
                self.plot_spatial()


    def load_data(self):
        filetypes = [
            ["*.tiff;*.tif;*.mp;*.h5", "Supported Files (*.tiff *.tif *.mp *.h5)"],
            ["*.tiff", "TIFF Files"],
            ["*.tif", "TIF Files"],
            ["*.mp", "MP Files"],
            ["*.h5", "H5 Files"],
            ["*.*", "All Files"]
        ]
        path = easygui.fileopenbox(msg="Open Mass Photometry file.", filetypes=filetypes, default="NN:/*.tiff;*.tif;*.mp;*.h5")
        if path:
            try:
                self.new_exp(load=True)
            except Exception:
                print('An error occured while clearing data and interface:')
                print(traceback.format_exc())
            if path[-5:] == '.tiff' or path[-4:] == '.tif':
                self.file_path = path
                self.window.title(f"OpenMASS Mass Photometry Analysis {VERSION} - '{path}'")
                try:
                    self.progress_win.window.destroy()
                except Exception:
                    pass
                self.progress_win = ProgressWin(title='Loading data', msg='Opening IFDs...')
                self.event_label['text'] = 'Opening IFDs...'
                self.event_label.update()
                raw_tif = Image.open(path)
                # print("TIF file shape:", np.shape(raw_tif))
                h, w = np.shape(raw_tif)
                self.resolution = (h, w)
                self.scale_factor = 136 / h
                self.canvas_offset = int(256 - w*self.scale_factor/2)
                tif_array = np.zeros((h, w, raw_tif.n_frames), dtype='float32')
                self.progress_win.label['text'] = 'Reading frames...'
                self.progress_win.progress['maximum'] = raw_tif.n_frames + 1
                for index in range(raw_tif.n_frames):
                    raw_tif.seek(index)
                    tif_array[:, :, index] = np.array(raw_tif)
                    if index % 10 == 0:
                        self.progress_win.progress.step(10)
                        self.event_label['text'] = f'Reading... ({index}/{raw_tif.n_frames})'
                        self.event_label.update()
                self.progress_win.handle_close()
                self.native_stack = tif_array.astype(np.float32)
            elif path[-3:] == '.mp' or path[-3:] == '.h5':
                self.file_path = path
                self.window.title(f"OpenMASS Mass Photometry Analysis {VERSION} - '{path}'")
                mpfile = h5py.File(path, 'r')
                try:
                    frames = np.asarray(mpfile['frame'])
                except Exception:
                    print("mp file - new version, h5 key = ['movie']['frame']")
                    frames = np.asarray(mpfile['movie']['frame'])
                h, w = np.shape(frames)[1:]
                self.resolution = (h, w)
                self.scale_factor = 136 / h
                self.canvas_offset = int(256 - w * self.scale_factor / 2)
                print(np.shape(frames))
                shape = np.shape(frames)
                self.native_stack = np.zeros(shape=(shape[1], shape[2], shape[0]), dtype=np.float32)
                try:
                    self.progress_win.window.destroy()
                except Exception:
                    pass
                self.progress_win = ProgressWin(title='Loading data', msg='Reading frames...')
                self.progress_win.progress['maximum'] = shape[0] + 1
                for idx in range(shape[0]):
                    self.native_stack[:, :, idx] = frames[idx]
                    if idx % 10 == 0:
                        self.progress_win.progress.step(10)
                        self.event_label['text'] = f'Reading... ({idx}/{shape[0]})'
                        self.event_label.update()
                self.progress_win.handle_close()
            self.event_label['text'] = ''
            self.event_label.update()
            self.slider_native['to'] = np.shape(self.native_stack)[2] - 1
            self.display_frame_native(0)
            self.native_frame_var.set(0)


    def display_frame_native(self, index):
        self.canvas_native.delete('all')
        display_time = int(float(self.persistence.get()))
        index = int(float(index))
        maximum = np.max(self.native_stack[:, :, index])
        minimum = np.min(self.native_stack[:, :, index])
        norm = mplib.colors.Normalize(vmin=minimum, vmax=maximum)
        frame_rgb = mplib.cm.gray(norm(self.native_stack[:, :, index]))
        frame_rgb = frame_rgb[:, :, :3] * 255
        PIL_image = Image.fromarray(frame_rgb.astype('uint8'), 'RGB')
        h, w = self.resolution
        sfx, sfy = int((136 / h) * w), 136
        PIL_image = PIL_image.resize((sfx, sfy), Image.NEAREST)
        global raw_img
        raw_img = ImageTk.PhotoImage(master=self.window, image=PIL_image)
        self.canvas_native.create_image(1+self.canvas_offset, 1, anchor="nw", image=raw_img)
        self.native_frame_label['text'] = str(index)

        if len(self.tracks) > 0 and self.overlay_tracks.get() == 1 and self.use_median.get() == 1:
            current_track = int(float(self.current_track.get())) - 1
            scale = self.scale_factor
            if index >= self.tracks[current_track].frames[0] and index <= self.tracks[current_track].frames[-1]:
                tcs = np.array(self.tracks[current_track].coords)
                xmin, ymin, xmax, ymax = np.min(tcs[:, 0]) - 1, np.min(tcs[:, 1]) - 1, np.max(tcs[:, 0]) + 1, np.max(tcs[:, 1]) + 1
                self.canvas_native.create_rectangle(xmin*scale+self.canvas_offset, ymin*scale, xmax*scale+self.canvas_offset, ymax*scale, outline="#ff6f33", width=2, dash=(4, 2))
            for idx in range(len(self.tracks)):
                col = self.track_cols[idx % len(self.track_cols)]
                if index >= self.tracks[idx].frames[0] and index <= self.tracks[idx].frames[-1]:
                    for i in range(len(self.tracks[idx].frames) - 1):
                        s, e = self.tracks[idx].coords[i], self.tracks[idx].coords[i + 1]
                        self.canvas_native.create_line(scale * s[0] + self.canvas_offset, scale * s[1], scale * e[0] + self.canvas_offset, scale * e[1], fill=col)

        if len(self.events) > 0 and self.show_boxes.get() == 1 and self.use_median.get() == 1:
            for idx in range(len(self.events)):
                if index > self.events[idx][0] - display_time and index < self.events[idx][0] + display_time:
                    col = 'yellow'
                    coords = self.events[idx][2][:2]
                    scale = self.scale_factor
                    self.canvas_native.create_rectangle((coords[0] - 3) * scale + self.canvas_offset, (coords[1] - 3) * scale, (coords[0] + 4) * scale + self.canvas_offset,
                                                             (coords[1] + 4) * scale, outline=col, width=1)

    def update_ratio_contrast(self):
        index = self.ratio_frame_var.get()
        self.display_frame_ratio(index)

    def ratio_popout(self):
        try:
            self.ratio_popout_win.handle_close()
        except Exception:
            """ Window was closed """
        try:
            self.resolution[0]
        except IndexError:
            return
        self.ratio_popout_win = RatioPopoutWin()
        self.is_popped_out = True
        self.display_frame_ratio(int(float(self.ratio_frame_var.get())))

    def display_frame_ratio(self, index):
        global _canvas
        if self.is_popped_out:
            _canvas = self.ratio_popout_win.canvas_ratiometric
        else:
            _canvas = self.canvas_ratiometric
        display_time = int(float(self.persistence.get()))
        self.canvas_ratiometric.delete("all")
        _canvas.delete("all")
        index = int(float(index))
        norm = mplib.colors.Normalize(vmin=-float(self.ratio_norm.get()), vmax=float(self.ratio_norm.get()))
        frame_rgb = mplib.cm.gray(norm(self.ratiometric_stack[:, :, index]))
        frame_rgb = frame_rgb[:, :, :3] * 255
        PIL_image = Image.fromarray(frame_rgb.astype('uint8'), 'RGB')
        h, w = self.resolution
        if self.is_popped_out:
            sfx, sfy = w*5, h*5
            offset = 0
        else:
            sfx, sfy = int((136 / h) * w), 136
            offset = self.canvas_offset
        PIL_image = PIL_image.resize((sfx, sfy), Image.NEAREST)
        global proc_img
        if self.is_popped_out:
            proc_img = ImageTk.PhotoImage(master=self.ratio_popout_win.window, image=PIL_image)
        else:
            proc_img = ImageTk.PhotoImage(master=self.window, image=PIL_image)
        _canvas.create_image(1+offset, 1, anchor="nw", image=proc_img)
        self.ratio_frame_label['text'] = str(index)
        if len(self.tracks) > 0 and self.overlay_tracks.get() == 1:
            current_track = int(float(self.current_track.get())) - 1
            if self.is_popped_out:
                scale = 5
            else:
                scale = self.scale_factor
            if index >= self.tracks[current_track].frames[0] and index <= self.tracks[current_track].frames[-1]:
                tcs = np.array(self.tracks[current_track].coords)
                xmin, ymin, xmax, ymax = np.min(tcs[:, 0]) - 1, np.min(tcs[:, 1]) - 1, np.max(tcs[:, 0]) + 1, np.max(tcs[:, 1]) + 1
                _canvas.create_rectangle(xmin*scale+offset, ymin*scale, xmax*scale+offset, ymax*scale, outline="#ff6f33", width=2, dash=(4, 2))
            for idx in range(len(self.tracks)):
                col = self.track_cols[idx % len(self.track_cols)]
                if index >= self.tracks[idx].frames[0] and index <= self.tracks[idx].frames[-1]:
                    for i in range(len(self.tracks[idx].frames) - 1):
                        s, e = self.tracks[idx].coords[i], self.tracks[idx].coords[i + 1]
                        _canvas.create_line(scale * s[0] + offset, scale * s[1], scale * e[0] + offset, scale * e[1], fill=col)
        if len(self.events) > 0 and self.show_boxes.get() == 1:
            for idx in range(len(self.events)):
                if self.sharpness_mc.get() == 0:
                    factor = 1
                else:
                    factor = self.sharpness[self.events[idx][0]]
                if index > self.events[idx][0] - display_time and index < self.events[idx][0] + display_time:
                    if idx == self.current_trace_event:
                        if self.events[idx][1] / factor in self.contrasts:
                            col = '#0036aa'
                            tc = 'black'
                        else:
                            col = '#990000'
                            tc = 'black'
                    else:
                        if self.events[idx][1] / factor in self.contrasts:
                            col = 'yellow'
                            tc = '#002f99'
                        else:
                            col = '#ff6f00'
                            tc = '#002f99'
                    coords, mass = self.events[idx][2][:2], self.events[idx][1]*self.calibration[0] / factor
                    if mass > 0:
                        mass += self.calibration[1]
                    else:
                        mass -= self.calibration[1]
                    if self.is_popped_out:
                        scale = 5
                    else:
                        scale = self.scale_factor
                    if coords[0] > w - 10:
                        tx_off = -20
                    else:
                        tx_off = 0
                    if coords[1] < 7:
                        ty_off = 11
                    else:
                        ty_off = 0
                    if self.scale_factor < 2:
                        size = 7
                    else:
                        size = 9
                    if self.is_popped_out:
                        size=10
                    _canvas.create_rectangle((coords[0] - 3) * scale + offset, (coords[1] - 3) * scale, (coords[0] + 4) * scale + offset,
                                                             (coords[1] + 4) * scale, outline=col, width=1)
                    _canvas.create_text(round(coords[0] + 9 + tx_off) * scale + offset, round(coords[1] - 4.5 + ty_off) * scale, fill=tc,
                                                        font=f"arial {size}", text=f'{str(round(mass))} kDa')

    def display_frame_proj(self, index):
        index = int(float(index))
        old_index = index
        index = int(index / self.internal_window)
        self.canvas_projection.delete("all")
        frame = np.zeros(shape=np.shape(self.zstack_adsorption[0]))
        if self.enable_adsorption.get() == 1:
            try:
                frame = frame + self.zstack_adsorption[index]
            except Exception:
                print(traceback.format_exc())
        if self.enable_desorption.get() == 1:
            try:
                frame = frame + self.zstack_desorption[index]
            except Exception:
                print(traceback.format_exc())
        maximum = np.max(frame)
        minimum = np.min(frame)
        norm = mplib.colors.Normalize(vmin=minimum, vmax=maximum)
        frame_rgb = mplib.cm.inferno(norm(frame))
        frame_rgb = frame_rgb[:, :, :3] * 255
        PIL_image = Image.fromarray(frame_rgb.astype('uint8'), 'RGB')
        h, w = self.resolution
        sfx, sfy = int((136 / h) * w), 136
        PIL_image = PIL_image.resize((sfx, sfy), Image.NEAREST)
        global zstack_img
        zstack_img = ImageTk.PhotoImage(master=self.window, image=PIL_image)
        self.canvas_projection.create_image(1+self.canvas_offset, 1, anchor="nw", image=zstack_img)
        self.proj_frame_label['text'] = f'F: {old_index} I:{index}'
        if len(self.events) > 0:
            for idx in range(len(self.events)):
                if self.events[idx][0] > index*self.internal_window and self.events[idx][0] <= index*self.internal_window+self.internal_window:
                    coords, mass = self.events[idx][2][:2], self.events[idx][1]*self.calibration[0]+self.calibration[1]
                    scale = self.scale_factor
                    if mass > 0 and self.enable_adsorption.get() == 1:
                        self.canvas_projection.create_rectangle((coords[0] - 3) * scale+self.canvas_offset, (coords[1] - 3) * scale, (coords[0] + 4) * scale+self.canvas_offset,
                                                                 (coords[1] + 4) * scale, outline='#004fff', width=1)
                    if mass < 0 and self.enable_desorption.get() == 1:
                        self.canvas_projection.create_rectangle((coords[0] - 3) * scale+self.canvas_offset, (coords[1] - 3) * scale, (coords[0] + 4) * scale+self.canvas_offset,
                                                                 (coords[1] + 4) * scale, outline='#00afff', width=1)

    def matplotlib_display(self):
        index = int(float(self.proj_frame_var.get()))
        old_index = index
        index = int(index / self.internal_window)
        frame = np.zeros(shape=np.shape(self.zstack_adsorption[0]))
        if self.enable_adsorption.get() == 1:
            frame = frame + self.zstack_adsorption[index]
        if self.enable_desorption.get() == 1:
            frame = frame + self.zstack_desorption[index]
        fsx, fsy = self.resolution
        fsx, fsy = fsx*self.scale_factor/100, fsy*self.scale_factor/100
        fig = plt.figure(figsize=(5.12, 1.36))
        ax = fig.add_subplot(111)
        fig.subplots_adjust(left=0.0, right=1.0, top=1.0, bottom=0.0)
        ax.imshow(frame, cmap=mplib.cm.inferno)
        fig.show()

    def motion_enabled(self):
        if self.correct_motion_var.get() == 1 and preferences['warn']['motion']:
            easygui.msgbox(title='Warning!', msg='Caution (Experimental feature): Motion correction attempts to remove artifacts caused by vibrations during data collection. It is suitable '
                                                 'when masses are less than 1000 kDa and helps improve mass resolution. At higher masses, sinc wiggles caused by the notch filter may become '
                                                 'apparent around high contrast objects. Use with care.')

    def process_new(self):
        avg = int(float(self.pre_avg.get()))
        self.internal_frame_average = avg
        shape = np.shape(self.native_stack)
        length = shape[2]
        try:
            self.progress_win.window.destroy()
        except Exception:
            pass
        self.progress_win = ProgressWin(title='Processing', msg='Generating ratiometric stack...')
        self.progress_win.progress['maximum'] = length - avg*2 + 20

        def update_label(frame_, length_, avg_):
            if frame_ % 20 == 0:
                self.progress_win.progress.step(20)
                self.event_label['text'] = f'Generating Ratiometric Stack... ({frame_}/{length_ - avg_ * 2 + 1})'
                self.event_label.update()
                if self.cancel_flag:
                    return False
            return True

        if self.use_median.get() == 0:
            result = np.zeros((shape[0], shape[1], shape[2] - avg * 2 + 1))
            for frame in range(length - avg*2 + 1):

                win1 = np.zeros((shape[0], shape[1]))
                for add in range(avg):
                    win1 = win1 + self.native_stack[:, :, frame + add]
                win1 = win1 / avg
                win1 = win1 / np.mean(win1)

                win2 = np.zeros((shape[0], shape[1]))
                for add in range(avg):
                    win2 = win2 + self.native_stack[:, :, frame + add + avg]
                win2 = win2 / avg
                win2 = win2 / np.mean(win2)

                result[:, :, frame] = win2 / win1 - 1
                # result[:, :, frame] = win1 - 1
                if frame % 20 == 0:
                    self.progress_win.progress.step(20)
                    self.event_label['text'] = f'Generating Ratiometric Stack... ({frame}/{length - avg*2 + 1})'
                    self.event_label.update()
                    if self.cancel_flag:
                        break
        else:
            self.progress_win.progress['maximum'] = length + 20
            result_array = np.zeros((shape[0], shape[1], shape[2]))
            result = iscam.ratiometric_median(self.native_stack, result_array, avg, length, callback=update_label)
            # for frame in range(avg+1, length - avg - 1):
            #     for x in range(shape[0]):
            #         for y in range(shape[1]):
            #             median = np.median(self.native_stack[x, y, frame - avg:frame + avg])
            #             result[x, y, frame-avg-1] = self.native_stack[x, y, frame] / median - 1
                # plt.imshow(result[:, :, frame-avg-1])
                # plt.show()
                # if frame % 20 == 0:
                #     self.progress_win.progress.step(20)
                #     self.event_label['text'] = f'Generating Ratiometric Stack... ({frame}/{length - avg * 2 + 1})'
                #     self.event_label.update()

        if self.cancel_flag:
            self.progress_win.handle_close()
            self.cancel_flag = False
            return

        if self.invert_ratiometric.get() == 1:
            result = -result

        self.ratiometric_stack = result
        self.ratio_norm.set(0.01)

        shape = np.shape(self.ratiometric_stack)[:2]
        h, w = shape
        projection = np.zeros(shape=shape)
        tif_limit = np.abs(np.clip(self.ratiometric_stack, -100, 0))
        max_index_matrix = np.argmax(tif_limit, axis=2)
        for x in range(w):
            for y in range(h):
                projection[y, x] = tif_limit[y, x, max_index_matrix[y, x]]
        tif_limit = np.abs(np.clip(self.ratiometric_stack, 0, 100))
        max_index_matrix = np.argmax(tif_limit, axis=2)

        try:
            self.progress_win.handle_close()
        except Exception:
            """ Already closed """

        for x in range(w):
            for y in range(h):
                projection[y, x] += tif_limit[y, x, max_index_matrix[y, x]]
        self.spatial_mask_image = projection

        self.slider_ratio['to'] = np.shape(result)[2] - 1
        self.export_start_frame['to'] = np.shape(result)[2] - 2
        self.export_end_frame['from_'] = 1
        self.export_end_frame['to'] = np.shape(result)[2]
        self.export_end_frame.set(np.shape(self.ratiometric_stack)[2])
        self.display_frame_ratio(0)
        self.event_label['text'] = ''
        self.event_label.update()
        self.temporal_mask_timeline = np.zeros(shape=np.shape(self.ratiometric_stack)[2])
        self.temporal_mask = []
        self.spatial_mask_binary = np.zeros(shape=shape)
        self.spatial_mask = []
        self.plot_temporal()
        self.plot_spatial()

        self.window.update()
        try:
            self.progress_win.window.destroy()
        except Exception:
            pass
        apodise = preferences['motion']['apodise']
        if self.use_median.get() == 1:
            apodise = False
        self.progress_win = ProgressWin(title='Processing', msg='Correcting motion...')
        self.progress_win.progress['maximum'] = 2
        self.progress_win.window.update()
        if self.correct_motion_var.get() == 1:
            self.progress_win.progress.step(1)
            self.progress_win.window.update()
            try:
                self.ratiometric_stack, spectrum = motion.correct_motion(self.ratiometric_stack, filter_quality=preferences['motion']['notch'], apodise=apodise)
            except Exception:
                print(traceback.format_exc())

        self.progress_win.handle_close()
        self.display_frame_ratio(0)

        sharp = []
        for idx in range(np.shape(self.ratiometric_stack)[2] - 1):
            sharp.append(np.std(self.native_stack[:, :, idx]))
        sharp = np.convolve(sharp, [1, 1, 1], mode="valid")
        sharp = sharp / np.max(sharp)
        # sharp = sharp ** 2
        sharp = list(sharp) + [1, 1, 1]
        self.sharpness = sharp
        self.max_contrast = np.max(np.abs(self.ratiometric_stack))
        self.plot_temporal()


    def process(self):
        avg = int(float(self.pre_avg.get()))
        shape = np.shape(self.native_stack)
        mean = np.mean(self.native_stack)
        length = shape[2]
        result = np.zeros((shape[0], shape[1], shape[2] - avg + 1))
        for frame in range(length - avg + 1):
            averaged_frame = np.zeros((shape[0], shape[1]))
            for add in range(avg):
                averaged_frame = averaged_frame + self.native_stack[:, :, frame + add]
            result[:, :, frame] = averaged_frame / avg
            if frame % 20 == 0:
                self.event_label['text'] = f'Applying Pre Window... ({frame}/{length-avg+1})'
                self.event_label.update()

        final = np.zeros(shape=(shape[0], shape[1], shape[2] - avg))
        for frame in range(length - avg):
            subtracted_frame = result[:, :, frame + 1] - result[:, :, frame]
            final[:, :, frame] = subtracted_frame
            self.event_label['text'] = f'Calculating Temporal Residual... ({frame}/{length - avg + 1})'
            self.event_label.update()

        avg = int(float(self.post_avg.get()))
        shape = np.shape(final)
        length = shape[2]
        result = np.zeros((shape[0], shape[1], shape[2] - avg + 1))
        for frame in range(length - avg + 1):
            averaged_frame = np.zeros((shape[0], shape[1]))
            for add in range(avg):
                averaged_frame = averaged_frame + final[:, :, frame + add]
            result[:, :, frame] = averaged_frame / avg
            self.event_label['text'] = f'Applying Post Window... ({frame}/{length - avg + 1})'
            self.event_label.update()

        # final = np.zeros(np.shape(result))
        # for frame in range(np.shape(result)[2]):
        #     processed = iscam.low_pass(result[:, :, frame], 1)
        #     final[:, :, frame] = processed

        # normalize contrast
        result = result / mean
        result = result * int(float(self.post_avg.get())) / (1+(0.16*int(float(self.post_avg.get()))/int(float(self.pre_avg.get()))))
        self.ratio_norm.set(round(-np.min(result)/4, 4))

        self.ratiometric_stack = result
        self.slider_ratio['to'] = np.shape(result)[2] - 1
        self.display_frame_ratio(0)
        self.event_label['text'] = ''
        self.event_label.update()
        self.plot_temporal()

    def calculate_laplacian_kernel(self):
        inner = float(self.lap_centre_sigma.get())
        outer = float(self.lap_outer_sigma.get())
        magnitude = float(self.lap_magnitude.get())
        kernel_1 = iscam.gauss(13, 13, 6, 6, inner, inner, magnitude, mode='protr')[:, :, 0]
        kernel_2 = iscam.gauss(13, 13, 6, 6, outer, outer, 1, mode='protr')[:, :, 0]
        self.laplacian_filter = kernel_2 - kernel_1
        self.display_laplacian()

    def display_laplacian(self):
        self.laplacian_canvas_adsorption.delete("all")
        norm = mplib.colors.Normalize(vmin=-float(self.lap_contrast.get()), vmax=float(self.lap_contrast.get()))
        frame_rgb = mplib.cm.gray(norm(self.laplacian_filter))
        frame_rgb = frame_rgb[:, :, :3] * 255
        PIL_image = Image.fromarray(frame_rgb.astype('uint8'), 'RGB')
        PIL_image = PIL_image.resize((130, 130), Image.NEAREST)
        global lap_ad
        lap_ad = ImageTk.PhotoImage(master=self.window, image=PIL_image)
        self.laplacian_canvas_adsorption.create_image(1, 1, anchor="nw", image=lap_ad)

        self.laplacian_canvas_desorption.delete("all")
        norm = mplib.colors.Normalize(vmin=-float(self.lap_contrast.get()), vmax=float(self.lap_contrast.get()))
        frame_rgb = mplib.cm.gray(norm(-self.laplacian_filter))
        frame_rgb = frame_rgb[:, :, :3] * 255
        PIL_image = Image.fromarray(frame_rgb.astype('uint8'), 'RGB')
        PIL_image = PIL_image.resize((130, 130), Image.NEAREST)
        global lap_des
        lap_des = ImageTk.PhotoImage(master=self.window, image=PIL_image)
        self.laplacian_canvas_desorption.create_image(1, 1, anchor="nw", image=lap_des)

        self.lgf_ad_figure.clf()
        self.lgf_ad_canvas.draw()
        self.lgf_des_figure.clf()
        self.lgf_des_canvas.draw()
        ax_ad = self.lgf_ad_figure.add_subplot(111)
        ax_des = self.lgf_des_figure.add_subplot(111)
        ax_ad.plot(self.laplacian_filter[:, 6], color='orange', marker='o', markersize=2, markeredgecolor='white')
        ax_des.plot(-self.laplacian_filter[:, 6], color='orange', marker='o', markersize=2, markeredgecolor='white')
        ax_ad.set_facecolor('#242428')
        ax_des.set_facecolor('#242428')
        ax_ad.grid(True, color='#3f3f3f')
        ax_des.grid(True, color='#3f3f3f')
        ax_ad.xaxis.set_major_locator(plt.MaxNLocator(integer=True))
        ax_ad.yaxis.set_major_locator(plt.MaxNLocator(integer=True))
        ax_des.xaxis.set_major_locator(plt.MaxNLocator(integer=True))
        ax_des.yaxis.set_major_locator(plt.MaxNLocator(integer=True))
        ax_ad.xaxis.set_major_locator(MultipleLocator(1))
        ax_des.xaxis.set_major_locator(MultipleLocator(1))
        self.lgf_ad_canvas.draw()
        self.lgf_des_canvas.draw()

    def get_contrasts_and_masses(self):
        self.current_trace_event = 0
        self.events = []
        self.contrasts = []
        self.masses = []
        self.zstack_adsorption, self.zstack_desorption = [], []
        self.delete_fits(auto=True)
        self.get_events(negative=True)
        if int(float(self.window_size.get())) == 1:
            self.cancel_flag = True
        if not self.cancel_flag:
            self.get_events(negative=False)
        if self.cancel_flag:
            self.cancel_flag = False
        if self.use_mass.get() == 1:
            try:
                cmin, cmax = round(np.percentile(self.masses, preferences['hist']['percentile low'])-5), round(np.percentile(self.masses, preferences['hist']['percentile hi'])+5)
                self.min_entry.delete(0, tk.END)
                self.max_entry.delete(0, tk.END)
                self.min_entry.insert(tk.END, cmin)
                self.max_entry.insert(tk.END, cmax)
            except Exception:
                print(traceback.format_exc())
        else:
            try:
                cmin, cmax = round(np.percentile(self.contrasts, preferences['hist']['percentile low']), 5), round(np.percentile(self.contrasts, preferences['hist']['percentile hi']), 5)
                self.min_entry.delete(0, tk.END)
                self.max_entry.delete(0, tk.END)
                self.min_entry.insert(tk.END, f'{cmin}')
                self.max_entry.insert(tk.END, f'{cmax}')
            except Exception:
                print(traceback.format_exc())

        self.internal_window = int(float(self.window_size.get())) + int(2*float(self.offset.get()))
        self.slider_proj['to'] = self.slider_ratio['to']
        self.display_frame_ratio(index=int(float(self.ratio_frame_var.get())))
        self.proj_frame_var.set(int(float(self.ratio_frame_var.get())))
        self.display_frame_proj(index=int(float(self.proj_frame_var.get())))
        self.update_fit_mode(load=True)
        self.plot_histogram(None)
        if preferences['event']['auto']:
            try:
                self.filter_plot_decide()
                self.auto_refinement()
            except Exception:
                print('Warning! Failure detected in Event Auto-refine:\n')
                print(traceback.format_exc())
        if self.fit_manually.get() == 0:
            if not int(float(self.window_size.get())) == 1:
                if preferences['hist']['auto fit']:
                    self.auto_fit(data_type='events')
                    self.plot_histogram(None)
        if len(self.events) > 0:
            self.current_trace_event = 0
            self.plot_trace()
            self.filter_plot_decide()

    def subpixel_mask(self, cx, cy):
        mask = np.array([
            [0, 0, 0, 0, 0],
            [0, 255, 255, 255, 0],
            [0, 255, 255, 255, 0],
            [0, 255, 255, 255, 0],
            [0, 0, 0, 0, 0]
        ]).astype(np.uint16)
        x0, y0 = cx - int(cx) + 2, cy - int(cy) + 2
        subpixel_target = iscam.gauss(5, 5, x0, y0, 0.3, 0.3, 1, mode="protr")[:, :, 0]
        subpixel_target = (subpixel_target / np.sum(subpixel_target)).astype(np.float64)
        subpixel_target = subpixel_target[::-1, ::-1]
        adjusted_mask = iscam.cv2.filter2D(mask, -1, subpixel_target, borderType=2)
        # plt.imshow(subpixel_target)
        # plt.show()
        return adjusted_mask / 255

    def get_events(self, negative=True):
        if negative:
            label_text = 'Detecting contrast events (adsorption)...'
        else:
            label_text = 'Detecting contrast events (desorption)...'
        window_length = int(float(self.window_size.get()))
        offset = int(float(self.offset.get()))
        traces = []
        shape = np.shape(self.ratiometric_stack)[:2]
        length = np.shape(self.ratiometric_stack)[2]
        h, w = shape
        norm_limit = np.max(np.abs(self.ratiometric_stack))

        use_lap = self.use_laplacian.get()
        use_low = self.use_low_pass.get()
        use_bin = self.binary_mask.get()
        c_thresh = float(self.detection_threshold.get())
        sigmin = float(self.min_sigma.get())
        sigmax = float(self.max_sigma.get())
        min_ec = float(self.eccentricity.get())
        nearest_neigh = float(self.nearest_neighbours.get())

        extended = self.extended_trace.get()
        extension_amt = int(float(self.extension_amount.get()))

        try:
            self.progress_win.window.destroy()
        except Exception:
            pass
        self.progress_win = ProgressWin(title='Analysing', msg=label_text)
        self.progress_win.progress['maximum'] = (int((length-window_length) / (window_length+offset*2))) + 1
        counter = 0
        if window_length == 1:
            update_point = 10
        else:
            update_point = 1
        for frame_point in range(offset+1, length-window_length, window_length+offset*2):
            if self.cancel_flag == True:
                break
            counter += 1

            if counter % update_point == 0:
                self.progress_win.progress.step(update_point)
                self.event_label["text"] = f'Detecting Contrast Events... ({frame_point} / {length-window_length})'
                self.event_label.update()
            if window_length <= 400 and self.dynamic_updates.get() == 1:
                if counter % int(preferences['hist']['update'] / (window_length)) == 0:
                    try:
                        self.convert_masses()
                        self.plot_histogram(None)
                    except Exception:
                        ''' Failed to plot distribution '''
            # img_max_projection = np.zeros(shape=shape)
            tif_limit = np.copy(self.ratiometric_stack[:, :, frame_point:frame_point+window_length])
            # plt.imshow(tif_limit[:, :, 0], cmap='gray', vmin=-0.02, vmax=0.02)
            # plt.show()
            if use_lap == 1:
                for fp in range(np.shape(tif_limit)[2]):
                    if negative:
                        tif_limit[:, :, fp] = iscam.cv2.filter2D(tif_limit[:, :, fp], -1, -self.laplacian_filter, borderType=2)
                    else:
                        tif_limit[:, :, fp] = iscam.cv2.filter2D(tif_limit[:, :, fp], -1, -self.laplacian_filter, borderType=2)
            if negative:
                tif_limit = np.clip(tif_limit, -100, 0)
            else:
                tif_limit = np.clip(tif_limit, 0, 100)
            tif_limit = np.abs(tif_limit)
            # plt.imshow(tif_limit[:, :, 0], cmap='gray', vmin=-0.0, vmax=0.02)
            # plt.show()
            # plt.imshow(tif_limit[:, :, 0], cmap='plasma')
            # plt.show()
            if use_lap == 1:
                for fp in range(np.shape(tif_limit)[2]):
                    tif_limit[:, :, fp] = iscam.cv2.filter2D(tif_limit[:, :, fp], -1, -self.laplacian_filter, borderType=2)
                tif_limit = np.clip(tif_limit, 0, 100)
            # plt.imshow(tif_limit[:, :, 0], cmap='gray', vmin=-0.0, vmax=0.02)
            # plt.show()
            # max_index_matrix = np.argmax(tif_limit, axis=2)
            # for x in range(w):
            #     for y in range(h):
            #         img_max_projection[y, x] = tif_limit[y, x, max_index_matrix[y, x]]
            img_max_projection = np.max(tif_limit, axis=2)
            # plt.imshow(img_max_projection, cmap='gray', vmin=-0.0, vmax=0.02)
            # plt.show()
            # img_mean_projection = np.mean(tif_limit, axis=2)
            # img_max_projection = img_max_projection + img_mean_projection
            # img_max_projection = img_max_projection / 2
            if use_low == 1:
                img_max_projection = iscam.low_pass(img_max_projection, 1.01, filter=False)
                img_max_projection = np.clip(img_max_projection, 0, 100)
            # plt.imshow(img_max_projection, cmap='gray', vmin=-0.0, vmax=0.02)
            # plt.show()
            # if self.use_global_norm.get() == 0:
            #     norm_limit = np.max(img_max_projection)
            # img_max_projection = img_max_projection / norm_limit
            # img_max_projection = img_max_projection * 255
            # img_max_projection = np.clip(img_max_projection, 0, 255)
            # img_max_projection = img_max_projection / 255
            # img_max_projection = img_max_projection * 254.9
            # img_max_projection = img_max_projection + 0.1
            # plt.imshow(img_max_projection)
            # plt.show()
            if negative:
                self.zstack_adsorption.append(img_max_projection)
            else:
                self.zstack_desorption.append(img_max_projection)
            av_mode = False
            # if self.average_mode.get() == 1:
            #     av_mode = True
            # events = iscam.find_events(view=0, img_array=img_max_projection, mode=av_mode,
            #                            threshold=float(self.detection_threshold.get()),
            #                            averaging_dist=2,
            #                            gauss_fit_residual_threshold=float(self.min_kernel.get()),
            #                            min_sigma_threshold=float(self.min_sigma.get()),
            #                            max_sigma_threshold=float(self.max_sigma.get()),
            #                            min_intensity=float(self.min_intensity.get()),
            #                            eccentricity_threshold=float(self.eccentricity.get()),
            #                            true_gauss_threshold=float(self.min_gauss.get()),
            #                            region=[[0, 0], [self.resolution[1], self.resolution[0]]],
            #                            inverted=False
            #                            )
            # spot data stored as [spot x, spot y, [sigma x, sigma y, amplitude, gaussian residual]]

            init_events = foci_detection.detect_and_fit_gaussian_optimized(img_max_projection, threshold=c_thresh)
            filtered_events = foci_detection.filter_results(init_events,
                                                            imgx=self.resolution[0],
                                                            imgy=self.resolution[1],
                                                            min_sigma=sigmin,
                                                            max_sigma=sigmax,
                                                            ecc_thresh=min_ec,
                                                            min_dist=nearest_neigh,
                                                            )
            # spot data is stored as [spot_x, spot_y, sigma_x, sigma_y, amplitude, baseline]
            # converting to old format to prevent unforseen errors:
            events = [[e[0], e[1], [e[2], e[3], e[4], e[5]]] for e in filtered_events]

            event_traces = []
            for event in events:
                if extended == 1 and event[0] > extension_amt + offset and event[0] < length-window_length - extension_amt - offset:
                    extension = extension_amt
                else:
                    extension = 0
                cx, cy = event[0], event[1]
                if use_bin == 0:
                    cuboid = self.ratiometric_stack[int(round(event[1]-1)):int(round(event[1]+2)), int(round(event[0]-1)):int(round(event[0]+2)), frame_point-offset-extension:frame_point+window_length+offset+extension]
                else:
                    cuboid = self.ratiometric_stack[int(event[1] - 2):int(event[1] + 3), int(event[0] - 2):int(event[0] + 3), frame_point - offset - extension:frame_point + window_length + offset + extension]
                # if not negative:
                #     cuboid = np.clip(cuboid, 0, 10)
                # else:
                #     cuboid = np.clip(cuboid, -10, 0)
                if not negative:
                    zp = np.max(np.abs(np.clip(cuboid, 0, 10)), axis=2)
                else:
                    zp = np.max(np.abs(np.clip(cuboid, -10, 0)), axis=2)
                cshape = np.shape(cuboid)
                dim = 5
                if (cshape[2] > 0 and cshape[0] == dim and cshape[1] == dim and use_bin == 1) or (use_bin == 0):
                    # zpm = np.median(zp)
                    # msk = np.zeros(shape=(dim, dim))
                    # for ix in range(dim):
                    #     for iy in range(dim):
                    #         if zp[iy, ix] > zpm:
                    #             msk[iy, ix] = 1
                    # msk = np.array(
                    #     [
                    #         [0, 1, 0],
                    #         [1, 1, 1],
                    #         [0, 1, 0]
                    #     ]
                    # )
                    if use_bin == 1:
                        msk = self.subpixel_mask(cx, cy)
                        msk_norm = np.sum(msk)
                    # plt.imshow(zp)
                    # plt.show()
                    # plt.imshow(msk)
                    # plt.title(f' mask sum: {msk_norm}')
                    # plt.show()
                    trace = []
                    add_trace = True
                    for idx in range(np.shape(cuboid)[2]):
                        try:
                            # tm = np.max(np.abs(cuboid[:, :, idx]))
                            # cont = []
                            # for val in np.ravel(cuboid[:, :, idx]):
                            #     if abs(val) >= 0.55*tm:
                            #         cont.append(val)
                            # time_point = np.median(cont)
                            if use_bin == 1:
                                time_point = np.sum(cuboid[:, :, idx]*msk) / msk_norm
                            else:
                                time_point = np.mean(cuboid[:, :, idx])
                            trace.append(time_point)
                        except Exception:
                            """ end of movie reached """
                    if add_trace:
                        # try:
                        #     trace = list(np.convolve(trace, [0.25, 0.5, 0.25], mode='full'))
                        # except Exception:
                        #     'trace is wrong shape'
                        traces.append(trace)
                        event_traces.append(trace)

            for idx in range(len(events)):
                try:
                    if negative:
                        c = np.min(event_traces[idx])
                    else:
                        c = np.max(event_traces[idx])
                    index = event_traces[idx].index(c)
                    if extended == 1 and event[0] > extension_amt + offset and event[0] < length - window_length - extension_amt - offset:
                        extension = extension_amt
                    else:
                        extension = 0
                    frame_ = index + frame_point - offset - extension
                    try:
                        g1, g2, r1, r2 = self.get_trace_gradients(event_traces[idx], index)
                    except Exception:
                        g1, g2, r1, r2 = None, None, None, None
                    self.events.append([frame_, -c, events[idx], event_traces[idx], None, [g1, g2, r1, r2]])
                    # print(self.events[-1])
                except Exception:
                    'error calculating trace'
                    print(traceback.format_exc())

        self.event_label["text"] = ''
        self.event_label.update()
        self.progress_win.handle_close()
        self.convert_masses()

        self.display_frame_ratio(index=int(float(self.ratio_frame_var.get())))

    def update_include(self):
        if self.include.get() == 0:
            self.events[self.current_trace_event][4] = False
        elif self.include.get() == 1:
            self.events[self.current_trace_event][4] = True
        self.convert_masses()
        self.plot_histogram(None)

    def set_include_checkbox(self):
        if self.current_trace_event is not None and len(self.events) > 0:
            if self.sharpness_mc.get() == 0:
                factor = 1
            else:
                factor = self.sharpness[self.events[self.current_trace_event][0]]
            if self.events[self.current_trace_event][1] / factor in self.contrasts:
                self.include_cb['command'] = None
                self.include.set(1)
                self.include_cb['command'] = self.update_include
            else:
                self.include_cb['command'] = None
                self.include.set(0)
                self.include_cb['command'] = self.update_include

    def convert_masses(self):
        self.contrasts = []
        for event in self.events:
            if event[4] is not None and not event[4]:
                continue
            elif event[4] is None:
                if self.temporal_mask_timeline[event[0]] == 0 and self.temp_invert.get() == 0:
                    if self.spatial_mask_binary[int(event[2][1]), int(event[2][0])] == 0 and self.spatial_invert.get() == 0:
                        self.contrasts.append(event[1])
                        if self.sharpness_mc.get() == 1:
                            self.contrasts[-1] = self.contrasts[-1] / self.sharpness[event[0]]
                    elif self.spatial_mask_binary[int(event[2][1]), int(event[2][0])] == 1 and self.spatial_invert.get() == 1:
                        self.contrasts.append(event[1])
                        if self.sharpness_mc.get() == 1:
                            self.contrasts[-1] = self.contrasts[-1] / self.sharpness[event[0]]
                elif self.temporal_mask_timeline[event[0]] == 1 and self.temp_invert.get() == 1:
                    if self.spatial_mask_binary[int(event[2][1]), int(event[2][0])] == 0 and self.spatial_invert.get() == 0:
                        self.contrasts.append(event[1])
                        if self.sharpness_mc.get() == 1:
                            self.contrasts[-1] = self.contrasts[-1] / self.sharpness[event[0]]
                    elif self.spatial_mask_binary[int(event[2][1]), int(event[2][0])] == 1 and self.spatial_invert.get() == 1:
                        self.contrasts.append(event[1])
                        if self.sharpness_mc.get() == 1:
                            self.contrasts[-1] = self.contrasts[-1] / self.sharpness[event[0]]
            elif event[4] is not None and event[4]:
                self.contrasts.append(event[1])
                if self.sharpness_mc.get() == 1:
                    self.contrasts[-1] = self.contrasts[-1] / self.sharpness[event[0]]

        self.set_include_checkbox()

        self.masses = []
        for idx in range(len(self.contrasts)):
            if self.contrasts[idx] > 0:
                self.masses.append(self.contrasts[idx]*self.calibration[0]+self.calibration[1])
            else:
                self.masses.append(self.contrasts[idx] * self.calibration[0] - self.calibration[1])
        self.display_frame_ratio(index=int(float(self.ratio_frame_var.get())))

    def event_clicked(self, event):
        if self.is_popped_out:
            x, y = event.x / 5, event.y / 5
        else:
            x, y = (event.x-self.canvas_offset)/self.scale_factor, event.y/self.scale_factor
        print(x, y)
        display_time = int(float(self.persistence.get()))
        if len(self.events) > 0:
            current_visible_events = []
            for idx, e in enumerate(self.events):
                if int(float(self.ratio_frame_var.get())) > e[0] - display_time and int(float(self.ratio_frame_var.get())) < e[0] + display_time:
                    current_visible_events.append(e)
            if len(current_visible_events) > 0:
                distances = []
                for idx, e in enumerate(current_visible_events):
                    distances.append(np.sqrt((x - e[2][0])**2 + (y - e[2][1])**2))
                index = distances.index(min(distances))
                event_index = self.events.index(current_visible_events[index])
                print(event_index)
                self.current_trace_event = event_index
                self.display_frame_ratio(index=self.ratio_frame_var.get())
                self.plot_trace()
        if len(self.tracks) > 0:
            distances = []
            indices = []
            for idx, track in enumerate(self.tracks):
                if track.frames[0] <= int(float(self.ratio_frame_var.get())) <= track.frames[-1]:
                    mean_x, mean_y = np.mean([c[0] for c in track.coords]), np.mean([c[1] for c in track.coords])
                    distances.append(np.sqrt((x - mean_x)**2 + (y - mean_y)**2))
                    indices.append(idx)
            idx = np.argmin(distances)
            try:
                bound_func = self.current_track['command']
                self.current_track['command'] = None
                self.current_track.set(indices[idx] + 1)
                self.current_track['command'] = bound_func
                self.update_tracks(change_frame=False)
            except Exception:
                print(traceback.format_exc())

    def update_trace_labels(self):
        index = self.current_trace_event
        try:
            contrast = self.events[self.current_trace_event][1]
            contrast_label = round(contrast, 4)
        except Exception:
            contrast_label = 'N/A'
        try:
            if contrast < 0:
                mass = contrast * self.calibration[0] - self.calibration[1]
            else:
                mass = contrast * self.calibration[0] + self.calibration[1]
            mass_label = round(mass)
        except Exception:
            mass_label = 'N/A'
        try:
            trace = self.events[self.current_trace_event][3]
            if contrast < 0:
                point = trace.index(max(trace))
            else:
                point = trace.index(min(trace))
            new_trace = []
            for idx in range(len(trace)):
                if idx <= point - int(float(self.pre_avg.get())) or idx >= point + int(float(self.pre_avg.get())):
                    new_trace.append(trace[idx])
            std = np.std(new_trace, ddof=1)
            std_label = round(abs(std), 5)
        except Exception:
            std_label = 'N/A'
        try:
            snr = contrast / std
            snr_label = round(abs(snr), 2)
        except Exception:
            snr_label = 'N/A'
        self.trace_label1['text']=f'Event: [{index}]'
        self.trace_label2['text']=f'Cont.: [{contrast_label}]'
        self.trace_label3['text']=f'Mass: [{mass_label} kDa]'
        self.trace_label4['text']=f'StD: [{std_label}]'
        self.trace_label5['text']=f'SNR: [{snr_label}]'

    def fit_minimum(self, trace, point):
        boundary = self.internal_frame_average
        sec1 = trace[point-boundary:point+1]
        sec2 = trace[point:point+boundary+1]
        x1, x2 = list(range(point-boundary, point+1)), list(range(point, point+boundary+1))
        g1, i1 = linear_regression(x1, sec1)
        g2, i2 = linear_regression(x2, sec2)
        return g1, i1, g2, i2, x1, x2

    def get_trace_gradients(self, trace, point):
        g1, i1, g2, i2, x1, x2 = self.fit_minimum(trace, point)
        d1 = [trace[x] for x in x1]
        d2 = [trace[x] for x in x2]
        f1 = [x * g1 + i1 for x in x1]
        f2 = [x * g2 + i2 for x in x2]
        r2_1 = r2_score(d1, f1)
        r2_2 = r2_score(d2, f2)
        return g1, g2, r2_1, r2_2

    def plot_trace(self):
        self.update_trace_labels()
        self.set_include_checkbox()
        self.trace_figure.clf()
        self.trace_canvas.draw()

        bg_col = "#333333"
        fig_col = "#333333"
        highlight_col = "#007fff"
        text_col = "#cccccc"
        grid_col = "#3f3f3f"

        axis_text = 'Time (frames)'

        self.trace_figure.set_facecolor(fig_col)
        self.trace_plotter = self.trace_figure.add_subplot(111)
        self.trace_plotter.set_facecolor(bg_col)
        self.trace_plotter.spines['bottom'].set_color(highlight_col)
        self.trace_plotter.spines['top'].set_color(fig_col)
        self.trace_plotter.spines['left'].set_color(highlight_col)
        self.trace_plotter.spines['right'].set_color(fig_col)
        self.trace_plotter.xaxis.label.set_color(text_col)
        self.trace_plotter.yaxis.label.set_color(text_col)
        self.trace_plotter.tick_params(axis='x', colors=text_col, labelsize=9)
        self.trace_plotter.tick_params(axis='y', colors=text_col, labelsize=9)
        self.trace_plotter.set_xlabel(axis_text, color=text_col, size=10)
        self.trace_plotter.set_ylabel('Contrast', color=text_col, size=10)
        self.trace_plotter.grid(color=grid_col)

        try:
            trace = self.events[self.current_trace_event][3]
            if self.events[self.current_trace_event][1] > 0:
                m = min(trace)
            else:
                m = max(trace)
            self.trace_plotter.plot(trace, marker='o', linewidth=1, markersize=3)
            i = trace.index(m)
            self.trace_plotter.plot(i, m, marker='o', markersize=4, color='red')
            self.trace_canvas.draw()
        except Exception:
            traceback.format_exc()
        try:
            g1, i1, g2, i2, x1, x2 = self.fit_minimum(trace, i)
            d1 = [trace[x] for x in x1]
            d2 = [trace[x] for x in x2]
            f1 = [x*g1+i1 for x in x1]
            f2 = [x*g2+i2 for x in x2]
            r2_1 = r2_score(d1, f1)
            r2_2 = r2_score(d2, f2)
            text_offset = m / 9
            self.trace_plotter.plot(x1, f1, color='white', linewidth=1, linestyle='--')
            self.trace_plotter.plot(x2, f2, color='white', linewidth=1, linestyle='--')
            self.trace_plotter.text(x1[0]+0.5, f1[0], f'r²={round(r2_1, 4)}', size=7, color='white')
            self.trace_plotter.text(x2[-1] + 0.5, f2[-1], f'r²={round(r2_2, 4)}', size=7, color='white')
            self.trace_plotter.text(x1[0] + 0.5, f1[0]-text_offset, f'g={round(g1, 4)}', size=7, color='orange')
            self.trace_plotter.text(x2[-1] + 0.5, f2[-1]-text_offset, f'g={round(g2, 4)}', size=7, color='orange')
            self.trace_canvas.draw()
        except Exception:
            traceback.format_exc()

    def plot_spatial(self):
        self.spatial_figure.clf()
        self.spatial_canvas.draw()

        bg_col = "#333333"
        fig_col = "#333333"
        highlight_col = "#007fff"
        text_col = "#cccccc"
        grid_col = "#1f1f1f"

        axis_text = 'x coordinate'

        self.spatial_figure.set_facecolor(fig_col)
        self.spatial_plotter = self.spatial_figure.add_subplot(111)
        self.spatial_plotter.set_facecolor(bg_col)
        self.spatial_plotter.spines['bottom'].set_color(highlight_col)
        self.spatial_plotter.spines['top'].set_color(highlight_col)
        self.spatial_plotter.spines['left'].set_color(highlight_col)
        self.spatial_plotter.spines['right'].set_color(highlight_col)
        self.spatial_plotter.xaxis.label.set_color(text_col)
        self.spatial_plotter.yaxis.label.set_color(text_col)
        self.spatial_plotter.tick_params(axis='x', colors=text_col, labelsize=9)
        self.spatial_plotter.tick_params(axis='y', colors=text_col, labelsize=9)
        self.spatial_plotter.set_xlabel(axis_text, color=text_col, size=10)
        self.spatial_plotter.set_ylabel("y coordinate", color=text_col, size=10)
        self.spatial_plotter.grid(color=grid_col)

        self.spatial_plotter.imshow(self.spatial_mask_image, cmap='inferno')

        self.spatial_canvas.draw()
        self.spatial_bind_mouse = self.spatial_figure.canvas.mpl_connect('button_press_event', self.get_init_spatial)


    def update_contrast_gradient(self):
        pass
        # data = []
        # for idx in range(len(self.events)):
        #     if self.events[idx][5][0] is not None:
        #         data.append(self.events[idx][5][0]*5)
        # plt.hist(data, bins=800)
        # plt.title('G1')
        # plt.show()
        #
        # data = []
        # for idx in range(len(self.events)):
        #     if self.events[idx][5][1] is not None:
        #         data.append(self.events[idx][5][1]*5)
        # plt.hist(data, bins=800)
        # plt.title('G2')
        # plt.show()
        #
        # data = []
        # for idx in range(len(self.events)):
        #     if self.events[idx][5][0] is not None and self.events[idx][5][1] is not None:
        #         data.append((self.events[idx][5][1] - self.events[idx][5][0])*2.5)
        # plt.hist(data, bins=800)
        # plt.title('mean')
        # plt.show()
        #
        # data = []
        # for idx in range(len(self.events)):
        #     if self.events[idx][5][0] is not None and self.events[idx][5][1] is not None:
        #         if self.events[idx][5][2] > self.events[idx][5][3]:
        #             data.append(self.events[idx][5][0] * -5)
        #         else:
        #             data.append(self.events[idx][5][1] * 5)
        # plt.hist(data, bins=800)
        # plt.title('best')
        # plt.show()

    def filter_reset(self):
        for ev in self.events:
            ev[4] = None
        self.filter_r2_threshold = None
        self.filter_grad_threshold = None
        self.filter_snr_threshold = None
        self.filter_plot_decide()
        self.convert_masses()
        self.plot_histogram(None)

    def get_filter_cursor(self, e):
        if self.filter_plotter.get_navigate_mode() is not None and e is not None:
            return

        if e is not None:
            x = e.xdata
            if self.filter_plot_mode.get() == 1:
                self.filter_r2_threshold = x
            elif self.filter_plot_mode.get() == 2:
                self.filter_grad_threshold = x
            elif self.filter_plot_mode.get() == 3:
                self.filter_snr_threshold = x

            self.filter_plot_decide()

        for ev in self.events:
            if ev[4] is not None:
                ev[4] = None
        for ev in self.events:
            if ev[5][0] is None or ev[5][1] is None or ev[5][2] is None or ev[5][3] is None:
                ev[4] = False
            else:
                if self.filter_r2_threshold is not None:
                    if ev[5][2] < self.filter_r2_threshold or ev[5][3] < self.filter_r2_threshold:
                        ev[4] = False
                if self.filter_grad_threshold is not None:
                    a, b = abs(ev[5][0]), abs(ev[5][1])
                    if a > b:
                        n = a
                    else:
                        n = b
                    percentage_diff = 100 * abs(ev[5][0] + ev[5][1]) / n
                    if percentage_diff > self.filter_grad_threshold:
                        ev[4] = False
                if self.filter_snr_threshold is not None:
                    ind = self.events.index(ev)
                    snr = self.get_event_snr(ind)
                    if snr < self.filter_snr_threshold:
                        ev[4] = False

        self.convert_masses()
        self.plot_histogram(None)

    def auto_refinement(self):
        self.filter_r2_threshold = preferences['event']['grad r2']
        self.filter_grad_threshold = preferences['event']['grad diff']
        self.filter_snr_threshold = preferences['event']['snr']
        self.get_filter_cursor(None)

    def filter_plot_decide(self):
        self.filter_bind_mouse = self.filter_figure.canvas.mpl_connect('button_press_event', self.get_filter_cursor)
        if self.filter_plot_mode.get() == 1:
            self.plot_filter_r2()
        elif self.filter_plot_mode.get() == 2:
            self.plot_filter_grads()
        elif self.filter_plot_mode.get() == 3:
            self.plot_filter_snr()

    def plot_filter_r2(self):
        data1 = []
        data2 = []
        for event in self.events:
            if event[5][2] is not None:
                data1.append(event[5][2])
            if event[5][3] is not None:
                data2.append(event[5][3])
        self.filter_figure.clf()
        self.filter_canvas.draw()
        bg_col = "#2f2f2f"
        fig_col = "#333333"
        highlight_col = "#007fff"
        text_col = "#cccccc"
        grid_col = "#373737"

        axis_text = 'r² - G1 (Orange) G2 (blue)'

        self.filter_figure.set_facecolor(fig_col)
        self.filter_plotter = self.filter_figure.add_subplot(111)
        self.filter_plotter.set_facecolor(bg_col)
        self.filter_plotter.spines['bottom'].set_color(highlight_col)
        self.filter_plotter.spines['top'].set_color(highlight_col)
        self.filter_plotter.spines['left'].set_color(highlight_col)
        self.filter_plotter.spines['right'].set_color(highlight_col)
        self.filter_plotter.xaxis.label.set_color(text_col)
        self.filter_plotter.yaxis.label.set_color(text_col)
        self.filter_plotter.tick_params(axis='x', colors=text_col, labelsize=8)
        self.filter_plotter.tick_params(axis='y', colors=text_col, labelsize=8)
        self.filter_plotter.set_xlabel(axis_text, color=text_col, size=8)
        self.filter_plotter.set_ylabel("Density", color=text_col, size=8)
        self.filter_plotter.grid(color=grid_col)

        self.filter_plotter.hist(data1, bins=400, density=True, alpha=0.5, ec="#ff7f00", color="#aa5f00", histtype='stepfilled')
        self.filter_plotter.hist(data2, bins=400, density=True, alpha=0.5, ec="#004fff", color="#0022aa", histtype='stepfilled')

        if self.filter_r2_threshold is not None:
            m = min(data1+data2)
            self.filter_plotter.axvspan(m, self.filter_r2_threshold, facecolor='#8833ff', alpha=0.5)
            self.filter_plotter.axvline(self.filter_r2_threshold, color="w", linestyle="--", linewidth=1)

        self.filter_canvas.draw()

    def plot_filter_grads(self):
        data = []
        for event in self.events:
            if event[5][0] is not None and event[5][1]:
                a, b = abs(event[5][0]), abs(event[5][1])
                if a > b:
                    n = a
                else:
                    n = b
                data.append(100*abs(event[5][0] + event[5][1]) / n)
        self.filter_figure.clf()
        self.filter_canvas.draw()
        bg_col = "#2f2f2f"
        fig_col = "#333333"
        highlight_col = "#007fff"
        text_col = "#cccccc"
        grid_col = "#373737"

        axis_text = 'Gradient diff. (%)'

        self.filter_figure.set_facecolor(fig_col)
        self.filter_plotter = self.filter_figure.add_subplot(111)
        self.filter_plotter.set_facecolor(bg_col)
        self.filter_plotter.spines['bottom'].set_color(highlight_col)
        self.filter_plotter.spines['top'].set_color(highlight_col)
        self.filter_plotter.spines['left'].set_color(highlight_col)
        self.filter_plotter.spines['right'].set_color(highlight_col)
        self.filter_plotter.xaxis.label.set_color(text_col)
        self.filter_plotter.yaxis.label.set_color(text_col)
        self.filter_plotter.tick_params(axis='x', colors=text_col, labelsize=8)
        self.filter_plotter.tick_params(axis='y', colors=text_col, labelsize=8)
        self.filter_plotter.set_xlabel(axis_text, color=text_col, size=8)
        self.filter_plotter.set_ylabel("Density", color=text_col, size=8)
        self.filter_plotter.grid(color=grid_col)

        self.filter_plotter.hist(data, bins=400, density=True, alpha=0.5, ec="#ff7f00", color="#aa5f00", histtype='stepfilled')
        if self.filter_grad_threshold is not None:
            m = max(data)
            self.filter_plotter.axvspan(m, self.filter_grad_threshold, facecolor='#8833ff', alpha=0.5)
            self.filter_plotter.axvline(self.filter_grad_threshold, color="w", linestyle="--", linewidth=1)

        self.filter_canvas.draw()

    def plot_filter_snr(self):
        data = []
        for idx in range(len(self.events)):
            data.append(self.get_event_snr(idx))

        self.filter_figure.clf()
        self.filter_canvas.draw()
        bg_col = "#2f2f2f"
        fig_col = "#333333"
        highlight_col = "#007fff"
        text_col = "#cccccc"
        grid_col = "#373737"

        axis_text = 'Signal / Noise'

        self.filter_figure.set_facecolor(fig_col)
        self.filter_plotter = self.filter_figure.add_subplot(111)
        self.filter_plotter.set_facecolor(bg_col)
        self.filter_plotter.spines['bottom'].set_color(highlight_col)
        self.filter_plotter.spines['top'].set_color(highlight_col)
        self.filter_plotter.spines['left'].set_color(highlight_col)
        self.filter_plotter.spines['right'].set_color(highlight_col)
        self.filter_plotter.xaxis.label.set_color(text_col)
        self.filter_plotter.yaxis.label.set_color(text_col)
        self.filter_plotter.tick_params(axis='x', colors=text_col, labelsize=8)
        self.filter_plotter.tick_params(axis='y', colors=text_col, labelsize=8)
        self.filter_plotter.set_xlabel(axis_text, color=text_col, size=8)
        self.filter_plotter.set_ylabel("Density", color=text_col, size=8)
        self.filter_plotter.grid(color=grid_col)

        self.filter_plotter.hist(data, bins=400, density=True, alpha=0.5, ec="#ff7f00", color="#aa5f00", histtype='stepfilled')
        if self.filter_snr_threshold is not None:
            m = min(data)
            self.filter_plotter.axvspan(m, self.filter_snr_threshold, facecolor='#8833ff', alpha=0.5)
            self.filter_plotter.axvline(self.filter_snr_threshold, color="w", linestyle="--", linewidth=1)

        self.filter_canvas.draw()

    def get_event_snr(self, index):
        try:
            contrast = self.events[index][1]
            trace = self.events[index][3]
            if contrast < 0:
                point = trace.index(max(trace))
            else:
                point = trace.index(min(trace))
            new_trace = []
            for idx in range(len(trace)):
                if idx <= point - int(float(self.pre_avg.get())) or idx >= point + int(float(self.pre_avg.get())):
                    new_trace.append(trace[idx])
            std = np.std(new_trace, ddof=1)
            snr = abs(contrast / std)
            return snr
        except Exception:
            traceback.format_exc()

    def plot_temporal(self):
        self.temp_figure.clf()
        self.temp_canvas.draw()

        bg_col = "#333333"
        fig_col = "#333333"
        highlight_col = "#007fff"
        text_col = "#cccccc"
        grid_col = "#3f3f3f"

        axis_text = 'Time (frames)'

        self.temp_figure.set_facecolor(fig_col)
        self.temp_plotter = self.temp_figure.add_subplot(111)
        self.temp_plotter.set_facecolor(bg_col)
        self.temp_plotter.spines['bottom'].set_color(highlight_col)
        self.temp_plotter.spines['top'].set_color(fig_col)
        self.temp_plotter.spines['left'].set_color(highlight_col)
        self.temp_plotter.spines['right'].set_color(fig_col)
        self.temp_plotter.xaxis.label.set_color(text_col)
        self.temp_plotter.yaxis.label.set_color(text_col)
        self.temp_plotter.tick_params(axis='x', colors=text_col, labelsize=9)
        self.temp_plotter.tick_params(axis='y', colors=text_col, labelsize=9)
        self.temp_plotter.set_xlabel(axis_text, color=text_col, size=10)
        self.temp_plotter.set_ylabel("Signal", color=text_col, size=10)
        self.temp_plotter.grid(color=grid_col)

        signal = []
        if not self.disp_sharpness.get() == 1:
            for idx in range(np.shape(self.ratiometric_stack)[2]):
                signal.append(np.max(np.abs(self.ratiometric_stack[:, :, idx])))

        else:
            signal = self.sharpness
            self.temp_plotter.set_ylabel("Sharpness", color=text_col, size=10)
        self.temp_plotter.plot(signal, linewidth=1)
        self.temp_plotter.set_xlim(0, len(signal))

        for idx in range(len(self.temporal_mask)):
            self.temp_plotter.axvspan(self.temporal_mask[idx][0], self.temporal_mask[idx][1], facecolor='#7711ff', alpha=0.5)

        self.temp_canvas.draw()
        self.temp_bind_mouse = self.temp_figure.canvas.mpl_connect('button_press_event', self.get_init_temp)

    def get_init_spatial(self, event):
        self.spatial_figure.canvas.mpl_disconnect(self.spatial_bind_mouse)
        self.spatial_bind_mouse = self.spatial_figure.canvas.mpl_connect('button_release_event', self.get_final_spatial)
        self.spatial_bind_motion = self.spatial_figure.canvas.mpl_connect('motion_notify_event', self.track_mouse_spatial)
        mouse_xpos, mouse_ypos = round(event.xdata), round(event.ydata)
        self.spatial_mask.append([])
        self.spatial_mask[-1].append([mouse_xpos, mouse_ypos])
        print(self.spatial_mask)

    def get_init_temp(self, event):
        self.temp_figure.canvas.mpl_disconnect(self.temp_bind_mouse)
        self.temp_figure.canvas.mpl_disconnect(self.capture_motion_temporal)
        self.temp_bind_mouse = self.temp_figure.canvas.mpl_connect('button_release_event', self.get_final_temp)
        self.temp_bind_motion = self.temp_figure.canvas.mpl_connect('motion_notify_event', self.track_mouse_temporal)
        mouse_xpos = int(event.xdata)
        self.temporal_mask.append([])
        self.temporal_mask[-1].append(mouse_xpos)

    def temporal_vline(self, event):
        try:
            self.temporal_vline_widget.remove()
            self.temporal_vline_text.remove()
        except Exception:
            """ Can't """
        try:
            self.temporal_vline_widget = self.temp_plotter.axvline(event.xdata, linestyle='--', linewidth=1, color='orange')
            self.temporal_vline_text = self.temp_plotter.text(event.xdata, self.max_contrast/1.2, f'{round(event.xdata)}',
                                                              color='white', fontsize=10)
        except Exception:
            """ Failed to plot """
        self.temp_canvas.draw()

    def get_final_spatial(self, event):
        self.spatial_figure.canvas.mpl_disconnect(self.spatial_bind_mouse)
        self.spatial_figure.canvas.mpl_disconnect(self.spatial_bind_motion)
        manual_modify = False
        choice = 0
        for ev in self.events:
            if ev[4] is not None:
                manual_modify = True
                break
        if manual_modify:
            choice = easygui.indexbox(title='Warning!', msg='Certain events have been manually included/discluded. Modifying spatial mask will '
                                                            'reset these events based on the mask. Continue?', choices=['Continue', 'Cancel'],
                                      default_choice=1, cancel_choice=1)
        self.spatial_bind_mouse = self.spatial_figure.canvas.mpl_connect('button_press_event', self.get_init_spatial)
        if choice == 0:
            try:
                self.spatial_mask[-1].append([round(event.xdata), round(event.ydata)])
                self.spatial_mask[-1].sort()
                if self.spatial_mask[-1][1][1] < self.spatial_mask[-1][0][1]:
                    self.spatial_mask[-1][1][1], self.spatial_mask[-1][0][1] = self.spatial_mask[-1][0][1], self.spatial_mask[-1][1][1]
            except TypeError:
                self.spatial_mask.pop(-1)
                self.spatial_rect.remove()
                self.spatial_canvas.draw()
            print(self.spatial_mask)
            try:
                self.spatial_rect.remove()
                x1, y1 = self.spatial_mask[-1][1][0] - self.spatial_mask[-1][0][0], self.spatial_mask[-1][1][1] - self.spatial_mask[-1][0][1]
                spatial_rect = mplib.patches.Rectangle(tuple(self.spatial_mask[-1][0]), x1, y1, linewidth=1, edgecolor=(0.35, 0.08, 1, 0.7), facecolor=(0.45, 0, 1, 0.5))
                self.spatial_plotter.add_patch(spatial_rect)
                self.spatial_canvas.draw()
            except Exception:
                'failed'
            self.spatial_mask_binary = np.zeros(shape=np.shape(self.ratiometric_stack)[:2])
            for idx in range(len(self.spatial_mask)):
                self.spatial_mask_binary[self.spatial_mask[idx][0][1]:self.spatial_mask[idx][1][1], self.spatial_mask[idx][0][0]:self.spatial_mask[idx][1][0]] = 1
            for ev in self.events:
                if ev[4] is not None:
                    ev[4] = None
            self.convert_masses()
            self.plot_histogram(None)
        elif choice == 1:
            try:
                self.spatial_mask.pop(-1)
                self.spatial_rect.remove()
                self.spatial_canvas.draw()
            except Exception:
                'failed'

    def get_final_temp(self, event):
        self.temp_figure.canvas.mpl_disconnect(self.temp_bind_mouse)
        self.temp_figure.canvas.mpl_disconnect(self.temp_bind_motion)
        self.capture_motion_temporal = self.temp_canvas.mpl_connect('motion_notify_event', self.temporal_vline)
        manual_modify = False
        choice = 0
        for ev in self.events:
            if ev[4] is not None:
                manual_modify = True
                break
        if manual_modify:
            choice = easygui.indexbox(title='Warning!', msg='Certain events have been manually included/discluded. Modifying temporal mask will '
                                                            'reset these events based on the mask. Continue?', choices=['Continue', 'Cancel'],
                                      default_choice=1, cancel_choice=1)
        self.temp_bind_mouse = self.temp_figure.canvas.mpl_connect('button_press_event', self.get_init_temp)
        if choice == 0:
            try:
                self.temporal_mask[-1].append(int(event.xdata))
                self.temporal_mask[-1].sort()
            except TypeError:
                self.temporal_mask.pop(-1)
                self.temp_rect.remove()
                self.temp_canvas.draw()

            try:
                self.temp_rect.remove()
                self.temp_plotter.axvspan(self.temporal_mask[-1][0], self.temporal_mask[-1][1], facecolor='#7711ff', alpha=0.5)
                self.temp_canvas.draw()
            except Exception:
                'Failed'
            self.temporal_mask_timeline = np.zeros(shape=np.shape(self.ratiometric_stack)[2])
            for idx in range(len(self.temporal_mask)):
                self.temporal_mask_timeline[self.temporal_mask[idx][0]:self.temporal_mask[idx][1]] = 1
            for ev in self.events:
                if ev[4] is not None:
                    ev[4] = None
            self.convert_masses()
            self.plot_histogram(None)
        else:
            try:
                self.temporal_mask.pop(-1)
                self.temp_rect.remove()
                self.temp_canvas.draw()
            except Exception:
                'Failed'
        try:
            self.temporal_vline_text2.remove()
            self.temp_canvas.draw()
        except Exception:
            'Failed'

    def delete_temp(self):
        self.temporal_mask_timeline = np.zeros(shape=np.shape(self.ratiometric_stack)[2])
        self.temporal_mask = []
        self.convert_masses()
        self.plot_temporal()
        self.plot_histogram(None)

    def invert_temp_command(self):
        self.convert_masses()
        self.plot_histogram(None)

    def track_mouse_spatial(self, event):
        try:
            x, y = round(event.xdata), round(event.ydata)
            x1, y1 = x - self.spatial_mask[-1][0][0], y - self.spatial_mask[-1][0][1]
        except TypeError:
            # self.spatial_rect.remove()
            # self.spatial_canvas.draw()
            return
        try:
            self.spatial_rect.remove()
        except Exception:
            """ failed """
        try:
            self.spatial_rect = mplib.patches.Rectangle(tuple(self.spatial_mask[-1][0]), x1, y1, linewidth=1, edgecolor=(0.35, 0.08, 1, 0.7), facecolor=(0.45, 0, 1, 0.5))
            self.spatial_plotter.add_patch(self.spatial_rect)
            self.spatial_canvas.draw()
        except Exception:
            'failed'

    def track_mouse_temporal(self, event):
        try:
            x = int(event.xdata)
        except TypeError:
            # self.temp_rect.remove()
            # self.temp_canvas.draw()
            return
        try:
            self.temp_rect.remove()
            self.temporal_vline_text2.remove()
        except Exception:
            """ failed """
        try:
            x1 = self.temporal_mask[-1][0]
            self.temp_rect = self.temp_plotter.axvspan(x1, x, facecolor='#8833ff', alpha=0.5)
            self.temporal_vline_text2 = self.temp_plotter.text(event.xdata, self.max_contrast / 1.4, f'{round(event.xdata)}',
                                                               color='#007fff', fontsize=10)
            self.temp_canvas.draw()
        except Exception:
            """ failed """

    def delete_spatial(self):
        self.spatial_mask_binary = np.zeros(shape=np.shape(self.ratiometric_stack)[:2])
        self.spatial_mask = []
        self.convert_masses()
        self.plot_spatial()
        self.plot_histogram(None)

    def invert_spatial_command(self):
        self.convert_masses()
        self.plot_histogram(None)

    def auto_fit(self, data_type='events'):
        assert data_type == 'events' or data_type == 'tracks', "data_type parameter declares whether event contrasts or track contrasts are to be fitted amd should be set to 'events' or 'tracks'."
        if data_type == 'events':
            data_to_fit = np.copy(self.contrasts)
        elif data_type == 'tracks':
            data_to_fit = np.copy(self.track_contrasts)

        positive_data_p = []
        negative_data_p = []

        for idx in range(len(data_to_fit)):
            if data_to_fit[idx] > 0:
                positive_data_p.append(data_to_fit[idx])
            elif data_to_fit[idx] < 0:
                negative_data_p.append(data_to_fit[idx])

        try:
            pos_1_percentile, pos_99_percentile = 0, np.percentile(positive_data_p, preferences['hist']['percentile hi'])
        except Exception:
            pos_1_percentile, pos_99_percentile = 0, 1
        try:
            neg_1_percentile, neg_99_percentile = np.percentile(negative_data_p, preferences['hist']['percentile low']), 0
        except Exception:
            neg_1_percentile, neg_99_percentile = -1, 0

        positive_data = []
        negative_data = []

        for idx in range(len(positive_data_p)):
            if pos_1_percentile < positive_data_p[idx] < pos_99_percentile:
                positive_data.append(positive_data_p[idx])

        for idx in range(len(negative_data_p)):
            if neg_1_percentile < negative_data_p[idx] < neg_99_percentile:
                negative_data.append(negative_data_p[idx])

        try:
            bins_pos = int((np.max(positive_data) - np.min(positive_data)) / 0.0002)
        except ValueError:
            ''' Array is empty '''
        try:
            bins_neg = int((np.max(negative_data) - np.min(negative_data)) / 0.0002)
        except ValueError:
            ''' Array is empty '''

        error = False
        err_trace = ""
        fits = []

        def progress_update():
            self.progress_win.progress.step(8)
            self.progress_win.window.update()

        try:
            self.progress_win.window.destroy()
        except Exception:
            pass
        self.progress_win = ProgressWin(title='Fitting', msg='Fitting skewed Gaussian mixture...')
        self.progress_win.progress['mode'] = 'indeterminate'

        try:
            positive_params = self.fit_hist_data(positive_data, comps='auto', bins=bins_pos, penalty=100, callback=progress_update)
            for idx in range(len(positive_params)):
                fits.append(positive_params[idx])
                fits[-1][0][2] = fits[-1][0][2] * (len(positive_data) / len(data_to_fit))
        except Exception:
            err_trace = traceback.format_exc()
            error = True
        try:
            negative_params = self.fit_hist_data(negative_data, comps='auto', bins=bins_neg, penalty=1000, callback=progress_update)
            for idx in range(len(negative_params)):
                fits.append(negative_params[idx])
                fits[-1][0][2] = fits[-1][0][2] * (len(negative_data) / len(data_to_fit))
        except Exception:
            err_trace = traceback.format_exc()
            error = True

        self.progress_win.handle_close()

        self.auto_fits = []
        self.auto_fits = fits
        self.plot_histogram(None)

        if error:
            print(err_trace)
            if preferences['warn']['fiterr']:
                easygui.msgbox(title='Error', msg=f'Error occurred while trying to autofit histogram:\n{str(err_trace)}')

        for idx in range(len(self.auto_fits)):
            print(self.auto_fits[idx])

    def fit_hist_data(self, data, comps, bins, penalty, callback, max_comp=4, max_iter=8000, n_init=5, optimizer='bic'):
        print(f"len(data) = {len(data)}")
        for idx in range(len(data)):
            data[idx] = data[idx] * 40_000  # rescale data to approximate mass - speeds up curve fit significantly
        if len(data) < 50:
            max_comp = 1
        elif 50 <= len(data) < 100:
            max_comp = 2
        elif 100 <= len(data) < 150:
            max_comp = 3
        else:
            pass
        params, cov, fit_func, rms_error = sgmm.fit_skewed_gaussian_mixture(data,
                                                                            num_components=comps,
                                                                            bins=bins,
                                                                            component_optimizer=optimizer,
                                                                            component_penalty=penalty,
                                                                            density=True,
                                                                            plot=False,
                                                                            callback=callback,
                                                                            max_components=max_comp,
                                                                            max_iter=max_iter,
                                                                            n_init=n_init,
                                                                            )
        num_components = len(params) // 4
        stats = []
        for i in range(num_components):
            amp = params[i * 4] * 40_000
            mean = params[i * 4 + 1] / 40_000
            stddev = params[i * 4 + 2] / 40_000
            skewness = params[i * 4 + 3]
            true_mean, true_std, true_skew = sgmm.get_stat_moments(mean, stddev, skewness)
            if amp < 0:
                raise ValueError("Negative amplitude fitted, discarding fits.")
            if stddev > 0 and np.min(data)/40_000 < true_mean < np.max(data)/40_000:
                stats.append([[mean, stddev, amp, skewness], [true_mean, true_std, true_skew, rms_error]])
        return stats

    def bins_mode(self, plot_type=None):
        if self.main_bins_mode.get() == 1:
            if plot_type is None:
                self.bins.set(preferences['hist']['default count'])
            self.bins['from_'] = 10
            self.bins['to'] = 5000
            self.bins['increment'] = 10
        else:
            if self.use_mass.get() == 1:
                self.bins.set(preferences['hist']['default mass'])
                self.bins['from_'] = 0.1
                self.bins['to'] = 100
                self.bins['increment'] = 0.1
            else:
                self.bins.set(preferences['hist']['default contrast'])
                self.bins['from_'] = 0.000002
                self.bins['to'] = 0.0025
                self.bins['increment'] = 0.000005
        self.plot_histogram(plot_type)

    def plot_histogram(self, null):
        self.capture_motion = self.figure.canvas.mpl_connect('motion_notify_event', self.figure_vline)
        if null != 'initialisation':
            if not self.imported_data:
                self.convert_masses()
            self.get_fits()
            try:
                self.get_track_masses()
            except Exception:
                """ No tracks """

        self.figure.clf()
        self.canvas.draw()

        bg_col = "#333333"
        fig_col = "#333333"
        highlight_col = "#007fff"
        text_col = "#cccccc"
        grid_col = "#444444"

        if self.use_mass.get() == 1:
            axis_text = 'Mass (kDa)'
        else:
            axis_text = 'Ratiometric Contrast'

        self.figure.set_facecolor(fig_col)
        self.plotter = self.figure.add_subplot(111)
        self.plotter.set_facecolor(bg_col)
        self.plotter.spines['bottom'].set_color(highlight_col)
        self.plotter.spines['top'].set_color(fig_col)
        self.plotter.spines['left'].set_color(highlight_col)
        self.plotter.spines['right'].set_color(fig_col)
        self.plotter.xaxis.label.set_color(text_col)
        self.plotter.yaxis.label.set_color(text_col)
        self.plotter.tick_params(axis='x', colors=text_col, labelsize=8)
        self.plotter.tick_params(axis='y', colors=text_col, labelsize=8)
        self.plotter.set_xlabel(axis_text, color=text_col, size=9)
        self.plotter.set_ylabel("Density (Normalized Counts)", color=text_col, size=9)
        self.plotter.set_title('Mass / Contrast', color=text_col, size=9)
        self.plotter.grid(color=grid_col)
        if null == 'initialisation':
            self.canvas.draw()
            return

        if self.track_override.get() == 0:
            if self.use_mass.get() == 1:
                data = self.masses
            else:
                data = self.contrasts
        else:
            if self.use_mass.get() == 1:
                data = self.track_masses
            else:
                data = self.track_contrasts

        if self.main_bins_mode.get() == 1:
            bins = int(float(self.bins.get()))
        elif self.main_bins_mode.get() == 2:
                bins = (np.max(data) - np.min(data)) / float(self.bins.get())

        if np.isnan(bins):
            print("NaN in distribution")
            bins = 200

        self.plotter.hist(data, bins=int(bins), density=True, alpha=0.7, ec="#559fff", color="#005faa", histtype='stepfilled')

        if len(self.fits) >= 1:
            self.plotter.plot(self.x_axis, self.fits[-1], color='#00ccff', linewidth=2, alpha=0.6)
            for idx in range(len(self.fits)-1):
                if self.fit_manually.get() == 1:
                    mass, width, amplitude = float(self.field_list[idx].split(",")[1][1:]), float(self.field_list[idx].split(",")[2][1:])/3, float(self.field_list[idx].split(",")[3][1:])
                else:
                    mass, width, amplitude = self.auto_fits[idx][1][0], self.auto_fits[idx][1][1]/3, np.max(self.fits[idx])
                self.plotter.plot(self.x_axis, self.fits[idx], color='#eeeeee', linewidth=1, alpha=0.6)
                if self.use_mass.get() == 1:
                    width = width * self.calibration[0]
                    width_string = f'\nσ = {round(width*3, 1)} kDa'
                    if self.fit_manually.get() == 1:
                        skew_string = ""
                        counts_string = ""
                    else:
                        skew_string = f"\nSkew = {round(self.auto_fits[idx][1][2], 2)}"
                        if self.track_override.get() == 0:
                            norm_condition = len(self.contrasts)
                        else:
                            norm_condition = len(self.track_contrasts)
                        counts_string = f"\n{round(np.trapz(self.fits[idx], self.x_axis) * norm_condition)} Counts"
                    if self.fit_manually.get() == 1:
                        amplitude_correction = self.calibration[0]
                    else:
                        amplitude_correction = 1
                    if mass > 0:
                        self.plotter.text(mass*self.calibration[0]+self.calibration[1] + width*1.25, (amplitude)/amplitude_correction, f'{round(mass*self.calibration[0]+self.calibration[1])} kDa'+width_string+skew_string+counts_string, color='white', fontsize=10)
                    else:
                        self.plotter.text(mass * self.calibration[0] - self.calibration[1] + width*1.25, (amplitude) / amplitude_correction, f'{round(mass * self.calibration[0] - self.calibration[1])} kDa'+width_string+skew_string+counts_string,
                                          color='white', fontsize=10)
                else:
                    self.plotter.text(mass + width, amplitude, f'{round(mass, 6)}', color='white', fontsize=10)

        xmin, xmax = float(self.min_entry.get()), float(self.max_entry.get())
        self.plotter.axvline(0, linewidth=1, color='grey')
        self.plotter.set_xlim(xmin=xmin, xmax=xmax)
        self.plotter.set_ylim(ymin=0, ymax=self.plotter.get_ylim()[1]*1.1)

        self.canvas.draw()

    def get_valid_events(self):
        valid = []
        for event in self.events:
            if event[4] is not None and not event[4] and self.track_use_custom.get() == 1:
                continue
            elif event[4] is None:
                if self.track_use_spatial.get() == 1 and self.track_use_temporal.get() == 1:
                    if self.temporal_mask_timeline[event[0]] == 0 and self.temp_invert.get() == 0:
                        if self.spatial_mask_binary[int(event[2][1]), int(event[2][0])] == 0 and self.spatial_invert.get() == 0:
                            valid.append(event)
                        elif self.spatial_mask_binary[int(event[2][1]), int(event[2][0])] == 1 and self.spatial_invert.get() == 1:
                            valid.append(event)
                    elif self.temporal_mask_timeline[event[0]] == 1 and self.temp_invert.get() == 1:
                        if self.spatial_mask_binary[int(event[2][1]), int(event[2][0])] == 0 and self.spatial_invert.get() == 0:
                            valid.append(event)
                        elif self.spatial_mask_binary[int(event[2][1]), int(event[2][0])] == 1 and self.spatial_invert.get() == 1:
                            valid.append(event)
                elif self.track_use_temporal.get() == 1 and self.track_use_spatial.get() == 0:
                    if self.temporal_mask_timeline[event[0]] == 0 and self.temp_invert.get() == 0:
                        valid.append(event)
                    elif self.temporal_mask_timeline[event[0]] == 1 and self.temp_invert.get() == 1:
                        valid.append(event)
                elif self.track_use_temporal.get() == 0 and self.track_use_spatial.get() == 1:
                    if self.spatial_mask_binary[int(event[2][1]), int(event[2][0])] == 0 and self.spatial_invert.get() == 0:
                        valid.append(event)
                    elif self.spatial_mask_binary[int(event[2][1]), int(event[2][0])] == 1 and self.spatial_invert.get() == 1:
                        valid.append(event)
                elif self.track_use_spatial.get() == 0 and self.track_use_temporal.get() == 0:
                    valid.append(event)
            elif event[4] is not None and event[4] and self.track_use_custom.get() == 1:
                valid.append(event)
            elif self.track_use_custom.get() == 0:
                valid.append(event)
        return valid

    def track_events(self):
        valid_events = list(self.get_valid_events())
        valid_events.sort()
        if len(valid_events) == 0 or len(valid_events) < int(float(self.min_path.get())):
            return

        active_track_list = []
        final_list = []

        active_track_list.append(Track())
        active_track_list[-1].events.append(valid_events[0])
        active_track_list[-1].frames.append(valid_events[0][0])
        active_track_list[-1].coords.append(valid_events[0][2][0:2])
        active_track_list[-1].contrasts.append(valid_events[0][1])
        if valid_events[0][1] < 0:
            active_track_list[-1].type = 'desorption'
        else:
            active_track_list[-1].type = 'adsorption'

        try:
            self.progress_win.window.destroy()
        except Exception:
            pass
        self.progress_win = ProgressWin(title='Analysing', msg='Building trajectories...')
        self.progress_win.progress['maximum'] = len(valid_events) + 11

        use_mc = self.sharpness_mc.get()
        max_disp = float(self.max_displacement.get())
        max_dark = int(float(self.max_dark.get()))
        min_path = int(float(self.min_path.get()))

        for idx, event in enumerate(valid_events[1:]):
            if self.cancel_flag:
                break

            coords = event[2][0:2]
            if event[1] < 0:
                event_type = 'desorption'
            else:
                event_type = 'adsorption'
            frame = event[0]
            if use_mc == 1:
                try:
                    contrast = event[1] / self.sharpness[event[0]]
                except Exception:
                    contrast = event[0]
                    print(f"Error. Could not determine sharpness of event {idx} in valid events list")
            else:
                contrast = event[1]
            new_tracks = []
            tracks_used = []
            event_used = False
            for index, track in enumerate(active_track_list):
                if track.same_type(event_type) and track.get_distance(coords) <= max_disp and frame - track.frames[-1] > 0 and frame - track.frames[-1] <= max_dark and event not in tracks_used:
                    track.events.append(event)
                    track.frames.append(frame)
                    track.coords.append(coords)
                    track.contrasts.append(contrast)
                    tracks_used.append(event)
                    event_used = True
            if not event_used:
                active_track_list.append(Track())
                active_track_list[-1].events.append(event)
                active_track_list[-1].frames.append(frame)
                active_track_list[-1].coords.append(coords)
                active_track_list[-1].contrasts.append(contrast)
                active_track_list[-1].type = event_type
            if idx % 10 == 0:
                self.progress_win.progress.step(10)
                self.progress_win.progress.update()
                self.event_label['text'] = f"Connecting event {idx+1} / {len(self.events)}"
            # remove old tracks to speed up computation
            if idx % 100 == 0:
                for index, track in enumerate(active_track_list):
                    if frame - track.frames[-1] > max_dark and track not in final_list:
                        final_list.append(track)
                for index, track in enumerate(final_list):
                    if track in active_track_list:
                        active_track_list.pop(active_track_list.index(track))

                sublist = []
                for index, track in enumerate(final_list):
                    if len(track.frames) >= min_path:
                        sublist.append(track)
                final_list = sublist

        self.event_label['text'] = ""
        if self.cancel_flag:
            self.cancel_flag = False
            self.progress_win.handle_close()
            return

        for index, track in enumerate(active_track_list):
            if track not in final_list:
                final_list.append(track)

        print(f'Found {len(final_list)} tracks...')
        self.tracks = []
        for _track in final_list:
            if len(_track.frames) >= min_path:
                self.tracks.append(_track)
        print(f'of which {len(self.tracks)} satisfy constraints')
        if len(self.tracks) > 0:
            self.current_track['state'] = tk.NORMAL
            self.current_track['to'] = len(self.tracks)
            self.progress_win.handle_close()
            self.current_track.set(1)
            self.recalculate_plateaus()
            self.update_tracks()
            self.plot_track_masses()
        else:
            self.current_track.set('N/A')
            self.current_track['state'] = tk.DISABLED
            self.progress_win.handle_close()
            self.track_figure.clf()
            self.track_canvas.draw()
            self.mass_figure.clf()
            self.mass_canvas.draw()
            self.time_figure.clf()
            self.time_canvas.draw()

    def get_track_masses(self):
        self.track_masses = []
        self.track_contrasts = []

        if not self.use_plateau_contrasts.get() == 1:
            for idx, track in enumerate(self.tracks):
                if len(track.frames) >= int(float(self.mass_min_track.get())):
                    filtered = []
                    filtered.append(track.contrasts[0])
                    distances = []
                    for i in range(1, len(track.contrasts)):
                        distances.append(np.sqrt((track.coords[i][0] - track.coords[i - 1][0]) ** 2 + (track.coords[i][1] - track.coords[i - 1][1]) ** 2))
                    if np.mean(distances) >= float(self.mass_min_disp.get()):
                        filtered = track.contrasts
                    if len(filtered) > 1:
                        std_filtered = []
                        mean = np.mean(filtered)
                        std = np.std(filtered)
                        for i in range(len(filtered)):
                            if np.abs(filtered[i] - mean) < float(self.mass_std_lim.get()) * std:
                                std_filtered.append(filtered[i])
                        contrast = np.mean(std_filtered)
                        if not np.isnan(contrast):
                            self.track_contrasts.append(contrast)
        else:
            try:
                for idx in self.get_filtered_indices():
                    track = self.tracks[idx]
                    if len(track.plateau_contrasts) > 0:
                        for c in track.plateau_contrasts:
                            if not np.isnan(c):
                                self.track_contrasts.append(c)
            except Exception:
                print(traceback.format_exc())
                print(f'Tracks are from old file and have no plateau attributes')
        for idx in range(len(self.track_contrasts)):
            if self.track_contrasts[idx] > 0:
                self.track_masses.append(self.track_contrasts[idx] * self.calibration[0] + self.calibration[1])
            else:
                self.track_masses.append(self.track_contrasts[idx] * self.calibration[0] - self.calibration[1])

    def get_filtered_indices(self):
        indices = []
        for idx, track in enumerate(self.tracks):
            if len(track.frames) >= int(float(self.mass_min_track.get())):
                filtered = []
                filtered.append(track.contrasts[0])
                distances = []
                for i in range(1, len(track.contrasts)):
                    distances.append(np.sqrt((track.coords[i][0] - track.coords[i - 1][0]) ** 2 + (track.coords[i][1] - track.coords[i - 1][1]) ** 2))
                if np.mean(distances) >= float(self.mass_min_disp.get()):
                    indices.append(idx)
        return indices

    def get_filtered_mean(self, data):
        std_filtered = []
        mean = np.mean(data)
        std = np.std(data)
        for i in range(len(data)):
            if np.abs(data[i] - mean) < float(self.mass_std_lim.get()) * std:
                std_filtered.append(data[i])
        return np.mean(std_filtered), std

    def plot_track_masses(self):
        self.get_track_masses()

        self.mass_figure.clf()
        self.mass_canvas.draw()

        bg_col = "#2f2f2f"
        box_col = "#007fff"
        text_col = "#cccccc"
        grid_col = "#3f3f3f"
        line_col = "lightgray"

        self.mass_figure.set_facecolor(bg_col)
        self.mass_plotter = self.mass_figure.add_subplot(111)
        self.mass_plotter.set_facecolor(bg_col)
        self.mass_plotter.spines['bottom'].set_color(box_col)
        self.mass_plotter.spines['top'].set_color(bg_col)
        self.mass_plotter.spines['left'].set_color(box_col)
        self.mass_plotter.spines['right'].set_color(bg_col)
        self.mass_plotter.xaxis.label.set_color(text_col)
        self.mass_plotter.yaxis.label.set_color(text_col)
        self.mass_plotter.tick_params(axis='x', colors=text_col, labelsize=8)
        self.mass_plotter.tick_params(axis='y', colors=text_col, labelsize=8)
        self.mass_plotter.grid(color=grid_col)
        self.mass_plotter.set_ylabel('Density', size=9, color=text_col)
        self.mass_plotter.set_title('Mass / Contrast', size=9, color=text_col)

        if self.track_use_mass.get() == 1:
            self.mass_plotter.hist(self.track_masses, bins=int(self.mass_bins.get()), density=True, alpha=0.9, ec="white", color="#ff6f00", histtype='stepfilled')
            self.mass_plotter.set_xlabel('Mass (kDa)')
        else:
            self.mass_plotter.hist(self.track_contrasts, bins=int(self.mass_bins.get()), density=True, alpha=0.9, ec="white", color="#ff6f00", histtype='stepfilled')
            self.mass_plotter.set_xlabel('Ratiometric Contrast')
        self.mass_canvas.draw()

    def plot_diffusivity(self):
        coefficients = []
        for idx in range(len(self.tracks)):
            try:
                TAUS, MSD = self.get_msd(idx)
                length = len(MSD)
                if length < 10:
                    prop = int(length // 2) + 1
                else:
                    prop = int(np.sqrt(length)) + 1
                gradient, intercept = linear_regression(TAUS[:prop], MSD[:prop])
                coeff = gradient / 4
                if coeff < 0:
                    coeff = 0
                coefficients.append(coeff)
            except Exception:
                print(f'Failed to calculate coefficient for track {idx + 1}')
        self.track_plotter.hist(coefficients, density=True, alpha=0.5, ec='#22afff', color='#229fff', bins=int(len(self.tracks)/1.5), histtype='stepfilled')
        self.track_plotter.set_xlabel("Diffusivity (µm² / s)", color='#cccccc', size=9)
        self.track_plotter.set_ylabel("Density", color='#cccccc', size=9)
        self.track_canvas.draw()
        
        # dist = []
        # for idx, track in enumerate(self.tracks):
        #     filtered = []
        #     filtered.append(track.contrasts[0])
        #     distances = []
        #     for i in range(1, len(track.contrasts)):
        #         distances.append(np.sqrt((track.coords[i][0] - track.coords[i-1][0])**2 + (track.coords[i][1] - track.coords[i-1][1])**2))
        #         if np.sqrt((track.coords[i][0] - track.coords[i-1][0])**2 + (track.coords[i][1] - track.coords[i-1][1])**2) > 2:
        #             filtered.append(track.contrasts[i])
        #     std_filtered = []
        #     mean = np.mean(filtered)
        #     std = np.std(filtered)
        #     for i in range(len(filtered)):
        #         if np.abs(filtered[i] - mean) < 1.1675*std:
        #             std_filtered.append(filtered[i])
        #     contrast = np.mean(std_filtered)
        #     dist.append(contrast)
        #     # plt.plot(track.contrasts[1:])
        #     # plt.plot(distances)
        #     # for ind in range(0, len(track.contrasts)):
        #     #     if track.contrasts[ind] not in filtered:
        #     #         plt.plot(ind, track.contrasts[ind], linewidth=0, color='red', marker='o', markersize=3)
        #     # plt.title(f'{idx + 1}')
        #     # plt.show()
        # plt.hist(dist, density=True, bins=100)
        # plt.xlim(-0.02, 0.02)
        # plt.title('Track mean contrasts')
        # plt.xlabel('Ratiometric Contrast')
        # plt.ylabel('Density')
        # plt.show()

    def update_tracks(self, **kwargs):
        if 'change_frame' in kwargs:
            change_frame = kwargs['change_frame']
        else:
            change_frame = True
        self.plot_track(change_frame=change_frame)
        self.plot_time_series()

    def match_multimers(self):
        if len(self.auto_fits) <= 2:
            easygui.msgbox(title='Error!', msg=f'Not enough components ({len(self.auto_fits)}). Manual Gaussians are not avaliable for this function.')
            return
        means = [m[1][0] for m in self.auto_fits]
        means = sorted(means)
        monomer_size = easygui.integerbox(title='Enter size', msg='Enter expected monomer mass (kDa).', lowerbound=15, upperbound=5000)
        multimers = easygui.enterbox(title='Enter mutltimers', msg='Enter expected multimerisms, separated by commas in peak order. Number of entries must equal number of peaks.')
        multimers = [float(m)*monomer_size for m in multimers.split(',')]
        grad, inter = linear_regression(means, multimers)
        plt.plot([means[0], means[-1]], [means[0]*grad+inter, means[-1]*grad+inter], linewidth=1, color='black', linestyle='--')
        plt.plot(means, multimers, marker='o', markersize=6, color='#007fff', linewidth=0)
        plt.xlabel('Ratiometric contrast')
        plt.ylabel('Mass of multimer (kDa)')
        plt.title(f"Gradient: {grad}, Intercept: {inter}")
        plt.show()

    def plot_time_hist(self):
        try:
            index = int(float(self.current_track.get())) - 1
        except Exception:
            return
        if len(self.tracks[index].plateau_contrasts) == 0:
            easygui.msgbox(title='Error!', msg='Current track has no plateau fits')
            return
        plt.hist(self.tracks[index].plateau_contrasts, density=False, histtype='stepfilled', color='orange', ec='black', bins=20)
        plt.xlabel('Ratioemtric contrast')
        plt.ylabel('Counts')
        plt.show()

    def recalculate_plateaus(self):
        indices = self.get_filtered_indices()
        for idx in indices:
            try:
                track = self.tracks[idx]
                data = track.contrasts
                if self.use_chung.get() == 1:
                    data = steps.chung_kennedy_filter(data, lambda_value=1)
                _, plateaus = steps.find_changepoints(data,
                                                      window_size=int(float(self.time_rolling_win_size.get())),
                                                      threshold=float(self.time_contrast_threshold.get())/100,
                                                      min_plateau=float(self.time_min_plateau.get())
                                                      )
                self.tracks[idx].plateaus = plateaus
                contrasts = []
                for index, plat in enumerate(plateaus):
                    mean_plateau_contrast, _ = self.get_filtered_mean(track.contrasts[plat[0]:plat[1]])
                    contrasts.append(mean_plateau_contrast)
                self.tracks[idx].plateau_contrasts = contrasts
                if np.std(contrasts) / np.mean(contrasts) < float(self.time_plateau_std_filter.get()):
                    cont, _ = self.get_filtered_mean(data)
                    self.tracks[idx].plateau_contrasts = [cont]
                    self.tracks[idx].plateaus = [[0, len(data)-1]]
            except Exception:
                print(traceback.format_exc())
                print(f'Error ocurred fitting plateaus to track {idx + 1}')
                # plt.plot(data)
                # plt.show()
        self.update_tracks()
        if self.use_plateau_contrasts.get() == 1:
            self.plot_histogram(None)
            self.plot_track_masses()

    def time_plot_filtered_trace(self):
        try:
            index = int(float(self.current_track.get())) - 1
        except (TypeError, ValueError):
            return
        track = self.tracks[index]

        ts = list(track.contrasts)
        filtered = steps.chung_kennedy_filter(ts, lambda_value=1)
        if int(float(self.time_rolling_win_size.get())) > 1:
            filtered = steps.rolling_mean(filtered, window_size=int(float(self.time_rolling_win_size.get())))
        diff = np.diff(filtered)
        max_diff = np.max(np.abs(diff))

        fig, (ax1, ax2, ax3) = plt.subplots(3, 1, figsize=(8, 9))
        fig.subplots_adjust(hspace=0.4, left=0.14, right=0.96, top=0.96, bottom=0.07)
        ax1.plot(ts, color='black')
        ax1.set_title(f'Track {index + 1} - Ratiometric Contrast vs Time', size=11)
        ax1.set_xlabel('Time (frames)', size=11)
        ax1.set_ylabel('Ratiometric Contrast', size=11)

        ax2.plot(filtered, color='black')
        ax2.set_title(f'Chung-Kennedy + Mov. Avrg.', size=11)
        ax2.set_xlabel('Time (frames)', size=11)
        ax2.set_ylabel('Cntrst. Filt.', size=11)

        ax3.plot(diff, color='red')
        ax3.set_title(f'Change (max-diff = {round(max_diff, 6)}  |  Threshold = {float(self.time_contrast_threshold.get())}).', size=11)
        ax3.set_xlabel('Time (frames)', size=11)
        ax3.set_ylabel('Cntrst. Filt. Diff.', size=11)
        ax3.axhline(0, linewidth=1, color='black')
        ax3.axhline(float(self.time_contrast_threshold.get()), linestyle='--', linewidth=1, color='black')
        ax3.axhline(-float(self.time_contrast_threshold.get()), linestyle='--', linewidth=1, color='black')

        plt.show()

    def plot_time_series(self):
        try:
            index = int(float(self.current_track.get())) - 1
        except (TypeError, ValueError):
            return

        self.time_figure.clf()
        self.time_canvas.draw()

        bg_col = "#2f2f2f"
        box_col = "#007fff"
        text_col = "#cccccc"
        grid_col = "#3f3f3f"
        line_col = "lightgray"

        self.time_figure.set_facecolor(bg_col)
        self.time_plotter1 = self.time_figure.add_subplot(211)
        self.time_plotter1.set_facecolor(bg_col)
        self.time_plotter1.spines['bottom'].set_color(box_col)
        self.time_plotter1.spines['top'].set_color(bg_col)
        self.time_plotter1.spines['left'].set_color(box_col)
        self.time_plotter1.spines['right'].set_color(bg_col)
        self.time_plotter1.xaxis.label.set_color(text_col)
        self.time_plotter1.yaxis.label.set_color(text_col)
        self.time_plotter1.tick_params(axis='x', colors=text_col, labelsize=8)
        self.time_plotter1.tick_params(axis='y', colors=text_col, labelsize=8)
        self.time_plotter1.grid(color=grid_col)
        # self.time_plotter1.set_xlabel('Time (frames)', size=9, color=text_col)
        self.time_plotter1.set_ylabel('Displacement (px)', size=9, color=text_col)

        self.time_plotter2 = self.time_figure.add_subplot(212)
        self.time_plotter2.set_facecolor(bg_col)
        self.time_plotter2.spines['bottom'].set_color(box_col)
        self.time_plotter2.spines['top'].set_color(bg_col)
        self.time_plotter2.spines['left'].set_color(box_col)
        self.time_plotter2.spines['right'].set_color(bg_col)
        self.time_plotter2.xaxis.label.set_color(text_col)
        self.time_plotter2.yaxis.label.set_color(text_col)
        self.time_plotter2.tick_params(axis='x', colors=text_col, labelsize=8)
        self.time_plotter2.tick_params(axis='y', colors=text_col, labelsize=8)
        self.time_plotter2.grid(color=grid_col)
        self.time_plotter2.set_xlabel('Time (frames)', size=9, color=text_col)
        self.time_plotter2.set_ylabel('Contrast', size=9, color=text_col)

        track = self.tracks[index]
        track.displacements = [np.nan]
        for idx in range(1, len(track.coords)):
            dist = np.sqrt((track.coords[idx][0] - track.coords[idx-1][0])**2 + (track.coords[idx][1] - track.coords[idx-1][1])**2)
            track.displacements.append(dist)
        mean_disp = np.nanmean(track.displacements)

        self.time_plotter1.plot(track.frames, track.displacements, linewidth=1, color='lightgray')
        self.time_plotter1.plot([track.frames[0], track.frames[-1]], [mean_disp, mean_disp], linewidth=1, color='#007fff', linestyle='--')
        self.time_plotter2.plot(track.frames, track.contrasts, linewidth=1, color='#ff6f00')

        mean, std = self.get_filtered_mean(track.contrasts)
        stdlim = std * float(self.mass_std_lim.get())
        low, hi = mean - stdlim, mean + stdlim
        self.time_plotter2.plot([track.frames[0], track.frames[-1]], [low, low], linewidth=1, linestyle='--', color='white', alpha=0.4, label='_nolegend_')
        self.time_plotter2.plot([track.frames[0], track.frames[-1]], [hi, hi], linewidth=1, linestyle='--', color='white', alpha=0.4, label='_nolegend_')
        self.time_plotter2.fill_between([track.frames[0], track.frames[-1]], [low, low], [hi, hi], color='white', alpha=0.2, label='_nolegend_')
        try:
            for idx, p in enumerate(track.plateaus):
                mean, _ = self.get_filtered_mean(track.contrasts[p[0]:p[1]])
                self.time_plotter2.plot([track.frames[p[0]], track.frames[p[1]]], [mean, mean], color='white', linewidth=1)
        except Exception:
            print(traceback.format_exc())
            print(f"Failed to plot time series fit for track {index + 1}")
        # plt.hist(track.contrasts, bins=100)
        # plt.show()

        self.time_plotter1.set_xlim(self.time_plotter2.get_xlim())

        self.time_canvas.draw()


    def plot_track(self, **kwargs):
        if 'change_frame' in kwargs:
            if not kwargs['change_frame']:
                change_frame = False
            else:
                change_frame = True
        else:
            change_frame = True
        try:
            index = int(float(self.current_track.get())) - 1
        except (TypeError, ValueError):
            return

        x, y = [c[0] for c in self.tracks[index].coords], [c[1] for c in self.tracks[index].coords]
        frames = [self.tracks[index].frames[0], self.tracks[index].frames[-1]]
        TAUS, MSD = self.get_msd(index)
        print(TAUS)
        print(MSD)
        event_type = f'{self.tracks[index].type}'
        event_type = event_type[0].upper() + event_type[1:]
        self.track_type_label['text'] = event_type

        self.track_figure.clf()
        self.track_canvas.draw()

        bg_col = "#2f2f2f"
        box_col = "#007fff"
        text_col = "#cccccc"
        grid_col = "#3f3f3f"
        line_col = "lightgray"

        self.track_figure.set_facecolor(bg_col)
        self.track_plotter = self.track_figure.add_subplot(111)
        self.track_plotter.set_facecolor(bg_col)
        self.track_plotter.spines['bottom'].set_color(box_col)
        self.track_plotter.spines['top'].set_color(bg_col)
        self.track_plotter.spines['left'].set_color(box_col)
        self.track_plotter.spines['right'].set_color(bg_col)
        self.track_plotter.xaxis.label.set_color(text_col)
        self.track_plotter.yaxis.label.set_color(text_col)
        self.track_plotter.tick_params(axis='x', colors=text_col, labelsize=8)
        self.track_plotter.tick_params(axis='y', colors=text_col, labelsize=8)
        self.track_plotter.grid(color=grid_col)
        if self.track_plot_mode.get() == 1:
            # if self.tracks[index].type == 'adsorption':
            #     tc = '#004faa'
            self.track_plotter.plot(x, y, linewidth=1, color=line_col)
            if self.annotate_tracks.get() == 1:
                self.track_plotter.text(x[0], y[0], f"Start {frames[0]}", size=8, color="#007fff")
                self.track_plotter.text(x[-1], y[-1], f"End {frames[1]}", size=8, color="orange")
            if change_frame:
                self.slider_ratio.set(self.tracks[index].frames[0])
            else:
                self.display_frame_ratio(int(float(self.ratio_frame_var.get())))
        elif self.track_plot_mode.get() == 2:
            mplstyle.use("fast")
            self.display_frame_ratio(int(float(self.ratio_frame_var.get())))
            indices = self.get_filtered_indices()
            for idx_t in indices:
                track = self.tracks[idx_t]
                if track.type == 'adsorption' and self.track_show_binding.get() == 1 and self.track_show_unbinding.get() == 0:
                    x, y = [c[0] for c in track.coords], [c[1] for c in track.coords]
                    frames = [track.frames[0], track.frames[-1]]
                    self.track_plotter.plot(x, y, linewidth=1, color=self.track_cols[self.tracks.index(track) % len(self.track_cols)])
                    if self.annotate_tracks.get() == 1:
                        self.track_plotter.text(x[0], y[0], f"{frames[0]}", size=8, color="#007fff")
                        self.track_plotter.text(x[-1], y[-1], f"{frames[1]}", size=8, color="orange")
                    self.track_type_label['text'] = 'Adsorption'
                elif track.type == 'desorption' and self.track_show_binding.get() == 0 and self.track_show_unbinding.get() == 1:
                    x, y = [c[0] for c in track.coords], [c[1] for c in track.coords]
                    frames = [track.frames[0], track.frames[-1]]
                    self.track_plotter.plot(x, y, linewidth=1)
                    if self.annotate_tracks.get() == 1:
                        self.track_plotter.text(x[0], y[0], f"{frames[0]}", size=8, color="#007fff")
                        self.track_plotter.text(x[-1], y[-1], f"{frames[1]}", size=8, color="orange")
                    self.track_type_label['text'] = 'Adsorption'
                elif self.track_show_binding.get() == 1 and self.track_show_unbinding.get() == 1:
                    x, y = [c[0] for c in track.coords], [c[1] for c in track.coords]
                    if track.type == 'adsorption':
                        self.track_plotter.plot(x, y, linewidth=1, color='#004faa')
                    else:
                        self.track_plotter.plot(x, y, linewidth=1, color='#ff7f22')
                    self.track_type_label['text'] = 'All Tracks'
                elif self.track_show_binding.get() == 0 and self.track_show_unbinding.get() == 0:
                    pass
            mplstyle.use("default")
        elif self.track_plot_mode.get() == 3:
            self.plot_diffusivity()
            return
        self.track_plotter.axis("square")
        self.track_plotter.set_ylim(max(self.track_plotter.get_ylim()),
                                    min(self.track_plotter.get_ylim()))

        self.track_plotter.set_xlabel("x displacement (px)", color=text_col, size=9)
        self.track_plotter.set_ylabel("y displacement (px)", color=text_col, size=9)
        self.track_canvas.draw()

    def get_msd(self, track_index):
        track = self.tracks[track_index]
        tau_series = dict()
        start, end = track.frames[0], track.frames[-1]
        max_tau = end - start - 1
        max_delta_n = len(track.frames) - 2
        for tau in range(1, max_tau):
            tau_series[tau] = []
        calib = float(self.pixel_size.get()) / 1000
        frame_time = float(self.frame_interval.get()) / 1000
        for delta_n in range(1, max_delta_n):
            for offset in range(max_delta_n - delta_n):
                coords_s, coords_e = track.coords[offset], track.coords[offset + delta_n]
                tau = track.frames[offset + delta_n] - track.frames[offset]
                SD = (coords_s[0]*calib - coords_e[0]*calib)**2 + (coords_s[1]*calib - coords_e[1]*calib)**2
                tau_series[tau].append(SD)
        TAUS = []
        MSD = []
        for tau in tau_series:
            if len(tau_series[tau]) > 0:
                MSD.append(np.mean(tau_series[tau]))
                TAUS.append(tau*frame_time)
            else:
                MSD.append(np.nan)
                TAUS.append(tau*frame_time)
        return TAUS, MSD

    def add_dataset(self):
        try:
            if len(self.contrasts) == 0:
                easygui.msgbox(title='Warning!', msg='No file loaded or analysed. Cannot create dataset.')
                return
            if not self.imported_data:
                path = os.path.split(self.file_path)[1]
            else:
                path = 'Imported Data'
            self.all_datasets.append(Dataset())
            self.dataset_index = len(self.all_datasets)
            self.all_datasets[-1].ID = f'Dataset-{self.dataset_index}'
            self.all_datasets[-1].name = f'Dataset {self.dataset_index}'
            self.all_datasets[-1].filename = path
            self.treeview.insert(parent='', index="end", iid=self.all_datasets[-1].ID, text=self.all_datasets[-1].name,
                                 values=(self.all_datasets[-1].filename, self.all_datasets[-1].info))
            self.treeview.see(self.all_datasets[-1].ID)
            self.treeview.focus(self.all_datasets[-1].ID)
            self.treeview.selection_set(self.all_datasets[-1].ID)
        except _tkinter.TclError:
            self.add_dataset()
        except Exception:
            print(traceback.format_exc())

    def delete_dataset(self):
        parent_ID = self.treeview.focus()
        print(parent_ID)
        if len(self.treeview.get_children(parent_ID)) == 0:
            if self.treeview.parent(parent_ID) != "":
                parent_ID = self.treeview.parent(parent_ID)
                print(parent_ID)
        index = 0
        for idx in range(len(self.all_datasets)):
            if self.all_datasets[idx].ID == parent_ID:
                index = idx
        try:
            self.all_datasets.pop(index)
            self.treeview.delete(parent_ID)
        except Exception:
            print(traceback.format_exc())
        if len(self.treeview.get_children('')) == 0:
            self.all_datasets = []

    def treeview_clicked(self, event):
        pass
        # print(event)
        # print(self.treeview.focus())

    def treeview_double_clicked(self, event):
        # print('\nDouble clicked!')
        # print(event.x, event.y)
        parent_ID = self.treeview.focus()
        # print(parent_ID)
        if parent_ID != '' and 'Dataset' in parent_ID:
            self.add_data_win = AddDatasetWin(parent_ID, event.x, event.y)

    def treeview_select(self):
        parent_ID = self.treeview.focus()
        if parent_ID != '' and 'Dataset' in parent_ID:
            self.add_data_win = AddDatasetWin(parent_ID, 400, 40)

    def fig_list_clicked(self, event):
        self.plot_figure()

    def fig_list_double_clicked(self, event):
        self.plot_figure()
        self.plot_fullsize_fig()

    def fig_list_right_clicked(self, event):
        parent_ID = self.figure_listbox.focus()
        if 'Figure' in parent_ID:
            self.dataset_temp_rebind()
            self.cancel_fig_select_data_btn.place(x=760, y=100, width=330, height=60)
        else:
            self.edit_plot_properties(parent_ID)

    def edit_plot_properties(self, event):
        pass

    def dataset_temp_rebind(self):
        self.window.bind('<Escape>', self.cancel_fig_select_data)
        self.analysis_notebook.select(self.analysis_tab2)
        self.treeview.unbind(self.bind_double_tree)
        self.bind_double_tree = self.treeview.bind('<Double-1>', self.fig_select_data)

    def cancel_fig_select_data(self, *args):
        self.window.unbind('<Escape>')
        self.analysis_notebook.select(self.analysis_tab3)
        self.treeview.unbind(self.bind_double_tree)
        self.bind_double_tree = self.treeview.bind('<Double-1>', self.treeview_double_clicked)
        self.cancel_fig_select_data_btn.place_forget()


    def create_figure(self):
        index = len(self.all_figures) + 1
        self.all_figures.append(Figure())
        self.all_figures[-1].ID = f'Figure-{index}'
        self.all_figures[-1].name = f'Figure {index}'
        self.all_figures[-1].title = 'New Figure'
        self.all_figures[-1].fig_params['xlabel'] = 'x axis label'
        self.all_figures[-1].fig_params['ylabel'] = 'y axis label'
        self.all_figures[-1].fig_params['axsize'] = 22
        self.all_figures[-1].fig_params['titlesize'] = 28
        self.all_figures[-1].fig_params['ansize'] = 20
        self.all_figures[-1].fig_params['binw'] = 0.0001
        self.all_figures[-1].fig_params['scale_inc'] = 0.0001
        self.all_figures[-1].fig_params['left'] = 0.15
        self.all_figures[-1].fig_params['right'] = 0.96
        self.all_figures[-1].fig_params['top'] = 0.92
        self.all_figures[-1].fig_params['bottom'] = 0.12
        self.all_figures[-1].fig_params['xmin'] = 0
        self.all_figures[-1].fig_params['xmax'] = 1
        try:
            self.figure_listbox.insert(parent='', index="end", iid=self.all_figures[-1].ID, text=self.all_figures[-1].name,
                                       values=(self.all_figures[-1].title, None))
            self.figure_listbox.see(self.all_figures[-1].ID)
        except _tkinter.TclError:
            print(traceback.format_exc())
            self.create_figure()

    def delete_figure(self):
        fig_ID = self.figure_listbox.focus()
        if len(self.figure_listbox.get_children(fig_ID)) == 0:
            fig_ID = self.figure_listbox.parent(fig_ID)
            if fig_ID == '':
                fig_ID = self.figure_listbox.focus()
        print('\n', fig_ID, '\n')

        self.figure_listbox.delete(fig_ID)

        index = 0
        for idx in range(len(self.all_figures)):
            if self.all_figures[idx].ID == fig_ID:
                index = idx
        self.all_figures.pop(index)
        self.fig_figure.clf()
        self.fig_canvas.draw()

        if len(self.figure_listbox.get_children('')) == 0:
            self.all_figures = []
        try:
            self.disp_3d_fig.place_forget()
        except Exception:
            'Unplaced'

    def fig_select_data(self, event):
        self.cancel_fig_select_data_btn.place_forget()
        parent_ID = self.figure_listbox.focus()
        ID_data = self.treeview.focus()
        if len(self.all_datasets) == 0:
            self.treeview.unbind(self.bind_double_tree)
            self.bind_double_tree = self.treeview.bind('<Double-1>', self.treeview_double_clicked)
            easygui.msgbox(title='Warning!', msg='No datasets available to add to figure.')
            return
        if 'Dataset' in ID_data:
            easygui.msgbox(title='Warning!', msg='Cannot add superset to figure. Please select a single data series to plot (child node).')
            self.treeview.unbind(self.bind_double_tree)
            self.bind_double_tree = self.treeview.bind('<Double-1>', self.treeview_double_clicked)
            return
        else:
            self.treeview.unbind(self.bind_double_tree)
            self.bind_double_tree = self.treeview.bind('<Double-1>', self.treeview_double_clicked)
            try:
                index = 0
                for idx in range(len(self.all_figures)):
                    if self.all_figures[idx].ID == parent_ID:
                        index = idx
                ID_data = self.treeview.focus()
                data_index = 0
                data_subindex = 0
                for idx in range(len(self.all_datasets)):
                    for subidx in range(len(self.all_datasets[idx].subsets)):
                        if self.all_datasets[idx].subsets[subidx].ID == ID_data:
                            data_index = idx
                            data_subindex = subidx
                self.all_figures[index].datasets.append(deepcopy(self.all_datasets[data_index].subsets[data_subindex]))
                new_ID = self.all_figures[index].datasets[-1].ID
                new_name = self.all_datasets[data_index].name
                new_info = self.all_figures[index].datasets[-1].name
                self.all_figures[index].datasets[-1].data_parent = new_name
                if not 'All Tracks' in new_info and not 'Time' in new_info:
                    try:
                        xmin, xmax = np.nanmin(self.all_figures[index].datasets[0].data), np.nanmax(self.all_figures[index].datasets[0].data)
                    except Exception:
                        print(traceback.format_exc())
                        print(' - could not calculate axis limits')
                    for idx in range(len(self.all_figures[index].datasets)):
                        min_data = np.nanmin(self.all_figures[index].datasets[idx].data)
                        max_data = np.nanmax(self.all_figures[index].datasets[idx].data)
                        if min_data < xmin and not np.isnan(min_data):
                            xmin = min_data
                        if max_data > xmax and not np.isnan(max_data):
                            xmax = max_data
                if 'Masses' in new_info:
                    self.all_figures[index].fig_params['xlabel'] = 'Mass (kDa)'
                    self.all_figures[index].fig_params['ylabel'] = 'Counts'
                    self.all_figures[index].fig_params['scale_inc'] = 20
                    self.all_figures[index].fig_params['xmin'] = round(xmin)
                    self.all_figures[index].fig_params['xmax'] = round(xmax)
                    self.all_figures[index].fig_params['binw'] = 5
                elif 'Cont' in new_info:
                    self.all_figures[index].fig_params['xlabel'] = 'Ratiometric Contrast'
                    self.all_figures[index].fig_params['ylabel'] = 'Counts'
                    self.all_figures[index].fig_params['scale_inc'] = 0.0001
                    self.all_figures[index].fig_params['xmin'] = round(xmin, 3)
                    self.all_figures[index].fig_params['xmax'] = round(xmax, 3)
                    self.all_figures[index].fig_params['binw'] = round(5 / self.calibration[0], 6)
                elif 'Diff' in new_info:
                    self.all_figures[index].fig_params['xlabel'] = 'Diffusivity (µm² / s)'
                    self.all_figures[index].fig_params['ylabel'] = 'Probability Density'
                    self.all_figures[index].fig_params['scale_inc'] = 0.01
                    self.all_figures[index].fig_params['xmin'] = round(xmin, 3)
                    self.all_figures[index].fig_params['xmax'] = round(xmax, 3)
                    self.all_figures[index].fig_params['binw'] = 0.001
                elif 'D vs Mass' in new_info:
                    self.all_figures[index].fig_params['xlabel'] = 'Mass (kDa)'
                    self.all_figures[index].fig_params['ylabel'] = 'Diffusivity (µm² / s)'
                    xmin, xmax = np.nanmin(self.all_figures[index].datasets[0].data[0]), np.nanmax(self.all_figures[index].datasets[0].data[0])
                    self.all_figures[index].fig_params['scale_inc'] = 20
                    self.all_figures[index].fig_params['xmin'] = round(xmin, 3)
                    self.all_figures[index].fig_params['xmax'] = round(xmax, 3)
                    self.all_figures[index].fig_params['ansize'] = 5

                elif 'Calib' in new_info:
                    self.all_figures[index].fig_params['xlabel'] = 'Calibrant Ratiometric Contrast'
                    self.all_figures[index].fig_params['ylabel'] = 'Calibrant Mass (kDa)'
                    self.all_figures[index].title = str(self.all_figures[index].datasets[idx].metadata)


                print(self.all_figures[index].ID+new_ID)
                self.all_figures[index]
                self.figure_listbox.insert(parent=parent_ID, index="end", iid=self.all_figures[index].ID[0:3]+self.all_figures[index].ID[6:]+new_ID, text=new_name,
                                           values=(new_info, None))
                self.figure_listbox.see(self.all_figures[index].ID)
                self.analysis_notebook.select(self.analysis_tab3)
                self.plot_figure()
            except Exception:
                err_text = traceback.format_exc()
                print(err_text)
                self.all_figures[index].datasets.pop(-1)
                easygui.msgbox(title='Warning!', msg=f'{err_text}')

    def update_figure(self):
        parent_ID = self.figure_listbox.focus()
        if not 'Figure' in parent_ID:
            parent_ID = self.figure_listbox.parent(parent_ID)
            self.figure_listbox.selection_set(parent_ID)
        index = 0
        for idx in range(len(self.all_figures)):
            if self.all_figures[idx].ID == parent_ID:
                index = idx
        self.all_figures[index].title = str(self.figure_title.get())
        if 'Figure' in parent_ID:
            self.figure_listbox.item(parent_ID, values=(self.all_figures[index].title, None))
        self.all_figures[index].fig_params['xlabel'] = str(self.xaxis_title.get())
        self.all_figures[index].fig_params['ylabel'] = str(self.yaxis_title.get())
        self.all_figures[index].fig_params['axsize'] = int(float(self.axis_size.get()))
        self.all_figures[index].fig_params['titlesize'] = int(float(self.title_size.get()))
        if float(self.fig_right_align.get()) < float(self.fig_left_align.get()):
            self.fig_right_align.set(float(self.fig_left_align.get()) + 0.1)
        if float(self.fig_top_align.get()) < float(self.fig_bottom_align.get()):
            self.fig_top_align.set(float(self.fig_bottom_align.get()) + 0.1)
        if float(self.fig_xmax.get()) < float(self.fig_xmin.get()):
            if self.all_figures[index].fig_params['scale_inc'] == 20:
                self.fig_xmax.set(float(self.fig_xmin.get()) + 20)
            else:
                self.fig_xmax.set(float(self.fig_xmin.get()) + 0.0001)
        self.all_figures[index].fig_params['left'] = float(self.fig_left_align.get())
        self.all_figures[index].fig_params['right'] = float(self.fig_right_align.get())
        self.all_figures[index].fig_params['top'] = float(self.fig_top_align.get())
        self.all_figures[index].fig_params['bottom'] = float(self.fig_bottom_align.get())
        self.all_figures[index].fig_params['xmin'] = float(self.fig_xmin.get())
        self.all_figures[index].fig_params['xmax'] = float(self.fig_xmax.get())
        self.all_figures[index].fig_params['binw'] = float(self.fig_bin_width.get())
        self.all_figures[index].fig_params['ansize'] = int(float(self.fig_an_size.get()))
        self.fig_figure.subplots_adjust(left=self.all_figures[index].fig_params['left'],
                                        right=self.all_figures[index].fig_params['right'],
                                        top=self.all_figures[index].fig_params['top'],
                                        bottom=self.all_figures[index].fig_params['bottom'])
        self.fig_xmin['increment'] = round(abs(float(self.fig_xmin.get()) / 10), 4)
        self.fig_xmax['increment'] = round(abs(float(self.fig_xmax.get()) / 10), 4)
        self.fig_bin_width['increment'] = round(abs(float(self.fig_bin_width.get()) / 10), 6)
        print(self.all_figures[index].datasets[0].name)
        self.plot_figure()

    def import_fig_params_to_GUI(self, param_dict, title):
        self.figure_title.delete(0, tk.END)
        self.xaxis_title.delete(0, tk.END)
        self.yaxis_title.delete(0, tk.END)
        self.figure_title.insert(0, title)
        self.xaxis_title.insert(0, param_dict['xlabel'])
        self.yaxis_title.insert(0, param_dict['ylabel'])
        self.axis_size.set(param_dict['axsize'])
        self.fig_an_size.set(param_dict['ansize'])
        self.title_size.set(param_dict['titlesize'])
        self.fig_left_align.set(param_dict['left'])
        self.fig_right_align.set(param_dict['right'])
        self.fig_top_align.set(param_dict['top'])
        self.fig_bottom_align.set(param_dict['bottom'])
        self.fig_xmin.set(param_dict['xmin'])
        self.fig_xmax.set(param_dict['xmax'])
        self.fig_bin_width.set(param_dict['binw'])
        if param_dict['scale_inc'] == 20:
            self.fig_xmin['from_'] = -10000
            self.fig_xmin['to'] = 10000
            self.fig_xmax['from_'] = -10000
            self.fig_xmax['to'] = 10000
            self.fig_bin_width['to'] = 200
            self.fig_bin_width['increment'] = 1
        else:
            self.fig_xmin['from_'] = -1
            self.fig_xmin['to'] = 1
            self.fig_xmax['from_'] = -1
            self.fig_xmax['to'] = 1
            self.fig_bin_width['to'] = 0.1
            self.fig_bin_width['increment'] = 0.00001
        self.fig_figure.subplots_adjust(left=param_dict['left'], right=param_dict['right'], top=param_dict['top'], bottom=param_dict['bottom'])

    def clear_all_datasets(self):
        choice = easygui.indexbox(title='Warning!', msg='Clear all datasets in the workspace.\nAre you sure?', choices=['Confirm Delete', 'Cancel'], default_choice=1)
        if choice == 0:
            for item in self.treeview.get_children():
                self.treeview.delete(item)
            self.all_datasets = []


    def clear_all_figures(self):
        choice = easygui.indexbox(title='Warning!', msg='Clear all figures in the workspace.\nAre you sure?', choices=['Confirm Delete', 'Cancel'], default_choice=1)
        if choice == 0:
            for item in self.figure_listbox.get_children():
                self.figure_listbox.delete(item)
            self.all_figures = []
            self.fig_figure.clf()
            self.fig_canvas.draw()
            try:
                self.disp_3d_fig.place_forget()
            except Exception:
                'Unplaced'

    def get_fig_params(self, index):
        param_dict = self.all_figures[index].fig_params
        title = self.all_figures[index].title
        xlabel, ylabel, axsize, titlesize, ansize = param_dict['xlabel'], param_dict['ylabel'], param_dict['axsize'], param_dict['titlesize'], param_dict['ansize']
        self.import_fig_params_to_GUI(param_dict, title)
        if 'Event Masses' in self.all_figures[index].datasets[0].name or 'Event Contrasts' in self.all_figures[index].datasets[0].name or\
                'Track Masses' in self.all_figures[index].datasets[0].name or 'Track Contrasts' in self.all_figures[index].datasets[0].name:
            try:
                self.disp_3d_fig.place(x=520, y=312, width=150)
            except Exception:
                'Placed'
        else:
            try:
                self.disp_3d_fig.place_forget()
            except Exception:
                'Unplaced'
        return title, xlabel, ylabel, axsize, titlesize, ansize

    def plot_figure(self):
        if len(self.all_figures) == 0:
            return
        selected_ID = self.figure_listbox.focus()
        index = 0
        if 'Figure' not in selected_ID:
            selected_ID = self.figure_listbox.parent(selected_ID)
        for idx in range(len(self.all_figures)):
            if self.all_figures[idx].ID == selected_ID:
                index = idx
        title, xlabel, ylabel, axsize, titlesize, ansize = self.get_fig_params(index)

        bg_col = "#ffffff"
        box_col = "black"
        text_col = "#333333"
        grid_col = "#bbbbbb"
        line_col = "black"
        face_col = '#ffffff'

        if len(self.all_figures[index].datasets) == 0:
            return

        self.fig_figure.clf()
        self.fig_canvas.draw()

        self.fig_figure.set_facecolor(bg_col)
        self.fig_plotter = self.fig_figure.add_subplot(111)
        self.fig_plotter.set_facecolor(face_col)
        self.fig_plotter.spines['bottom'].set_color(box_col)
        self.fig_plotter.spines['top'].set_color(bg_col)
        self.fig_plotter.spines['left'].set_color(box_col)
        self.fig_plotter.spines['right'].set_color(bg_col)
        for axis in ['top', 'bottom', 'left', 'right']:
            self.fig_plotter.spines[axis].set_linewidth(2)
        self.fig_plotter.tick_params(width=2)
        self.fig_plotter.xaxis.label.set_color(text_col)
        self.fig_plotter.yaxis.label.set_color(text_col)
        self.fig_plotter.tick_params(axis='x', colors=text_col, labelsize=axsize)
        self.fig_plotter.tick_params(axis='y', colors=text_col, labelsize=axsize)
        if self.fig_grid.get() == 1:
            self.fig_plotter.grid(color=grid_col)
        self.fig_plotter.set_xlabel(xlabel, size=axsize, color=text_col)
        self.fig_plotter.set_ylabel(ylabel, size=axsize, color=text_col)
        self.fig_plotter.set_title(title, size=titlesize, color=text_col)
        cycle = plt.rcParams['axes.prop_cycle']
        cols = cycle.by_key()['color']
        xmin, xmax = self.all_figures[index].fig_params['xmin'], self.all_figures[index].fig_params['xmax']
        leg = []
        image = False
        for idx in range(len(self.all_figures[index].datasets)):
            if 'Ratio' in self.all_figures[index].datasets[idx].name:
                image = True
                imshape = np.shape(self.all_figures[index].datasets[idx].data)
            data = self.all_figures[index].datasets[idx].data
            x_ax = self.all_figures[index].datasets[idx].x_axis
            fits = self.all_figures[index].datasets[idx].fits
            moments = self.all_figures[index].datasets[idx].fit_vals
            try:
                calib = self.all_figures[index].datasets[idx].calib
            except Exception:
                calib = [self.calibration[0], self.calibration[1]]
            fit_col = cols[int(idx % len(cols))]
            if 'Ratio' not in self.all_figures[index].datasets[idx].name:
                leg.append(self.all_figures[index].datasets[idx].name)

            if 'Contrasts' in self.all_figures[index].datasets[idx].name or 'Masses' in self.all_figures[index].datasets[idx].name or 'Diff' in self.all_figures[index].datasets[idx].name:
                data_xmin, data_xmax = np.min(data), np.max(data)
                bins = round((data_xmax - data_xmin) / self.all_figures[index].fig_params['binw'])
                if bins == 0:
                    print('Error due to no calibration, setting bins manually to 10')
                    bins = 10
                if self.fig_use_density.get() == 0:
                    density = True
                    factor = 1
                else:
                    density = False
                self.fig_plotter.hist(data, bins=bins, density=density, alpha=0.5, ec="black", histtype='stepfilled')

                if len(fits) >= 1:
                    if self.fig_use_density.get() == 1:
                        factor = self.all_figures[index].fig_params['binw'] * len(data)
                    self.fig_plotter.plot(x_ax, fits[-1] * factor, color=fit_col, linewidth=2, label='_nolegend_', alpha=0.9)
                    for idx2 in range(len(fits) - 1):
                        if len(moments[idx2]) == 3:
                            mass, width, amplitude = moments[idx2]
                            skew_string = ""
                            true_mean, true_sigma = mass, width
                        else:
                            mass, width, amplitude = moments[idx2][0][:-1]
                            true_mean, true_sigma, skew = moments[idx2][1][:-1]
                            skew_string = f"\nSkew = {round(skew, 3)}"
                        counts_string = f"\n{round(np.trapz(fits[idx2], x_ax)*len(data))} Counts"
                        self.fig_plotter.plot(x_ax, fits[idx2]*factor, color=fit_col, linewidth=2, label='_nolegend_', alpha=0.6)
                        if self.fig_annotations.get() == 1:
                            if 'Masses' in self.all_figures[index].datasets[idx].name:
                                width_string = f'\nσ = {round(true_sigma, 1)} kDa'
                                # self.fig_plotter.text(mass + width * 1.25, amplitude * factor, f'{round(mass)} kDa' + width_string, color=fit_col, fontsize=ansize)
                                self.fig_plotter.annotate(f'{round(true_mean)} kDa' + width_string + skew_string + counts_string, xy=(true_mean, (np.max(fits[idx2]) + np.max(fits[-1])*0.03)*factor),
                                                          xycoords='data', xytext=(40, 15), textcoords='offset points', arrowprops=dict(arrowstyle="-",
                                                                                                                                        connectionstyle=f"angle,angleA=0,angleB=90,rad=5",
                                                                                                                                        color=fit_col, linewidth=1),
                                                          fontsize=ansize, color=fit_col).draggable()
                            else:
                                width_string = f'\nσ = {round(width, 6)}'
                                self.fig_plotter.annotate(f'{round(true_mean, 5)}' + width_string + skew_string + counts_string, xy=(true_mean, (np.max(fits[idx2]) + np.max(fits[-1])*0.03)*factor),
                                                          xycoords='data', xytext=(40, 15), textcoords='offset points', arrowprops=dict(arrowstyle="-",
                                                                                                                                        connectionstyle=f"angle,angleA=0,angleB=90,rad=5",
                                                                                                                                        color=fit_col, linewidth=1),
                                                          fontsize=ansize, color=fit_col).draggable()

            elif 'Coords' in self.all_figures[index].datasets[idx].name:
                x, y = [c[0] for c in data], [c[1] for c in data]
                self.fig_plotter.plot(x, y, linewidth=2)
                self.fig_plotter.axis("square")

            elif 'All Tracks' in self.all_figures[index].datasets[idx].name:
                for tr_idx, track in enumerate(data):
                    x, y = [c[0] for c in track], [c[1] for c in track]
                    self.fig_plotter.plot(x, y, linewidth=1)
                self.fig_plotter.axis("square")


            elif 'Time' in self.all_figures[index].datasets[idx].name:
                plats = None
                stdlim_stored = float(self.mass_std_lim.get())
                if len(data) == 2:
                    data, plats = data
                elif len(data) == 3:
                    data, plats, stdlim_stored = data
                self.fig_plotter.plot(data, linewidth=2)
                if self.fig_annotations.get() == 1:
                    mean, std = self.get_filtered_mean(data)
                    stdlim = std * stdlim_stored
                    low, hi = mean - stdlim, mean + stdlim
                    self.fig_plotter.plot([0, len(data)-1], [low, low], linewidth=1, linestyle='--', color=fit_col, alpha=0.5, label='_nolegend_')
                    self.fig_plotter.plot([0, len(data) - 1], [hi, hi], linewidth=1, linestyle='--', color=fit_col, alpha=0.5, label='_nolegend_')
                    self.fig_plotter.plot([0, len(data) - 1], [mean, mean], linewidth=1, color=fit_col, alpha=0.8, label='_nolegend_')
                    self.fig_plotter.fill_between([0, len(data)-1], [low, low], [hi, hi], color=fit_col, alpha=0.2, label='_nolegend_')
                    if plats is not None:
                        try:
                            for p in plats:
                                mean, _ = self.get_filtered_mean(data[p[0]:p[1]])
                                w = ansize // 5
                                self.fig_plotter.plot([p[0], p[1]], [mean, mean], color='black', linewidth=w, label='_nolegend_')
                        except Exception:
                            print(traceback.format_exc())

            elif 'Ratio' in self.all_figures[index].datasets[idx].name:
                self.fig_plotter.imshow(data, cmap='gray', vmin=-float(self.ratio_norm.get()), vmax=float(self.ratio_norm.get()))

            elif 'Calib' in self.all_figures[index].datasets[idx].name:
                dx, dy = data[0], data[1]
                fx1, fx2 = 0, data[0][-1] * 1.2
                fy1, fy2 = fx1*calib[0] + calib[-1], fx2*calib[0] + calib[-1]
                self.fig_plotter.plot([fx1, fx2], [fy1, fy2], color='black', linewidth=2, linestyle='--')
                self.fig_plotter.plot(dx, dy, linewidth=0, marker='o', markersize=12, color='#007fff')
                toffy = abs(dy[-1]*0.93 - dy[-1])
                if self.fig_annotations.get() == 1:
                    for tidx in range(len(dx)):
                        self.fig_plotter.text(dx[tidx], dy[tidx] - toffy, f'{round(dx[tidx], 5)}, {round(dy[tidx])} kDa', fontsize=ansize, color='black')
                self.fig_plotter.set_xlim(fx1, data[0][-1] * 1.5)

            elif 'D vs Mass' in self.all_figures[index].datasets[idx].name:
                self.fig_plotter.plot(data[0], data[1], linewidth=0, marker='o', markersize=ansize)

        if 'Contrasts' in self.all_figures[index].datasets[idx].name or 'Masses' in self.all_figures[index].datasets[idx].name or \
                'Diff' in self.all_figures[index].datasets[idx].name or 'D vs Mass' in self.all_figures[index].datasets[idx].name:
            self.fig_plotter.set_xlim(xmin=xmin, xmax=xmax)
            if self.fig_annotations.get() == 1:
                self.fig_plotter.set_ylim(ymin=0, ymax=self.fig_plotter.get_ylim()[1]*1.2)
        elif 'Coords' in self.all_figures[index].datasets[idx].name or 'All Tracks' in self.all_figures[index].datasets[idx].name:
            self.fig_plotter.set_ylim(max(self.fig_plotter.get_ylim()), min(self.fig_plotter.get_ylim()))
        if self.fig_legend.get() == 1:
            self.fig_plotter.legend(leg, loc='best', fontsize=ansize, draggable=True)
        if image:
            self.fig_plotter.set_xlim(-0.5, imshape[1] - 0.5)
            self.fig_plotter.set_ylim(imshape[0] - 0.5, -0.5)
            print(f'Limits set to image shape {imshape}')
        self.fig_canvas.draw()

    def plot_fullsize_fig(self):
        print('Plotting fullsize figure')
        fig_figure = plt.figure(figsize=(8.4, 7.06), dpi=100)

        if len(self.all_figures) == 0:
            return
        selected_ID = self.figure_listbox.focus()
        index = 0
        if 'Figure' not in selected_ID:
            selected_ID = self.figure_listbox.parent(selected_ID)
        for idx in range(len(self.all_figures)):
            if self.all_figures[idx].ID == selected_ID:
                index = idx
        title, xlabel, ylabel, axsize, titlesize, ansize = self.get_fig_params(index)

        bg_col = "#ffffff"
        box_col = "black"
        text_col = "#333333"
        grid_col = "#bbbbbb"
        line_col = "black"
        face_col = '#ffffff'

        if len(self.all_figures[index].datasets) == 0:
            return

        fig_figure.set_facecolor(bg_col)
        fig_plotter = fig_figure.add_subplot(111)
        fig_figure.subplots_adjust(left=self.all_figures[index].fig_params['left'],
                                   right=self.all_figures[index].fig_params['right'],
                                   top=self.all_figures[index].fig_params['top'],
                                   bottom=self.all_figures[index].fig_params['bottom'])
        fig_plotter.set_facecolor(face_col)
        fig_plotter.spines['bottom'].set_color(box_col)
        fig_plotter.spines['top'].set_color(bg_col)
        fig_plotter.spines['left'].set_color(box_col)
        fig_plotter.spines['right'].set_color(bg_col)
        for axis in ['top', 'bottom', 'left', 'right']:
            fig_plotter.spines[axis].set_linewidth(2)
        fig_plotter.tick_params(width=2)
        fig_plotter.xaxis.label.set_color(text_col)
        fig_plotter.yaxis.label.set_color(text_col)
        fig_plotter.tick_params(axis='x', colors=text_col, labelsize=axsize)
        fig_plotter.tick_params(axis='y', colors=text_col, labelsize=axsize)
        if self.fig_grid.get() == 1:
            fig_plotter.grid(color=grid_col)
        fig_plotter.set_xlabel(xlabel, size=axsize, color=text_col)
        fig_plotter.set_ylabel(ylabel, size=axsize, color=text_col)
        fig_plotter.set_title(title, size=titlesize, color=text_col)
        cycle = plt.rcParams['axes.prop_cycle']
        cols = cycle.by_key()['color']
        xmin, xmax = self.all_figures[index].fig_params['xmin'], self.all_figures[index].fig_params['xmax']
        leg = []
        image = False
        for idx in range(len(self.all_figures[index].datasets)):
            if 'Ratio' in self.all_figures[index].datasets[idx].name:
                image = True
                imshape = np.shape(self.all_figures[index].datasets[idx].data)
            data = self.all_figures[index].datasets[idx].data
            x_ax = self.all_figures[index].datasets[idx].x_axis
            fits = self.all_figures[index].datasets[idx].fits
            moments = self.all_figures[index].datasets[idx].fit_vals
            try:
                calib = self.all_figures[index].datasets[idx].calib
            except Exception:
                calib = [self.calibration[0], self.calibration[1]]
            fit_col = cols[int(idx % len(cols))]
            if 'Ratio' not in self.all_figures[index].datasets[idx].name:
                leg.append(self.all_figures[index].datasets[idx].name)

            if 'Contrasts' in self.all_figures[index].datasets[idx].name or 'Masses' in self.all_figures[index].datasets[idx].name or 'Diff' in self.all_figures[index].datasets[idx].name:
                data_xmin, data_xmax = np.min(data), np.max(data)
                bins = round((data_xmax - data_xmin) / self.all_figures[index].fig_params['binw'])
                if self.fig_use_density.get() == 0:
                    density = True
                    factor = 1
                else:
                    density = False
                fig_plotter.hist(data, bins=bins, density=density, alpha=0.5, ec="black", histtype='stepfilled')

                if len(fits) >= 1:
                    if self.fig_use_density.get() == 1:
                        factor = self.all_figures[index].fig_params['binw'] * len(data)
                    fig_plotter.plot(x_ax, fits[-1] * factor, color=fit_col, linewidth=2, label='_nolegend_', alpha=0.9)
                    for idx2 in range(len(fits) - 1):
                        if len(moments[idx2]) == 3:
                            mass, width, amplitude = moments[idx2]
                            skew_string = ""
                            true_mean, true_sigma = mass, width
                        else:
                            mass, width, amplitude = moments[idx2][0][:-1]
                            true_mean, true_sigma, skew = moments[idx2][1][:-1]
                            skew_string = f"\nSkew = {round(skew, 3)}"
                        counts_string = f"\n{round(np.trapz(fits[idx2], x_ax) * len(data))} Counts"
                        fig_plotter.plot(x_ax, fits[idx2] * factor, color=fit_col, linewidth=2, label='_nolegend_', alpha=0.6)
                        if self.fig_annotations.get() == 1:
                            if 'Masses' in self.all_figures[index].datasets[idx].name:
                                width_string = f'\nσ = {round(true_sigma, 1)} kDa'
                                # self.fig_plotter.text(mass + width * 1.25, amplitude * factor, f'{round(mass)} kDa' + width_string, color=fit_col, fontsize=ansize)
                                fig_plotter.annotate(f'{round(true_mean)} kDa' + width_string + skew_string + counts_string,
                                                          xy=(true_mean, (np.max(fits[idx2]) + np.max(fits[-1]) * 0.03) * factor),
                                                          xycoords='data', xytext=(40, 15), textcoords='offset points', arrowprops=dict(arrowstyle="-",
                                                                                                                                        connectionstyle="angle,angleA=0,angleB=90,rad=5",
                                                                                                                                        color=fit_col, linewidth=1),
                                                          fontsize=ansize, color=fit_col).draggable()
                            else:
                                width_string = f'\nσ = {round(width, 6)}'
                                fig_plotter.annotate(f'{round(true_mean, 5)}' + width_string + skew_string + counts_string,
                                                          xy=(true_mean, (np.max(fits[idx2]) + np.max(fits[-1]) * 0.03) * factor),
                                                          xycoords='data', xytext=(40, 15), textcoords='offset points', arrowprops=dict(arrowstyle="-",
                                                                                                                                        connectionstyle="angle,angleA=0,angleB=90,rad=5",
                                                                                                                                        color=fit_col, linewidth=1),
                                                          fontsize=ansize, color=fit_col).draggable()

            elif 'Coords' in self.all_figures[index].datasets[idx].name:
                x, y = [c[0] for c in data], [c[1] for c in data]
                fig_plotter.plot(x, y, linewidth=1)
                fig_plotter.axis("square")

            elif 'All Tracks' in self.all_figures[index].datasets[idx].name:
                for tr_idx, track in enumerate(data):
                    x, y = [c[0] for c in track], [c[1] for c in track]
                    fig_plotter.plot(x, y, linewidth=1)
                fig_plotter.axis("square")

            elif 'Time' in self.all_figures[index].datasets[idx].name:
                plats = None
                stdlim_stored = float(self.mass_std_lim.get())
                if len(data) == 2:
                    data, plats = data
                elif len(data) == 3:
                    data, plats, stdlim_stored = data
                fig_plotter.plot(data, linewidth=2)
                if self.fig_annotations.get() == 1:
                    mean, std = self.get_filtered_mean(data)
                    stdlim = std * stdlim_stored
                    low, hi = mean - stdlim, mean + stdlim
                    fig_plotter.plot([0, len(data)-1], [low, low], linewidth=1, linestyle='--', color=fit_col, alpha=0.5, label='_nolegend_')
                    fig_plotter.plot([0, len(data) - 1], [hi, hi], linewidth=1, linestyle='--', color=fit_col, alpha=0.5, label='_nolegend_')
                    fig_plotter.plot([0, len(data) - 1], [mean, mean], linewidth=1, color=fit_col, alpha=0.8, label='_nolegend_')
                    fig_plotter.fill_between([0, len(data)-1], [low, low], [hi, hi], color=fit_col, alpha=0.2, label='_nolegend_')
                    if plats is not None:
                        try:
                            for p in plats:
                                mean, _ = self.get_filtered_mean(data[p[0]:p[1]])
                                w = ansize // 5
                                fig_plotter.plot([p[0], p[1]], [mean, mean], color='black', linewidth=w, label='_nolegend_')
                        except Exception:
                            print(traceback.format_exc())

            elif 'Ratio' in self.all_figures[index].datasets[idx].name:
                fig_plotter.imshow(data, cmap='gray', vmin=-float(self.ratio_norm.get()), vmax=float(self.ratio_norm.get()))

            elif 'Calib' in self.all_figures[index].datasets[idx].name:
                dx, dy = data[0], data[1]
                fx1, fx2 = 0, data[0][-1] * 1.2
                fy1, fy2 = fx1*calib[0] + calib[-1], fx2*calib[0] + calib[-1]
                fig_plotter.plot([fx1, fx2], [fy1, fy2], color='black', linewidth=2, linestyle='--')
                fig_plotter.plot(dx, dy, linewidth=0, marker='o', markersize=12, color='#007fff')
                toffy = abs(dy[-1]*0.93 - dy[-1])
                if self.fig_annotations.get() == 1:
                    for tidx in range(len(dx)):
                        fig_plotter.text(dx[tidx], dy[tidx] - toffy, f'{round(dx[tidx], 5)}, {round(dy[tidx])} kDa', fontsize=ansize, color='black')
                fig_plotter.set_xlim(fx1, data[0][-1] * 1.5)

            elif 'D vs Mass' in self.all_figures[index].datasets[idx].name:
                fig_plotter.plot(data[0], data[1], linewidth=0, marker='o', markersize=ansize)

        if 'Contrasts' in self.all_figures[index].datasets[idx].name or 'Masses' in self.all_figures[index].datasets[idx].name or \
                'Diff' in self.all_figures[index].datasets[idx].name or 'D vs Mass' in self.all_figures[index].datasets[idx].name:
            fig_plotter.set_xlim(xmin=xmin, xmax=xmax)
            if self.fig_annotations.get() == 1:
                fig_plotter.set_ylim(ymin=0, ymax=fig_plotter.get_ylim()[1] * 1.2)
        elif 'Coords' in self.all_figures[index].datasets[idx].name or 'All Tracks' in self.all_figures[index].datasets[idx].name:
            fig_plotter.set_ylim(max(fig_plotter.get_ylim()), min(fig_plotter.get_ylim()))
        if self.fig_legend.get() == 1:
            fig_plotter.legend(leg, loc='best', fontsize=ansize, draggable=True)
        if image:
            fig_plotter.set_xlim(-0.5, imshape[1] - 0.5)
            fig_plotter.set_ylim(imshape[0] - 0.5, -0.5)
            print(f'Limits set to image shape {imshape}')
        fig_figure.show()

    def plot_3D_hist(self):
        selected_ID = self.figure_listbox.focus()
        index = 0
        if 'Figure' not in selected_ID:
            selected_ID = self.figure_listbox.parent(selected_ID)
        for idx in range(len(self.all_figures)):
            if self.all_figures[idx].ID == selected_ID:
                index = idx
        data_s = self.all_figures[index].datasets
        title, xlabel, ylabel, axsize, titlesize, ansize = self.get_fig_params(index)
        xmin, xmax = self.all_figures[index].fig_params['xmin'], self.all_figures[index].fig_params['xmax']

        datasets = []
        gaussian_fits = []
        min_data, max_data = 0, 0
        for idx, datum in enumerate(data_s):
            td = []
            for t in datum.data:
                if xmin < t < xmax:
                    td.append(t)
            datasets.append(td)
            if np.nanmin(datasets[-1]) < min_data:
                min_data = np.nanmin(datasets[-1])
            if np.nanmax(datasets[-1]) > max_data:
                max_data = np.nanmax(datasets[-1])
            stop_index1 = 0
            for stpi in range(len(datum.x_axis)):
                if datum.x_axis[stpi] > xmin:
                    stop_index1 = stpi
                    break
            stop_index2 = len(datum.x_axis) - 1
            for stpi in range(len(datum.x_axis)):
                if datum.x_axis[stpi] > xmax:
                    stop_index2 = stpi
                    break
            x_ax = datum.x_axis[stop_index1:stop_index2]
            ft = datum.fits
            fts = [f[stop_index1:stop_index2] for f in ft]
            gaussian_fits.append([x_ax, fts])
            print(stop_index1, stop_index2, len(x_ax), len(datum.x_axis))

        bins = round((max_data - min_data) / self.all_figures[index].fig_params['binw'])
        x_off = self.all_figures[index].fig_params['binw'] / 2  # x offset since histogram is left aligned due to bar3D on 3D axes but should be centre aligned

        cycle = plt.rcParams['axes.prop_cycle']
        colors = cycle.by_key()['color']

        fig = plt.figure(figsize=(8, 8))
        ax = fig.add_subplot(111, projection='3d')

        bins = np.linspace(min_data, max_data, bins)  # Shared bin edges
        for i, (data, (x_fit, y_fit)) in enumerate(zip(datasets, gaussian_fits)):
            color = colors[i % len(colors)]

            hist, bin_edges = np.histogram(data, bins=bins, range=(xmin, xmax))

            x_m_idx = 0
            for idx in range(len(bin_edges)):
                if bin_edges[idx] < xmin:
                    x_m_idx = idx
            hist = hist[x_m_idx:]
            bin_edges = bin_edges[x_m_idx:]

            binw = bin_edges[1] - bin_edges[0]
            xpos = (bin_edges[:-1] + bin_edges[1:]) / 2
            xpos = xpos - x_off
            ypos = np.full_like(xpos, i * 5)
            zpos = np.zeros_like(xpos)
            dx = np.diff(bin_edges)
            dy = 0.1
            dz = hist
            ax.bar3d(xpos, ypos, zpos, dx, dy, dz, color=color, alpha=0.3)

            # Plot Gaussian fit
            try:
                ax.plot(x_fit, np.full_like(x_fit, ypos[0]), y_fit[-1] * len(data) * binw, color=color, linewidth=2, alpha=0.9)  # label=f'Time ({i *10} - {(i+1) * 10}) seconds.')
            except Exception:
                'No Fit'
            for idx in range(len(y_fit) - 1):
                ax.plot(x_fit, np.full_like(x_fit, ypos[0]), y_fit[idx] * len(data) * binw, color=color, linewidth=2, alpha=0.5)

        ax.set_title(title, size=titlesize//2)
        ax.set_xlabel(xlabel, size=axsize//2)
        ax.set_zlabel('Counts', size=axsize//2)
        ax.set_yticks([i * 5 for i in range(len(datasets))])
        ax.set_yticklabels([f'{n.data_parent}' for n in data_s], fontsize=axsize//2)
        ax.set_xlim(xmin, xmax)
        ax.tick_params(axis='x', pad=0)
        ax.tick_params(axis='y', pad=1)
        ax.tick_params(axis='z', pad=3)
        # for label in ax.xaxis.get_ticklabels():
        #     label.set_ha('left')
        for label in ax.yaxis.get_ticklabels():
            label.set_ha('left')
        # for label in ax.zaxis.get_ticklabels():
        #     label.set_ha('left')
        plt.show()

    def save_datafigs(self):
        path = easygui.filesavebox(title='Save datasets and figures workspace as .dsf file', default='N://*.dsf', filetypes=['.dsf'])
        if path:
            base = os.path.basename(path)
            base = os.path.splitext(base)[0]
            base = base + '.dsf'
            rootname = os.path.split(path)[0]
            path = os.path.join(rootname, base)

            with open(path, 'wb') as datafile:
                pickle.dump(self.all_datasets, datafile)
                pickle.dump(self.all_figures, datafile)
                pickle.dump(self.calibration, datafile)
                pickle.dump(self.fig_use_density.get(), datafile)
                pickle.dump(self.fig_annotations.get(), datafile)
                pickle.dump(self.fig_grid.get(), datafile)

    def load_datafigs(self):
        path = easygui.fileopenbox(title='Open datasets and figures workspace (.dsf) file', default='N://*.dsf', filetypes=['.dsf'])
        if path:
            with open(path, 'rb') as datafile:
                self.all_datasets = pickle.load(datafile)
                self.all_figures = pickle.load(datafile)
                try:
                    calibration = pickle.load(datafile)
                    self.calibration = calibration
                    self.fig_use_density.set(pickle.load(datafile))
                    self.fig_annotations.set(pickle.load(datafile))
                    self.fig_grid.set(pickle.load(datafile))
                except Exception:
                    ''' No calib was present in dsf file'''
                    print(traceback.format_exc())
            self.clear_treeviews()
            try:
                self.populate_treeviews()
                self.figure_listbox.selection_set(self.all_figures[0].ID)
                self.fig_list_clicked(None)
            except IndexError:
                print('Datasets loaded, No figures to display.')
            except Exception:
                easygui.msgbox(title='Error!', msg=traceback.format_exc())
            self.update_calibration_dsf()
            self.analysis_notebook.select(self.analysis_tab3)


    def clear_treeviews(self):
        for item in self.treeview.get_children():
            self.treeview.delete(item)
        for item in self.figure_listbox.get_children():
            self.figure_listbox.delete(item)

    def populate_treeviews(self):
        for idx, item in enumerate(self.all_datasets):
            self.treeview.insert(parent='', index="end", iid=item.ID, text=item.name,
                                 values=(item.filename, item.info))
            for idx2, subitem in enumerate(item.subsets):
                self.treeview.insert(parent=item.ID, index="end", iid=subitem.ID, text=subitem.name,
                                     values=(subitem.name.split('_')[1], subitem.info))

        for idx, item in enumerate(self.all_figures):
            self.figure_listbox.insert(parent='', index="end", iid=item.ID, text=item.name,
                                       values=(item.title, None))
            for idx2, subitem in enumerate(item.datasets):
                self.figure_listbox.insert(parent=item.ID, index="end", iid=item.ID[0:3] + item.ID[6:] + subitem.ID, text=subitem.data_parent,
                                           values=(subitem.name, None))

    def export_ratiometric_tiff(self):
        shape = np.shape(self.ratiometric_stack)
        path = easygui.filesavebox(title="Save ratiometric movie as .tiff", default= "NN:/ratiometric movie.tiff", filetypes=["*.tiff"])
        if not path:
            return
        base = os.path.basename(path)
        base = os.path.splitext(base)[0]
        base = base + '.tiff'
        rootname = os.path.split(path)[0]
        path = os.path.join(rootname, base)

        try:
            self.progress_win.window.destroy()
        except Exception:
            pass
        self.progress_win = ProgressWin("Preparing movie...", "Converting format...")
        self.progress_win.cancel_button.place_forget()
        self.progress_win.progress["maximum"] = shape[2] + 1

        movie = np.zeros((shape[2], shape[0], shape[1]), dtype=np.float32)
        for frame in range(shape[2]):
            movie[frame, :, :] = self.ratiometric_stack[:, :, frame]
            self.progress_win.progress.step(1)
            self.progress_win.progress.update()

        self.progress_win.label["text"] = "Saving movie..."
        self.progress_win.label.update()
        self.progress_win.window.title("Saving...")
        try:
            with tifffile.TiffWriter(path) as tif:
                tif.write(movie)
        except:
            print("Warning error occured: Full traceback shown below:")
            print(traceback.format_exc())
            easygui.msgbox(title="Error!", msg="An error occurred while attempting to export tif.\n" + str(traceback.format_exc()))

        try:
            self.progress_win.handle_close()
        except:
            """ Window already closed """

    def export_ratiometric_mov(self):
        height, width, length = np.shape(self.ratiometric_stack)
        upscale = 4

        filename = easygui.filesavebox(title='Save Ratiometric Movie...', filetypes=['.mp4'], default='N://untitled movie.mp4')
        base = os.path.basename(filename)
        base = os.path.splitext(base)[0]
        base = base + '.mp4'
        rootname = os.path.split(filename)[0]
        filename = os.path.join(rootname, base)
        print(f"Writing movie to '{filename}'")
        codec_id = "H264"
        fourcc = cv2.VideoWriter_fourcc(*codec_id)
        out = cv2.VideoWriter(filename, fourcc=fourcc, fps=30, frameSize=(width * upscale, height * upscale))

        try:
            self.progress_win.window.destroy()
        except Exception:
            pass
        self.progress_win = ProgressWin(title='Exporting', msg='Writing movie...')
        self.progress_win.progress['maximum'] = length + 1
        for frame in range(length):
            self.progress_win.progress.step(1)
            self.progress_win.window.update()
            norm = mplib.colors.Normalize(vmin=-float(self.ratio_norm.get()), vmax=float(self.ratio_norm.get()))
            frame_rgb = mplib.cm.gray(norm(self.ratiometric_stack[:, :, frame]))
            frame_rgb = frame_rgb[:, :, :3] * 255
            frame_rgb = frame_rgb.astype(np.uint8)
            frame_rgb = frame_rgb.repeat(repeats=upscale, axis=0).repeat(repeats=upscale, axis=1)
            out.write(frame_rgb)
        out.release()
        self.progress_win.handle_close()
        easygui.msgbox(title='Done!', msg='Movie has been exported.')

    def export_ratiometric_movtrack(self):
        index = self.current_track.get()
        if index == 'N/A':
            easygui.msgbox(title='Error!', msg='Cannot export movie with track overlay becuase no tracks were found')
            return

        filename = easygui.filesavebox(title='Save Ratiometric Movie...', filetypes=['.mp4'], default='N://untitled movie.mp4')
        base = os.path.basename(filename)
        base = os.path.splitext(base)[0]
        base = base + '.mp4'
        rootname = os.path.split(filename)[0]
        filename = os.path.join(rootname, base)

        index = int(float(index)) - 1
        track = self.tracks[index]
        x, y = [c[0] for c in track.coords], [c[1] for c in track.coords]
        start, end = track.frames[0], track.frames[-1]
        movie = np.copy(self.ratiometric_stack[:, :, start:end+1])
        height, width, length = np.shape(movie)

        new = []
        try:
            self.progress_win.window.destroy()
        except Exception:
            pass
        self.progress_win = ProgressWin(title='Exporting', msg='Rendering movie...')
        self.progress_win.progress['maximum'] = length + 1
        for idx in range(length):
            self.progress_win.progress.step(1)
            self.progress_win.window.update()
            fig = plt.figure(figsize=(round(width / 16, 2), round(height / 16, 2)), dpi=100)
            fig.subplots_adjust(left=0, bottom=0, right=1, top=1)
            ax = fig.add_subplot(111)
            ax.imshow(movie[:, :, idx], cmap='gray', vmin=-float(self.ratio_norm.get()), vmax=float(self.ratio_norm.get()))
            mass = ''
            try:
                fr_num = idx + start
                if fr_num in track.frames:
                    d_idx = track.frames.index(fr_num)
                    xc, yc = x[d_idx], y[d_idx]
                    mass = str(round(track.contrasts[d_idx]*self.calibration[0] + self.calibration[1]))
                    c = plt.Circle((xc, yc), 3.6, fill=False, color="#007fff", linewidth=4, alpha=0.6)
                    ax.add_artist(c)
            except Exception:
                ''' Error ocurred '''
                print(traceback.format_exc())
            ax.plot(x, y, color='orange', linewidth=1)
            ax.text(10, 20, f'Frame {idx + 1}\n{mass} kDa', fontsize=20, color='orange')
            # ax.set_xlabel('x coord')
            # ax.set_ylabel('y coord')
            io_buf = io.BytesIO()
            fig.savefig(io_buf, format='raw', dpi=100)
            io_buf.seek(0)
            img_arr = np.reshape(np.frombuffer(io_buf.getvalue(), dtype=np.uint8),
                                 newshape=(int(fig.bbox.bounds[3]), int(fig.bbox.bounds[2]), -1))
            io_buf.close()
            plt.close(fig)
            new.append(img_arr)
        self.progress_win.handle_close()

        length, height, width, chns = np.shape(new)
        upscale = 1
        codec_id = "H264"
        fourcc = cv2.VideoWriter_fourcc(*codec_id)
        out = cv2.VideoWriter(filename, fourcc=fourcc, fps=30, frameSize=(width * upscale, height * upscale))
        try:
            self.progress_win.window.destroy()
        except Exception:
            pass
        self.progress_win = ProgressWin(title='Exporting', msg='Writing movie...')
        self.progress_win.progress['maximum'] = length + 1
        print(f"Writing movie to '{filename}'")
        for frame in range(length):
            self.progress_win.progress.step(1)
            self.progress_win.window.update()
            new_fr = new[frame][:, :, :3]
            new_fr = cv2.cvtColor(new_fr, cv2.COLOR_BGR2RGB)
            out.write(new_fr)
        out.release()
        self.progress_win.handle_close()
        easygui.msgbox(title='Done!', msg='Movie has been exported.')

    def movie_preview(self):
        self.export_ratiometric_movtracks(preview=True)

    def export_ratiometric_movtracks(self, preview=False):
        index = self.current_track.get()
        if index == 'N/A':
            easygui.msgbox(title='Error!', msg='Cannot export movie with track overlay becuase no tracks were found')
            return

        if not preview:
            filename = easygui.filesavebox(title='Save Ratiometric Movie...', filetypes=['.mp4'], default='N://untitled movie.mp4')
            if not filename:
                return
            base = os.path.basename(filename)
            base = os.path.splitext(base)[0]
            base = base + '.mp4'
            rootname = os.path.split(filename)[0]
            filename = os.path.join(rootname, base)

        height, width, length = np.shape(self.ratiometric_stack)
        length = int(float(self.export_end_frame.get())) - int(float(self.export_start_frame.get()))
        start_fr = int(float(self.export_start_frame.get()))

        fig = plt.figure(figsize=(round(width / 16, 2), round(height / 16, 2)), dpi=100)
        fig.subplots_adjust(left=0, bottom=0, right=1, top=1)
        ax = fig.add_subplot(111)
        ax.imshow(self.ratiometric_stack[:, :, 0])
        io_buf = io.BytesIO()
        fig.savefig(io_buf, format='raw', dpi=100)
        io_buf.seek(0)
        img_arr = np.reshape(np.frombuffer(io_buf.getvalue(), dtype=np.uint8),
                             newshape=(int(fig.bbox.bounds[3]), int(fig.bbox.bounds[2]), -1))
        io_buf.close()
        plt.close(fig)

        m_height, m_width, _ = np.shape(img_arr)
        codec_id = "H264"
        fourcc = cv2.VideoWriter_fourcc(*codec_id)
        if not preview:
            out = cv2.VideoWriter(filename, fourcc=fourcc, fps=int(float(self.export_fps.get())), frameSize=(m_width, m_height))
            print(f"Writing movie to '{filename}'")

        try:
            self.progress_win.window.destroy()
        except Exception:
            pass
        self.progress_win = ProgressWin(title='Exporting', msg='Preparing data...')
        self.progress_win.progress['maximum'] = length + 1

        x_ax, fits, fit_func = self.return_fits()
        all_moments = self.return_moments()
        filtered_indices = self.get_filtered_indices()

        if preview:
            start_fr = int(self.ratio_frame_var.get())
            length = 1

        track_list_frames = []
        for idx in range(start_fr, start_fr + length):
            tr_list = []
            if not self.export_apply_filter.get() == 1:
                for index, track in enumerate(self.tracks):
                    if idx >= track.frames[0] and idx <= track.frames[-1]:
                        tcol = None
                        if self.export_histogram_vid.get() == 1:
                            if len(all_moments) > 0:
                                mean_mass = self.get_filtered_mean(track.contrasts)[0] * self.calibration[0] + self.calibration[1]
                                vals = []
                                for fit_idx, moments in enumerate(all_moments):
                                    args = [mean_mass] + moments
                                    vals.append(fit_func(*args))
                                best_fit = np.argmax(np.array(vals))
                                tcol = self.track_cols[best_fit % len(self.track_cols)]
                        tr_list.append([index, track, tcol])
            else:
                for index in filtered_indices:
                    track = self.tracks[index]
                    if idx >= track.frames[0] and idx <= track.frames[-1]:
                        tcol = None
                        if self.export_histogram_vid.get() == 1:
                            if len(all_moments) > 0:
                                mean_mass = self.get_filtered_mean(track.contrasts)[0] * self.calibration[0] + self.calibration[1]
                                vals = []
                                for fit_idx, moments in enumerate(all_moments):
                                    args = [mean_mass] + moments
                                    vals.append(fit_func(*args))
                                best_fit = np.argmax(np.array(vals))
                                tcol = self.track_cols[best_fit % len(self.track_cols)]
                        tr_list.append([index, track, tcol])
            track_list_frames.append(tr_list)
            if idx % 10 == 0:
                self.progress_win.progress.step(10)
                self.event_label['text'] = f"Preparing track frame bins...   {round((idx - start_fr) / length * 100, 1)}%"
                self.window.update()

        self.progress_win.handle_close()

        self.progress_win = ProgressWin(title='Exporting', msg='Rendering movie...')
        self.progress_win.progress['maximum'] = length + 1

        for idx in range(start_fr, start_fr + length):
            if self.cancel_flag:
                self.cancel_flag = False
                break
            try:
                self.progress_win.progress.step(1)
            except _tkinter.TclError:
                pass
            self.event_label['text'] = f"Rendering movie...   {round((idx - start_fr) / length * 100, 1)}%"
            self.progress_win.window.update()
            fig = plt.figure(figsize=(round(width / 16, 2), round(height / 16, 2)), dpi=100)
            fig.subplots_adjust(left=0, bottom=0, right=1, top=1)
            ax = fig.add_subplot(111)
            ax.imshow(self.ratiometric_stack[:, :, idx], cmap='gray', vmin=-float(self.ratio_norm.get()), vmax=float(self.ratio_norm.get()))

            if self.export_histogram_vid.get() == 1:
                ax2 = fig.add_axes([float(self.export_left_align.get()), float(self.export_bottom_align.get()), float(self.export_xsize.get()), float(self.export_ysize.get())])
                ax2.patch.set_alpha(0.01)
                binw = float(self.mov_bin_width.get())
                bins = int(np.max((self.track_masses) - np.min(self.track_masses)) / binw)
                ax2.hist(self.track_masses, color='white', alpha=0.5, ec='white', histtype='stepfilled', bins=bins, density=False)
                ax2.set_xlabel('Mass (kDa)', color='orange', size=8)
                ax2.set_ylabel('Counts', color='orange', size=8)
                ax2.tick_params(axis='x', colors='orange', labelsize=8)
                ax2.tick_params(axis='y', colors='orange', labelsize=8)
                ax2.spines['bottom'].set_color('black')
                ax2.spines['top'].set_alpha(0.01)
                ax2.spines['left'].set_color('black')
                ax2.spines['right'].set_alpha(0.01)
                ax2.plot(x_ax, fits[-1]*len(self.track_masses)*binw, linewidth=1, color='black', alpha=0.4)
                for index in range(len(fits) - 1):
                    col = self.track_cols[index % len(self.track_cols)]
                    ax2.plot(x_ax, fits[index]*len(self.track_masses)*binw, color=col, linewidth=1, alpha=0.9)
                ax2.set_xlim([x_ax[0], x_ax[-1]])

            mass = ''
            for track_data in track_list_frames[idx - start_fr]:
                index, track, fcol = track_data
                col = self.track_cols[index % len(self.track_cols)]
                if fcol is not None:
                    col = fcol
                x, y = [c[0] for c in track.coords], [c[1] for c in track.coords]
                ax.plot(x, y, color=col, linewidth=1, alpha=float(self.export_opacity.get()))
                try:
                    fr_num = idx
                    if fr_num in track.frames:
                        d_idx = track.frames.index(fr_num)
                        xc, yc = x[d_idx], y[d_idx]
                        mass = str(round(track.contrasts[d_idx]*self.calibration[0] + self.calibration[1]))
                        c = plt.Circle((xc, yc), 3.5, fill=False, color=col, linewidth=4, alpha=round(float(self.export_opacity.get()) * 0.75, 2))
                        ax.add_artist(c)
                        if self.export_masses.get() == 1:
                            tcol = col
                            ax.text(xc+5, yc-4, f'{mass} kDa', fontsize=int(float(self.export_mass_size.get())), color=tcol)
                except Exception:
                    ''' Error ocurred '''
            if self.export_frames.get() == 1:
                ax.text(4, 7, f'Frame {idx}', fontsize=18, color='orange')
            # ax.set_xlabel('x coord')
            # ax.set_ylabel('y coord')
            io_buf = io.BytesIO()
            fig.savefig(io_buf, format='raw', dpi=100)
            io_buf.seek(0)
            img_arr = np.reshape(np.frombuffer(io_buf.getvalue(), dtype=np.uint8),
                                 newshape=(int(fig.bbox.bounds[3]), int(fig.bbox.bounds[2]), -1))
            io_buf.close()
            plt.close(fig)
            new_fr = img_arr[:, :, :3]
            new_fr = cv2.cvtColor(new_fr, cv2.COLOR_BGR2RGB)
            if not preview:
                out.write(new_fr)
        self.progress_win.handle_close()
        self.event_label['text'] = ""
        if not preview:
            out.release()
            easygui.msgbox(title='Done!', msg='Movie has been exported.')
        else:
            plt.imshow(cv2.cvtColor(new_fr, cv2.COLOR_RGB2BGR))
            plt.title(f'Frame {start_fr} preview')
            plt.show()

    @staticmethod
    def create_export_sheet():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        return wb

    def export_events_binding(self):
        if len(self.events) == 0:
            easygui.msgbox(title='Error!', msg='No events to export!')
            return
        data_name = self.export_events_selection.get()
        data_type = self.export_events_format.get()
        if '.pydat' in data_type:
            path = easygui.filesavebox(title='Save events data as bytes serialised (pickled) python object (.pydat)',
                                       default='NN:/*.pydat', filetypes=['*.pydat'])
            if path:
                base = os.path.basename(path)
                base = os.path.splitext(base)[0]
                base = base + '.pydat'
                rootname = os.path.split(path)[0]
                path = os.path.join(rootname, base)
                if data_name.lower() == 'events':
                    data = self.events
                elif data_name.lower() == 'contrasts':
                    data = self.contrasts
                elif data_name.lower() == 'masses':
                    data = self.masses
                with open(path, 'wb') as export_file:
                    pickle.dump(data, export_file)
                easygui.msgbox(title='Success!', msg='Data has been exported.')
        elif '.xlsx' in data_type:
            path = easygui.filesavebox(title='Save events data as Excel workbook (.xlsx)',
                                       default='NN:/*.xlsx', filetypes=['*.xlsx'])
            if path:
                base = os.path.basename(path)
                base = os.path.splitext(base)[0]
                base = base + '.xlsx'
                rootname = os.path.split(path)[0]
                path = os.path.join(rootname, base)
                wb = self.create_export_sheet()
                if data_name.lower() == 'events':
                    # events is of the form [frame_, -c, events[idx], event_traces[idx], None, [g1, g2, r1, r2]]
                    ws = wb['Sheet1']
                    heads = ['Frame', 'Ratiometric Contrast', 'Loc x', 'Loc y', 'Sigma x', 'Sigma y', 'Amplitude', 'Baseline', 'Ascending grad.', 'Descending grad.',
                             'Ascending r^2', 'Descending r^2', 'Masked']
                    for idx, head in enumerate(heads):
                        ws.cell(row=1, column=idx + 1).value = head
                    try:
                        self.progress_win.window.destroy()
                    except Exception:
                        pass
                    self.progress_win = ProgressWin(title='Exporting', msg='Writing events to Excel sheet...')
                    self.progress_win.progress['maximum'] = len(self.events) + 1
                    for idx, event in enumerate(self.events):
                        gauss_params = event[2]
                        trace_fit = event[5]
                        ws.cell(row=3 + idx, column=1).value = event[0]
                        ws.cell(row=3 + idx, column=2).value = -event[1]
                        ws.cell(row=3 + idx, column=3).value = gauss_params[0]
                        ws.cell(row=3 + idx, column=4).value = gauss_params[1]
                        ws.cell(row=3 + idx, column=5).value = gauss_params[2][0]
                        ws.cell(row=3 + idx, column=6).value = gauss_params[2][1]
                        ws.cell(row=3 + idx, column=7).value = gauss_params[2][2]
                        ws.cell(row=3 + idx, column=8).value = gauss_params[2][3]
                        ws.cell(row=3 + idx, column=9).value = trace_fit[0]
                        ws.cell(row=3 + idx, column=10).value = trace_fit[1]
                        ws.cell(row=3 + idx, column=11).value = trace_fit[2]
                        ws.cell(row=3 + idx, column=12).value = trace_fit[3]
                        ws.cell(row=3 + idx, column=13).value = event[4]
                        if idx % 10 == 0:
                            self.progress_win.progress.step(10)
                            self.progress_win.window.update()
                    self.progress_win.handle_close()
                    try:
                        wb.save(path)
                        easygui.msgbox(title='Success!', msg='Data has been exported.')
                    except PermissionError:
                        easygui.msgbox(title='Error!', msg='Access Denied. Make sure you are not saving to a file that is currently open in excel.')
                elif data_name.lower() == 'contrasts':
                    ws = wb['Sheet1']
                    ws.cell(row=1, column=1).value = 'Contrasts'
                    for idx, c in enumerate(self.contrasts):
                        ws.cell(row=3 + idx, column=1).value = -c
                    try:
                        wb.save(path)
                        easygui.msgbox(title='Success!', msg='Data has been exported.')
                    except PermissionError:
                        easygui.msgbox(title='Error!', msg='Access Denied. Make sure you are not saving to a file that is currently open in excel.')
                elif data_name.lower() == 'masses':
                    ws = wb['Sheet1']
                    ws.cell(row=1, column=1).value = 'Masses'
                    for idx, m in enumerate(self.masses):
                        ws.cell(row=3 + idx, column=1).value = m
                    try:
                        wb.save(path)
                        easygui.msgbox(title='Success!', msg='Data has been exported.')
                    except PermissionError:
                        easygui.msgbox(title='Error!', msg='Access Denied. Make sure you are not saving to a file that is currently open in excel.')

    def export_tracks_binding(self):
        if len(self.tracks) == 0:
            easygui.msgbox(title='Error!', msg='No tracks or track data to export!')
            return
        data_name = self.export_tracks_selection.get()
        data_type = self.export_tracks_format.get()
        if '.pydat' in data_type:
            path = easygui.filesavebox(title='Save tracks data as bytes serialised (pickled) python object (.pydat)',
                                       default='NN:/*.pydat', filetypes=['*.pydat'])
            if path:
                base = os.path.basename(path)
                base = os.path.splitext(base)[0]
                base = base + '.pydat'
                rootname = os.path.split(path)[0]
                path = os.path.join(rootname, base)
                if data_name.lower() == 'tracks':
                    data = self.tracks
                elif data_name.lower() == 'track contrasts':
                    data = self.track_contrasts
                elif data_name.lower() == 'track masses':
                    data = self.track_masses
                with open(path, 'wb') as export_file:
                    pickle.dump(data, export_file)
                easygui.msgbox(title='Success!', msg='Data has been exported.')
        elif '.xlsx' in data_type:
            path = easygui.filesavebox(title='Save tracks data as Excel workbook (.xlsx)',
                                       default='NN:/*.xlsx', filetypes=['*.xlsx'])
            if path:
                base = os.path.basename(path)
                base = os.path.splitext(base)[0]
                base = base + '.xlsx'
                rootname = os.path.split(path)[0]
                path = os.path.join(rootname, base)
                wb = self.create_export_sheet()
                if data_name.lower() == 'tracks':
                    sheets = [wb['Sheet1']]
                    for idx in range(len(self.tracks) // 1000):
                        sheets.append(wb.create_sheet(f'Sheet{idx + 2}'))
                    heads = ['Frame', 'Loc x', 'Loc y', 'Contrast', 'Displacement']
                    try:
                        self.progress_win.window.destroy()
                    except Exception:
                        pass
                    self.progress_win = ProgressWin(title='Exporting', msg='Writing tracks to Excel sheet...')
                    self.progress_win.progress['maximum'] = len(self.tracks) + 1
                    for tr_idx, track in enumerate(self.tracks):
                        sheet_index = int(tr_idx // 1000)
                        sheet_column = int(tr_idx % 1000)

                        track.displacements = [np.nan]
                        for idx in range(1, len(track.coords)):
                            dist = np.sqrt((track.coords[idx][0] - track.coords[idx - 1][0]) ** 2 + (track.coords[idx][1] - track.coords[idx - 1][1]) ** 2)
                            track.displacements.append(dist)

                        for idx, head in enumerate(heads):
                            if idx == 0:
                                sheets[sheet_index].cell(row=1, column=sheet_column*6 + 1).value = f'Track {tr_idx + 1}'
                            sheets[sheet_index].cell(row=2, column=sheet_column*6 + 1 + idx).value = head

                        for idx, fr in enumerate(track.frames):
                            sheets[sheet_index].cell(row=4 + idx, column=sheet_column*6 + 1).value = fr
                        for idx, co in enumerate(track.coords):
                            sheets[sheet_index].cell(row=4 + idx, column=sheet_column*6 + 2).value = co[0]
                            sheets[sheet_index].cell(row=4 + idx, column=sheet_column*6 + 3).value = co[1]
                        for idx, con in enumerate(track.contrasts):
                            sheets[sheet_index].cell(row=4 + idx, column=sheet_column*6 + 4).value = con
                        for idx, di in enumerate(track.displacements):
                            sheets[sheet_index].cell(row=4 + idx, column=sheet_column*6 + 5).value = di

                        if tr_idx % 10 == 0:
                            self.progress_win.progress.step(10)
                            self.progress_win.window.update()
                    self.progress_win.handle_close()
                    try:
                        wb.save(path)
                        easygui.msgbox(title='Success!', msg='Data has been exported.')
                    except PermissionError:
                        easygui.msgbox(title='Error!', msg='Access Denied. Make sure you are not saving to a file that is currently open in excel.')
                elif data_name.lower() == 'track contrasts':
                    ws = wb['Sheet1']
                    ws.cell(row=1, column=1).value = 'Track Contrasts'
                    for idx, c in enumerate(self.track_contrasts):
                        ws.cell(row=3 + idx, column=1).value = -c
                    try:
                        wb.save(path)
                        easygui.msgbox(title='Success!', msg='Data has been exported.')
                    except PermissionError:
                        easygui.msgbox(title='Error!', msg='Access Denied. Make sure you are not saving to a file that is currently open in excel.')
                elif data_name.lower() == 'track masses':
                    ws = wb['Sheet1']
                    ws.cell(row=1, column=1).value = 'Track Masses'
                    for idx, m in enumerate(self.track_masses):
                        ws.cell(row=3 + idx, column=1).value = m
                    try:
                        wb.save(path)
                        easygui.msgbox(title='Success!', msg='Data has been exported.')
                    except PermissionError:
                        easygui.msgbox(title='Error!', msg='Access Denied. Make sure you are not saving to a file that is currently open in excel.')

    def export_hist_data(self):
        parent_ID = self.figure_listbox.focus()
        index = 0
        for idx in range(len(self.all_figures)):
            if self.all_figures[idx].ID == parent_ID:
                index = idx
        data = self.all_figures[index].datasets
        if len(data) == 0:
            easygui.msgbox(title='Error', msg='Figure Contains no data to export')
            return
        path = easygui.filesavebox(title='Save events data as bytes serialised (pickled) python object (.pydat)',
                                   default='NN:/*.pydat', filetypes=['*.pydat'])
        if path:
            base = os.path.basename(path)
            base = os.path.splitext(base)[0]
            base = base + '.pydat'
            rootname = os.path.split(path)[0]
            path = os.path.join(rootname, base)

            with open(path, 'wb') as export_file:
                pickle.dump(data, export_file)
            easygui.msgbox(title='Success!', msg='Data has been exported.')

    def restore_task_win(self):
        try:
            self.progress_win.restore()
        except Exception:
            ''' Failed '''

    def cancel_task(self):
        self.cancel_flag = True

    def open_preferences(self):
        self.preferences_win = PreferencesWin()

    def about(self):
        self.about_win = AboutWin(third_party=False)

    def about_third(self):
        self.about_win = AboutWin(third_party=True)

    def handle_close(self):
        choice = easygui.indexbox(title='Closing', msg='                                    Terminate Process. Unsaved data will be lost!\n\n\n                                                          Are you sure?',
                                  choices=['                      Cancel                   ', 'Confirm Terminate Process'], default_choice=0, cancel_choice=0)
        if choice is None or choice == 0:
            return
        try:
            self.mass_calib_win.handle_close()
        except Exception:
            print(traceback.format_exc())
        try:
            self.drift_win.handle_close()
        except Exception:
            print(traceback.format_exc())
        try:
            self.drift_win.progress_win.window.destroy()
        except Exception:
            pass
        try:
            self.add_data_win.handle_close()
        except Exception:
            print(traceback.format_exc())
        try:
            self.progress_win.window.destroy()
        except Exception:
            print(traceback.format_exc())
        try:
            self.profile_win.window.destroy()
        except Exception:
            print(traceback.format_exc())
        try:
            self.preferences_win.window.destroy()
        except Exception:
            print(traceback.format_exc())
        try:
            self.gauss_opt_win.window.destroy()
        except Exception:
            print(traceback.format_exc())
        try:
            self.gauss_preview_win.window.destroy()
        except Exception:
            print(traceback.format_exc())
        try:
            self.about_win.window.destroy()
        except Exception:
            print(traceback.format_exc())
        try:
            self.window.destroy()
        except Exception:
            print(traceback.format_exc())
        sys.exit()


class Track:
    def __init__(self):
        self.events = []
        self.frames = []
        self.coords = []
        self.contrasts = []
        self.displacements = []
        self.type = None
        self.ended = False
        self.plateaus = []
        self.plateau_contrasts = []

    def get_distance(self, other):
        return np.sqrt((self.coords[-1][0] - other[0])**2 + (self.coords[-1][1] - other[1])**2)

    def same_type(self, other):
        if self.type == other:
            return True
        else:
            return False


class Dataset:
    def __init__(self):
        self.subsets = []
        self.ID = None
        self.name = ''
        self.filename = ''
        self.info = 'Parent Data Superset'


class DataSubset:
    def __init__(self):
        self.data = None
        self.fits = []
        self.x_axis = None
        self.fit_vals = None
        self.ID = None
        self.parent = None
        self.data_parent = None
        self.name = ''
        self.type = ''
        self.info = ''
        self.metadata = ''
        self.calib = [0, 1]


class Figure:
    def __init__(self):
        self.datasets = []
        self.plot_modes = []
        self.plot_params = []
        self.ID = None
        self.name = ''
        self.title = ''
        self.fig_params = dict()


class AddDatasetWin:
    def __init__(self, parent_ID, winx, winy):
        self.window = tk.Toplevel()
        self.window.title('Add Data')
        self.window.protocol("WM_DELETE_WINDOW", self.handle_close)
        self.window.attributes('-topmost', True)
        self.window.iconbitmap("icons/luxmp_logo.ico")
        winx += 100
        winy += 650
        if winy >= 800:
            winy -= 250
        self.parent_ID = parent_ID
        self.window.geometry(f'320x200+{winx}+{winy}')
        self.window.resizable(False, False)
        self.window['bg'] = '#333333'

        self.data_selection = tk.StringVar(master=self.window)
        self.field_list = ['Event Contrasts', 'Event Masses', 'Track Contrasts', 'Track Masses', f'Track Coords {root.current_track.get()}', f'Time Series {root.current_track.get()}',
                           f'Time Series Mass {root.current_track.get()}', 'Ratiometric Frame', 'Track Diffusivity', 'Calibration', 'D vs Mass', 'All Tracks']
        if root.imported_data:
            self.field_list = ['Event Contrasts', 'Event Masses']
        self.selector = ttk.OptionMenu(self.window, self.data_selection, 'select data', *self.field_list, command=self.enable_confirm)
        ttk.Label(master=self.window, text='Data type').place(x=20, y=20)
        self.selector.place(x=100, y=16, width=210, height=33)

        self.confirm_button = ttk.Button(master=self.window, text='Add Dataset', command=self.add, state=tk.DISABLED)
        self.confirm_button.place(x=20, y=150, width=120)
        self.cancel_button = ttk.Button(master=self.window, text='Cancel', command=self.handle_close)
        self.cancel_button.place(x=160, y=150, width=120)

    def enable_confirm(self, null):
        self.confirm_button['state'] = tk.NORMAL

    def add(self):
        index = 0
        for idx in range(len(root.all_datasets)):
            if root.all_datasets[idx].ID == self.parent_ID:
                index = idx
        data = self.get_data()
        if data is None:
            return
        if 'Contrasts' in self.data_selection.get():
           x_ax, fits, fit_vals = self.get_fits(mass=False)
        elif 'Masses' in self.data_selection.get():
            x_ax, fits, fit_vals = self.get_fits(mass=True)
        else:
            x_ax, fits, fit_vals = None, [], None
        root.all_datasets[index].subsets.append(DataSubset())
        root.all_datasets[index].subsets[-1].ID = f'D{index}_{len(root.all_datasets[index].subsets)}_{self.data_selection.get()}'
        root.all_datasets[index].subsets[-1].parent = self.parent_ID
        if 'Calib' in self.data_selection.get():
            meta = data.pop(-1)
            root.all_datasets[index].subsets[-1].metadata = meta
        root.all_datasets[index].subsets[-1].data = data
        root.all_datasets[index].subsets[-1].x_axis = x_ax
        root.all_datasets[index].subsets[-1].fits = fits
        root.all_datasets[index].subsets[-1].fit_vals = fit_vals
        root.all_datasets[index].subsets[-1].calib = [root.calibration[0], root.calibration[1]]
        root.all_datasets[index].subsets[-1].name = f'{len(root.all_datasets[index].subsets)}_{self.data_selection.get()}'
        root.all_datasets[index].subsets[-1].type = f'{self.data_selection.get()}'
        if not 'All Tracks' in self.data_selection.get():
            if not 'Time' in self.data_selection.get():
                root.all_datasets[index].subsets[-1].info = f'{len(np.ravel(data))} data points. {round(sys.getsizeof(data)/1000, 1)} KB'
            else:
                if 'Mass' in self.data_selection.get():
                    root.all_datasets[index].subsets[-1].info = f'{len(np.ravel(data[0]))} data points. {round(sys.getsizeof(data) / 1000, 1)} KB'
                else:
                    root.all_datasets[index].subsets[-1].info = f'{len(data[0])} data points. {round(sys.getsizeof(data) / 1000, 1)} KB'
        elif 'All Tracks' in self.data_selection.get():
            root.all_datasets[index].subsets[-1].info = f'{len(data)} tracks. {round(sys.getsizeof(data) / 1000, 1)} KB'
        root.treeview.insert(parent=self.parent_ID, index="end", iid=root.all_datasets[index].subsets[-1].ID, text=root.all_datasets[index].subsets[-1].name,
                             values=(self.data_selection.get(), root.all_datasets[index].subsets[-1].info))
        root.treeview.see(root.all_datasets[index].subsets[-1].ID)
        self.handle_close()

    def get_data(self):
        data_dict = {
            0: root.contrasts,
            1: root.masses,
            2: root.track_contrasts,
            3: root.track_masses,
            4: 'trackcoords',
            5: 'trackconts',
            6: 'timemass',
            7: 'ratioframe',
            8: 'diff',
            9: 'calib',
            10: 'DvM',
            11: 'alltracks',
        }

        if data_dict[self.field_list.index(self.data_selection.get())] == 'diff':
            data = self.get_diffusivity()
        elif data_dict[self.field_list.index(self.data_selection.get())] == 'calib':
            data = self.get_calibration()
        elif data_dict[self.field_list.index(self.data_selection.get())] == 'timemass':
            data = self.get_timeseries_mass()
        elif data_dict[self.field_list.index(self.data_selection.get())] == 'ratioframe':
            data = root.ratiometric_stack[:, :, int(float(root.ratio_frame_var.get()))]
        elif data_dict[self.field_list.index(self.data_selection.get())] == 'DvM':
            data = self.get_diff_vs_mass()
        elif data_dict[self.field_list.index(self.data_selection.get())] == 'trackcoords':
            try:
                data = root.tracks[int(float(root.current_track.get())) - 1].coords
            except ValueError:
                data = None
        elif data_dict[self.field_list.index(self.data_selection.get())] == 'trackconts':
            try:
                data = [root.tracks[int(float(root.current_track.get())) - 1].contrasts, root.tracks[int(float(root.current_track.get())) - 1].plateaus, float(root.mass_std_lim.get())]
            except ValueError:
                print('Error')
                data = None
        elif data_dict[self.field_list.index(self.data_selection.get())] == 'alltracks':
            try:
                data = self.get_all_tracks()
            except ValueError:
                data = None
        else:
            data = data_dict[self.field_list.index(self.data_selection.get())]
        return data

    @staticmethod
    def get_diffusivity():
        coefficients = []
        try:
            root.progress_win.window.destroy()
        except Exception:
            pass
        root.progress_win = ProgressWin(title='Working', msg='Calculating coefficients')
        root.progress_win.progress['maximum'] = len(root.tracks) + 1
        for idx in range(len(root.tracks)):
            root.progress_win.progress.step(1)
            root.progress_win.window.update()
            try:
                TAUS, MSD = root.get_msd(idx)
                length = len(MSD)
                if length < 10:
                    prop = int(length // 2) + 1
                else:
                    prop = int(np.sqrt(length)) + 1
                gradient, intercept = linear_regression(TAUS[:prop], MSD[:prop])
                if gradient < 0:
                    gradient = 0
                coefficients.append(gradient / 4)
            except Exception:
                print(f'Failed to calculate coefficient for track {idx + 1}')
        try:
            root.progress_win.handle_close()
        except Exception:
            ''' Failed '''
        final_coeff = []
        for coeff in coefficients:
            if not np.isnan(coeff):
                final_coeff.append(coeff)
        return final_coeff

    @staticmethod
    def get_diff_vs_mass():
        indices = root.get_filtered_indices()
        coefficients = []
        masses = []
        try:
            root.progress_win.window.destroy()
        except Exception:
            pass
        root.progress_win = ProgressWin(title='Working', msg='Calculating coefficients')
        root.progress_win.progress['maximum'] = len(root.tracks) + 1
        for idx in range(len(root.tracks)):
            root.progress_win.progress.step(1)
            root.progress_win.window.update()
            if idx in indices:
                try:
                    TAUS, MSD = root.get_msd(idx)
                    length = len(MSD)
                    if length < 10:
                        prop = int(length // 2) + 1
                    else:
                        prop = int(np.sqrt(length)) + 1
                    gradient, intercept = linear_regression(TAUS[:prop], MSD[:prop])
                    if gradient < 0:
                        gradient = 0
                    coefficients.append(gradient / 4)
                except Exception:
                    print(f'Failed to calculate coefficient for track {idx + 1}')
                mean, std = root.get_filtered_mean(root.tracks[idx].contrasts)
                masses.append(mean * root.calibration[0] + root.calibration[1])
        try:
            root.progress_win.handle_close()
        except Exception:
            ''' Failed '''
        final_coeff = []
        final_mass = []
        for idx in range(len(coefficients)):
            if not np.isnan(coefficients[idx]):
                final_coeff.append(coefficients[idx])
                final_mass.append(masses[idx])
        return [final_mass, final_coeff]

    @staticmethod
    def get_calibration():
        try:
            return [root.mass_calib_win.contrasts, root.mass_calib_win.masses, root.mass_calib_win.calibration_label['text']]
        except Exception:
            easygui.msgbox(title='Error!', msg='Could not find data points. Please open the mass calibration plot window')
            return None

    @staticmethod
    def get_timeseries_mass():
        data = root.tracks[int(float(root.current_track.get())) - 1].contrasts
        plats = root.tracks[int(float(root.current_track.get())) - 1].plateaus
        new = []
        for idx in range(len(data)):
            new.append(data[idx] * root.calibration[0] + root.calibration[1])
        return [new, plats, float(root.mass_std_lim.get())]

    def get_fits(self, mass=False):
        if root.fit_manually.get() == 0:
            x_axis, fits, moments = self.get_skewed_fits(mass=mass)
            return x_axis, fits, moments
        components = int(float(root.number_of_fits.get()))
        moments = []
        for idx in range(components):
            try:
                mean, std, amp = float(root.field_list[idx].split(",")[1][1:]), float(root.field_list[idx].split(",")[2][1:]), float(root.field_list[idx].split(",")[3][1:])
                if mass:
                    if mean > 0:
                        moments.append([mean * root.calibration[0] + root.calibration[1], std * root.calibration[0], (amp) / root.calibration[0]])
                    else:
                        moments.append([mean * root.calibration[0] - root.calibration[1], std * root.calibration[0], (amp) / root.calibration[0]])
                else:
                    moments.append([mean, std, amp])
            except Exception:
                'fit not available'
                print(traceback.format_exc())
        try:
            if mass:
                start = np.min(root.masses)
                end = np.max(root.masses)
            else:
                start = np.min(root.contrasts)
                end = np.max(root.contrasts)
            x_axis = np.linspace(start, end, 2000)
        except ValueError:
            print(traceback.format_exc())
        fits = []
        for fit_index in range(len(moments)):
            fit = root.gauss(x_axis, moments[fit_index][2], moments[fit_index][0], moments[fit_index][1])
            fits.append(fit)
        if len(fits) >= 1:
            total_fit = np.copy(fits[0])
            for idx in range(1, len(fits)):
                total_fit += fits[idx]
            fits.append(total_fit)
        return x_axis, fits, moments

    def get_skewed_fits(self, mass=False):
        moments = []
        for idx in range(len(root.auto_fits)):
            if not mass:
                moments.append([
                    [root.auto_fits[idx][0][0], root.auto_fits[idx][0][1], root.auto_fits[idx][0][2], root.auto_fits[idx][0][3]],
                    [root.auto_fits[idx][1][0], root.auto_fits[idx][1][1], root.auto_fits[idx][1][2], root.auto_fits[idx][0][3]]
                ])
            else:
                if root.auto_fits[idx][0][0] > 0:
                    moments.append([
                        [root.auto_fits[idx][0][0]*root.calibration[0]+root.calibration[1], root.auto_fits[idx][0][1]*root.calibration[0], root.auto_fits[idx][0][2] / root.calibration[0], root.auto_fits[idx][0][3]],
                        [root.auto_fits[idx][1][0]*root.calibration[0]+root.calibration[1], root.auto_fits[idx][1][1]*root.calibration[0], root.auto_fits[idx][1][2], root.auto_fits[idx][0][3]]
                    ])
                else:
                    moments.append([
                        [root.auto_fits[idx][0][0]*root.calibration[0]-root.calibration[1], root.auto_fits[idx][0][1]*root.calibration[0], root.auto_fits[idx][0][2] / root.calibration[0], root.auto_fits[idx][0][3]],
                        [root.auto_fits[idx][1][0]*root.calibration[0]-root.calibration[1], root.auto_fits[idx][1][1]*root.calibration[0], root.auto_fits[idx][1][2], root.auto_fits[idx][0][3]]
                    ])
        try:
            if mass:
                start = np.min(root.masses)
                end = np.max(root.masses)
            else:
                start = np.min(root.contrasts)
                end = np.max(root.contrasts)
            x_axis = np.linspace(start, end, 2000)
        except ValueError:
            print(traceback.format_exc())
        fits = []
        for fit_index in range(len(moments)):
            fit = sgmm.skew_gaussian(x_axis, moments[fit_index][0][2], moments[fit_index][0][0], moments[fit_index][0][1], moments[fit_index][0][3])
            fits.append(fit)
        if len(fits) >= 1:
            total_fit = np.copy(fits[0])
            for idx in range(1, len(fits)):
                total_fit += fits[idx]
            fits.append(total_fit)
        return x_axis, fits, moments

    def get_all_tracks(self):
        data = []
        for idx in root.get_filtered_indices():
            track = root.tracks[idx]
            data.append(track.coords)
        return data

    def handle_close(self):
        try:
            self.window.destroy()
        except Exception:
            ''' Failed '''


class ProgressWin:
    def __init__(self, title, msg, cancel_func=None):
        self.title = title
        self.msg = msg
        self.window = tk.Toplevel()
        self.window.title(self.title)
        self.window["bg"] = "#282828"
        self.window.protocol("WM_DELETE_WINDOW", self.handle_close)
        self.window.attributes('-topmost', True)
        self.window.geometry("300x120+800+400")
        self.window.resizable(False, False)
        self.window.overrideredirect(1)

        self.title_bar = tk.Label(master=self.window, bg="#007fff", fg="white", text=title, justify="left", anchor=tk.W)
        self.title_bar.place(x=0, y=0, width=300, height=24)
        self.progress = ttk.Progressbar(master=self.window, orient="horizontal", mode="determinate", length=280)
        self.label = tk.Label(master=self.window, text=msg, bg="#282828", fg="white", anchor=tk.NW, justify="left")
        self.cancel_button = ttk.Button(master=self.window, text="Cancel", command=root.cancel_task)
        if cancel_func:
            self.cancel_button["command"] = cancel_func
        self.progress.place(x=10, y=30)
        self.label.place(x=30, y=48)
        self.cancel_button.place(x=110, y=82)
        self.minimize_button = tk.Button(master=self.window, text="_", relief=tk.FLAT, bg="#444444", fg="white", padx=2,
                                         command=self.minimize)
        self.minimize_button.place(x=278, y=2, height=20)
        self.title_bar.bind("<B1-Motion>", self.move_window)
        self.title_bar.bind("<Button-1>", self.get_init_click)
        self.title_click = [None, None]

    def get_init_click(self, event):
        self.title_click = [event.x, event.y]

    def move_window(self, event):
        x = event.x + self.window.winfo_x()
        y = event.y + self.window.winfo_y()
        self.window.geometry(f"300x120+{x - self.title_click[0]}+{y - self.title_click[1]}")

    def minimize(self):
        try:
            root.restore_minimized_task_button.place(x=968, y=0, width=140, height=25)
        except Exception:
            """ Failed """
        self.window.state("withdrawn")

    def restore(self):
        self.window.state("normal")
        try:
            root.restore_minimized_task_button.place_forget()
        except Exception:
            """ Failed """

    def handle_close(self):
        try:
            self.progress.stop()
        except Exception:
            """ No progress bar """
        try:
            self.window.destroy()
        except Exception:
            """ Window was closed """
        try:
            root.restore_minimized_task_button.place_forget()
        except Exception:
            """ Failed """


class MassCalibWin:
    def __init__(self, contrasts, number):
        self.window = tk.Toplevel()
        self.window.title('Create Mass Calibration')
        self.window.protocol("WM_DELETE_WINDOW", self.handle_close)
        self.window.attributes('-topmost', True)
        self.window.attributes('-topmost', False)
        self.window.geometry('1090x420')
        self.window.resizable(False, False)
        self.window.iconbitmap("icons/luxmp_logo.ico")
        self.window['bg'] = '#333333'

        self.parameter_frame = ttk.LabelFrame(master=self.window, text=" Calibrants ", padding=(6, 5))
        self.parameter_frame.place(x=8, y=4, width=450, height=404)

        self.frame = tk.Frame(master=self.window)
        self.frame.place(x=470, y=13)
        self.figure = plt.Figure(figsize=(6.1, 3.95), dpi=100)
        self.figure.set_facecolor("#292929")
        self.figure.subplots_adjust(top=0.92, bottom=0.12, left=0.12, right=0.96)

        self.canvas = FigureCanvasTkAgg(self.figure, master=self.frame)
        self.canvas.draw()
        self.canvas.get_tk_widget().pack(side="top", fill="both", expand=1)

        self.contrasts = []
        self.masses = []

        self.r2 = None
        self.mass_err = None

        self.get_series(contrasts, number)
        self.options = list(self.masses)
        self.widgets = []
        for idx in range(number):
            self.widgets.append(
                [
                    ttk.Entry(master=self.parameter_frame),
                    ttk.Spinbox(master=self.parameter_frame, from_=40, to=5000, increment=1, command=self.update_mass_list),
                    ttk.Combobox(master=self.parameter_frame, values=list(calibrants.keys())),
                ]
            )
            self.widgets[-1][0].insert(0, self.contrasts[idx])
            self.widgets[-1][0]['state'] = tk.DISABLED
            self.widgets[-1][1].set(self.masses[idx])
            self.widgets[-1][1].bind('<<SpinboxSelected>>', self.update_mass_list)
            self.widgets[-1][1].bind('<Return>', self.update_mass_list)
            self.widgets[-1][2].current()
            self.widgets[-1][2].insert(0, list(calibrants.keys())[idx])
            self.widgets[-1][2]['state'] = "readonly"
            self.widgets[-1][2].bind("<FocusIn>", self.defocus)
            self.widgets[-1][2].bind('<<ComboboxSelected>>', self.update_calibs)

            self.widgets[-1][0].place(x=10, y=idx*36+30, width=100, height=31)
            self.widgets[-1][1].place(x=120, y=idx*36+30, width=120, height=31)
            self.widgets[-1][2].place(x=250, y=idx * 36 + 30, width=120, height=31)

        self.calibration_label = tk.Label(master=self.frame, text='', bg='#292929', fg='#cccccc')
        self.calibration_label.place(x=74, y=4)
        self.fit_params = []
        self.line = []
        self.comp_fit = []
        self.update_mass_list(None, None)

        tk.Label(master=self.parameter_frame, bg='#333333', fg='#cccccc', text='Contrast                 Mass (kDa)               Calibrant', anchor=tk.NW).place(x=30, y=2)
        self.replot = ttk.Button(master=self.parameter_frame, text='Refresh', command=lambda: self.update_mass_list(None, None))
        self.replot.place(x=5, y=340, width=80)

        self.use = ttk.Button(master=self.parameter_frame, text='Use Calibration and Return', command=self.save_and_exit)
        self.use.place(x=95, y=340, width=180)

        self.manage = ttk.Button(master=self.parameter_frame, text='Manage calibrants', command=self.manage_calibrants)
        self.manage.place(x=290, y=340, width=140)

        self.manage_pane = ttk.LabelFrame(master=self.window, text=" Manage Calibrants ", padding=(6, 5))
        ttk.Label(master=self.manage_pane, text='Calibrant Name               Mass (kDa)').place(x=32, y=60)
        self.calib_name = ttk.Entry(master=self.manage_pane)
        self.calib_name.place(x=10, y=95, width=140)
        self.calib_mass = ttk.Spinbox(master=self.manage_pane, from_=40, to=5000, increment=1)
        self.calib_mass.place(x=170, y=95, width=120)
        self.calib_mass.set(100)
        self.calib_mass['state'] = "readonly"
        self.add_button = ttk.Button(master=self.manage_pane, text='Add Calibrant', command=self.add_calibrant)
        self.add_button.place(x=300, y=95, width=120)

        self.restore_calibrants_button = ttk.Button(master=self.manage_pane, text='Restore Default Calibrants', command=self.restore_calibrants)
        self.restore_calibrants_button.place(x=100, y=330, width=180)

        self.done_button = ttk.Button(master=self.manage_pane, text='Done', command=self.done_manage)
        self.done_button.place(x=300, y=330, width=100)

        self.plot()

    def add_calibrant(self):
        try:
            int(float(self.calib_mass.get()))
        except Exception:
            easygui.msgbox(title='Error!', msg='Invalid calibrant mass.')
            return
        if len(self.calib_name.get()) == 0 or str(self.calib_name.get()).isspace():
            easygui.msgbox(title='Error!', msg='Invalid calibrant name.')
            return
        calibrants[self.calib_name.get()] = int(float(self.calib_mass.get()))
        with open(os.path.join(os.getcwd(), "calibrants.dat"), "wb") as file:
            pickle.dump(calibrants, file)
        easygui.msgbox(title='Done', msg='Calibrant has been added.')

    def restore_calibrants(self):
        global calibrants
        calibrants = {
            "NM1": 66,
            "NM2": 146,
            "NM3": 480,
            "NM4": 1048,
            "DYN1": 96,
            "DYN2": 192,
            "DYN4": 384,
            "DYN6": 576,
            "DYN8": 768,
        }
        with open(os.path.join(os.getcwd(), "calibrants.dat"), "wb") as file:
            pickle.dump(calibrants, file)
        root.create_calibration()

    def manage_calibrants(self):
        self.manage_pane.place(x=8, y=4, width=450, height=404)

    def done_manage(self):
        root.create_calibration()

    def defocus(self, event):
        event.widget.selection_clear()

    def save_and_exit(self):
        warning = ""
        if self.r2 < 0.9:
            warning += "\nr² is less than 0.9."
        if self.mass_err > 10:
            warning += "\nMean mass error is greater then 10%."
        if warning != "":
            warning = "Warning! This mass calibration may be unreliable for the following reason(s):\n" + warning + "\n\nAre you sure you wish to keep it?"
            choice = easygui.indexbox(title='Warning!', msg=warning, choices=['Back', 'Ignore'], default_choice=0, cancel_choice=0)
            if choice == 0:
                return
        try:
            root.calibration[0] = self.fit_params[0]
            root.calibration[1] = self.fit_params[1]
            root.mass_entry['state'] = tk.NORMAL
            root.mass_entry.delete(0, tk.END)
            root.mass_entry.insert(0, f'G: {round(self.fit_params[0], 1)}, I: {round(self.fit_params[1],1)}')
            root.mass_entry['state'] = tk.DISABLED
            root.plot_histogram(None)
            root.update_trace_labels()
            root.display_frame_ratio(index=int(float(root.ratio_frame_var.get())))
        except Exception:
            traceback.format_exc()
        self.handle_close()

    def update_calibs(self, *args):
        for idx in range(len(self.widgets)):
            if not self.widgets[idx][2].get() == 'Other':
                self.widgets[idx][1].set(calibrants[self.widgets[idx][2].get()])
        self.update_mass_list()

    def update_mass_list(self, *args):
        calibs = [calibrants[k] for k in calibrants]
        self.masses = []
        for idx in range(len(self.widgets)):
            if not self.widgets[idx][1].get() == '':
                self.masses.append(float(self.widgets[idx][1].get()))
            if int(float(self.widgets[idx][1].get())) not in calibs:
                self.widgets[idx][2]['state'] = tk.NORMAL
                self.widgets[idx][2].delete(0, tk.END)
                self.widgets[idx][2].insert(0, 'Other')
                self.widgets[idx][2]['state'] = "readonly"

        print(self.masses)
        self.fit()
        self.plot()
        return True

    def get_series(self, data_list, number):
        masses = [calibrants[k] for k in calibrants]
        if number > 8:
            number = 8
        for idx in range(number):
            try:
                self.contrasts.append(data_list[idx])
            except Exception:
                traceback.format_exc()
        self.masses = masses
        self.contrasts.sort()

    def handle_close(self):
        try:
            self.window.destroy()
        except Exception:
            traceback.format_exc()

    def fit(self):
        try:
            gradient, intercept = linear_regression(self.contrasts, self.masses)
            self.fit_params = [gradient, intercept]
            self.line = [intercept, self.contrasts[-1]*1.1*gradient+intercept]
            self.comp_fit = []
            for con in self.contrasts:
                self.comp_fit.append(con*gradient+intercept)
            r2 = r2_score(self.masses, self.comp_fit)
            errors = []
            for idx, c in enumerate(self.contrasts):
                percentage_error = abs(((c * gradient + intercept) - self.masses[idx]) / self.masses[idx])*100
                errors.append(percentage_error)
            print(errors)
            mean_percentage_error = np.mean(errors)
            self.calibration_label['text'] = f'Gradient = {round(gradient)}       Intercept = {round(intercept)}       r² = {round(r2, 5)}       Average mass error: {round(mean_percentage_error, 2)}%'
            self.calibration_label.update()
            self.r2 = r2
            self.mass_err = mean_percentage_error
        except Exception:
            traceback.format_exc()

    def plot(self):
        self.figure.clf()
        self.canvas.draw()

        bg_col = "#222222"
        fig_col = "#292929"
        highlight_col = "#007fff"
        text_col = "#cccccc"

        self.figure.set_facecolor(fig_col)
        plotter = self.figure.add_subplot(111)
        plotter.set_facecolor(bg_col)
        plotter.spines['bottom'].set_color(highlight_col)
        plotter.spines['top'].set_color(highlight_col)
        plotter.spines['left'].set_color(highlight_col)
        plotter.spines['right'].set_color(highlight_col)
        plotter.xaxis.label.set_color(text_col)
        plotter.yaxis.label.set_color(text_col)
        plotter.tick_params(axis='x', colors=text_col, labelsize=9)
        plotter.tick_params(axis='y', colors=text_col, labelsize=9)
        plotter.set_xlabel('Calibrant Ratiometric Contrast', color=text_col, size=10)
        plotter.set_ylabel("Calibrant Mass (kDa)", color=text_col, size=10)
        plotter.plot([0, self.contrasts[-1] * 1.1], self.line, color='white', linewidth=1, linestyle='--')
        plotter.plot(self.contrasts, self.masses, color='#007fff', linewidth=0, marker='o', markersize=8)
        plotter.set_xlim(xmin=0, xmax=self.contrasts[-1] * 1.1)

        self.canvas.draw()


class DriftCorrectionWin:
    def __init__(self):
        self.window = tk.Toplevel()
        self.window.title('Drift Correction by Asymmetric Kernel Convolution')
        self.window.protocol("WM_DELETE_WINDOW", self.handle_close)
        self.window.attributes('-topmost', True)
        self.window.attributes('-topmost', False)
        self.window.geometry('1280x720')
        self.window.resizable(False, False)
        self.window.iconbitmap("icons/luxmp_logo.ico")
        self.window['bg'] = '#333333'

        self.native_stack = root.native_stack
        self.resolution = root.resolution
        h, w = self.resolution
        if w/h > 620/272:
            self.vertical_adjust = True
        else:
            self.vertical_adjust = False
        if not self.vertical_adjust:
            self.scale_factor = 272 / h
            self.canvas_offset = int(310 - w * self.scale_factor / 2)
        else:
            self.scale_factor = 620 / w
            self.canvas_offset = int(136 - h * self.scale_factor / 2)
            print(self.canvas_offset)
        self.corrected_stack = None

        tk.Frame(master=self.window, bg='#777777').place(x=50, y=18, width=1220, height=1)
        tk.Label(master=self.window, bg='#333333', fg='#cccccc', text='Native').place(x=5, y=6, width=50)
        tk.Label(master=self.window, bg='#333333', fg='#cccccc', text='Drift Corrected').place(x=646, y=6, width=100)
        self.canvas_native = tk.Canvas(master=self.window, width=620, height=272, bg='#222222', highlightthickness=1, highlightbackground='#111111')
        self.canvas_native.place(x=10, y=30)
        self.canvas_corrected = tk.Canvas(master=self.window, width=620, height=272, bg='#222222', highlightthickness=1, highlightbackground='#111111')
        self.canvas_corrected.place(x=648, y=30)

        self.native_frame_var = tk.IntVar(master=self.window)
        self.native_frame_var.set(0)
        self.slider_native = ttk.Scale(master=self.window, length=512, from_=0, to=root.slider_native['to'], variable=self.native_frame_var, command=self.display_frame_native)
        self.slider_native.place(x=10, y=312, width=562)
        self.native_frame_label = tk.Label(master=self.window, bg='#333333', fg='#cccccc', text='', anchor=tk.NW)
        self.native_frame_label.place(x=578, y=310, width=50)

        self.corrected_frame_var = tk.IntVar(master=self.window)
        self.corrected_frame_var.set(0)
        self.slider_corrected = ttk.Scale(master=self.window, length=512, from_=0, to=0, variable=self.corrected_frame_var, command=self.display_frame_corrected)
        self.slider_corrected.place(x=648, y=312, width=562)
        self.corrected_frame_label = tk.Label(master=self.window, bg='#333333', fg='#cccccc', text='', anchor=tk.NW)
        self.corrected_frame_label.place(x=1218, y=310, width=50)

        self.display_frame_native(0)

        self.initial_frame = tk.Frame(master=self.window)
        self.initial_frame.place(x=10, y=340)
        self.initial_figure = plt.Figure(figsize=(4.5, 3.7), dpi=100)
        self.initial_figure.set_facecolor("#282828")
        self.initial_figure.subplots_adjust(top=0.95, bottom=0.12, left=0.11, right=0.95)

        self.initial_canvas = FigureCanvasTkAgg(self.initial_figure, master=self.initial_frame)
        self.initial_canvas.draw()
        self.initial_canvas.get_tk_widget().pack(side="top", fill="both", expand=1)
        self.initial_plotter = None

        self.final_frame = tk.Frame(master=self.window)
        self.final_frame.place(x=470, y=340)
        self.final_figure = plt.Figure(figsize=(4.5, 3.7), dpi=100)
        self.final_figure.set_facecolor("#282828")
        self.final_figure.subplots_adjust(top=0.98, bottom=0.12, left=0.11, right=0.96)

        self.final_canvas = FigureCanvasTkAgg(self.final_figure, master=self.final_frame)
        self.final_canvas.draw()
        self.final_canvas.get_tk_widget().pack(side="top", fill="both", expand=1)
        self.final_plotter = None

        self.initial = None
        self.final = None
        self.initial_spots = None
        self.final_spots = None
        self.x_vect, self.y_vect = None, None

        try:
            self.progress_win.window.destroy()
        except Exception:
            pass
        self.progress_win = None

        tk.Label(master=self.window, bg='#333333', fg='#cccccc', text='Threshold', anchor=tk.NW).place(x=940, y=350, width=220)
        self.threshold = ttk.Spinbox(master=self.window, from_=0.3, to=0.9, increment=0.01, command=self.plot_frames)
        self.threshold.set(0.65)
        self.threshold.place(x=1020, y=345, width=80)
        self.threshold.bind("<Return>", lambda a: self.plot_frames())

        self.scan_button = ttk.Button(master=self.window, text='Scan Frames', command=self.detect_drift)
        self.scan_button.place(x=1110, y=345, width=150)

        self.correct_button = ttk.Button(master=self.window, text='Correct Drift', command=self.correct_drift)
        self.correct_button.place(x=1110, y=385, width=150)

        self.use_button = ttk.Button(master=self.window, text='Apply Correction and Return', command=self.done)
        self.use_button.place(x=1000, y=660)

        self.window.update()
        self.plot_frames()

    def done(self):
        root.native_stack = self.corrected_stack
        self.handle_close()

    def plot_frames(self):
        self.initial_figure.clf()
        self.initial_canvas.draw()
        self.final_figure.clf()
        self.final_canvas.draw()

        bg_col = "#2f2f2f"
        box_col = "#007fff"
        text_col = "#cccccc"
        grid_col = "#3f3f3f"
        line_col = "lightgray"

        self.initial_figure.set_facecolor(bg_col)
        self.initial_plotter = self.initial_figure.add_subplot(111)
        self.initial_plotter.set_facecolor(bg_col)
        self.initial_plotter.spines['bottom'].set_color(box_col)
        self.initial_plotter.spines['top'].set_color(bg_col)
        self.initial_plotter.spines['left'].set_color(box_col)
        self.initial_plotter.spines['right'].set_color(bg_col)
        self.initial_plotter.xaxis.label.set_color(text_col)
        self.initial_plotter.yaxis.label.set_color(text_col)
        self.initial_plotter.tick_params(axis='x', colors=text_col, labelsize=8)
        self.initial_plotter.tick_params(axis='y', colors=text_col, labelsize=8)
        self.initial_plotter.grid(color=grid_col)
        self.initial_plotter.set_xlabel('x initial (px)', size=9, color=text_col)
        self.initial_plotter.set_ylabel('y initial (px)', size=9, color=text_col)

        self.final_figure.set_facecolor(bg_col)
        self.final_plotter = self.final_figure.add_subplot(111)
        self.final_plotter.set_facecolor(bg_col)
        self.final_plotter.spines['bottom'].set_color(box_col)
        self.final_plotter.spines['top'].set_color(bg_col)
        self.final_plotter.spines['left'].set_color(box_col)
        self.final_plotter.spines['right'].set_color(bg_col)
        self.final_plotter.xaxis.label.set_color(text_col)
        self.final_plotter.yaxis.label.set_color(text_col)
        self.final_plotter.tick_params(axis='x', colors=text_col, labelsize=8)
        self.final_plotter.tick_params(axis='y', colors=text_col, labelsize=8)
        self.final_plotter.grid(color=grid_col)
        self.final_plotter.set_xlabel('x final (px)', size=9, color=text_col)
        self.final_plotter.set_ylabel('y final (px)', size=9, color=text_col)

        shape = np.shape(self.native_stack)

        img = self.native_stack[:, :, 0]
        m = np.max(img)
        img = np.clip(img, m * float(self.threshold.get()), m)
        m = np.min(img)
        img = img - m
        img = iscam.low_pass(img, 1, filter=iscam.large_filter)
        m = np.max(img)
        self.initial = np.clip(img, 0, m)
        self.initial = self.initial / m * 255

        img2 = self.native_stack[:, :, shape[2]-1]
        m = np.max(img2)
        img2 = np.clip(img2, m * float(self.threshold.get()), m)
        m = np.min(img2)
        img2 = img2 - m
        img2 = iscam.low_pass(img2, 1, filter=iscam.large_filter)
        m = np.max(img2)
        self.final = np.clip(img2, 0, m)
        self.final = self.final / m * 255

        self.initial_plotter.imshow(self.initial, cmap='inferno')
        self.initial_canvas.draw()
        self.final_plotter.imshow(self.final, cmap='inferno')
        self.final_canvas.draw()

        self.initial_spots, self.final_spots = None, None

    def detect_drift(self):
        shape = np.shape(self.native_stack)
        self.initial_spots = iscam.find_events(view=0, img_array=self.initial, mode=True,
                                               threshold=1.2,
                                               averaging_dist=2,
                                               gauss_fit_residual_threshold=8,
                                               min_sigma_threshold=0.6,
                                               max_sigma_threshold=2.5,
                                               min_intensity=1,
                                               eccentricity_threshold=0.7,
                                               true_gauss_threshold=4,
                                               region=[[0, 0], [self.resolution[1], self.resolution[0]]],
                                               inverted=False
                                               )
        self.final_spots = iscam.find_events(view=0, img_array=self.final, mode=True,
                                             threshold=1.2,
                                             averaging_dist=2,
                                             gauss_fit_residual_threshold=8,
                                             min_sigma_threshold=0.6,
                                             max_sigma_threshold=2.5,
                                             min_intensity=1,
                                             eccentricity_threshold=0.7,
                                             true_gauss_threshold=4,
                                             region=[[0, 0], [self.resolution[1], self.resolution[0]]],
                                             inverted=False
                                             )
        for i in range(len(self.initial_spots)):
            coords = self.initial_spots
            c = plt.Circle((coords[i][0], coords[i][1]), 4, fill=False, color="white")
            self.initial_plotter.text(coords[i][0] + 4, coords[i][1] - 4, str(i + 1), size=8, color="white")
            self.initial_plotter.add_artist(c)
            self.initial_canvas.draw()

        for i in range(len(self.final_spots)):
            coords = self.final_spots
            c = plt.Circle((coords[i][0], coords[i][1]), 4, fill=False, color="white")
            self.final_plotter.text(coords[i][0] + 4, coords[i][1] - 4, str(i + 1), size=8, color="white")
            self.final_plotter.add_artist(c)
            self.final_canvas.draw()

        pairs = []
        for idx in range(len(self.initial_spots)):
            coords = self.initial_spots[idx][0], self.initial_spots[idx][1]
            distances = []
            indices = []
            for idx2 in range(len(self.final_spots)):
                coords2 = self.final_spots[idx2][0], self.final_spots[idx2][1]
                dist = np.sqrt((coords2[0] - coords[0])**2 + (coords2[1] - coords[1])**2)
                distances.append(dist)
                indices.append(idx2)
            min_dist = min(distances)
            if min_dist < 2.5:
                min_dist_index = distances.index(min_dist)
                coords2 = self.final_spots[indices[min_dist_index]][0], self.final_spots[indices[min_dist_index]][1]
                pairs.append([coords, coords2])

        x_dists, y_dists = [], []
        for idx in range(len(pairs)):
            print(pairs[idx])
            x_dists.append(pairs[idx][1][0] - pairs[idx][0][0])
            y_dists.append(pairs[idx][1][1] - pairs[idx][0][1])

        self.x_vect = np.mean(x_dists) / shape[2]
        self.y_vect = np.mean(y_dists) / shape[2]

        print(self.x_vect)
        print(self.y_vect)

    def subsample_2x2_average(self, array):
        # Reshape the array to group 2x2 blocks
        reshaped = array.reshape(array.shape[0] // 2, 2, array.shape[1] // 2, 2)
        # Average the blocks
        subsampled = reshaped.mean(axis=(1, 3))
        return subsampled

    def correct_drift(self):
        shape = np.shape(self.native_stack)
        try:
            self.progress_win.window.destroy()
        except Exception:
            pass
        self.progress_win = ProgressWin(title='Correcting Drift', msg='Supersampling x4...')
        self.progress_win.progress['maximum'] = shape[2] + 11

        supersampled = np.zeros((shape[0]*4, shape[1]*4, shape[2]), dtype=np.uint16)
        for idx in range(shape[2]):
            supersampled[:, :, idx] = self.native_stack[:, :, idx].repeat(4, axis=0).repeat(4, axis=1)
            if idx % 10 == 0:
                self.progress_win.progress.step(10)
                self.progress_win.progress.update()

        self.progress_win.handle_close()

        try:
            self.progress_win.window.destroy()
        except Exception:
            pass
        self.progress_win = ProgressWin(title='Correcting Drift', msg='Applying asymmetric convolution...')
        self.progress_win.progress['maximum'] = shape[2] + 11
        self.progress_win.window.update()

        x_off, y_off = self.x_vect*4, self.y_vect*4
        fig = plt.figure(figsize=(5.12, 5.12))
        ax = fig.add_subplot(111)
        fig.subplots_adjust(left=0.0, right=1.0, top=1.0, bottom=0.0)
        for idx in range(shape[2]):
            kernel = np.clip(iscam.gauss(15, 15, 7 + x_off, 7 + y_off, 0.6, 0.6, 1, mode='protr')[:, :, 0], 0.01, 0.2)-0.01
            if idx % 10 == 0:
                self.progress_win.progress.step(10)
                self.progress_win.progress.update()
            x_off = x_off + self.x_vect*4
            y_off = y_off + self.y_vect*4
            kernel = kernel / np.sum(kernel)
            if idx % 100 == 0:
                try:
                    fig.clear()
                    ax = fig.add_subplot(111)
                    ax.imshow(kernel, cmap=mplib.cm.inferno)
                    coords = [7 + x_off, 7+y_off]
                    c2 = plt.Circle((7, 7), 2, fill=False, color="gray", linestyle='--')
                    c = plt.Circle((coords[0], coords[1]), 2, fill=False, color="white")
                    ax.add_artist(c2)
                    ax.add_artist(c)
                    ax.arrow(7, 7, x_off, y_off, head_width=0.18, head_length=0.25, fc='blue', ec='blue', linewidth=2, length_includes_head=True, overhang=0.08)
                    ax.text(7+x_off+2, 7+y_off-2, f'{round(float(root.pixel_size.get())*np.sqrt(x_off**2+y_off**2)/4, 1)} nm', size=16, color='blue')
                    fig.canvas.draw()
                    plt.show(block=False)
                except Exception:
                    '''Failed'''
            # print(f"type of image: {supersampled[:, :, idx].dtype}")
            # print(f"type of kernel: {kernel.dtype}")
            supersampled[:, :, idx] = iscam.cv2.filter2D(supersampled[:, :, idx], -1, kernel, borderType=2)
        self.progress_win.handle_close()

        self.progress_win = ProgressWin(title='Correcting Drift', msg='Downsampling...')
        self.progress_win.progress['maximum'] = shape[2] + 11
        self.progress_win.window.update()

        self.corrected_stack = np.copy(self.native_stack)
        for idx in range(shape[2]):
            frame = self.subsample_2x2_average(supersampled[:, :, idx])
            self.corrected_stack[:, :, idx] = self.subsample_2x2_average(frame)
            if idx % 10 == 0:
                self.progress_win.progress.step(10)
                self.progress_win.progress.update()

        self.progress_win.handle_close()
        self.slider_corrected['to'] = shape[2] - 1
        self.display_frame_corrected(0)

        print(np.max(self.native_stack), np.mean(self.native_stack))
        print(np.max(self.corrected_stack), np.mean(self.corrected_stack))

    def display_frame_native(self, index):
        index = int(float(index))
        maximum = np.max(self.native_stack[:, :, index])
        minimum = np.min(self.native_stack[:, :, index])
        norm = mplib.colors.Normalize(vmin=minimum, vmax=maximum)
        frame_rgb = mplib.cm.gray(norm(self.native_stack[:, :, index]))
        frame_rgb = frame_rgb[:, :, :3] * 255
        PIL_image = Image.fromarray(frame_rgb.astype('uint8'), 'RGB')
        h, w = self.resolution
        if not self.vertical_adjust:
            sfx, sfy = int((272 / h) * w), 272
        else:
            sfx, sfy = 620, int((620 / w) * h)
        PIL_image = PIL_image.resize((sfx, sfy), Image.NEAREST)
        global raw_img2
        raw_img2 = ImageTk.PhotoImage(master=self.window, image=PIL_image)
        if not self.vertical_adjust:
            self.canvas_native.create_image(1 + self.canvas_offset, 1, anchor="nw", image=raw_img2)
        else:
            self.canvas_native.create_image(1, 1 + self.canvas_offset, anchor="nw", image=raw_img2)
        self.native_frame_label['text'] = str(index)

    def display_frame_corrected(self, index):
        index = int(float(index))
        maximum = np.max(self.corrected_stack[:, :, index])
        minimum = np.min(self.corrected_stack[:, :, index])
        norm = mplib.colors.Normalize(vmin=minimum, vmax=maximum)
        frame_rgb = mplib.cm.gray(norm(self.corrected_stack[:, :, index]))
        frame_rgb = frame_rgb[:, :, :3] * 255
        PIL_image = Image.fromarray(frame_rgb.astype('uint8'), 'RGB')
        h, w = self.resolution
        if not self.vertical_adjust:
            sfx, sfy = int((272 / h) * w), 272
        else:
            sfx, sfy = 620, int((620 / w) * h)
        PIL_image = PIL_image.resize((sfx, sfy), Image.NEAREST)
        global corr_img
        corr_img = ImageTk.PhotoImage(master=self.window, image=PIL_image)
        if not self.vertical_adjust:
            self.canvas_corrected.create_image(1 + self.canvas_offset, 1, anchor="nw", image=corr_img)
        else:
            self.canvas_corrected.create_image(1, 1 + self.canvas_offset, anchor="nw", image=corr_img)
        self.corrected_frame_label['text'] = str(index)

    def handle_close(self):
        del(self.native_stack)
        del(self.corrected_stack)
        try:
            self.progress_win.window.destroy()
        except Exception:
            ''' Failed '''
        try:
            self.window.destroy()
        except Exception:
            ''' Failed '''


class ProfileWin:
    def __init__(self, cwd):
        self.window = tk.Toplevel()
        self.window.title(f"Analysis Profiles")
        self.window.geometry('800x200+550+380')
        self.window.resizable(False, False)
        self.window.protocol("WM_DELETE_WINDOW", self.handle_close)
        self.window.iconbitmap("icons/luxmp_logo.ico")
        self.profiles_folder = cwd + r"\Profiles"

        self.profiles = self.get_profiles()
        print(self.profiles)
        self.current_profile = tk.StringVar(master=self.window)
        self.plist = [p[:-4] for p in self.profiles]
        self.profile_menu = ttk.Combobox(master=self.window, values=self.plist)
        self.profile_menu.insert(0, 'Select Analysis Profile')
        self.profile_menu['state'] = 'readonly'
        self.profile_menu.bind('<<ComboboxSelected>>', self.release_widget)
        self.profile_menu.bind("<FocusIn>", self.defocus)
        self.profile_menu.place(x=20, y=10, width=620, height=34)
        self.path_type = None  # is init None, can be 'selected' or 'browsed' determines whether loading uses full path or joins loaded paths with folder

        self.browse_btn = ttk.Button(master=self.window, text='Browse', command=self.browse)
        self.browse_btn.place(x=660, y=10, width=120, height=34)

        self.load_btn = ttk.Button(master=self.window, text='Load Profile', command=self.load_profile)
        self.load_btn.place(x=20, y=54, width=400, height=34)

        self.presets_btn = ttk.Button(master=self.window, text='Presets', command=self.presets)
        self.presets_btn.place(x=440, y=54, width=200, height=34)

        self.delete_btn = ttk.Button(master=self.window, text='Delete Profile', command=self.delete_profile)
        self.delete_btn.place(x=660, y=54, width=120, height=34)

        self.save_btn = ttk.Button(master=self.window, text='Save New Profile', command=self.save_current)
        self.save_btn.place(x=510, y=154, width=130, height=34)

        self.browse_btn_2 = ttk.Button(master=self.window, text='Browse', command=self.browse_save)
        self.browse_btn_2.place(x=660, y=154, width=120, height=34)

        self.save_path_type = 'selected'

        self.save_name = ttk.Entry(master=self.window)
        self.save_name.insert(0, 'New Profile')
        self.save_name.bind("<Key>", self.update_save_type)
        self.save_name.place(x=20, y=154, width=480, height=34)


    def save_current(self):
        profile_name = self.save_name.get()
        if self.save_path_type == 'selected':
            full_path = os.path.join(self.profiles_folder, profile_name + ".prf")
        elif self.save_path_type == 'browsed':
            full_path = profile_name
        self.handle_close()
        root.save_analysis_profile(full_path)

    def load_profile(self):
        profile_name = self.profile_menu.get()
        if self.path_type == None:
            easygui.msgbox(title='Error!', msg='Cannot load profile when none is selected.')
            return
        elif self.path_type == 'selected':
            full_path = os.path.join(self.profiles_folder, profile_name + ".prf")
        elif self.path_type == 'browsed':
            full_path = profile_name
        self.handle_close()
        root.load_analysis_profile(full_path)

    def presets(self):
        self.handle_close()
        root.select_experiment_type()

    def delete_profile(self):
        profile_name = self.profile_menu.get()
        if self.path_type == None:
            easygui.msgbox(title='Error!', msg='Cannot delete profile when none is selected.')
            return
        elif self.path_type == 'selected':
            full_path = os.path.join(self.profiles_folder, profile_name + ".prf")
        elif self.path_type == 'browsed':
            full_path = profile_name
        choice = easygui.indexbox(title='Warning!', msg=f"Confirm delete analysis profile '{profile_name}'? This action cannot be undone.",
                                  choices=['Cancel', 'Confirm Delete'], default_choice=0, cancel_choice=0)
        if not choice or choice is None or choice == 0:
            return
        elif choice == 1:
            try:
                os.remove(full_path)
                self.update_combobox()
            except Exception:
                easygui.msgbox(title='Error!', msg=f'Cannot delete file:\n\n{traceback.format_exc()}')

    def update_combobox(self):
        self.profile_menu.place_forget()
        self.profile_menu.destroy()
        self.profiles = self.get_profiles()
        print(self.profiles)
        self.current_profile = tk.StringVar(master=self.window)
        self.plist = [p[:-4] for p in self.profiles]
        self.profile_menu = ttk.Combobox(master=self.window, values=self.plist)
        self.profile_menu.insert(0, 'Select Analysis Profile')
        self.profile_menu['state'] = 'readonly'
        self.profile_menu.bind('<<ComboboxSelected>>', self.release_widget)
        self.profile_menu.place(x=20, y=10, width=620, height=34)
        self.path_type = None  # is init None, can be 'selected' or 'browsed' determines whether loading uses full path or joins loaded paths with folder

    def defocus(self, event):
        event.widget.selection_clear()

    def release_widget(self, *args):
        self.path_type = 'selected'

    def update_save_type(self, *args):
        self.save_path_type = 'selected'

    def browse(self):
        path = easygui.fileopenbox(title="Open OpenMASS analysis profile '.prf' file", default=self.profiles_folder+r'\*.prf')
        if path:
            self.profile_menu['state'] = tk.NORMAL
            self.path_type = 'browsed'
            self.profile_menu.delete(0, tk.END)
            self.profile_menu.insert(0, path)
            if path[-4:] != '.prf':
                easygui.msgbox(title='Warning!', msg='This is not a .prf file. Loading it may crash the software.')

    def browse_save(self):
        path = easygui.filesavebox(title="Save OpenMASS analysis profile '.prf' file", default=self.profiles_folder+r'\*.prf')
        if path:
            if path[-4:] != '.prf':
                path = path + '.prf'
            self.profile_menu['state'] = tk.NORMAL
            self.save_path_type = 'browsed'
            self.save_name.delete(0, tk.END)
            self.save_name.insert(0, path)

    def get_profiles(self):
        profiles = []
        for file in os.listdir(self.profiles_folder):
            if file.endswith('.prf'):
                profiles.append(os.path.split(file)[1])
        return profiles

    def handle_close(self):
        try:
            self.window.destroy()
        except Exception:
            print(traceback.format_exc())


class PreferencesWin:
    def __init__(self):
        self.window = tk.Toplevel()
        self.window.title(f"Preferences")
        self.window.geometry('660x400+680+250')
        self.window.protocol("WM_DELETE_WINDOW", self.handle_close)
        self.window.attributes('-topmost', True)
        self.window.resizable(False, False)
        self.window.iconbitmap("icons/luxmp_logo.ico")

        self.preferences_notebook = ttk.Notebook(master=self.window)
        self.preferences_notebook.pack(side='top', fill='both', expand=True, padx=10, pady=10)
        self.pref_tab1 = ttk.Frame(master=self.preferences_notebook)
        self.preferences_notebook.add(self.pref_tab1, text='Start-up Behaviour')
        self.pref_tab2 = ttk.Frame(master=self.preferences_notebook)
        self.preferences_notebook.add(self.pref_tab2, text='Histogram')
        self.pref_tab3 = ttk.Frame(master=self.preferences_notebook)
        self.preferences_notebook.add(self.pref_tab3, text='Event Refinement')
        self.pref_tab4 = ttk.Frame(master=self.preferences_notebook)
        self.preferences_notebook.add(self.pref_tab4, text='Warnings')
        self.pref_tab5 = ttk.Frame(master=self.preferences_notebook)
        self.preferences_notebook.add(self.pref_tab5, text='Advanced')

        self.ribbon = ttk.Frame(master=self.window)
        self.ribbon.pack(side='bottom', fill='x', expand=False, padx=0, pady=10)

        self.done_btn = ttk.Button(master=self.ribbon, text='Save and Close', command=self.done)
        self.done_btn.pack(side='left', fill='x', expand=True, padx=10)
        self.reboot_btn = ttk.Button(master=self.ribbon, text='Save and Restart OpenMASS', command=self.reboot)
        self.reboot_btn.pack(side='left', fill='x', expand=True, padx=10)
        self.restore_btn = ttk.Button(master=self.ribbon, text='Restore Defaults', command=self.restore)
        self.restore_btn.pack(side='left', fill='x', expand=True, padx=10)
        self.cancel_btn = ttk.Button(master=self.ribbon, text='Cancel', command=self.handle_close)
        self.cancel_btn.pack(side='left', fill='x', expand=True, padx=10)

        self.converter = field_pref_converter = {
            "Choose Experiment": "experiment",
            "Open Analysis Profiles": "profile",
            "Default to Landing Assay": "landing",
            "Default to Dynamic Tracking": "tracking",
            "Default to Bin Width": "width",
            "Default to Bin Count": "count",
            "Advanced Preview": 3,
            "Fit Multiple Peaks": 1,
            "Fit Single Peaks": 2,
        }

        # because key-value pairs are exactly unique this is valid:
        pref_field_converter = dict()
        for key in field_pref_converter:
            pref_field_converter[field_pref_converter[key]] = key

        ttk.Label(master=self.pref_tab1, text='Select the default start-up behaviour. This determines whether you will be prompted to choose an \n'
                                              'experiment or analysis profile or default to landing assays or tracking when OpenMASS starts.').place(x=10, y=10)
        ttk.Label(master=self.pref_tab1, text='Select behaviour on start-up').place(x=50, y=95)
        self.start_behaviour = tk.StringVar(master=self.window)
        field_list = list(field_pref_converter.keys())[:4]
        self.start_behaviour_menu = ttk.OptionMenu(self.pref_tab1, self.start_behaviour, pref_field_converter[preferences['start']], *field_list)
        self.start_behaviour_menu['menu'].configure(bg='#555555', relief=tk.RIDGE, fg='white')
        self.start_behaviour_menu.place(x=260, y=90, width=260)

        ttk.Label(master=self.pref_tab2, text='Default bin mode').place(x=10, y=25)
        self.hist_bin_mode = tk.StringVar(master=self.window)
        field_list = list(field_pref_converter.keys())[4:6]
        self.hist_bin_mode_menu = ttk.OptionMenu(self.pref_tab2, self.hist_bin_mode, pref_field_converter[preferences['hist']['bin type']], *field_list)
        self.hist_bin_mode_menu['menu'].configure(bg='#555555', relief=tk.RIDGE, fg='white')
        self.hist_bin_mode_menu.place(x=125, y=20, width=160)

        ttk.Label(master=self.pref_tab2, text='Default peak fitting mode').place(x=310, y=25)
        self.hist_fit_mode = tk.StringVar(master=self.window)
        field_list = list(field_pref_converter.keys())[6:9]
        self.hist_fit_mode_menu = ttk.OptionMenu(self.pref_tab2, self.hist_fit_mode, pref_field_converter[preferences['hist']['fit mode']], *field_list)
        self.hist_fit_mode_menu['menu'].configure(bg='#555555', relief=tk.RIDGE, fg='white')
        self.hist_fit_mode_menu.place(x=470, y=20, width=150)

        tk.Label(master=self.pref_tab2, bg='#333333', fg='#cccccc', text='Default mass bin width', anchor=tk.NW).place(x=10, y=85)
        self.default_mass_width = ttk.Spinbox(master=self.pref_tab2, from_=1, to=50, increment=1)
        self.default_mass_width.set(preferences['hist']['default mass'])
        self.default_mass_width.place(x=170, y=80, width=115)

        tk.Label(master=self.pref_tab2, bg='#333333', fg='#cccccc', text='Default contrast bin width', anchor=tk.NW).place(x=10, y=125)
        self.default_contrast_width = ttk.Spinbox(master=self.pref_tab2, from_=0.000025, to=0.00125, increment=0.000025)
        self.default_contrast_width.set(preferences['hist']['default contrast'])
        self.default_contrast_width.place(x=170, y=120, width=115)

        tk.Label(master=self.pref_tab2, bg='#333333', fg='#cccccc', text='Default number of bins', anchor=tk.NW).place(x=10, y=165)
        self.default_bin_count = ttk.Spinbox(master=self.pref_tab2, from_=40, to=1000, increment=10)
        self.default_bin_count.set(preferences['hist']['default count'])
        self.default_bin_count.place(x=170, y=160, width=115)

        tk.Label(master=self.pref_tab2, bg='#333333', fg='#cccccc', text='Update histogram every n frames', anchor=tk.NW).place(x=310, y=85)
        self.update_hist = ttk.Spinbox(master=self.pref_tab2, from_=100, to=5000, increment=100)
        self.update_hist.set(preferences['hist']['update'])
        self.update_hist.place(x=530, y=80, width=90)

        tk.Label(master=self.pref_tab2, bg='#333333', fg='#cccccc', text='Auto-range percentile low', anchor=tk.NW).place(x=310, y=125)
        self.hist_low = ttk.Spinbox(master=self.pref_tab2, from_=0, to=5, increment=0.01)
        self.hist_low.set(preferences['hist']['percentile low'])
        self.hist_low.place(x=530, y=120, width=90)

        tk.Label(master=self.pref_tab2, bg='#333333', fg='#cccccc', text='Auto-range percentile high', anchor=tk.NW).place(x=310, y=165)
        self.hist_high = ttk.Spinbox(master=self.pref_tab2, from_=95, to=100, increment=0.01)
        self.hist_high.set(preferences['hist']['percentile hi'])
        self.hist_high.place(x=530, y=160, width=90)

        self.auto_fit_hist = tk.BooleanVar(master=self.window)
        self.auto_fit_hist.set(preferences['hist']['auto fit'])
        self.auto_fit_hist_label = ttk.Label(master=self.pref_tab2, text='Auto fit histogram after detecting events')
        self.auto_fit_hist_label.target_var = self.auto_fit_hist
        self.auto_fit_hist_label.place(x=310, y=255)
        self.auto_fit_hist_label.bind('<Button-1>', self.label_binding)
        self.auto_fit_hist_cb = ttk.Checkbutton(master=self.pref_tab2, text='', variable=self.auto_fit_hist, onvalue=True, offvalue=False, style='Switch.TCheckbutton')
        self.auto_fit_hist_cb.place(x=560, y=252)

        self.auto_refine = tk.BooleanVar(master=self.window)
        self.auto_refine.set(preferences['event']['auto'])
        self.auto_refine_label = ttk.Label(master=self.pref_tab3, text='Automatically apply event refinement after detecting events')
        self.auto_refine_label.target_var = self.auto_refine
        self.auto_refine_label.place(x=20, y=25)
        self.auto_refine_label.bind('<Button-1>', self.label_binding)
        self.auto_refine_cb = ttk.Checkbutton(master=self.pref_tab3, text='', variable=self.auto_refine, onvalue=True, offvalue=False, style='Switch.TCheckbutton')
        self.auto_refine_cb.place(x=400, y=22)

        tk.Label(master=self.pref_tab3, bg='#333333', fg='#cccccc', text='Auto gradient r² threshold', anchor=tk.NW).place(x=20, y=65)
        self.auto_gradient_r2 = ttk.Spinbox(master=self.pref_tab3, from_=0.5, to=0.99, increment=0.01)
        self.auto_gradient_r2.set(preferences['event']['grad r2'])
        self.auto_gradient_r2.place(x=400, y=60, width=200)

        tk.Label(master=self.pref_tab3, bg='#333333', fg='#cccccc', text='Auto gradient diff. threshold', anchor=tk.NW).place(x=20, y=105)
        self.auto_gradient_diff = ttk.Spinbox(master=self.pref_tab3, from_=20, to=99, increment=1)
        self.auto_gradient_diff.set(preferences['event']['grad diff'])
        self.auto_gradient_diff.place(x=400, y=100, width=200)

        tk.Label(master=self.pref_tab3, bg='#333333', fg='#cccccc', text='Auto trace signal / noise threshold', anchor=tk.NW).place(x=20, y=145)
        self.auto_trace_snr = ttk.Spinbox(master=self.pref_tab3, from_=1, to=5, increment=0.1)
        self.auto_trace_snr.set(preferences['event']['snr'])
        self.auto_trace_snr.place(x=400, y=140, width=200)

        ttk.Label(master=self.pref_tab4, text='Decide whether warnings pertaining to the use of experimental features or from verbose solver errors\n'
                                              'should be displayed. If you are a new user, we recommend leaving these in their default state.').place(x=10, y=10)

        self.warn_drift = tk.BooleanVar(master=self.window)
        self.warn_drift.set(preferences['warn']['drift'])
        self.warn_drift_label = ttk.Label(master=self.pref_tab4, text='Show warning when attempting lateral drift correction')
        self.warn_drift_label.target_var = self.warn_drift
        self.warn_drift_label.place(x=20, y=85)
        self.warn_drift_label.bind('<Button-1>', self.label_binding)
        self.warn_drift_cb = ttk.Checkbutton(master=self.pref_tab4, text='', variable=self.warn_drift, onvalue=True, offvalue=False, style='Switch.TCheckbutton')
        self.warn_drift_cb.place(x=500, y=82)

        self.warn_motion = tk.BooleanVar(master=self.window)
        self.warn_motion.set(preferences['warn']['motion'])
        self.warn_motion_label = ttk.Label(master=self.pref_tab4, text='Show warning when enabling vibrational motion correction')
        self.warn_motion_label.target_var = self.warn_motion
        self.warn_motion_label.place(x=20, y=125)
        self.warn_motion_label.bind('<Button-1>', self.label_binding)
        self.warn_motion_cb = ttk.Checkbutton(master=self.pref_tab4, text='', variable=self.warn_motion, onvalue=True, offvalue=False, style='Switch.TCheckbutton')
        self.warn_motion_cb.place(x=500, y=122)

        self.warn_gauss = tk.BooleanVar(master=self.window)
        self.warn_gauss.set(preferences['warn']['gauss'])
        self.warn_gauss_label = ttk.Label(master=self.pref_tab4, text='Show warning when enabling manual Gaussian fitting')
        self.warn_gauss_label.target_var = self.warn_gauss
        self.warn_gauss_label.place(x=20, y=165)
        self.warn_gauss_label.bind('<Button-1>', self.label_binding)
        self.warn_gauss_cb = ttk.Checkbutton(master=self.pref_tab4, text='', variable=self.warn_gauss, onvalue=True, offvalue=False, style='Switch.TCheckbutton')
        self.warn_gauss_cb.place(x=500, y=162)

        self.warn_fiterr = tk.BooleanVar(master=self.window)
        self.warn_fiterr.set(preferences['warn']['fiterr'])
        self.warn_fiterr_label = ttk.Label(master=self.pref_tab4, text='Show python tracebacks for fit solver runtime errors')
        self.warn_fiterr_label.target_var = self.warn_fiterr
        self.warn_fiterr_label.place(x=20, y=205)
        self.warn_fiterr_label.bind('<Button-1>', self.label_binding)
        self.warn_fiterr_cb = ttk.Checkbutton(master=self.pref_tab4, text='', variable=self.warn_fiterr, onvalue=True, offvalue=False, style='Switch.TCheckbutton')
        self.warn_fiterr_cb.place(x=500, y=202)

        ttk.Label(master=self.pref_tab5, text='Notch filter and apodisation control. Warning! These settings can adversely affect analysis if \n'
                                              'used incorrectly. For expert users only.').place(x=10, y=10)

        self.apodise = tk.BooleanVar(master=self.window)
        self.apodise.set(preferences['motion']['apodise'])
        self.apodise_label = ttk.Label(master=self.pref_tab5, text='Use data-driven apodisation function to supress ringing')
        self.apodise_label.target_var = self.apodise
        self.apodise_label.place(x=20, y=85)
        self.apodise_label.bind('<Button-1>', self.label_binding)
        self.apodise_cb = ttk.Checkbutton(master=self.pref_tab5, text='', variable=self.apodise, onvalue=True, offvalue=False, style='Switch.TCheckbutton')
        self.apodise_cb.place(x=500, y=82)

        tk.Label(master=self.pref_tab5, bg='#333333', fg='#cccccc', text='Notch filter sharpness', anchor=tk.NW).place(x=20, y=165)
        self.notch_sharpness = ttk.Spinbox(master=self.pref_tab5, from_=3, to=50, increment=1)
        self.notch_sharpness.set(preferences['motion']['notch'])
        self.notch_sharpness.place(x=170, y=160, width=100)

    def handle_close(self):
        try:
            self.window.destroy()
        except Exception:
            print('Failed to close pref. window. window may already be closed')
            print(traceback.format_exc())

    @staticmethod
    def label_binding(event):
        event.widget.target_var.set(not event.widget.target_var.get())

    def done(self):
        validate = self.get_prefs_from_UI()
        if not validate:
            return
        self.save_prefs()
        self.handle_close()

    def reboot(self):
        validate = self.get_prefs_from_UI()
        if not validate:
            return
        self.save_prefs()
        self.handle_close()
        reset()

    def restore(self):
        choice = easygui.indexbox(title='Warning!', msg='Warning! Restoring defaults will overwrite any current changes to preferences. Are you sure?',
                                  choices=['                   Cancel                  ', 'Confirm Restore Defaults'], cancel_choice=0, default_choice=0)
        if choice is None or choice == 0:
            return
        global preferences
        preferences = init_preferences()
        try:
            self.save_prefs()
            self.handle_close()
            root.open_preferences()
        except Exception:
            easygui.msgbox(title='Error!', msg=traceback.format_exc())

    def get_prefs_from_UI(self):
        global preferences
        prefs = deepcopy(preferences)
        prefs['start'] = self.converter[self.start_behaviour.get()]
        prefs['hist']['bin type'] = self.converter[self.hist_bin_mode.get()]
        prefs['hist']['fit mode'] = self.converter[self.hist_fit_mode.get()]
        prefs['hist']['auto fit'] = self.auto_fit_hist.get()
        prefs['event']['auto'] = self.auto_refine.get()
        prefs['warn']['drift'] = self.warn_drift.get()
        prefs['warn']['motion'] = self.warn_motion.get()
        prefs['warn']['gauss'] = self.warn_gauss.get()
        prefs['warn']['fiterr'] = self.warn_fiterr.get()
        prefs['motion']['apodise'] = self.apodise.get()
        try:
            p = float(self.default_mass_width.get())
            if not 0.1 <= p <= 50:
                raise ValueError(f'Value {p} was not in expected range.')
            prefs['hist']['default mass'] = p
        except Exception:
            self.validation_error("Default mass bin width")
            return False
        try:
            p = float(self.default_contrast_width.get())
            if not 0.000001 <= p <= 0.005:
                raise ValueError(f'Value {p} was not in expected range.')
            prefs['hist']['default contrast'] = p
        except Exception:
            self.validation_error("Default contrast bin width")
            return False
        try:
            p = int(float(self.default_bin_count.get()))
            if not 10 <= p <= 2000:
                raise ValueError(f'Value {p} was not in expected range.')
            prefs['hist']['default count'] = p
        except Exception:
            self.validation_error("Default number of bins")
            return False
        try:
            p = int(float(self.update_hist.get()))
            if not 10 <= p <= 5000:
                raise ValueError(f'Value {p} was not in expected range.')
            prefs['hist']['update'] = p
        except Exception:
            self.validation_error("Update histogram every n frames")
            return False
        try:
            p = float(self.hist_low.get())
            if not 0 <= p <= 10:
                raise ValueError(f'Value {p} was not in expected range.')
            prefs['hist']['percentile low'] = p
        except Exception:
            self.validation_error("Auto-range percentile low")
            return False
        try:
            p = float(self.hist_high.get())
            if not 90 <= p <= 100:
                raise ValueError(f'Value {p} was not in expected range.')
            prefs['hist']['percentile hi'] = p
        except Exception:
            self.validation_error("Auto-range percentile high")
            return False
        try:
            p = float(self.auto_gradient_r2.get())
            if not 0.1 <= p <= 0.99:
                raise ValueError(f'Value {p} was not in expected range.')
            prefs['event']['grad r2'] = p
        except Exception:
            self.validation_error("Auto gradient r² threshold")
            return False
        try:
            p = float(self.auto_gradient_diff.get())
            if not 10 <= p <= 100:
                raise ValueError(f'Value {p} was not in expected range.')
            prefs['event']['grad diff'] = p
        except Exception:
            self.validation_error("Auto gradient diff. threshold")
            return False
        try:
            p = float(self.auto_trace_snr.get())
            if not 1 <= p <= 10:
                raise ValueError(f'Value {p} was not in expected range.')
            prefs['event']['snr'] = p
        except Exception:
            self.validation_error("Auto trace signal / noise threshold")
            return False
        try:
            p = float(self.notch_sharpness.get())
            if not 3 <= p <= 50:
                raise ValueError(f'Value {p} was not in expected range.')
            prefs['motion']['notch'] = p
        except Exception:
            self.validation_error("Notch filter sharpness")
            return False

        preferences = prefs
        return True

    @staticmethod
    def validation_error(name):
        print(traceback.format_exc())
        easygui.msgbox(title='Error!', msg=f"Cannot save updated preferences because Parameter '{name}' either cannot "
                                           f"be interpreted as a number or the value is out of bounds.\n\n"
                                           f"Please correct the parameter and try again.")

    @staticmethod
    def save_prefs():
        with open('configs/preferences.dat', 'wb') as f:
            pickle.dump(preferences, f)


class GaussOptWin:
    def __init__(self):
        self.window = tk.Toplevel()
        self.window.title(f"Skewed GMM Options")
        self.window.geometry('500x330+700+290')
        self.window.protocol("WM_DELETE_WINDOW", self.handle_close)
        self.window.attributes('-topmost', True)
        self.window.resizable(False, False)
        self.window.iconbitmap("icons/luxmp_logo.ico")

        self.params = self.get_params()

        tk.Label(master=self.window, bg='#333333', fg='#cccccc', text='Maximimum components to optimize', anchor=tk.NW).place(x=20, y=14, width=350)
        self.max_comp = ttk.Spinbox(master=self.window, from_=1, to=5, increment=1)
        self.max_comp.set(self.params["max fits"])
        self.max_comp.bind("<Return>", self.update_params)
        self.max_comp.place(x=360, y=10, width=120)

        tk.Label(master=self.window, bg='#333333', fg='#cccccc', text='Initial k-means random states', anchor=tk.NW).place(x=20, y=64, width=350)
        self.n_init = ttk.Spinbox(master=self.window, from_=1, to=20, increment=1)
        self.n_init.set(self.params["n_init"])
        self.n_init.bind("<Return>", self.update_params)
        self.n_init.place(x=360, y=60, width=120)

        tk.Label(master=self.window, bg='#333333', fg='#cccccc', text='Maximum function evaluations during fitting', anchor=tk.NW).place(x=20, y=114, width=350)
        self.max_iter = ttk.Spinbox(master=self.window, from_=1000, to=30000, increment=1000)
        self.max_iter.set(self.params["maxiter"])
        self.max_iter.bind("<Return>", self.update_params)
        self.max_iter.place(x=360, y=110, width=120)

        tk.Label(master=self.window, bg='#333333', fg='#cccccc', text='Component overfitting regularization penalty', anchor=tk.NW).place(x=20, y=164, width=350)
        self.penalty = ttk.Spinbox(master=self.window, from_=0, to=2000, increment=10)
        self.penalty.set(self.params["penalty"])
        self.penalty.bind("<Return>", self.update_params)
        self.penalty.place(x=360, y=160, width=120)

        tk.Label(master=self.window, bg='#333333', fg='#cccccc', text='Information criterion for parameter optimization', anchor=tk.NW).place(x=20, y=214, width=350)
        self.criterion = tk.StringVar(master=self.window)
        plist = ['bic', 'aic']
        self.criterion = ttk.Combobox(master=self.window, values=plist)
        self.criterion.insert(0, self.params["optimizer"])
        self.criterion['state'] = 'readonly'
        self.criterion.bind("<FocusIn>", self.defocus)
        self.criterion.place(x=360, y=210, width=120)

        self.close_button = ttk.Button(master=self.window, text='Done', command=lambda: self.update_params(close=True))
        self.close_button.place(x=20, y=280, width=230)

        self.apply_button = ttk.Button(master=self.window, text='Apply', command=self.update_params)
        self.apply_button.place(x=260, y=280, width=70)

        self.restore_button = ttk.Button(master=self.window, text='Restore Defaults', command=self.restore_defaults)
        self.restore_button.place(x=360, y=280, width=120)

    def defocus(self, event):
        event.widget.selection_clear()

    def get_params(self):
        print(gauss_settings)
        return gauss_settings

    def update_params(self, close=False):
        try:
            global gauss_settings
            gauss_settings['max fits'] = int(float(self.max_comp.get()))
            gauss_settings['n_init'] = int(float(self.n_init.get()))
            gauss_settings['maxiter'] = int(float(self.max_iter.get()))
            gauss_settings['penalty'] = int(float(self.penalty.get()))
            gauss_settings['optimizer'] = str(self.criterion.get())
            self.save_settings()
            if close:
                self.handle_close()
        except Exception:
            easygui.msgbox(title='Error!', msg='Error occurred while updating settings. Ensure parameters are valid.')

    @staticmethod
    def save_settings():
        with open(os.path.join(root.cwd, "configs/gauss_config.dat"), "wb") as file:
            pickle.dump(gauss_settings, file)

    def restore_defaults(self):
        global gauss_settings
        gauss_settings = init_gauss_settings()
        self.max_comp.set(gauss_settings['max fits'])
        self.n_init.set(gauss_settings['n_init'])
        self.max_iter.set(gauss_settings['maxiter'])
        self.penalty.set(gauss_settings['penalty'])
        self.criterion['state'] = tk.NORMAL
        self.criterion.delete(0, tk.END)
        self.criterion.insert(0, gauss_settings['optimizer'])
        self.criterion['state'] = 'readonly'
        self.save_settings()

    def handle_close(self):
        try:
            self.window.destroy()
        except Exception:
            """ Failed """


class GaussPreviewWin:
    def __init__(self, data, bins, amp_factor, params):
        self.window = tk.Toplevel()
        self.window.geometry('1100x500+50+380')
        self.window.title('Advanced Fitting Preview')
        self.window.attributes('-topmost', True)
        self.window.attributes('-topmost', False)
        self.window.resizable(False, False)
        self.window.protocol("WM_DELETE_WINDOW", self.handle_close)
        self.window.iconbitmap("icons/luxmp_logo.ico")

        self.data = data
        self.fits = []

        self.frame = tk.Frame(master=self.window)
        self.frame.place(x=5, y=5)
        self.figure = plt.Figure(figsize=(7, 4.95), dpi=100)
        self.figure.subplots_adjust(top=0.97, bottom=0.07, left=0.09, right=0.98)
        self.plotter = None

        self.canvas = FigureCanvasTkAgg(self.figure, master=self.frame)
        self.canvas.draw()
        self.canvas.get_tk_widget().pack(side="top", fill="both", expand=1)

        self.figure.set_facecolor("#333333")
        self.figure.subplots_adjust(top=0.98, bottom=0.09, left=0.1, right=0.98)

        ttk.Label(master=self.window, text='Bins').place(x=730, y=20)
        self.bin_width = ttk.Spinbox(master=self.window, from_=10, to=1000, increment=5, command=self.plot)
        self.bin_width.set(int(bins))
        self.bin_width.bind("<Return>", self.plot)
        self.bin_width.place(x=770, y=15, width=80)

        components, max_fits, max_iter, penalty, optimizer, n_init = params

        ttk.Label(master=self.window, text='Components to fit').place(x=870, y=20)
        self.components = ttk.Combobox(master=self.window, values=['auto', 1, 2, 3, 4, 5])
        self.components.delete(0, tk.END)
        self.components.insert(0, components)
        self.components['state'] = 'readonly'
        self.components.bind("<FocusIn>", self.defocus)
        self.components.place(x=995, y=15, width=80)

        ttk.Label(master=self.window, text='Number of fits to try (auto)').place(x=730, y=80)
        self.max_fits = ttk.Spinbox(master=self.window, from_=1, to=5, increment=1)
        self.max_fits.set(max_fits)
        self.max_fits.place(x=920, y=75, width=150)

        ttk.Label(master=self.window, text='Information criterion (auto)').place(x=730, y=120)
        self.optimizer = ttk.Combobox(master=self.window, values=['bic', 'aic'])
        self.optimizer.delete(0, tk.END)
        self.optimizer.insert(0, optimizer)
        self.optimizer['state'] = 'readonly'
        self.optimizer.bind("<FocusIn>", self.defocus)
        self.optimizer.place(x=920, y=115, width=150)

        ttk.Label(master=self.window, text='Overfitting penalty (auto)').place(x=730, y=160)
        self.penalty = ttk.Spinbox(master=self.window, from_=-500, to=2000, increment=10)
        self.penalty.set(penalty)
        self.penalty.place(x=920, y=155, width=150)

        ttk.Label(master=self.window, text='Curve fit maximum iterations').place(x=730, y=200)
        self.max_iter = ttk.Spinbox(master=self.window, from_=1000, to=30000, increment=1000)
        self.max_iter.set(max_iter)
        self.max_iter.place(x=920, y=195, width=150)

        ttk.Label(master=self.window, text='K-means initial states').place(x=730, y=240)
        self.n_init = ttk.Spinbox(master=self.window, from_=1, to=20, increment=1)
        self.n_init.set(n_init)
        self.n_init.place(x=920, y=235, width=150)

        self.use_symmetric_gaussian = tk.IntVar(master=self.window)
        self.use_symmetric_gaussian.set(0)
        self.use_symmetric_gaussian_cb = ttk.Checkbutton(master=self.window, text='Symmetric Gaussian', onvalue=1, offvalue=0, variable=self.use_symmetric_gaussian)
        self.use_symmetric_gaussian_cb.place(x=915, y=280)

        ttk.Label(master=self.window, text='Status').place(x=730, y=340)
        self.status = tk.StringVar(master=self.window)
        self.status_bar = ttk.Entry(master=self.window, textvariable=self.status)
        self.status_bar.place(x=780, y=335, width=290)
        self.status.set('Fits calculated successfully.')
        self.status_bar['state'] = tk.DISABLED

        self.refit_button = ttk.Button(master=self.window, text='Refit', command=self.get_fits_UI)
        self.refit_button.place(x=920, y=375, width=150)

        self.defaults_button = ttk.Button(master=self.window, text='Reset Defaults', command=self.restore_defaults)
        self.defaults_button.place(x=730, y=375, width=150)

        self.done_button = ttk.Button(master=self.window, text='Done', command=self.done)
        self.done_button.place(x=730, y=430, width=150)

        self.cancel_button = ttk.Button(master=self.window, text='Cancel', command=self.handle_close)
        self.cancel_button.place(x=920, y=430, width=150)

        self.UI_params = None
        self.amp_factor = amp_factor
        self.x_axis = []
        self.fits = []
        self.stats = []
        self.window.update()
        try:
            self.init()
        except Exception:
            try:
                root.progress_win.handle_close()
                self.plot()
            except Exception:
                'Failed'
            self.status_bar['state'] = tk.NORMAL
            self.status.set('Runtime error in solver! Please try again.')
            self.status_bar['state'] = tk.DISABLED
            easygui.msgbox(title='Error', msg=f'Failed to calculate initial fits:\n\n{traceback.format_exc()}')

    def done(self):
        root.check_fit_replace()
        for idx, fit in enumerate(self.stats):
            fit[0][2] = fit[0][2] * self.amp_factor
            root.auto_fits.append(fit)
        self.handle_close()

    def restore_defaults(self):
        self.max_fits.set(gauss_settings['max fits'])
        self.max_iter.set(gauss_settings['maxiter'])
        self.penalty.set(gauss_settings['penalty'])
        self.n_init.set(gauss_settings['n_init'])
        self.optimizer['state'] = tk.NORMAL
        self.optimizer.delete(0, tk.END)
        self.optimizer.insert(0, gauss_settings['optimizer'])
        self.optimizer['state'] = 'readonly'

    def get_fits_UI(self):
        self.status_bar['state'] = tk.NORMAL
        try:
            self.status.set('Fits calculated successfully.')
            self.get_fits()
        except Exception:
            print(traceback.format_exc())
            self.status.set('Runtime error in solver! Please try again.')
            try:
                root.progress_win.handle_close()
                self.plot()
            except Exception:
                'Failed'
        self.status_bar['state'] = tk.DISABLED

    def defocus(self, event):
        event.widget.selection_clear()

    def return_values(self, UI_mode=False):
        if self.components.get() == 'auto' or UI_mode:
            comps = 'auto'
        else:
            comps = int(float(self.components.get()))
        maxf = int(float(self.max_fits.get()))
        maxit = int(float(self.max_iter.get()))
        pen = int(float(self.penalty.get()))
        opt = self.optimizer.get()
        ninit = int(float(self.n_init.get()))

        return comps, maxf, maxit, pen, opt, ninit

    def init(self):
        self.xmin = np.min(self.data)
        self.xmax = np.max(self.data)
        self.canvas.draw()
        self.get_fits()

    def plot(self, *args):
        self.figure.clf()
        self.canvas.draw()

        bg_col = "#323238"
        fig_col = "#333333"
        highlight_col = "#007fff"
        text_col = "#cccccc"
        grid_col = "#444444"

        self.figure.set_facecolor(fig_col)
        self.plotter = self.figure.add_subplot(111)
        self.plotter.set_facecolor(bg_col)
        self.plotter.spines['bottom'].set_color(highlight_col)
        self.plotter.spines['top'].set_color(fig_col)
        self.plotter.spines['left'].set_color(highlight_col)
        self.plotter.spines['right'].set_color(fig_col)
        self.plotter.xaxis.label.set_color(text_col)
        self.plotter.yaxis.label.set_color(text_col)
        self.plotter.tick_params(axis='x', colors=text_col, labelsize=8)
        self.plotter.tick_params(axis='y', colors=text_col, labelsize=8)
        self.plotter.set_xlabel("Ratiometric Contrast", color=text_col, size=9)
        self.plotter.set_ylabel("Probaility Density", color=text_col, size=9)
        self.plotter.grid(color=grid_col)

        bins = int(float(self.bin_width.get()))
        self.plotter.hist(self.data, density=True, bins=bins, color='orange', ec='orange', alpha=0.4, histtype='stepfilled')
        self.plotter.hist(self.data, density=True, bins=bins, color='orange', ec='orange', alpha=0.9, histtype='step')
        if len(self.fits) > 0:
            full = np.copy(self.fits[0])
            for idx in range(1, len(self.fits)):
                full = full + self.fits[idx]
            for fit in self.fits:
                self.plotter.plot(self.x_axis, fit, color='orange', alpha=0.7, linewidth=2)
            self.plotter.plot(self.x_axis, full, color='orange', alpha=1, linewidth=2)
            for stats in self.stats:
                mean = stats[1][0]
                self.plotter.axvline(mean, color='white', alpha=0.5, linewidth=1, linestyle='--')
        self.canvas.draw()

    def fit_hist_data(self, data, comps, bins, penalty=100, callback=None, max_comp=4, max_iter=8000, n_init=5, optimizer='bic'):
        print(f"len(data) = {len(data)}")
        data = list(data)
        for idx in range(len(data)):
            data[idx] = data[idx] * 40_000  # rescale data to approximate mass - speeds up curve fit significantly
        root.progress_win
        if comps == 'auto':
            if len(data) < 50:
                max_comp = 1
            elif 50 <= len(data) < 100:
                max_comp = 2
            elif 100 <= len(data) < 200:
                max_comp = 3
            elif 200 <= len(data) < 400:
                max_comp = 4
            else:
                pass
        if self.use_symmetric_gaussian.get() == 0:
            sym = False
        else:
            sym = True
        params, cov, fit_func, rms_error = sgmm.fit_skewed_gaussian_mixture(data,
                                                                            num_components=comps,
                                                                            bins=bins,
                                                                            component_optimizer=optimizer,
                                                                            component_penalty=penalty,
                                                                            density=True,
                                                                            plot=False,
                                                                            callback=callback,
                                                                            max_components=max_comp,
                                                                            max_iter=max_iter,
                                                                            n_init=n_init,
                                                                            sym=sym,
                                                                            )
        num_components = len(params) // 4
        print(len(params))
        print(params)
        stats = []
        for i in range(num_components):
            amp = params[i * 4] * 40_000
            mean = params[i * 4 + 1] / 40_000
            stddev = abs(params[i * 4 + 2]) / 40_000
            skewness = params[i * 4 + 3]
            if self.use_symmetric_gaussian.get() == 1:
                skewness = 0
            true_mean, true_std, true_skew = sgmm.get_stat_moments(mean, stddev, skewness)
            if amp < 0:
                raise ValueError("Negative amplitude fitted, discarding fits.")
            # print(np.min(self.data), true_mean)
            # print(np.max(self.data), true_mean)
            if stddev > 0 and np.min(self.data) < true_mean < np.max(self.data):
                stats.append([[mean, stddev, amp, skewness], [true_mean, true_std, true_skew, rms_error]])
        return stats

    def get_fits(self):
        bins = int(float(self.bin_width.get()))
        self.x_axis = np.linspace(self.xmin, self.xmax, 5000)

        def progress_update():
            root.progress_win.progress.step(8)
            root.progress_win.window.update()

        try:
            root.progress_win.window.destroy()
        except Exception:
            pass
        root.progress_win = ProgressWin(title='Fitting', msg='Fitting skewed Gaussian mixture...')
        root.progress_win.progress['mode'] = 'indeterminate'
        comps, maxf, maxit, pen, opt, ninit = self.return_values()
        print(comps, maxf, maxit, pen, opt, ninit)
        self.stats = self.fit_hist_data(self.data, comps, bins,
                                        penalty=pen,
                                        max_comp=maxf,
                                        max_iter=maxit,
                                        n_init=ninit,
                                        optimizer=opt,
                                        callback=progress_update,
                                        )
        self.window.update()
        self.window.attributes('-topmost', True)
        self.window.attributes('-topmost', False)
        print(len(self.stats))
        print(self.stats)
        if comps != "auto":
            if len(self.stats) < comps:
                easygui.msgbox(title='Warning!', msg='Unable to fit all specified components to data since paramaters were out of range. Check if number of components selected fits the distribution.'
                                                     'If this error persists, try changing the bins since unoptimal bin count may affect solver accuracy.')
        if len(self.stats) == 0:
            self.status_bar['state'] = tk.NORMAL
            self.status.set('Runtime error in solver! Please try again.')
            self.status_bar['state'] = tk.DISABLED
        root.progress_win.handle_close()
        self.fits = []
        for i in range(len(self.stats)):
            self.fits.append(sgmm.skew_gaussian(self.x_axis, self.stats[i][0][2], self.stats[i][0][0], self.stats[i][0][1], self.stats[i][0][3]))
        self.plot()

    def handle_close(self, auto=False):
        self.UI_params = self.return_values(UI_mode=True)
        try:
            self.window.destroy()
        except Exception:
            """ Failed """
        if not auto:
            try:
                root.plot_histogram(None)
            except Exception:
                """ Failed """


class RatioPopoutWin:
    def __init__(self):
        self.res = root.resolution
        self.window = tk.Toplevel()
        self.window.geometry(f"{self.res[1]*5+2}x{self.res[0]*5+2}+540+300")
        self.window.resizable(False, False)
        self.window.title('Ratiometric Stack')
        self.window.attributes('-topmost', True)
        self.window.protocol("WM_DELETE_WINDOW", self.handle_close)
        self.window.iconbitmap("icons/luxmp_logo.ico")
        self.canvas_ratiometric = tk.Canvas(master=self.window, width=self.res[1]*5, height=self.res[0]*5, bg='#222222', highlightthickness=1, highlightbackground='#111111')
        self.canvas_ratiometric.place(x=0, y=0)
        self.canvas_ratiometric.bind("<Button-1>", root.event_clicked)

    def handle_close(self):
        root.is_popped_out = False
        root.display_frame_ratio(int(float(root.ratio_frame_var.get())))
        try:
            self.window.destroy()
        except Exception:
            """ Failed """


class hover(Hovertip):
    def __init__(self, widget, text, *, hover_delay=1000,
                 bg="#222230", fg="#55afff",
                 font=("Segoe UI", 9), wraplength=300):
        super().__init__(widget, text, hover_delay=hover_delay)
        self._style = dict(bg=bg, fg=fg, font=font,
                           padx=6, pady=4,
                           relief="solid", borderwidth=1,
                           wraplength=wraplength, anchor='nw', justify='left')

    def showcontents(self):
        tk.Label(self.tipwindow, text=self.text, **self._style).pack()

    def showtip(self, text=None):
        """Position tooltip near the mouse cursor with automatic
        flipping in X and Y to keep it on-screen."""

        if text:
            self.text = text

        if self.tipwindow:
            return

        widget = self.anchor_widget

        # Mouse pointer
        mouse_x = widget.winfo_pointerx()
        mouse_y = widget.winfo_pointery()

        # Screen size
        screen_w = widget.winfo_screenwidth()
        screen_h = widget.winfo_screenheight()

        # Tooltip size estimates
        tip_w = self._style.get("wraplength", 300)
        # Height estimate (approx; tooltip grows with text but this is safe)
        tip_h = 80

        pad = 10  # cursor offset

        # -------- HORIZONTAL PLACEMENT --------
        # Default: place to the right of the cursor (top-left corner anchored)
        x = mouse_x + pad
        anchor_right = False

        # If it overflows to the right, flip to the left
        if x + tip_w > screen_w:
            x = mouse_x - tip_w - pad
            anchor_right = True

        # -------- VERTICAL PLACEMENT --------
        # Default: below cursor
        y = mouse_y + pad
        anchor_bottom = False

        # If it overflows below, flip above
        if y + tip_h > screen_h:
            y = mouse_y - tip_h - pad
            anchor_bottom = True

        # Create tooltip window
        self.tipwindow = tw = tk.Toplevel(widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")

        self.showcontents()


class AboutWin:
    def __init__(self, third_party=False):
        self.window = tk.Toplevel()
        if not third_party:
            self.window.title('About OpenMASS')
        else:
            self.window.title('Third Party Licences')
        self.window.geometry('800x600+500+180')
        self.window.resizable(False, False)
        self.window.iconbitmap("icons/luxmp_logo.ico")
        self.text = self.get_text(third_party)
        self.frame = ttk.Frame(master=self.window)
        self.frame.pack(fill='both', expand=True, padx=8, pady=8)
        self.text_widget = tk.Text(master=self.frame, wrap='word', width=102, height=len(self.text.split('\n')), relief=tk.FLAT)
        self.text_widget.pack(side='left', fill='both', expand=True)
        self.scrollbar = ttk.Scrollbar(master=self.frame, orient='vertical', command=self.text_widget.yview)
        self.scrollbar.pack(side='right', fill='y')
        self.text_widget.config(yscrollcommand=self.scrollbar.set)
        self.text_widget.insert(tk.END, self.text)
        self.text_widget['state'] = tk.DISABLED

    def get_text(self, third_party):
        if not third_party:
            LICENCE_TEXT = f"""
            =======================================================================
            OpenMASS (version {VERSION[1:]})
            Copyright (c) 2024 - 2025: Maximilian F. K. Wills. All Rights Reserved.
    
            Our Research Website: https://www.spliceselect.org/
    
            -----------------------------------------------------------------------
    
            OpenMASS is made available under the following licence:
    
            Creative Commons Attribution-NonCommercial 4.0 International Public
            License
    
            By exercising the Licensed Rights (defined below), You accept and agree
            to be bound by the terms and conditions of this Creative Commons
            Attribution-NonCommercial 4.0 International Public License ("Public
            License"). To the extent this Public License may be interpreted as a
            contract, You are granted the Licensed Rights in consideration of Your
            acceptance of these terms and conditions, and the Licensor grants You
            such rights in consideration of benefits the Licensor receives from
            making the Licensed Material available under these terms and
            conditions.
    
    
            Section 1 -- Definitions.
    
              a. Adapted Material means material subject to Copyright and Similar
                 Rights that is derived from or based upon the Licensed Material
                 and in which the Licensed Material is translated, altered,
                 arranged, transformed, or otherwise modified in a manner requiring
                 permission under the Copyright and Similar Rights held by the
                 Licensor. For purposes of this Public License, where the Licensed
                 Material is a musical work, performance, or sound recording,
                 Adapted Material is always produced where the Licensed Material is
                 synched in timed relation with a moving image.
    
              b. Adapter's License means the license You apply to Your Copyright
                 and Similar Rights in Your contributions to Adapted Material in
                 accordance with the terms and conditions of this Public License.
    
              c. Copyright and Similar Rights means copyright and/or similar rights
                 closely related to copyright including, without limitation,
                 performance, broadcast, sound recording, and Sui Generis Database
                 Rights, without regard to how the rights are labeled or
                 categorized. For purposes of this Public License, the rights
                 specified in Section 2(b)(1)-(2) are not Copyright and Similar
                 Rights.
              d. Effective Technological Measures means those measures that, in the
                 absence of proper authority, may not be circumvented under laws
                 fulfilling obligations under Article 11 of the WIPO Copyright
                 Treaty adopted on December 20, 1996, and/or similar international
                 agreements.
    
              e. Exceptions and Limitations means fair use, fair dealing, and/or
                 any other exception or limitation to Copyright and Similar Rights
                 that applies to Your use of the Licensed Material.
    
              f. Licensed Material means the artistic or literary work, database,
                 or other material to which the Licensor applied this Public
                 License.
    
              g. Licensed Rights means the rights granted to You subject to the
                 terms and conditions of this Public License, which are limited to
                 all Copyright and Similar Rights that apply to Your use of the
                 Licensed Material and that the Licensor has authority to license.
    
              h. Licensor means the individual(s) or entity(ies) granting rights
                 under this Public License.
    
              i. NonCommercial means not primarily intended for or directed towards
                 commercial advantage or monetary compensation. For purposes of
                 this Public License, the exchange of the Licensed Material for
                 other material subject to Copyright and Similar Rights by digital
                 file-sharing or similar means is NonCommercial provided there is
                 no payment of monetary compensation in connection with the
                 exchange.
    
              j. Share means to provide material to the public by any means or
                 process that requires permission under the Licensed Rights, such
                 as reproduction, public display, public performance, distribution,
                 dissemination, communication, or importation, and to make material
                 available to the public including in ways that members of the
                 public may access the material from a place and at a time
                 individually chosen by them.
    
              k. Sui Generis Database Rights means rights other than copyright
                 resulting from Directive 96/9/EC of the European Parliament and of
                 the Council of 11 March 1996 on the legal protection of databases,
                 as amended and/or succeeded, as well as other essentially
                 equivalent rights anywhere in the world.
    
              l. You means the individual or entity exercising the Licensed Rights
                 under this Public License. Your has a corresponding meaning.
    
    
            Section 2 -- Scope.
    
              a. License grant.
    
                   1. Subject to the terms and conditions of this Public License,
                      the Licensor hereby grants You a worldwide, royalty-free,
                      non-sublicensable, non-exclusive, irrevocable license to
                      exercise the Licensed Rights in the Licensed Material to:
    
                        a. reproduce and Share the Licensed Material, in whole or
                           in part, for NonCommercial purposes only; and
    
                        b. produce, reproduce, and Share Adapted Material for
                           NonCommercial purposes only.
    
                   2. Exceptions and Limitations. For the avoidance of doubt, where
                      Exceptions and Limitations apply to Your use, this Public
                      License does not apply, and You do not need to comply with
                      its terms and conditions.
    
                   3. Term. The term of this Public License is specified in Section
                      6(a).
    
                   4. Media and formats; technical modifications allowed. The
                      Licensor authorizes You to exercise the Licensed Rights in
                      all media and formats whether now known or hereafter created,
                      and to make technical modifications necessary to do so. The
                      Licensor waives and/or agrees not to assert any right or
                      authority to forbid You from making technical modifications
                      necessary to exercise the Licensed Rights, including
                      technical modifications necessary to circumvent Effective
                      Technological Measures. For purposes of this Public License,
                      simply making modifications authorized by this Section 2(a)
                      (4) never produces Adapted Material.
    
                   5. Downstream recipients.
    
                        a. Offer from the Licensor -- Licensed Material. Every
                           recipient of the Licensed Material automatically
                           receives an offer from the Licensor to exercise the
                           Licensed Rights under the terms and conditions of this
                           Public License.
    
                        b. No downstream restrictions. You may not offer or impose
                           any additional or different terms or conditions on, or
                           apply any Effective Technological Measures to, the
                           Licensed Material if doing so restricts exercise of the
                           Licensed Rights by any recipient of the Licensed
                           Material.
    
                   6. No endorsement. Nothing in this Public License constitutes or
                      may be construed as permission to assert or imply that You
                      are, or that Your use of the Licensed Material is, connected
                      with, or sponsored, endorsed, or granted official status by,
                      the Licensor or others designated to receive attribution as
                      provided in Section 3(a)(1)(A)(i).
    
              b. Other rights.
    
                   1. Moral rights, such as the right of integrity, are not
                      licensed under this Public License, nor are publicity,
                      privacy, and/or other similar personality rights; however, to
                      the extent possible, the Licensor waives and/or agrees not to
                      assert any such rights held by the Licensor to the limited
                      extent necessary to allow You to exercise the Licensed
                      Rights, but not otherwise.
    
                   2. Patent and trademark rights are not licensed under this
                      Public License.
    
                   3. To the extent possible, the Licensor waives any right to
                      collect royalties from You for the exercise of the Licensed
                      Rights, whether directly or through a collecting society
                      under any voluntary or waivable statutory or compulsory
                      licensing scheme. In all other cases the Licensor expressly
                      reserves any right to collect such royalties, including when
                      the Licensed Material is used other than for NonCommercial
                      purposes.
    
    
            Section 3 -- License Conditions.
    
            Your exercise of the Licensed Rights is expressly made subject to the
            following conditions.
    
              a. Attribution.
    
                   1. If You Share the Licensed Material (including in modified
                      form), You must:
    
                        a. retain the following if it is supplied by the Licensor
                           with the Licensed Material:
    
                             i. identification of the creator(s) of the Licensed
                                Material and any others designated to receive
                                attribution, in any reasonable manner requested by
                                the Licensor (including by pseudonym if
                                designated);
    
                            ii. a copyright notice;
    
                           iii. a notice that refers to this Public License;
    
                            iv. a notice that refers to the disclaimer of
                                warranties;
    
                             v. a URI or hyperlink to the Licensed Material to the
                                extent reasonably practicable;
    
                        b. indicate if You modified the Licensed Material and
                           retain an indication of any previous modifications; and
    
                        c. indicate the Licensed Material is licensed under this
                           Public License, and include the text of, or the URI or
                           hyperlink to, this Public License.
    
                   2. You may satisfy the conditions in Section 3(a)(1) in any
                      reasonable manner based on the medium, means, and context in
                      which You Share the Licensed Material. For example, it may be
                      reasonable to satisfy the conditions by providing a URI or
                      hyperlink to a resource that includes the required
                      information.
    
                   3. If requested by the Licensor, You must remove any of the
                      information required by Section 3(a)(1)(A) to the extent
                      reasonably practicable.
    
                   4. If You Share Adapted Material You produce, the Adapter's
                      License You apply must not prevent recipients of the Adapted
                      Material from complying with this Public License.
    
    
            Section 4 -- Sui Generis Database Rights.
    
            Where the Licensed Rights include Sui Generis Database Rights that
            apply to Your use of the Licensed Material:
    
              a. for the avoidance of doubt, Section 2(a)(1) grants You the right
                 to extract, reuse, reproduce, and Share all or a substantial
                 portion of the contents of the database for NonCommercial purposes
                 only;
    
              b. if You include all or a substantial portion of the database
                 contents in a database in which You have Sui Generis Database
                 Rights, then the database in which You have Sui Generis Database
                 Rights (but not its individual contents) is Adapted Material; and
    
              c. You must comply with the conditions in Section 3(a) if You Share
                 all or a substantial portion of the contents of the database.
    
            For the avoidance of doubt, this Section 4 supplements and does not
            replace Your obligations under this Public License where the Licensed
            Rights include other Copyright and Similar Rights.
    
    
            Section 5 -- Disclaimer of Warranties and Limitation of Liability.
    
              a. UNLESS OTHERWISE SEPARATELY UNDERTAKEN BY THE LICENSOR, TO THE
                 EXTENT POSSIBLE, THE LICENSOR OFFERS THE LICENSED MATERIAL AS-IS
                 AND AS-AVAILABLE, AND MAKES NO REPRESENTATIONS OR WARRANTIES OF
                 ANY KIND CONCERNING THE LICENSED MATERIAL, WHETHER EXPRESS,
                 IMPLIED, STATUTORY, OR OTHER. THIS INCLUDES, WITHOUT LIMITATION,
                 WARRANTIES OF TITLE, MERCHANTABILITY, FITNESS FOR A PARTICULAR
                 PURPOSE, NON-INFRINGEMENT, ABSENCE OF LATENT OR OTHER DEFECTS,
                 ACCURACY, OR THE PRESENCE OR ABSENCE OF ERRORS, WHETHER OR NOT
                 KNOWN OR DISCOVERABLE. WHERE DISCLAIMERS OF WARRANTIES ARE NOT
                 ALLOWED IN FULL OR IN PART, THIS DISCLAIMER MAY NOT APPLY TO YOU.
    
              b. TO THE EXTENT POSSIBLE, IN NO EVENT WILL THE LICENSOR BE LIABLE
                 TO YOU ON ANY LEGAL THEORY (INCLUDING, WITHOUT LIMITATION,
                 NEGLIGENCE) OR OTHERWISE FOR ANY DIRECT, SPECIAL, INDIRECT,
                 INCIDENTAL, CONSEQUENTIAL, PUNITIVE, EXEMPLARY, OR OTHER LOSSES,
                 COSTS, EXPENSES, OR DAMAGES ARISING OUT OF THIS PUBLIC LICENSE OR
                 USE OF THE LICENSED MATERIAL, EVEN IF THE LICENSOR HAS BEEN
                 ADVISED OF THE POSSIBILITY OF SUCH LOSSES, COSTS, EXPENSES, OR
                 DAMAGES. WHERE A LIMITATION OF LIABILITY IS NOT ALLOWED IN FULL OR
                 IN PART, THIS LIMITATION MAY NOT APPLY TO YOU.
    
              c. The disclaimer of warranties and limitation of liability provided
                 above shall be interpreted in a manner that, to the extent
                 possible, most closely approximates an absolute disclaimer and
                 waiver of all liability.
    
    
            Section 6 -- Term and Termination.
    
              a. This Public License applies for the term of the Copyright and
                 Similar Rights licensed here. However, if You fail to comply with
                 this Public License, then Your rights under this Public License
                 terminate automatically.
    
              b. Where Your right to use the Licensed Material has terminated under
                 Section 6(a), it reinstates:
    
                   1. automatically as of the date the violation is cured, provided
                      it is cured within 30 days of Your discovery of the
                      violation; or
    
                   2. upon express reinstatement by the Licensor.
    
                 For the avoidance of doubt, this Section 6(b) does not affect any
                 right the Licensor may have to seek remedies for Your violations
                 of this Public License.
    
              c. For the avoidance of doubt, the Licensor may also offer the
                 Licensed Material under separate terms or conditions or stop
                 distributing the Licensed Material at any time; however, doing so
                 will not terminate this Public License.
    
              d. Sections 1, 5, 6, 7, and 8 survive termination of this Public
                 License.
    
    
            Section 7 -- Other Terms and Conditions.
    
              a. The Licensor shall not be bound by any additional or different
                 terms or conditions communicated by You unless expressly agreed.
    
              b. Any arrangements, understandings, or agreements regarding the
                 Licensed Material not stated herein are separate from and
                 independent of the terms and conditions of this Public License.
    
    
            Section 8 -- Interpretation.
    
              a. For the avoidance of doubt, this Public License does not, and
                 shall not be interpreted to, reduce, limit, restrict, or impose
                 conditions on any use of the Licensed Material that could lawfully
                 be made without permission under this Public License.
    
              b. To the extent possible, if any provision of this Public License is
                 deemed unenforceable, it shall be automatically reformed to the
                 minimum extent necessary to make it enforceable. If the provision
                 cannot be reformed, it shall be severed from this Public License
                 without affecting the enforceability of the remaining terms and
                 conditions.
    
              c. No term or condition of this Public License will be waived and no
                 failure to comply consented to unless expressly agreed to by the
                 Licensor.
    
              d. Nothing in this Public License constitutes or may be interpreted
                 as a limitation upon, or waiver of, any privileges and immunities
                 that apply to the Licensor or You, including from the legal
                 processes of any jurisdiction or authority.
    
            =================================================================================
            =================================================================================
    
    
            If this version of OpenMASS is distributed in binary form, the above
            licence applies only to the executable. All libraries in source or binary
            form are subject to their own respective licences.
            
            For more information, click 'About' --> 'Third Party Licences'
    
            
            """
        else:
            LICENCE_TEXT = licences.THIRD_PARTY_LICENCES
        return LICENCE_TEXT


def linear_regression(series_x, series_y):
    x_sqrd_sum = 0
    xy_sum = 0
    x_sum = 0
    y_sum = 0
    for ind in range(len(series_x)):
        x_sqrd_sum += series_x[ind] ** 2
        xy_sum += series_x[ind] * series_y[ind]
        x_sum += series_x[ind]
        y_sum += series_y[ind]

    gradient = (len(series_x) * xy_sum - x_sum * y_sum) / (len(series_x) * x_sqrd_sum - x_sum**2)
    intercept = (y_sum - gradient*x_sum) / len(series_x)

    return gradient, intercept

def reset():
    global root
    print('reset action intiated')
    try:
        if root.progress_win is not None:
            root.progress_win.handle_close()
        root.window.destroy()
        root = Root()
        root.window.mainloop()
    except Exception:
        print(traceback.format_exc())


def load_calibrants():
    with open(os.path.join(os.getcwd(), "configs/calibrants.dat"), "rb") as file:
        calibs = pickle.load(file)
    return calibs


def load_preferences():
    with open(os.path.join(os.getcwd(), "configs/preferences.dat"), "rb") as file:
        preferences = pickle.load(file)
    return preferences


def load_gauss_settings():
    with open(os.path.join(os.getcwd(), "configs/gauss_config.dat"), "rb") as file:
        gauss_settings = pickle.load(file)
    return gauss_settings


def init_preferences():
    return {
        "start": "experiment",
        "hist": {
            "bin type": "width",
            "default mass": 10,
            "default contrast": 0.00025,
            "default count": 800,
            "update": 400,
            "fit mode": 2,
            "auto fit": True,
            "percentile low": 0.1,
            "percentile hi": 99.8,
        },
        "event": {
            "auto": False,
            "grad r2": 0.9,
            "grad diff": 40,
            "snr": 3,
        },
        "warn": {
            "drift": True,
            "motion": True,
            "gauss": True,
            "fiterr": False,
        },
        "motion": {
            "apodise": False,
            "notch": 40,
        },
    }


def init_gauss_settings():
    settings = {
        "penalty": 0,
        "optimizer": "bic",
        "maxiter": 10_000,
        "n_init": 10,
        "max fits": 4,
    }
    return settings


if __name__ != '__main__':
    loading_win.withdraw()


if __name__ == '__main__':
    _canvas = None
    loading_label['text'] = 'Building Interface...'
    loading_label.update()
    try:
        calibrants = load_calibrants()
    except Exception:
        easygui.msgbox(title='Warning!', msg='Unable to load calibrants; using internal defaults.')
        calibrants = {
            "NM1": 66,
            "NM2": 146,
            "NM3": 480,
            "NM4": 1048,
            "DYN1": 96,
            "DYN2": 192,
            "DYN4": 384,
            "DYN6": 576,
            "DYN8": 768,
        }
    try:
        preferences = load_preferences()
    except Exception:
        preferences = init_preferences()
        easygui.msgbox(title='Error!', msg=f'Unable to load application preferences:\n\n{traceback.format_exc()}\n\nUsing defaults.')
    try:
        gauss_settings = load_gauss_settings()
    except Exception:
        gauss_settings = init_gauss_settings()
        easygui.msgbox(title='Error!', msg=f'Unable to load settings for skewed Gaussian mixture model:\n\n{traceback.format_exc()}\n\nUsing defaults.')
    try:
        root = Root()
    except Exception:
        easygui.msgbox(title='Error!', msg=f"Failed to load interface:\n\n{str(traceback.format_exc())}")
    root.window.mainloop()

